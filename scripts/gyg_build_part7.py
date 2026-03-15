"""Part 7: Value sheet - DCF and SOTP valuation."""
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.load_workbook('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')
ws = wb['Value']

CL = get_column_letter

# ============================================================
# MARKET SNAPSHOT (R2-R9)
# ============================================================
ws['B2'] = 'DCF Valuation'
ws['B4'] = 'Current Share Price (AUD)'
ws['C4'] = 28.00

ws['B5'] = 'Shares Outstanding (#m)'
ws['C5'] = 101.0

ws['B6'] = 'Market Cap (AUDm)'
ws['C6'] = '=C4*C5'

ws['B7'] = 'Net Cash (AUDm)'
ws['C7'] = '=-INDEX(Annual!D:P,MATCH("BS-Cash",Annual!A:A,0),MATCH(2025,Annual!$1:$1,0))'  # Net cash from FY25

ws['B8'] = 'Market EV (AUDm)'
ws['C8'] = '=C6-C7'

ws['B9'] = 'Valuation Date'
ws['C9'] = datetime(2025, 12, 31)

# ============================================================
# WACC INPUTS (R11-R22)
# ============================================================
ws['B11'] = 'WACC Inputs'
ws['B12'] = 'Risk-free Rate'
ws['C12'] = 0.04
ws['B13'] = 'Equity Risk Premium'
ws['C13'] = 0.06
ws['B14'] = 'Beta'
ws['C14'] = 1.2
ws['B15'] = 'Cost of Equity'
ws['C15'] = '=C12+C13*C14'
ws['B16'] = 'Pre-tax Cost of Debt'
ws['C16'] = 0  # No debt
ws['B17'] = 'Tax Rate'
ws['C17'] = 0.30
ws['B18'] = 'After-tax Cost of Debt'
ws['C18'] = '=C16*(1-C17)'
ws['B19'] = 'Debt Weight (D/(D+E))'
ws['C19'] = 0  # No debt
ws['B20'] = 'WACC'
ws['C20'] = '=C15*(1-C19)+C18*C19'
ws['B21'] = 'Terminal Growth Rate'
ws['C21'] = 0.025
ws['B22'] = 'Stub Period'
ws['C22'] = '=(DATE(2026,6,30)-C9)/365.25'

# ============================================================
# FCF PROJECTION (R24-R39)
# ============================================================
ws['B24'] = 'FCF Projection'

# Period labels: FY26E through FY35E
for i, c in enumerate(range(4, 14)):  # cols D-M = 10 years
    yr = 2026 + i
    ws.cell(24, c).value = f'FY{str(yr)[2:]}E'

# R25: EBITDA = pull from Annual Stat-Statutory EBITDA
ws['B25'] = 'EBITDA'
ws['C25'] = 'AUDm'
for i, c in enumerate(range(4, 14)):
    yr = 2026 + i
    ws.cell(25, c).value = f"=INDEX(Annual!$D:$P,MATCH(\"Stat-Statutory EBITDA\",Annual!$A:$A,0),MATCH({yr},Annual!$1:$1,0))"

# R26: less D&A
ws['B26'] = 'less D&A'
ws['C26'] = 'AUDm'
for i, c in enumerate(range(4, 14)):
    yr = 2026 + i
    ws.cell(26, c).value = f"=INDEX(Annual!$D:$P,MATCH(\"DA-Total DA\",Annual!$A:$A,0),MATCH({yr},Annual!$1:$1,0))"

# R27: EBIT
ws['B27'] = 'EBIT'
ws['C27'] = 'AUDm'
for c in range(4, 14):
    col = CL(c)
    ws.cell(27, c).value = f'={col}25+{col}26'

# R28: less Tax on EBIT
ws['B28'] = 'less Tax on EBIT'
ws['C28'] = 'AUDm'
for c in range(4, 14):
    col = CL(c)
    ws.cell(28, c).value = f'=IF({col}27>0,-{col}27*$C$17,0)'

# R29: NOPAT
ws['B29'] = 'NOPAT'
ws['C29'] = 'AUDm'
for c in range(4, 14):
    col = CL(c)
    ws.cell(29, c).value = f'={col}27+{col}28'

# R30: plus D&A
ws['B30'] = 'plus D&A'
ws['C30'] = 'AUDm'
for c in range(4, 14):
    col = CL(c)
    ws.cell(30, c).value = f'=-{col}26'

# R31: less Capex
ws['B31'] = 'less Capex'
ws['C31'] = 'AUDm'
for i, c in enumerate(range(4, 14)):
    yr = 2026 + i
    ws.cell(31, c).value = f"=INDEX(Annual!$D:$P,MATCH(\"CF-Capex PPE\",Annual!$A:$A,0),MATCH({yr},Annual!$1:$1,0))"

# R32: less WC Change
ws['B32'] = 'less WC Change'
ws['C32'] = 'AUDm'
for i, c in enumerate(range(4, 14)):
    yr = 2026 + i
    ws.cell(32, c).value = f"=INDEX(Annual!$D:$P,MATCH(\"CF-WC Change\",Annual!$A:$A,0),MATCH({yr},Annual!$1:$1,0))"

# R33: FCFF
ws['B33'] = 'FCFF'
ws['C33'] = 'AUDm'
for c in range(4, 14):
    col = CL(c)
    ws.cell(33, c).value = f'={col}29+{col}30+{col}31+{col}32'

# R34: Normalised FCFF (capex=D&A in terminal year)
ws['B34'] = 'Normalised FCFF (capex=D&A)'
ws['C34'] = 'AUDm'
# Only in terminal year column (M=col 13)
ws.cell(34, 13).value = '=M29'  # NOPAT (capex=D&A cancels out)

# R35: Terminal Value
ws['B35'] = 'Terminal Value'
ws.cell(35, 13).value = '=M34*(1+$C$21)/($C$20-$C$21)'

# R37: Discount Factors
ws['B37'] = 'Discount Factor'
for i, c in enumerate(range(4, 14)):
    ws.cell(37, c).value = f'=1/(1+$C$20)^($C$22+{i})'

# R38: PV of FCFF
ws['B38'] = 'PV of FCFF'
for c in range(4, 14):
    col = CL(c)
    ws.cell(38, c).value = f'={col}33*{col}37'

# R39: PV of Terminal Value
ws['B39'] = 'PV of Terminal Value'
ws.cell(39, 13).value = '=M35*M37'

# ============================================================
# DCF SUMMARY (R41-R49)
# ============================================================
ws['B41'] = 'Sum of PV of FCFs'
ws['C41'] = '=SUM(D38:M38)'

ws['B42'] = 'PV of Terminal Value'
ws['C42'] = '=M39'

ws['B43'] = 'Enterprise Value'
ws['C43'] = '=C41+C42'

ws['B45'] = 'less Net Debt'
ws['C45'] = '=C7'  # Net cash (negative = add)

ws['B46'] = 'less Lease Liabilities'
# Pull lease liabilities from FY25
ws['C46'] = "=-INDEX(Annual!D:P,MATCH(\"BS-Lease Liabilities\",Annual!A:A,0),MATCH(2025,Annual!$1:$1,0))"

ws['B47'] = 'Equity Value'
ws['C47'] = '=C43+C45+C46'

ws['B48'] = 'Per Share Value (AUD)'
ws['C48'] = '=C47/C5'

ws['B49'] = 'Upside / Downside'
ws['C49'] = '=C48/C4-1'

# ============================================================
# EV/EBITDA SOTP (R52-R68)
# ============================================================
ws['B52'] = 'EV/EBITDA SOTP'
ws['B54'] = 'Select FY:'
ws['C54'] = 'FY27E'

ws['B56'] = 'Segment'
ws['C56'] = 'FY27E EBITDA'
ws['D56'] = 'Multiple'
ws['E56'] = 'Implied EV'

# Australia
ws['B57'] = 'Australia'
ws['C57'] = "=INDEX(Annual!$D:$P,MATCH(\"EBITDA-Australia EBITDA\",Annual!$A:$A,0),MATCH(2027,Annual!$1:$1,0))"
ws['D57'] = 25  # Premium QSR multiple
ws['E57'] = '=C57*D57'

# US
ws['B58'] = 'US'
ws['C58'] = "=INDEX(Annual!$D:$P,MATCH(\"EBITDA-US EBITDA\",Annual!$A:$A,0),MATCH(2027,Annual!$1:$1,0))"
ws['D58'] = 0  # Loss-making
ws['E58'] = '=C58*D58'

# Corporate
ws['B59'] = 'Corporate'
ws['C59'] = 0
ws['D59'] = '=IF(C57+C58=0,0,(D57*C57+D58*C58)/(C57+C58))'
ws['E59'] = '=C59*D59'

ws['B61'] = 'Group EV'
ws['E61'] = '=SUM(E57:E59)'

ws['B62'] = 'less Net Debt'
ws['E62'] = '=C45'

ws['B63'] = 'less Lease Liabilities'
ws['E63'] = '=C46'

ws['B64'] = 'Equity Value'
ws['E64'] = '=E61+E62+E63'

ws['B65'] = 'Per Share Value (AUD)'
ws['E65'] = '=E64/C5'

ws['B66'] = 'Upside / Downside'
ws['E66'] = '=E65/C4-1'

ws['B68'] = 'Implied Group EV/EBITDA'
ws['E68'] = '=IF(C57+C58=0,"",E61/(C57+C58))'

wb.save('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')
print("Part 7 complete: Value sheet set up")

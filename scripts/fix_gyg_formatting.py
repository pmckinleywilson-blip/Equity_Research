"""
Fix ALL formatting issues in GYG Model.xlsx

Part 1: Fix BS/CF/OpFCF/ROIC forecast formatting (Annual rows 105-191, cols G-P)
Part 2: Fix P&L subtotal formatting across ALL sheets
Part 3: Fix section header fill coverage
Part 4: Fix remaining 'General' format cells in data columns
"""

import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy

FILEPATH = '/home/pmwilson/Project_Equities/GYG/Models/GYG Model.xlsx'

# Units-to-format mapping
UNITS_FORMAT = {
    'A$m': '#,##0.0',
    '%': '0.0%',
    '% YoY': '0.0%',
    'cps': '0.000',
    '#': '#,##0',
    'm': '0.0',
    'x': '0.0',
}

thin_side = Side(style='thin')
no_side = Side(style=None)


def get_fill_rgb(cell):
    """Get the fill RGB string for a cell, or None."""
    f = cell.fill
    if f and f.patternType and f.patternType != 'none':
        fg = f.fgColor
        if fg and fg.rgb and fg.rgb not in ('00000000', '0'):
            return fg.rgb
    return None


def is_section_header(ws, row):
    """Row with light blue FFC5D9F1 fill on col B."""
    return get_fill_rgb(ws.cell(row, 2)) == 'FFC5D9F1'


def is_subsection_header(ws, row):
    """Row with grey FFD9D9D9 fill on col B."""
    return get_fill_rgb(ws.cell(row, 2)) == 'FFD9D9D9'


def is_subheader(ws, row):
    """Bold + underline on col B (e.g. 'Revenue', 'CFO')."""
    font = ws.cell(row, 2).font
    return font.bold and font.underline == 'single'


def is_subtotal_row(ws, row):
    """Bold on col B, NOT section header, NOT subheader."""
    font = ws.cell(row, 2).font
    if not font.bold:
        return False
    if is_section_header(ws, row):
        return False
    if is_subsection_header(ws, row):
        return False
    if is_subheader(ws, row):
        return False
    return True


def apply_bold_borders(cell, keep_color=True, keep_fmt=True):
    """Set bold + thin top/bottom border, preserving color and format."""
    old_font = cell.font
    if keep_color:
        cell.font = Font(
            name=old_font.name, size=old_font.size, bold=True,
            italic=old_font.italic, underline=old_font.underline,
            color=copy(old_font.color) if old_font.color else None,
        )
    else:
        cell.font = Font(
            name=old_font.name, size=old_font.size, bold=True,
            italic=old_font.italic, underline=old_font.underline,
        )
    old_border = cell.border
    cell.border = Border(
        left=copy(old_border.left) if old_border.left else no_side,
        right=copy(old_border.right) if old_border.right else no_side,
        top=thin_side,
        bottom=thin_side,
    )


wb = openpyxl.load_workbook(FILEPATH)

# ============================================================
# PART 1: Fix BS/CF/OpFCF/ROIC forecast formatting
# Annual sheet, rows 105-191, cols G(7) through P(16)
# ============================================================
print("=== PART 1: Fix Annual forecast formatting rows 105-191 ===")
ws = wb['Annual']
part1_count = 0
for row in range(105, 192):
    col_f = ws.cell(row, 6)  # Column F = last actual
    f_fmt = col_f.number_format
    f_bold = col_f.font.bold
    f_border_top = col_f.border.top.style if col_f.border.top else None
    f_border_bot = col_f.border.bottom.style if col_f.border.bottom else None

    for col in range(7, 17):  # G=7 through P=16
        cell = ws.cell(row, col)
        if cell.value is None and f_fmt == 'General':
            continue  # skip empty rows

        # Copy number format from col F
        if f_fmt != 'General':
            cell.number_format = f_fmt

        # Copy bold if col F is bold
        if f_bold:
            old_font = cell.font
            cell.font = Font(
                name=old_font.name, size=old_font.size, bold=True,
                italic=old_font.italic, underline=old_font.underline,
                color=copy(old_font.color) if old_font.color else None,
            )

        # Copy borders if col F has thin top/bottom
        if f_border_top == 'thin' or f_border_bot == 'thin':
            old_border = cell.border
            cell.border = Border(
                left=copy(old_border.left) if old_border.left else no_side,
                right=copy(old_border.right) if old_border.right else no_side,
                top=thin_side if f_border_top == 'thin' else copy(old_border.top) if old_border.top else no_side,
                bottom=thin_side if f_border_bot == 'thin' else copy(old_border.bottom) if old_border.bottom else no_side,
            )

        part1_count += 1

print(f"  Fixed {part1_count} cells")

# ============================================================
# PART 2: Fix P&L subtotal formatting across ALL sheets
# ============================================================
print("\n=== PART 2: Fix subtotal formatting ===")

sheet_configs = {
    'Annual': {'data_start': 4, 'data_end': 16},      # D=4 through P=16
    'HY & Segments': {'data_start': 4, 'data_end': 29},  # D=4 through AC=29
}

part2_count = 0
for sheet_name, cfg in sheet_configs.items():
    ws = wb[sheet_name]
    for row in range(1, ws.max_row + 1):
        if is_subtotal_row(ws, row):
            for col in range(cfg['data_start'], cfg['data_end'] + 1):
                cell = ws.cell(row, col)
                apply_bold_borders(cell, keep_color=True)
                part2_count += 1

print(f"  Fixed {part2_count} cells across both sheets")

# ============================================================
# PART 3: Fix section header fill coverage
# ============================================================
print("\n=== PART 3: Fix section header fill coverage ===")

part3_count = 0
for sheet_name, cfg in sheet_configs.items():
    ws = wb[sheet_name]
    last_col = cfg['data_end']
    for row in range(1, ws.max_row + 1):
        fill_rgb = get_fill_rgb(ws.cell(row, 2))
        if fill_rgb in ('FFC5D9F1', 'FFD9D9D9'):
            the_fill = PatternFill(start_color=fill_rgb, end_color=fill_rgb, fill_type='solid')
            for col in range(1, last_col + 1):
                ws.cell(row, col).fill = the_fill
                part3_count += 1

print(f"  Applied fill to {part3_count} cells")

# ============================================================
# PART 4: Fix remaining 'General' format cells in data columns
# ============================================================
print("\n=== PART 4: Fix remaining General format cells ===")

part4_count = 0
for sheet_name, cfg in sheet_configs.items():
    ws = wb[sheet_name]
    for row in range(1, ws.max_row + 1):
        for col in range(cfg['data_start'], cfg['data_end'] + 1):
            cell = ws.cell(row, col)
            if cell.number_format == 'General' and cell.value is not None:
                # Get units from col C
                units = ws.cell(row, 3).value
                if units and units in UNITS_FORMAT:
                    cell.number_format = UNITS_FORMAT[units]
                    part4_count += 1

# Also check Value sheet
ws_val = wb['Value']
for row in range(1, ws_val.max_row + 1):
    for col in range(4, ws_val.max_column + 1):
        cell = ws_val.cell(row, col)
        if cell.number_format == 'General' and cell.value is not None:
            units = ws_val.cell(row, 3).value
            if units and units in UNITS_FORMAT:
                cell.number_format = UNITS_FORMAT[units]
                part4_count += 1

print(f"  Fixed {part4_count} cells with General format")

# ============================================================
# SAVE
# ============================================================
wb.save(FILEPATH)
print(f"\nSaved to {FILEPATH}")

# ============================================================
# VERIFICATION
# ============================================================
print("\n=== VERIFICATION ===")
wb2 = openpyxl.load_workbook(FILEPATH)

# 1. Check Annual row 114 (Total Assets)
ws = wb2['Annual']
print("\n1. Annual Row 114 (Total Assets):")
for col in [4, 7, 10, 16]:
    c = ws.cell(114, col)
    print(f"   Col {get_column_letter(col)}: fmt={c.number_format} bold={c.font.bold} "
          f"border_top={c.border.top.style} border_bot={c.border.bottom.style}")

# 2. Check Annual row 153 (Net OCF)
print("\n2. Annual Row 153 (Net Operating Cash Flow):")
for col in [4, 7, 10, 16]:
    c = ws.cell(153, col)
    print(f"   Col {get_column_letter(col)}: fmt={c.number_format} bold={c.font.bold} "
          f"border_top={c.border.top.style} border_bot={c.border.bottom.style}")

# 3. Count remaining General format cells in data columns
general_count = 0
for sheet_name, cfg in sheet_configs.items():
    ws_check = wb2[sheet_name]
    for row in range(1, ws_check.max_row + 1):
        for col in range(cfg['data_start'], cfg['data_end'] + 1):
            cell = ws_check.cell(row, col)
            if cell.number_format == 'General' and cell.value is not None:
                general_count += 1
                if general_count <= 10:
                    print(f"   Remaining General: {sheet_name} row {row} col {get_column_letter(col)} val={cell.value}")

print(f"\n3. Remaining General format cells in data columns: {general_count}")

# 4. Count subtotal rows with missing bold/borders
missing_count = 0
for sheet_name, cfg in sheet_configs.items():
    ws_check = wb2[sheet_name]
    for row in range(1, ws_check.max_row + 1):
        font_b = ws_check.cell(row, 2).font
        if not font_b.bold:
            continue
        fill_rgb = get_fill_rgb(ws_check.cell(row, 2))
        if fill_rgb in ('FFC5D9F1', 'FFD9D9D9'):
            continue
        if font_b.underline == 'single':
            continue
        # This is a subtotal row - check data cells
        for col in range(cfg['data_start'], cfg['data_end'] + 1):
            cell = ws_check.cell(row, col)
            if not cell.font.bold or cell.border.top.style != 'thin' or cell.border.bottom.style != 'thin':
                missing_count += 1

print(f"4. Subtotal cells with missing bold/borders: {missing_count}")

wb2.close()
print("\nDone!")

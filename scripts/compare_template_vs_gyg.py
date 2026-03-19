"""
Compare HY_model_template.xlsx against GYG Model.xlsx
Systematic quality review of template convention adherence.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import sys


def safe_color(color_obj):
    """Extract color RGB safely."""
    if color_obj is None:
        return None
    try:
        rgb = color_obj.rgb
        if rgb and isinstance(rgb, str) and rgb != '00000000':
            return rgb
        return None
    except (TypeError, AttributeError):
        return None


def border_style(side):
    """Extract border style."""
    if side and side.style:
        return side.style
    return None


def get_row_format(ws, row, data_col):
    """Extract formatting for a row (label cell + data cell)."""
    b_cell = ws.cell(row, 2)
    d_cell = ws.cell(row, data_col)

    return {
        'label': b_cell.value,
        'key': ws.cell(row, 1).value,
        'units': ws.cell(row, 3).value,
        # Label cell formatting
        'b_bold': b_cell.font.bold if b_cell.font.bold else False,
        'b_color': safe_color(b_cell.font.color),
        'b_size': b_cell.font.size,
        'b_fill': safe_color(b_cell.fill.fgColor) if b_cell.fill else None,
        'b_border_top': border_style(b_cell.border.top) if b_cell.border else None,
        'b_border_bottom': border_style(b_cell.border.bottom) if b_cell.border else None,
        # Data cell formatting
        'd_bold': d_cell.font.bold if d_cell.font.bold else False,
        'd_color': safe_color(d_cell.font.color),
        'd_size': d_cell.font.size,
        'd_fill': safe_color(d_cell.fill.fgColor) if d_cell.fill else None,
        'd_border_top': border_style(d_cell.border.top) if d_cell.border else None,
        'd_border_bottom': border_style(d_cell.border.bottom) if d_cell.border else None,
        'd_numfmt': d_cell.number_format,
    }


def classify_row(key, label):
    """Classify a row by its structural role based on the template convention."""
    if key is None and label is None:
        return 'blank'
    if key is None and label is not None:
        # Could be section header, subtotal, analytical, or section banner
        return 'computed_or_header'
    if key is not None:
        return 'data_row'
    return 'unknown'


def build_section_map(ws, max_row):
    """Build a list of (key_prefix, label, row) for matching between files."""
    rows = []
    for r in range(1, max_row + 1):
        key = ws.cell(r, 1).value
        label = ws.cell(r, 2).value
        units = ws.cell(r, 3).value
        rows.append((r, key, label, units))
    return rows


def match_rows_by_key(t_rows, g_rows):
    """Match rows between template and GYG by column A key.
    Returns list of (template_row_info, gyg_row_info) tuples.
    Unmatched rows get None for the missing side."""

    # Build key -> row mapping for keyed rows
    t_keyed = {}
    g_keyed = {}
    for r, key, label, units in t_rows:
        if key:
            t_keyed[key] = (r, key, label, units)
    for r, key, label, units in g_rows:
        if key:
            g_keyed[key] = (r, key, label, units)

    return t_keyed, g_keyed


def print_separator(char='=', width=120):
    print(char * width)


def print_header(text):
    print()
    print_separator()
    print(f"  {text}")
    print_separator()
    print()


def compare_sheet_structure(sheet_name, ws_t, ws_g, t_data_col, g_data_col):
    """Compare a sheet between template and GYG model."""

    print_header(f"SHEET: {sheet_name}")

    # ─── 1. HEADER ROWS (1-4) ───
    print("1. HEADER ROWS (Rows 1-4)")
    print("-" * 80)
    issues_found = False
    for r in range(1, 5):
        for c in range(1, max(ws_t.max_column, ws_g.max_column) + 1):
            t_val = ws_t.cell(r, c).value if c <= ws_t.max_column else None
            g_val = ws_g.cell(r, c).value if c <= ws_g.max_column else None
            # Skip expected company-name differences
            if r in (2, 3) and c == 2:
                continue
            # Skip zone label year differences (row 1) — expected
            if r == 1:
                continue
            # Skip FY label differences in row 3 col D+ — expected
            if r == 3 and c >= 4:
                continue
            if t_val != g_val:
                col_letter = get_column_letter(c)
                print(f"  DIFF Row {r}, Col {col_letter}: Template={t_val!r}, GYG={g_val!r}")
                issues_found = True

    # Check header formatting
    for r in [2, 3]:
        t_cell = ws_t.cell(r, 2)
        g_cell = ws_g.cell(r, 2)
        t_bold = t_cell.font.bold if t_cell.font.bold else False
        g_bold = g_cell.font.bold if g_cell.font.bold else False
        t_size = t_cell.font.size
        g_size = g_cell.font.size
        if t_bold != g_bold:
            print(f"  FMT Row {r} Col B: bold Template={t_bold}, GYG={g_bold}")
            issues_found = True
        if t_size != g_size:
            print(f"  FMT Row {r} Col B: font size Template={t_size}, GYG={g_size}")
            issues_found = True

    if not issues_found:
        print("  OK - Header rows match template conventions")
    print()

    # ─── 2. COLUMN LAYOUT ───
    print("2. COLUMN LAYOUT")
    print("-" * 80)
    print(f"  Template: {ws_t.max_column} columns")
    print(f"  GYG:      {ws_g.max_column} columns")

    # Check zone labels (row 1) pattern
    t_zones = [(c, ws_t.cell(1, c).value) for c in range(1, ws_t.max_column + 1) if ws_t.cell(1, c).value]
    g_zones = [(c, ws_g.cell(1, c).value) for c in range(1, ws_g.max_column + 1) if ws_g.cell(1, c).value]
    print(f"  Template zone labels: {len(t_zones)} entries, cols {t_zones[0][0]}-{t_zones[-1][0]}" if t_zones else "  Template zone labels: none")
    print(f"  GYG zone labels:     {len(g_zones)} entries, cols {g_zones[0][0]}-{g_zones[-1][0]}" if g_zones else "  GYG zone labels: none")

    # Check if zone label format matches (should always start at col D = col 4)
    if t_zones and g_zones:
        if t_zones[0][0] != g_zones[0][0]:
            print(f"  WARNING: Zone labels start at different columns: Template col {t_zones[0][0]}, GYG col {g_zones[0][0]}")

    # Check zone label formatting
    if t_zones and g_zones:
        t_zone_cell = ws_t.cell(1, t_zones[0][0])
        g_zone_cell = ws_g.cell(1, g_zones[0][0])
        t_zfill = safe_color(t_zone_cell.fill.fgColor)
        g_zfill = safe_color(g_zone_cell.fill.fgColor)
        t_zfont_color = safe_color(t_zone_cell.font.color)
        g_zfont_color = safe_color(g_zone_cell.font.color)
        if t_zfill != g_zfill:
            print(f"  DIFF Zone label fill: Template={t_zfill}, GYG={g_zfill}")
        if t_zfont_color != g_zfont_color:
            print(f"  DIFF Zone label font color: Template={t_zfont_color}, GYG={g_zfont_color}")

    # Check for trailing empty columns in GYG
    trailing_empty = 0
    for c in range(ws_g.max_column, 0, -1):
        col_has_data = False
        for r in range(1, min(ws_g.max_row + 1, 10)):
            if ws_g.cell(r, c).value is not None:
                col_has_data = True
                break
        if col_has_data:
            break
        trailing_empty += 1
    if trailing_empty > 0:
        print(f"  WARNING: GYG has {trailing_empty} trailing empty columns (max_column={ws_g.max_column}, "
              f"last data col={ws_g.max_column - trailing_empty})")
    print()

    # ─── 3. ROW STRUCTURE — KEYED ROWS ───
    print("3. ROW STRUCTURE — KEYED ROWS (Column A)")
    print("-" * 80)

    t_rows = build_section_map(ws_t, ws_t.max_row)
    g_rows = build_section_map(ws_g, ws_g.max_row)

    t_keyed, g_keyed = match_rows_by_key(t_rows, g_rows)

    # Find keys in template but not in GYG
    missing_from_gyg = set(t_keyed.keys()) - set(g_keyed.keys())
    extra_in_gyg = set(g_keyed.keys()) - set(t_keyed.keys())
    common_keys = set(t_keyed.keys()) & set(g_keyed.keys())

    if missing_from_gyg:
        print("  TEMPLATE KEYS MISSING FROM GYG (rows that should exist):")
        for k in sorted(missing_from_gyg):
            r, _, label, units = t_keyed[k]
            print(f"    Row {r:3d}: Key={k!r:35s} Label={label!r}")
    else:
        print("  All template keyed rows have equivalents in GYG (keys adapted for GYG company)")

    if extra_in_gyg:
        print()
        print("  GYG-SPECIFIC KEYS (not in template — expected for different company):")
        for k in sorted(extra_in_gyg):
            r, _, label, units = g_keyed[k]
            print(f"    Row {r:3d}: Key={k!r:35s} Label={label!r}")

    # Check key format convention
    print()
    print("  KEY FORMAT CHECK (should follow PREFIX-Description pattern):")
    key_issues = []
    for k in sorted(g_keyed.keys()):
        if '-' not in k:
            key_issues.append(f"    Key {k!r} has no hyphen separator")
        parts = k.split('-', 1)
        if len(parts) == 2:
            prefix = parts[0]
            # Check prefix matches template convention
            valid_prefixes = {'Rev', 'COGS', 'GP', 'OPEX', 'EBITDA', 'Stat', 'DA',
                              'EBIT', 'Int', 'PBT', 'Tax', 'NPAT', 'EPS', 'Div',
                              'KPI', 'BS', 'CF'}
            if prefix not in valid_prefixes:
                key_issues.append(f"    Key {k!r} uses non-standard prefix '{prefix}'")
    if key_issues:
        for issue in key_issues:
            print(issue)
    else:
        print("    All keys follow standard prefix convention")
    print()

    # ─── 4. ROW STRUCTURE — SECTION HEADERS & ANALYTICAL ROWS ───
    print("4. SECTION STRUCTURE — HEADERS, SUBTOTALS, ANALYTICAL ROWS")
    print("-" * 80)

    # Build section-by-section comparison using non-keyed rows
    # Match by label for common structural rows
    template_structural = {}
    gyg_structural = {}

    for r, key, label, units in t_rows:
        if key is None and label is not None:
            template_structural[label] = (r, label, units)
    for r, key, label, units in g_rows:
        if key is None and label is not None:
            gyg_structural[label] = (r, label, units)

    # Common labels (structural rows both files share)
    common_structural = set(template_structural.keys()) & set(gyg_structural.keys())
    template_only = set(template_structural.keys()) - set(gyg_structural.keys())
    gyg_only = set(gyg_structural.keys()) - set(template_structural.keys())

    # Filter out expected company-specific differences
    expected_template_only = set()
    expected_gyg_only = set()
    for label in template_only:
        # Template company-specific labels
        if any(x in label.lower() for x in ['steel', 'metals', 'vulcan', 'vsl']):
            expected_template_only.add(label)
    for label in gyg_only:
        if any(x in label.lower() for x in ['drive thru', 'strip', 'franchise', 'australia',
                                              'gyg', 'guzman', 'network', 'segment',
                                              'corp ', 'restaurant', 'food', 'marketing',
                                              'admin', 'aasb', 'impairment', 'reacquired',
                                              'term deposit', 'lease receivable',
                                              'us ', 'asia', 'dt ', 'other auw',
                                              'au ', 'g&a']):
            expected_gyg_only.add(label)

    unexpected_template_only = template_only - expected_template_only
    unexpected_gyg_only = gyg_only - expected_gyg_only

    if unexpected_template_only:
        print("  TEMPLATE STRUCTURAL ROWS MISSING FROM GYG (potential quality issue):")
        for label in sorted(unexpected_template_only):
            r, _, units = template_structural[label]
            print(f"    Row {r:3d}: {label!r:50s} Units={units!r}")

    if unexpected_gyg_only:
        print()
        print("  GYG STRUCTURAL ROWS NOT IN TEMPLATE (verify these are intentional):")
        for label in sorted(unexpected_gyg_only):
            r, _, units = gyg_structural[label]
            print(f"    Row {r:3d}: {label!r:50s} Units={units!r}")

    # Show expected company-specific differences
    if expected_template_only:
        print()
        print(f"  Template-specific structural rows (expected, {len(expected_template_only)} rows): "
              f"{', '.join(sorted(expected_template_only)[:5])}{'...' if len(expected_template_only) > 5 else ''}")
    if expected_gyg_only:
        print(f"  GYG-specific structural rows (expected, {len(expected_gyg_only)} rows): "
              f"{', '.join(sorted(expected_gyg_only)[:5])}{'...' if len(expected_gyg_only) > 5 else ''}")
    print()

    # ─── 5. FORMATTING COMPARISON ───
    print("5. FORMATTING COMPARISON (label col B + data col)")
    print(f"   Template data col: {t_data_col} ({get_column_letter(t_data_col)})")
    print(f"   GYG data col:      {g_data_col} ({get_column_letter(g_data_col)})")
    print("-" * 80)

    # Compare formatting for common keyed rows
    fmt_issues = []
    for key in sorted(common_keys):
        t_r = t_keyed[key][0]
        g_r = g_keyed[key][0]
        t_fmt = get_row_format(ws_t, t_r, t_data_col)
        g_fmt = get_row_format(ws_g, g_r, g_data_col)

        label = t_fmt['label'] or g_fmt['label']
        row_issues = []

        # Label cell formatting
        if t_fmt['b_bold'] != g_fmt['b_bold']:
            row_issues.append(f"Label bold: T={t_fmt['b_bold']}, G={g_fmt['b_bold']}")
        if t_fmt['b_fill'] != g_fmt['b_fill']:
            row_issues.append(f"Label fill: T={t_fmt['b_fill']}, G={g_fmt['b_fill']}")
        if t_fmt['b_border_top'] != g_fmt['b_border_top']:
            row_issues.append(f"Label border-top: T={t_fmt['b_border_top']}, G={g_fmt['b_border_top']}")
        if t_fmt['b_border_bottom'] != g_fmt['b_border_bottom']:
            row_issues.append(f"Label border-bottom: T={t_fmt['b_border_bottom']}, G={g_fmt['b_border_bottom']}")

        # Data cell formatting
        if t_fmt['d_bold'] != g_fmt['d_bold']:
            row_issues.append(f"Data bold: T={t_fmt['d_bold']}, G={g_fmt['d_bold']}")
        if t_fmt['d_color'] != g_fmt['d_color']:
            row_issues.append(f"Data font color: T={t_fmt['d_color']}, G={g_fmt['d_color']}")
        if t_fmt['d_fill'] != g_fmt['d_fill']:
            row_issues.append(f"Data fill: T={t_fmt['d_fill']}, G={g_fmt['d_fill']}")
        if t_fmt['d_border_top'] != g_fmt['d_border_top']:
            row_issues.append(f"Data border-top: T={t_fmt['d_border_top']}, G={g_fmt['d_border_top']}")
        if t_fmt['d_border_bottom'] != g_fmt['d_border_bottom']:
            row_issues.append(f"Data border-bottom: T={t_fmt['d_border_bottom']}, G={g_fmt['d_border_bottom']}")
        if t_fmt['d_numfmt'] != g_fmt['d_numfmt']:
            row_issues.append(f"Number format: T={t_fmt['d_numfmt']!r}, G={g_fmt['d_numfmt']!r}")

        if row_issues:
            fmt_issues.append((key, label, t_r, g_r, row_issues))

    if fmt_issues:
        for key, label, t_r, g_r, issues in fmt_issues:
            print(f"  Key={key!r} (T-row {t_r}, G-row {g_r}): {label!r}")
            for issue in issues:
                print(f"    - {issue}")
            print()
    else:
        print("  All common keyed rows have matching formatting")
        print()

    # ─── 5b. FORMATTING — STRUCTURAL ROWS ───
    print("  STRUCTURAL ROW FORMATTING (section headers, subtotals, analytical):")
    print("  " + "-" * 78)

    struct_fmt_issues = []
    for label in sorted(common_structural):
        t_r = template_structural[label][0]
        g_r = gyg_structural[label][0]
        t_fmt = get_row_format(ws_t, t_r, t_data_col)
        g_fmt = get_row_format(ws_g, g_r, g_data_col)

        row_issues = []
        if t_fmt['b_bold'] != g_fmt['b_bold']:
            row_issues.append(f"Label bold: T={t_fmt['b_bold']}, G={g_fmt['b_bold']}")
        if t_fmt['b_fill'] != g_fmt['b_fill']:
            row_issues.append(f"Label fill: T={t_fmt['b_fill']}, G={g_fmt['b_fill']}")
        if t_fmt['b_border_top'] != g_fmt['b_border_top']:
            row_issues.append(f"Label border-top: T={t_fmt['b_border_top']}, G={g_fmt['b_border_top']}")
        if t_fmt['b_border_bottom'] != g_fmt['b_border_bottom']:
            row_issues.append(f"Label border-bottom: T={t_fmt['b_border_bottom']}, G={g_fmt['b_border_bottom']}")
        if t_fmt['d_bold'] != g_fmt['d_bold']:
            row_issues.append(f"Data bold: T={t_fmt['d_bold']}, G={g_fmt['d_bold']}")
        if t_fmt['d_fill'] != g_fmt['d_fill']:
            row_issues.append(f"Data fill: T={t_fmt['d_fill']}, G={g_fmt['d_fill']}")
        if t_fmt['d_border_top'] != g_fmt['d_border_top']:
            row_issues.append(f"Data border-top: T={t_fmt['d_border_top']}, G={g_fmt['d_border_top']}")
        if t_fmt['d_border_bottom'] != g_fmt['d_border_bottom']:
            row_issues.append(f"Data border-bottom: T={t_fmt['d_border_bottom']}, G={g_fmt['d_border_bottom']}")
        if t_fmt['d_numfmt'] != g_fmt['d_numfmt']:
            row_issues.append(f"Number format: T={t_fmt['d_numfmt']!r}, G={g_fmt['d_numfmt']!r}")

        if row_issues:
            struct_fmt_issues.append((label, t_r, g_r, row_issues))

    if struct_fmt_issues:
        for label, t_r, g_r, issues in struct_fmt_issues:
            print(f"  '{label}' (T-row {t_r}, G-row {g_r}):")
            for issue in issues:
                print(f"    - {issue}")
    else:
        print("  All common structural rows have matching formatting")
    print()

    # ─── 6. BLANK ROW SPACING ───
    print("6. BLANK ROW SPACING")
    print("-" * 80)

    # Build section sequence for both: identify section headers and count blanks before them
    def get_section_pattern(ws, max_row):
        """Returns list of (section_label, blanks_before) for section headers."""
        pattern = []
        blank_count = 0
        for r in range(1, max_row + 1):
            key = ws.cell(r, 1).value
            label = ws.cell(r, 2).value
            if key is None and label is None:
                blank_count += 1
            else:
                is_section = False
                b_cell = ws.cell(r, 2)
                if b_cell.font.bold and key is None:
                    # Check if it's a section header (not a subtotal with borders)
                    has_border = border_style(b_cell.border.top) or border_style(b_cell.border.bottom)
                    if not has_border:
                        is_section = True

                if is_section and blank_count > 0:
                    pattern.append((label, blank_count, r))
                blank_count = 0
        return pattern

    t_pattern = get_section_pattern(ws_t, ws_t.max_row)
    g_pattern = get_section_pattern(ws_g, ws_g.max_row)

    # Match by label
    t_spacing = {p[0]: p[1] for p in t_pattern}
    g_spacing = {p[0]: p[1] for p in g_pattern}
    common_sections = set(t_spacing.keys()) & set(g_spacing.keys())

    spacing_issues = []
    for label in sorted(common_sections):
        if t_spacing[label] != g_spacing[label]:
            spacing_issues.append((label, t_spacing[label], g_spacing[label]))

    if spacing_issues:
        print("  SPACING DIFFERENCES (blank rows before section header):")
        for label, t_blanks, g_blanks in spacing_issues:
            print(f"    '{label}': Template={t_blanks} blank(s), GYG={g_blanks} blank(s)")
    else:
        print("  Blank row spacing matches template convention for all common sections")
    print()

    # ─── 7. FORMULA PATTERNS ───
    print("7. FORMULA PATTERNS (subtotals and analytical rows)")
    print("-" * 80)

    # Compare formulas for common keyed rows that should have formulas (subtotals)
    subtotal_keys = [k for k in common_keys if any(
        x in k for x in ['Total', 'Gross Profit', 'Underlying', 'Statutory', 'Net Finance',
                          'Net OCF', 'Net Change']
    )]

    formula_issues = []
    for key in sorted(subtotal_keys):
        t_r = t_keyed[key][0]
        g_r = g_keyed[key][0]

        # Check formula in first forecast column
        t_cell = ws_t.cell(t_r, t_data_col)
        g_cell = ws_g.cell(g_r, g_data_col)

        t_formula = t_cell.value if isinstance(t_cell.value, str) and t_cell.value.startswith('=') else None
        g_formula = g_cell.value if isinstance(g_cell.value, str) and g_cell.value.startswith('=') else None

        if t_formula and not g_formula:
            formula_issues.append((key, t_r, g_r, f"Template has formula, GYG has value/empty. "
                                   f"T={t_formula!r}, G={g_cell.value!r}"))
        elif not t_formula and g_formula:
            pass  # GYG having formulas where template has values is fine
        elif t_formula and g_formula:
            # Compare formula structure (normalize row references)
            pass  # Both have formulas — OK

    if formula_issues:
        for key, t_r, g_r, desc in formula_issues:
            print(f"  Key={key!r} (T-row {t_r}, G-row {g_r}): {desc}")
    else:
        print("  Formula patterns appear consistent")
    print()


def compare_value_sheet(ws_t, ws_g):
    """Compare the Value sheet specifically."""
    print_header("SHEET: Value")

    print("1. STRUCTURE COMPARISON")
    print("-" * 80)

    # Row-by-row comparison
    max_row = max(ws_t.max_row, ws_g.max_row)
    issues = []
    for r in range(1, max_row + 1):
        t_label = ws_t.cell(r, 2).value if r <= ws_t.max_row else None
        g_label = ws_g.cell(r, 2).value if r <= ws_g.max_row else None

        if t_label != g_label:
            # Skip expected currency differences
            if t_label and g_label and t_label.replace('NZD', 'AUD') == g_label.replace('NZD', 'AUD'):
                continue
            # Skip expected company-name differences
            if t_label and g_label:
                t_clean = t_label.replace('NZD', '').replace('AUD', '')
                g_clean = g_label.replace('NZD', '').replace('AUD', '')
                if t_clean == g_clean:
                    continue
            issues.append((r, t_label, g_label))

    if issues:
        print("  LABEL DIFFERENCES (excluding expected currency/company swaps):")
        for r, t_label, g_label in issues:
            print(f"    Row {r:3d}: Template={t_label!r:40s} GYG={g_label!r}")
    else:
        print("  All Value sheet labels match (accounting for currency differences)")
    print()

    # Check formulas
    print("2. FORMULA COMPARISON")
    print("-" * 80)
    formula_issues = []
    for r in range(1, max_row + 1):
        t_val = ws_t.cell(r, 3).value if r <= ws_t.max_row else None
        g_val = ws_g.cell(r, 3).value if r <= ws_g.max_row else None

        t_is_formula = isinstance(t_val, str) and t_val.startswith('=')
        g_is_formula = isinstance(g_val, str) and g_val.startswith('=')

        if t_is_formula and g_is_formula:
            # Compare structure, ignoring column range differences ($D:$P vs $D:$R)
            t_norm = t_val.replace('$P', '$X').replace('$R', '$X')
            g_norm = g_val.replace('$P', '$X').replace('$R', '$X')
            if t_norm != g_norm:
                formula_issues.append((r, ws_t.cell(r, 2).value, t_val, g_val, 'formula_diff'))
        elif t_is_formula and not g_is_formula:
            formula_issues.append((r, ws_t.cell(r, 2).value, t_val, g_val, 'missing_formula'))
        elif not t_is_formula and g_is_formula:
            formula_issues.append((r, ws_t.cell(r, 2).value, t_val, g_val, 'extra_formula'))
        elif t_val != g_val:
            # Value differences
            t_label = ws_t.cell(r, 2).value
            g_label = ws_g.cell(r, 2).value
            # Skip expected differences
            if t_label == g_label and t_val != g_val:
                # Check if it's just different company-specific values
                if isinstance(t_val, (int, float)) and isinstance(g_val, (int, float)):
                    pass  # Different values expected
                elif isinstance(t_val, str) and isinstance(g_val, str):
                    if t_val.replace('NZD', '').replace('AUD', '') != g_val.replace('NZD', '').replace('AUD', ''):
                        formula_issues.append((r, t_label, t_val, g_val, 'value_diff'))

    if formula_issues:
        for r, label, t_val, g_val, issue_type in formula_issues:
            type_desc = {
                'formula_diff': 'FORMULA STRUCTURE DIFFERS',
                'missing_formula': 'TEMPLATE HAS FORMULA, GYG DOES NOT',
                'extra_formula': 'GYG HAS FORMULA, TEMPLATE DOES NOT',
                'value_diff': 'VALUE DIFFERS',
            }[issue_type]
            print(f"  Row {r:3d} ({label}): {type_desc}")
            if isinstance(t_val, str) and len(t_val) > 80:
                print(f"    Template: {t_val[:80]}...")
            else:
                print(f"    Template: {t_val!r}")
            if isinstance(g_val, str) and len(g_val) > 80:
                print(f"    GYG:      {g_val[:80]}...")
            else:
                print(f"    GYG:      {g_val!r}")
            print()
    else:
        print("  All formulas match template pattern")
    print()

    # Check units column (C) for NZD vs AUD consistency
    print("3. CURRENCY LABEL CHECK (Column C)")
    print("-" * 80)
    currency_issues = []
    for r in range(1, max_row + 1):
        g_units = ws_g.cell(r, 3).value if r <= ws_g.max_row else None
        t_units = ws_t.cell(r, 3).value if r <= ws_t.max_row else None
        if isinstance(g_units, str) and 'NZD' in g_units:
            currency_issues.append((r, ws_g.cell(r, 2).value, g_units, t_units))
        elif isinstance(g_units, str) and 'NZD' in g_units:
            currency_issues.append((r, ws_g.cell(r, 2).value, g_units, t_units))

    if currency_issues:
        print("  CURRENCY LABEL ISSUES (GYG should use AUD, not NZD):")
        for r, label, g_units, t_units in currency_issues:
            print(f"    Row {r:3d}: {label!r:40s} Units={g_units!r} (Template had {t_units!r})")
    else:
        print("  OK (or no NZD labels found where AUD expected)")
    print()

    # Check formatting
    print("4. VALUE SHEET FORMATTING")
    print("-" * 80)
    fmt_issues = []
    for r in range(1, max_row + 1):
        for c in [2, 3]:
            t_cell = ws_t.cell(r, c) if r <= ws_t.max_row else None
            g_cell = ws_g.cell(r, c) if r <= ws_g.max_row else None
            if t_cell and g_cell and t_cell.value and g_cell.value:
                t_bold = t_cell.font.bold if t_cell.font.bold else False
                g_bold = g_cell.font.bold if g_cell.font.bold else False
                if t_bold != g_bold:
                    fmt_issues.append((r, c, ws_t.cell(r, 2).value,
                                       f"Bold: T={t_bold}, G={g_bold}"))
                t_fill = safe_color(t_cell.fill.fgColor)
                g_fill = safe_color(g_cell.fill.fgColor)
                if t_fill != g_fill:
                    fmt_issues.append((r, c, ws_t.cell(r, 2).value,
                                       f"Fill: T={t_fill}, G={g_fill}"))

    if fmt_issues:
        for r, c, label, desc in fmt_issues:
            col_letter = get_column_letter(c)
            print(f"  Row {r:3d} Col {col_letter} ({label}): {desc}")
    else:
        print("  Formatting matches")
    print()


def check_gyg_specific_issues(ws_g, sheet_name, g_data_col):
    """Check for GYG-specific issues that don't have a template counterpart."""
    print(f"  GYG-SPECIFIC CHECKS ({sheet_name}):")
    print("  " + "-" * 78)

    # Check for rows with formatting on the label but missing formatting on data cells
    data_fmt_issues = []
    for r in range(5, ws_g.max_row + 1):
        label = ws_g.cell(r, 2).value
        if label is None:
            continue

        b_cell = ws_g.cell(r, 2)
        is_subtotal = (b_cell.font.bold if b_cell.font.bold else False) and \
                      (border_style(b_cell.border.top) or border_style(b_cell.border.bottom))

        if is_subtotal:
            # Check that data cell also has bold + borders
            d_cell = ws_g.cell(r, g_data_col)
            d_bold = d_cell.font.bold if d_cell.font.bold else False
            d_border_top = border_style(d_cell.border.top)
            d_border_bottom = border_style(d_cell.border.bottom)

            if not d_bold:
                data_fmt_issues.append((r, label, "Data cell NOT bold (label is bold+bordered)"))
            if not d_border_top and border_style(b_cell.border.top):
                data_fmt_issues.append((r, label, "Data cell missing top border"))
            if not d_border_bottom and border_style(b_cell.border.bottom):
                data_fmt_issues.append((r, label, "Data cell missing bottom border"))

        # Check section banner fill
        is_banner = (b_cell.font.bold if b_cell.font.bold else False) and \
                    safe_color(b_cell.fill.fgColor)
        if is_banner:
            d_cell = ws_g.cell(r, g_data_col)
            d_fill = safe_color(d_cell.fill.fgColor)
            b_fill = safe_color(b_cell.fill.fgColor)
            if b_fill and not d_fill:
                data_fmt_issues.append((r, label, f"Label has fill {b_fill} but data cell has no fill"))

    if data_fmt_issues:
        for r, label, desc in data_fmt_issues:
            print(f"    Row {r:3d} ({label}): {desc}")
    else:
        print("    No data-cell formatting inconsistencies found")
    print()


def check_number_format_consistency(ws_g, sheet_name, g_data_col):
    """Check that number formats are applied consistently across all data columns."""
    print(f"  NUMBER FORMAT CONSISTENCY CHECK ({sheet_name}):")
    print("  " + "-" * 78)

    # For each row with a number format in the reference data col, check all data cols match
    issues = []
    # Find data column range
    start_col = 4
    end_col = g_data_col

    for r in range(5, ws_g.max_row + 1):
        label = ws_g.cell(r, 2).value
        if not label:
            continue

        ref_fmt = ws_g.cell(r, start_col).number_format
        if ref_fmt == 'General':
            # Check if any other col has a format
            for c in range(start_col + 1, end_col + 1):
                cell_fmt = ws_g.cell(r, c).number_format
                if cell_fmt != 'General':
                    issues.append((r, label, f"Col {start_col} is General but col {c} is {cell_fmt!r}"))
                    break
        else:
            for c in range(start_col + 1, end_col + 1):
                cell_fmt = ws_g.cell(r, c).number_format
                if cell_fmt != ref_fmt and cell_fmt != 'General':
                    issues.append((r, label, f"Col {start_col} is {ref_fmt!r} but col {c} is {cell_fmt!r}"))
                    break

    if issues:
        for r, label, desc in issues:
            print(f"    Row {r:3d} ({label}): {desc}")
    else:
        print("    Number formats consistent across all data columns")
    print()


def main():
    print("=" * 120)
    print("  TEMPLATE vs GYG MODEL — QUALITY REVIEW REPORT")
    print("  Template: .claude/templates/HY_model_template.xlsx")
    print("  Model:    GYG/Models/GYG Model.xlsx")
    print("=" * 120)

    wb_t = openpyxl.load_workbook('.claude/templates/HY_model_template.xlsx')
    wb_g = openpyxl.load_workbook('GYG/Models/GYG Model.xlsx')

    # ─── ANNUAL SHEET ───
    ws_t = wb_t['Annual']
    ws_g = wb_g['Annual']
    # Use last actual year data column for formatting comparison
    # Template: col 16 (FY35E), but better to use a col with data — col 6 (FY25A)
    # GYG: col 8 (FY25A)
    compare_sheet_structure('Annual', ws_t, ws_g, t_data_col=6, g_data_col=8)
    check_gyg_specific_issues(ws_g, 'Annual', g_data_col=8)
    check_number_format_consistency(ws_g, 'Annual', g_data_col=18)

    # ─── HY & SEGMENTS SHEET ───
    ws_t = wb_t['HY & Segments']
    ws_g = wb_g['HY & Segments']
    compare_sheet_structure('HY & Segments', ws_t, ws_g, t_data_col=8, g_data_col=12)
    check_gyg_specific_issues(ws_g, 'HY & Segments', g_data_col=12)
    check_number_format_consistency(ws_g, 'HY & Segments', g_data_col=33)

    # ─── VALUE SHEET ───
    ws_t = wb_t['Value']
    ws_g = wb_g['Value']
    compare_value_sheet(ws_t, ws_g)

    # ─── SUMMARY ───
    print_header("SUMMARY OF KEY QUALITY ISSUES")
    print("""
This report identifies differences between the HY_model_template and the GYG Model.
Expected differences (company name, segment names, currency, column count) are noted
but not flagged as issues. Focus on:

  1. FORMATTING: Subtotal rows (bold + borders) should be consistent between label
     and data cells. Section banners (blue fill) should extend across all data columns.

  2. NUMBER FORMATS: Every data row should have the correct number format applied
     consistently across all data columns.

  3. STRUCTURAL INTEGRITY: The GYG model should follow the same section ordering
     and blank-row spacing as the template.

  4. VALUE SHEET: Formulas should match template patterns. Currency references
     should be AUD not NZD.

  5. KEY FORMAT: Column A keys should follow the PREFIX-Description convention
     established in the template.
""")


if __name__ == '__main__':
    main()

"""Part 4: HY & Segments sheet - labels, structure, and historical data."""
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')
ws = wb['HY & Segments']

# --- Headers ---
ws['B2'] = 'GYG Segments (Half-Year)'
ws['B3'] = 'Guzman y Gomez (GYG.AX)'

# Clear all data cells
for r in range(5, ws.max_row + 1):
    for c in range(4, 30):
        ws.cell(r, c).value = None

# Column mapping for HY sheet:
# D(4)=1H23, E(5)=2H23, F(6)=1H24, G(7)=2H24, H(8)=1H25, I(9)=2H25
# J(10)=1H26, K(11)=2H26, L(12)=1H27 ... AC(29)=2H35

# ============================================================
# P&L SECTION - rewrite labels
# ============================================================

# R5: P&L header
# R6: Revenue header

# R7: Total Revenue (statutory)
ws.cell(7, 1).value = 'Rev-Total Revenue'
ws.cell(7, 2).value = 'Total Revenue'
ws.cell(7, 3).value = 'AUDm'

# R8: Other Revenue
ws.cell(8, 1).value = 'Rev-Other Revenue'
ws.cell(8, 2).value = 'Other Revenue & Income'
ws.cell(8, 3).value = 'AUDm'

# R9: Total Revenue incl Other
ws.cell(9, 1).value = 'Rev-Total Rev Incl Other'
ws.cell(9, 2).value = 'Total Revenue (incl Other)'
ws.cell(9, 3).value = 'AUDm'

ws.cell(10, 2).value = 'Revenue Growth (YoY)'
ws.cell(10, 3).value = '% YoY'

# R12: COGS header
ws.cell(13, 1).value = 'COGS-Food & Packaging'
ws.cell(13, 2).value = 'Cost of Food & Packaging'
ws.cell(13, 3).value = 'AUDm'

ws.cell(14, 1).value = 'COGS-Employee Benefits'
ws.cell(14, 2).value = 'Employee Benefits'
ws.cell(14, 3).value = 'AUDm'

ws.cell(15, 1).value = 'COGS-Other Costs'
ws.cell(15, 2).value = 'Admin, Marketing & Other'
ws.cell(15, 3).value = 'AUDm'

ws.cell(16, 1).value = 'COGS-Total Costs'
ws.cell(16, 2).value = 'Total Operating Costs'
ws.cell(16, 3).value = 'AUDm'

# R18: Stat EBITDA header
ws.cell(18, 2).value = 'Statutory EBITDA'

ws.cell(19, 1).value = 'GP-Statutory EBITDA'
ws.cell(19, 2).value = 'Statutory EBITDA'
ws.cell(19, 3).value = 'AUDm'

ws.cell(20, 2).value = None; ws.cell(20, 3).value = None; ws.cell(20, 1).value = None
ws.cell(21, 2).value = None; ws.cell(21, 3).value = None; ws.cell(21, 1).value = None
ws.cell(22, 2).value = None; ws.cell(22, 3).value = None; ws.cell(22, 1).value = None
ws.cell(23, 2).value = None; ws.cell(23, 3).value = None
ws.cell(24, 2).value = None; ws.cell(24, 3).value = None
ws.cell(25, 2).value = None; ws.cell(25, 3).value = None
ws.cell(26, 2).value = None; ws.cell(26, 3).value = None

# R28: OpEx -> not needed, clear
ws.cell(28, 2).value = 'Segment EBITDA'
for r in range(29, 37):
    ws.cell(r, 1).value = None; ws.cell(r, 2).value = None; ws.cell(r, 3).value = None

# R37: EBITDA section header
ws.cell(37, 2).value = 'Segment EBITDA'

ws.cell(38, 1).value = 'EBITDA-Australia EBITDA'
ws.cell(38, 2).value = 'Australia Segment EBITDA'
ws.cell(38, 3).value = 'AUDm'

ws.cell(39, 1).value = 'EBITDA-US EBITDA'
ws.cell(39, 2).value = 'US Segment EBITDA'
ws.cell(39, 3).value = 'AUDm'

ws.cell(40, 1).value = 'EBITDA-Corporate EBITDA'
ws.cell(40, 2).value = 'Corporate / Eliminations'
ws.cell(40, 3).value = 'AUDm'

ws.cell(41, 2).value = None; ws.cell(41, 3).value = None

ws.cell(42, 1).value = 'EBITDA-Underlying EBITDA'
ws.cell(42, 2).value = 'Group Segment Underlying EBITDA'
ws.cell(42, 3).value = 'AUDm'

ws.cell(43, 2).value = 'Seg EBITDA Growth (YoY)'
ws.cell(43, 3).value = '% YoY'

ws.cell(44, 2).value = 'Seg EBITDA Margin'
ws.cell(44, 3).value = '%'
ws.cell(45, 2).value = None; ws.cell(45, 3).value = None
ws.cell(46, 2).value = None; ws.cell(46, 3).value = None

# Statutory adjustments
ws.cell(48, 1).value = 'Stat-AASB16'
ws.cell(48, 2).value = 'AASB 16 Lease Impact'
ws.cell(48, 3).value = 'AUDm'

ws.cell(49, 1).value = 'Stat-SBP'
ws.cell(49, 2).value = 'Share-Based Payments'
ws.cell(49, 3).value = 'AUDm'

# Add Other Adj row - repurpose R50
# But R49 was Stat-Statutory EBITDA in VSL. Need to shift.
# Actually R48=Sig Items, R49=Stat EBITDA in VSL. We have:
# R48=AASB16, R49=SBP. Need R50 for Other, R51 for Stat EBITDA.
# But R51 was D&A header. Let me use:
# R48=AASB16, R49=SBP+Other combined, and compute Stat EBITDA differently
# Or: just combine SBP and Other into one line
# Actually, let me keep it simple: R48=AASB16, R49=Stat-Statutory EBITDA
# and put SBP+Other into a note or compute from the bridge

# Revised: R48=AASB16, R49=Stat-Statutory EBITDA (= Seg EBITDA + AASB16 - SBP + Other)
ws.cell(48, 1).value = 'Stat-AASB16'
ws.cell(48, 2).value = 'AASB 16 / SBP / Other Adjustments'
ws.cell(48, 3).value = 'AUDm'

ws.cell(49, 1).value = 'Stat-Statutory EBITDA'
ws.cell(49, 2).value = 'Statutory EBITDA'
ws.cell(49, 3).value = 'AUDm'

# D&A section
ws.cell(51, 2).value = 'D&A'
ws.cell(52, 1).value = 'DA-Total DA'
ws.cell(52, 2).value = 'Depreciation & Amortisation'
ws.cell(52, 3).value = 'AUDm'

ws.cell(53, 1).value = None; ws.cell(53, 2).value = None; ws.cell(53, 3).value = None
ws.cell(54, 1).value = None; ws.cell(54, 2).value = None; ws.cell(54, 3).value = None
ws.cell(55, 2).value = 'D&A / Revenue'; ws.cell(55, 3).value = '%'

# EBIT
ws.cell(57, 2).value = 'EBIT'
ws.cell(58, 1).value = 'EBIT-EBIT'
ws.cell(58, 2).value = 'EBIT (Operating Profit)'
ws.cell(58, 3).value = 'AUDm'

ws.cell(59, 2).value = 'EBIT Growth (YoY)'; ws.cell(59, 3).value = '% YoY'
ws.cell(60, 2).value = 'EBIT Margin'; ws.cell(60, 3).value = '%'

# Interest
ws.cell(62, 2).value = 'Interest'
ws.cell(63, 1).value = 'Int-Finance Income'
ws.cell(63, 2).value = 'Finance Income'
ws.cell(63, 3).value = 'AUDm'

ws.cell(64, 1).value = 'Int-Finance Costs'
ws.cell(64, 2).value = 'Finance Costs'
ws.cell(64, 3).value = 'AUDm'

ws.cell(65, 1).value = None; ws.cell(65, 2).value = None; ws.cell(65, 3).value = None

ws.cell(66, 1).value = 'Int-Net Finance Costs'
ws.cell(66, 2).value = 'Net Finance Costs'
ws.cell(66, 3).value = 'AUDm'

ws.cell(67, 2).value = None; ws.cell(67, 3).value = None
ws.cell(68, 2).value = None; ws.cell(68, 3).value = None
ws.cell(69, 2).value = None; ws.cell(69, 3).value = None

# PBT, Tax, NPAT
ws.cell(71, 1).value = 'PBT-PBT'
ws.cell(71, 2).value = 'PBT'
ws.cell(71, 3).value = 'AUDm'

ws.cell(72, 1).value = 'Tax-Tax Expense'
ws.cell(72, 2).value = 'Tax Expense'
ws.cell(72, 3).value = 'AUDm'

ws.cell(73, 2).value = 'Tax Rate'; ws.cell(73, 3).value = '%'

ws.cell(74, 1).value = None; ws.cell(74, 2).value = None; ws.cell(74, 3).value = None

ws.cell(75, 1).value = 'NPAT-Statutory NPAT'
ws.cell(75, 2).value = 'NPAT'
ws.cell(75, 3).value = 'AUDm'

ws.cell(76, 1).value = None; ws.cell(76, 2).value = None; ws.cell(76, 3).value = None
ws.cell(77, 1).value = None; ws.cell(77, 2).value = None; ws.cell(77, 3).value = None
ws.cell(78, 2).value = 'NPAT Growth (YoY)'; ws.cell(78, 3).value = '% YoY'

# KPIs
ws.cell(80, 2).value = 'Operating Metrics'

ws.cell(81, 1).value = 'KPI-Total Restaurants'
ws.cell(81, 2).value = 'Total Restaurants'
ws.cell(81, 3).value = '#'

ws.cell(82, 1).value = 'KPI-Aus Corp Restaurants'
ws.cell(82, 2).value = 'Aus Corporate Restaurants'
ws.cell(82, 3).value = '#'

ws.cell(83, 1).value = 'KPI-Aus Fran Restaurants'
ws.cell(83, 2).value = 'Aus Franchise Restaurants'
ws.cell(83, 3).value = '#'

ws.cell(84, 1).value = 'KPI-US Restaurants'
ws.cell(84, 2).value = 'US Restaurants'
ws.cell(84, 3).value = '#'

ws.cell(85, 1).value = 'KPI-Aus Network Sales'
ws.cell(85, 2).value = 'Australia Network Sales'
ws.cell(85, 3).value = 'AUDm'

ws.cell(86, 1).value = 'KPI-US Network Sales'
ws.cell(86, 2).value = 'US Network Sales'
ws.cell(86, 3).value = 'AUDm'

ws.cell(87, 1).value = 'KPI-Comp Sales Growth'
ws.cell(87, 2).value = 'Comp Sales Growth'
ws.cell(87, 3).value = '%'

ws.cell(88, 1).value = None; ws.cell(88, 2).value = None; ws.cell(88, 3).value = None
ws.cell(89, 1).value = None; ws.cell(89, 2).value = None; ws.cell(89, 3).value = None
ws.cell(90, 1).value = None; ws.cell(90, 2).value = None; ws.cell(90, 3).value = None

# ============================================================
# SEGMENT FORECAST SECTIONS (replacing Steel/Metals)
# ============================================================

# R92: Australia Segment Forecast header
ws.cell(92, 2).value = 'Segment Forecast - Australia'

ws.cell(93, 2).value = 'Drive Thru Restaurants (#)'; ws.cell(93, 3).value = '#'
ws.cell(94, 2).value = 'Strip Restaurants (#)'; ws.cell(94, 3).value = '#'
ws.cell(95, 2).value = 'Other Restaurants (#)'; ws.cell(95, 3).value = '#'
ws.cell(96, 2).value = 'Total Australia Restaurants (#)'; ws.cell(96, 3).value = '#'
ws.cell(97, 2).value = 'DT AUV ($m)'; ws.cell(97, 3).value = 'AUDm'
ws.cell(98, 2).value = 'Strip AUV ($m)'; ws.cell(98, 3).value = 'AUDm'
ws.cell(99, 2).value = 'Other AUV ($m)'; ws.cell(99, 3).value = 'AUDm'
ws.cell(100, 2).value = 'Network Sales - Drive Thru'; ws.cell(100, 3).value = 'AUDm'
ws.cell(101, 2).value = 'Network Sales - Strip'; ws.cell(101, 3).value = 'AUDm'
ws.cell(102, 2).value = 'Network Sales - Other'; ws.cell(102, 3).value = 'AUDm'
ws.cell(103, 2).value = 'Total Australia Network Sales'; ws.cell(103, 3).value = 'AUDm'
ws.cell(104, 2).value = 'Corporate Restaurant Sales'; ws.cell(104, 3).value = 'AUDm'
ws.cell(105, 2).value = 'Corporate Restaurant Margin (%)'; ws.cell(105, 3).value = '%'
ws.cell(106, 2).value = 'Corporate Restaurant Margin ($)'; ws.cell(106, 3).value = 'AUDm'
ws.cell(107, 2).value = 'Franchise Network Sales'; ws.cell(107, 3).value = 'AUDm'
ws.cell(108, 2).value = 'Implied Franchise Royalty Rate (%)'; ws.cell(108, 3).value = '%'
ws.cell(109, 2).value = 'Franchise & Other Revenue'; ws.cell(109, 3).value = 'AUDm'
ws.cell(110, 2).value = 'G&A Costs'; ws.cell(110, 3).value = 'AUDm'
ws.cell(111, 2).value = 'G&A as % of Network Sales'; ws.cell(111, 3).value = '%'
ws.cell(112, 2).value = 'Australia Segment EBITDA'; ws.cell(112, 3).value = 'AUDm'
ws.cell(113, 2).value = 'Seg EBITDA as % of NS'; ws.cell(113, 3).value = '%'

# Clear old column A codes for segment forecast rows
for r in range(92, 117):
    ws.cell(r, 1).value = None

# R115: US Segment Forecast header (was Metals)
ws.cell(115, 2).value = 'Segment Forecast - US'
ws.cell(116, 2).value = 'US Restaurants (#)'; ws.cell(116, 3).value = '#'

# We're running out of rows (max 116). Let me check if we can extend.
# The VSL sheet had 116 rows. We need more for US forecast.
# We have rows 92-116 = 25 rows. Australia uses 92-113 = 22 rows.
# US needs: header + restaurants + NS + Corp Sales + Corp Margin % + Corp Margin $ +
# Fran Rev + G&A + Seg EBITDA = 9 rows (115-123). But sheet only goes to 116.
# Need to extend. Just write beyond max_row.

# Actually let me compress Australia to fit both in available space:
# Australia: R92 header, R93-R112 data (20 rows including EBITDA)
# US: R114 header, R115-R116 + extend to R122 (8 rows)
# This works since openpyxl will auto-extend.

ws.cell(114, 2).value = None; ws.cell(114, 3).value = None  # blank separator

ws.cell(115, 2).value = 'Segment Forecast - US'
ws.cell(116, 2).value = 'US Restaurants (#)'; ws.cell(116, 3).value = '#'
ws.cell(117, 2).value = 'US Network Sales'; ws.cell(117, 3).value = 'AUDm'
ws.cell(118, 2).value = 'Corporate Restaurant Sales'; ws.cell(118, 3).value = 'AUDm'
ws.cell(119, 2).value = 'Corporate Restaurant Margin (%)'; ws.cell(119, 3).value = '%'
ws.cell(120, 2).value = 'Corporate Restaurant Margin ($)'; ws.cell(120, 3).value = 'AUDm'
ws.cell(121, 2).value = 'Franchise & Other Revenue'; ws.cell(121, 3).value = 'AUDm'
ws.cell(122, 2).value = 'G&A Costs'; ws.cell(122, 3).value = 'AUDm'
ws.cell(123, 2).value = 'US Segment EBITDA'; ws.cell(123, 3).value = 'AUDm'

wb.save('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')
print("Part 4 complete: HY sheet labels set up")

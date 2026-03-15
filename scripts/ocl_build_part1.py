"""Part 1: Copy template, update headers, column layout, period labels."""
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date

SRC = '/home/pmwilson/Project_Equities/.claude/templates/HY_model_template.xlsx'
DST = '/home/pmwilson/Project_Equities/OCL/Models/OCL Model.xlsx'

shutil.copy2(SRC, DST)
wb = openpyxl.load_workbook(DST)

# === ANNUAL SHEET ===
ws = wb['Annual']

# Row 2: title
ws['B2'] = 'OCL Model Summary'

# Row 3: company
ws['B3'] = 'Objective Corporation (OCL.AX)'

# Column layout: FY21A-FY30E in columns D-M (10 years)
# Row 1: year integers
years = list(range(2021, 2031))  # FY21 to FY30
data_cols_annual = list(range(4, 14))  # D=4 to M=13

for i, (col, yr) in enumerate(zip(data_cols_annual, years)):
    ws.cell(row=1, column=col, value=yr)
    
    # Row 3: period labels
    if yr <= 2025:
        label = f'FY{yr-2000}A'
    else:
        label = f'FY{yr-2000}E'
    ws.cell(row=3, column=col, value=label)
    
    # Row 4: period end dates (30 June)
    ws.cell(row=4, column=col, value=datetime(yr, 6, 30))

# Clear old columns N-P (cols 14-16) that were in the 13-col template
for col in range(14, 17):
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=col).value = None

# Row 2 banners: Actual over FY21-FY25 (D-H), Forecast over FY26E-FY30E (I-M)
white_bold = Font(bold=True, color='FFFFFF')
dark_blue = PatternFill(start_color='FF002060', end_color='FF002060', fill_type='solid')
mid_blue = PatternFill(start_color='FF0070C0', end_color='FF0070C0', fill_type='solid')

# Clear row 2 data cells first
for col in range(4, 14):
    ws.cell(row=2, column=col).value = None
    ws.cell(row=2, column=col).font = Font()
    ws.cell(row=2, column=col).fill = PatternFill()

ws.cell(row=2, column=4, value='Actual ---------->')
ws.cell(row=2, column=4).font = white_bold
ws.cell(row=2, column=4).fill = dark_blue
for col in range(5, 9):  # E-H
    ws.cell(row=2, column=col).fill = dark_blue

ws.cell(row=2, column=9, value='Forecast ----->')
ws.cell(row=2, column=9).font = white_bold
ws.cell(row=2, column=9).fill = mid_blue
for col in range(10, 14):  # J-M
    ws.cell(row=2, column=col).fill = mid_blue

# Update unit labels to A$m
for row in range(1, ws.max_row + 1):
    c = ws.cell(row=row, column=3).value
    if c and 'NZDm' in str(c):
        ws.cell(row=row, column=3, value='A$m')
    elif c and 'NZDps' in str(c):
        ws.cell(row=row, column=3, value='A$ps')
    elif c and 'NZD' in str(c) and 'NZDm' not in str(c) and 'NZDps' not in str(c):
        ws.cell(row=row, column=3, value=str(c).replace('NZD', 'A$'))

# === HY & SEGMENTS SHEET ===
ws2 = wb['HY & Segments']

ws2['B2'] = 'OCL Segments (Half-Year)'
ws2['B3'] = 'Objective Corporation (OCL.AX)'

# Column layout: 1H21-2H30 = 20 half-year columns in D-W
# Template had D-AC (26 cols for 13 years). We need 10 years = 20 cols D-W
hy_years = list(range(2021, 2031))
col_idx = 4  # start at D
hy_col_map = {}  # (half, year) -> col

for yr in hy_years:
    yr2 = yr - 2000
    # 1H
    ws2.cell(row=1, column=col_idx, value=yr)
    ws2.cell(row=3, column=col_idx, value=f'1H{yr2}')
    ws2.cell(row=4, column=col_idx, value=datetime(yr-1, 12, 31))  # 1H ends 31 Dec prior year
    hy_col_map[('1H', yr)] = col_idx
    col_idx += 1
    
    # 2H
    ws2.cell(row=1, column=col_idx, value=yr)
    ws2.cell(row=3, column=col_idx, value=f'2H{yr2}')
    ws2.cell(row=4, column=col_idx, value=datetime(yr, 6, 30))  # 2H ends 30 June
    hy_col_map[('2H', yr)] = col_idx
    col_idx += 1

# Clear old columns beyond W (col 23) up to AC (col 29)
for col in range(24, 30):
    for row in range(1, ws2.max_row + 1):
        ws2.cell(row=row, column=col).value = None

# Row 2 banners: Actual 1H21-1H26 (D-K), Forecast from 2H26E (L onwards)
# 1H21=D, 2H21=E, 1H22=F, 2H22=G, 1H23=H, 2H23=I, 1H24=J, 2H24=K, 1H25=L, 2H25=M
# 1H26=N, 2H26=O... wait let me recalculate
# D=1H21, E=2H21, F=1H22, G=2H22, H=1H23, I=2H23, J=1H24, K=2H24
# L=1H25, M=2H25, N=1H26, O=2H26E, P=1H27E, Q=2H27E...
# Actuals: 1H21 through 1H26 = D through N (11 cols)
# Forecast: 2H26E through 2H30E = O through W (9 cols)

for col in range(4, 24):
    ws2.cell(row=2, column=col).value = None
    ws2.cell(row=2, column=col).font = Font()
    ws2.cell(row=2, column=col).fill = PatternFill()

ws2.cell(row=2, column=4, value='Actual ---------->')
ws2.cell(row=2, column=4).font = white_bold
ws2.cell(row=2, column=4).fill = dark_blue
for col in range(5, 15):  # E through N (1H26)
    ws2.cell(row=2, column=col).fill = dark_blue

ws2.cell(row=2, column=15, value='Forecast ----->')  # O = 2H26E
ws2.cell(row=2, column=15).font = white_bold
ws2.cell(row=2, column=15).fill = mid_blue
for col in range(16, 24):  # P through W
    ws2.cell(row=2, column=col).fill = mid_blue

# Update unit labels on HY sheet
for row in range(1, ws2.max_row + 1):
    c = ws2.cell(row=row, column=3).value
    if c and 'NZDm' in str(c):
        ws2.cell(row=row, column=3, value='A$m')
    elif c and 'NZDps' in str(c):
        ws2.cell(row=row, column=3, value='A$ps')
    elif c and 'NZD' in str(c):
        ws2.cell(row=row, column=3, value=str(c).replace('NZD', 'A$'))

wb.save(DST)
print('Part 1 complete: template copied, headers and column layout updated')
print(f'Annual: {len(data_cols_annual)} columns (FY21-FY30)')
print(f'HY: {len(hy_col_map)} half-year columns (1H21-2H30)')

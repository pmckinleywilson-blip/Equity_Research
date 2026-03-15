"""
GYG Annual Sheet Structure Build
================================
Modifies the VSL template Annual sheet for GYG (Guzman y Gomez).
Uses INSERT and DELETE approach — never clears/rebuilds.

Strategy: Do all insertions bottom-to-top first, then deletions top-to-bottom.
This preserves row references during the process.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from copy import copy
import datetime

# ─── Load workbook ───
SRC = '/home/pmwilson/Project_Equities/GYG/Models/GYG Model.xlsx'
wb = openpyxl.load_workbook(SRC)
ws = wb['Annual']

MAX_COL = 16  # Column P

# ─── Helper functions ───

def copy_cell_style(src_cell, dst_cell):
    """Copy all formatting from src to dst."""
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.number_format = src_cell.number_format
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.protection = copy(src_cell.protection)


def copy_row_format(ws, src_row, dst_row, max_col=MAX_COL):
    """Copy formatting from src_row to dst_row for all columns."""
    for col in range(1, max_col + 1):
        copy_cell_style(ws.cell(src_row, col), ws.cell(dst_row, col))


def insert_data_row(ws, row, key, label, unit, fmt_source_row, max_col=MAX_COL):
    """Set up a data row: col A = key, col B = label, col C = unit, format from source."""
    copy_row_format(ws, fmt_source_row, row, max_col)
    ws.cell(row, 1).value = key
    ws.cell(row, 2).value = label
    ws.cell(row, 3).value = unit


def insert_header_row(ws, row, label, fmt_source_row, max_col=MAX_COL):
    """Set up a section header row: col B = label (bold), format from source."""
    copy_row_format(ws, fmt_source_row, row, max_col)
    ws.cell(row, 2).value = label
    # Ensure bold
    f = copy(ws.cell(row, 2).font)
    ws.cell(row, 2).font = Font(name=f.name, size=f.sz, bold=True, italic=f.i,
                                 color=f.color, family=f.family)


def insert_subtotal_row(ws, row, key, label, unit, fmt_source_row, max_col=MAX_COL):
    """Set up a subtotal row (bold, thin bottom border on data cols)."""
    copy_row_format(ws, fmt_source_row, row, max_col)
    ws.cell(row, 1).value = key
    ws.cell(row, 2).value = label
    ws.cell(row, 3).value = unit
    # Ensure bold B
    f = copy(ws.cell(row, 2).font)
    ws.cell(row, 2).font = Font(name=f.name, size=f.sz, bold=True, italic=f.i,
                                 color=f.color, family=f.family)
    # Ensure thin bottom border on data cols
    for col in range(4, max_col + 1):
        old_border = ws.cell(row, col).border
        ws.cell(row, col).border = Border(
            left=copy(old_border.left),
            right=copy(old_border.right),
            top=copy(old_border.top),
            bottom=Side(style='thin'),
            diagonal=copy(old_border.diagonal)
        )


def insert_ratio_row(ws, row, label, unit, fmt_source_row, max_col=MAX_COL):
    """Set up an analytical/ratio row (no key in col A)."""
    copy_row_format(ws, fmt_source_row, row, max_col)
    ws.cell(row, 1).value = None
    ws.cell(row, 2).value = label
    ws.cell(row, 3).value = unit


def insert_blank_row(ws, row, fmt_source_row=None, max_col=MAX_COL):
    """Clear a row (or copy format from source if given)."""
    if fmt_source_row:
        copy_row_format(ws, fmt_source_row, row, max_col)
    for col in range(1, max_col + 1):
        ws.cell(row, col).value = None


def set_sum_formula(ws, row, sum_start, sum_end, col_start=4, col_end=MAX_COL):
    """Set SUM formula across data columns."""
    for col in range(col_start, col_end + 1):
        cl = get_column_letter(col)
        ws.cell(row, col).value = f'=SUM({cl}{sum_start}:{cl}{sum_end})'


def rename_row(ws, row, key=None, label=None, unit=None):
    """Rename key/label/unit on an existing row (None = don't change)."""
    if key is not None:
        ws.cell(row, 1).value = key
    if label is not None:
        ws.cell(row, 2).value = label
    if unit is not None:
        ws.cell(row, 3).value = unit


def clear_data(ws, row, col_start=4, col_end=MAX_COL):
    """Clear data values from a row but keep formatting."""
    for col in range(col_start, col_end + 1):
        ws.cell(row, col).value = None


# ═══════════════════════════════════════════════════════════════════
# PHASE 1: INSERT NEW COLUMN (for FY22A)
# ═══════════════════════════════════════════════════════════════════
# Current: D=FY23A, E=FY24A, F=FY25A, G=FY26E ... P=FY35E
# Need:    D=FY22A, E=FY23A, F=FY24A, G=FY25A, H=FY26E ... Q=FY35E
# Insert column at D, shifting everything right

ws.insert_cols(4, 1)
MAX_COL = 17  # Now extends to column Q

# Set up the new column D (FY22A) headers
ws.cell(1, 4).value = 2022
ws.cell(3, 4).value = 'FY22A'
ws.cell(4, 4).value = datetime.datetime(2022, 6, 30)

# Copy formatting from col E (which was the old col D) for header rows
for row in range(1, 5):
    copy_cell_style(ws.cell(row, 5), ws.cell(row, 4))

# Copy formatting from col E for ALL data rows
for row in range(5, ws.max_row + 1):
    copy_cell_style(ws.cell(row, 5), ws.cell(row, 4))
    # Clear the data in new col D (it's a new actuals column, no data yet)
    ws.cell(row, 4).value = None

# Restore header values that got cleared
ws.cell(1, 4).value = 2022
ws.cell(3, 4).value = 'FY22A'
ws.cell(4, 4).value = datetime.datetime(2022, 6, 30)

# Update row 2: "Actual ---------->" now spans D-G (4 cols), "Forecast ----->" starts at H
# Old: D2 = "Actual ---------->", H2 = "Forecast ----->"
# After insert: D2 still has it, but forecast label shifted to I. Fix:
ws.cell(2, 4).value = 'Actual ---------->'
ws.cell(2, 5).value = None
ws.cell(2, 6).value = None
ws.cell(2, 7).value = None
# Find and set forecast label - it should be at col H now (was G, shifted to H)
ws.cell(2, 8).value = 'Forecast ----->'
# Clear any stale label
ws.cell(2, 9).value = None

# Update year headers for shifted columns (E onwards should be correct from the shift,
# but let me verify and set explicitly)
years_and_labels = [
    (4, 2022, 'FY22A', datetime.datetime(2022, 6, 30)),
    (5, 2023, 'FY23A', datetime.datetime(2023, 6, 30)),
    (6, 2024, 'FY24A', datetime.datetime(2024, 6, 30)),
    (7, 2025, 'FY25A', datetime.datetime(2025, 6, 30)),
    (8, 2026, 'FY26E', datetime.datetime(2026, 6, 30)),
    (9, 2027, 'FY27E', datetime.datetime(2027, 6, 30)),
    (10, 2028, 'FY28E', datetime.datetime(2028, 6, 30)),
    (11, 2029, 'FY29E', datetime.datetime(2029, 6, 30)),
    (12, 2030, 'FY30E', datetime.datetime(2030, 6, 30)),
    (13, 2031, 'FY31E', datetime.datetime(2031, 6, 30)),
    (14, 2032, 'FY32E', datetime.datetime(2032, 6, 30)),
    (15, 2033, 'FY33E', datetime.datetime(2033, 6, 30)),
    (16, 2034, 'FY34E', datetime.datetime(2034, 6, 30)),
    (17, 2035, 'FY35E', datetime.datetime(2035, 6, 30)),
]
for col, year, label, dt in years_and_labels:
    ws.cell(1, col).value = year
    ws.cell(3, col).value = label
    ws.cell(4, col).value = dt

print("Phase 1 complete: Column inserted for FY22A")

# ═══════════════════════════════════════════════════════════════════
# PHASE 2: ROW INSERTIONS (bottom to top to preserve references)
# ═══════════════════════════════════════════════════════════════════
# Track current row positions as we insert. Start from bottom.

# Current state after column insert (row numbers unchanged):
# Row 7: Rev-Steel Revenue
# Row 8: Rev-Metals Revenue
# Row 9: Rev-Total Revenue
# Row 10: Revenue Growth
# ...
# Row 94: Operating Metrics header
# Row 95-106: KPI rows

# --- INSERT PLAN (bottom to top) ---
# I'll track cumulative row shifts as we go.

# ─── BS insertions (rows 110-area) ───
# After col insert, BS rows are unchanged. Let me do BS inserts first since they're lowest.

# BS-Contract Liabilities: insert after Trade Payables (row 125), before Other Liabilities (row 126)
# But first I need to find where things are after any shifts. Let's go strictly bottom-to-top.

# 1. BS: Insert "BS-DTA" after Intangibles (row 114) → insert at row 115, shifts ROU Assets etc down
#    But I also need "Finance Lease Receivables" after Inventories (row 112) → insert at row 113
#    And "Term Deposits" after Cash (row 110) → insert at row 111
#    And "Contract Liabilities" after Trade Payables
#
# Let me do these bottom-to-top within BS:

# a) Contract Liabilities: after Trade Payables (currently row 125)
#    Insert 1 row at 126 (before current Other Liabilities at 126)
ws.insert_rows(126, 1)
# Now: row 126 = new (Contract Liabilities), row 127 = Other Liabilities, etc.
# Everything below shifted by 1
insert_data_row(ws, 126, 'BS-Contract Liabilities', 'Contract Liabilities', 'AUDm', 125, MAX_COL)
clear_data(ws, 126)
print("Inserted BS-Contract Liabilities at row 126")

# b) DTA: after Intangibles (row 114) → insert at 115
ws.insert_rows(115, 1)
# Now: row 115 = new (DTA), row 116 = ROU Assets, etc.
# Everything below shifted by 1 more (cumulative +2 from original)
insert_data_row(ws, 115, 'BS-DTA', 'Deferred Tax Assets', 'AUDm', 114, MAX_COL)
clear_data(ws, 115)
print("Inserted BS-DTA at row 115")

# c) Finance Lease Receivables: after Inventories (row 112, still at 112) → insert at 113
ws.insert_rows(113, 1)
# Now: row 113 = new, row 114 = PPE (was 113), etc.
# Cumulative +3 from original below this point
insert_data_row(ws, 113, 'BS-Finance Lease Rec', 'Finance Lease Receivables', 'AUDm', 112, MAX_COL)
clear_data(ws, 113)
print("Inserted BS-Finance Lease Rec at row 113")

# d) Term Deposits: after Cash (row 110) → insert at 111
ws.insert_rows(111, 1)
# Now: row 111 = new, row 112 = Trade Receivables (was 111), etc.
# Cumulative +4 from original below this point
insert_data_row(ws, 111, 'BS-Term Deposits', 'Funds in Term Deposits', 'AUDm', 110, MAX_COL)
clear_data(ws, 111)
print("Inserted BS-Term Deposits at row 111")

# ─── KPIs section: delete old and insert new ───
# Original KPI rows were 95-106 (12 rows). After +4 shifts, they're at 99-110.
# Wait no — the BS inserts were all BELOW row 94 (the Operating Metrics header).
# Actually row 94 is ABOVE row 110. Let me re-check.
# Row 94 = Operating Metrics header, rows 95-106 = KPIs, row 107 = blank, row 108 = BS header
# The BS inserts at rows 111, 113, 115, 126 are all below row 108.
# So KPI rows 95-106 are NOT shifted by BS inserts. Correct.
# But BS insert at 111 shifts everything at 111+ down. 113 shifts 113+ down, etc.
# KPIs are at 95-106, which is ABOVE 111. So KPIs are unaffected. Good.

# For KPIs, I'll DELETE old rows then INSERT new ones.
# Current KPI rows: 95-106 (12 rows). Delete them, then insert new GYG KPIs.
# Actually, let me handle KPIs during the deletion phase and just insert what I need.
# For now, let me continue with other insertions.

# ─── Interest section insertions ───
# Original rows 58-65. These are above row 94, so unaffected by BS inserts.
# Need to insert:
#   - "Int-Lease Receivable" after Int-Interest Income (row 59) → at row 60
#   - "Int-Other Income" after that → at row 61
# Current:
#   59: Int-Interest Income
#   60: Int-Lease Interest
#   61: Int-Bank Interest
#   62: Net Finance Costs
#   63: Interest Income Rate
#   64: Lease Interest Rate
#   65: Bank Interest Rate

# Insert 2 rows at 60 (shifting Lease Interest and below down)
ws.insert_rows(60, 2)
# Now:
#   59: Int-Interest Income (unchanged)
#   60: NEW - Lease Receivable Interest
#   61: NEW - Other Finance Income
#   62: Int-Lease Interest (was 60)
#   63: Int-Bank Interest (was 61)
#   64: Net Finance Costs (was 62)
#   65: Interest Income Rate (was 63)
#   66: Lease Interest Rate (was 64)
#   67: Bank Interest Rate (was 65)

insert_data_row(ws, 60, 'Int-Lease Receivable', 'Lease Receivable Interest', 'AUDm', 59, MAX_COL)
clear_data(ws, 60)
insert_data_row(ws, 61, 'Int-Other Income', 'Other Finance Income', 'AUDm', 59, MAX_COL)
clear_data(ws, 61)
print("Inserted Int-Lease Receivable at row 60, Int-Other Income at row 61")

# Everything below shifted +2 more. Cumulative shifts from original:
# Rows 60+ shifted by +2 (interest inserts)
# Rows 95+ (KPIs): now at 97+ (wait, 95 is above? No, 95 > 60. So yes, +2.)
# Actually let me just track carefully. After interest inserts:
# Original row 60 → now row 62 (Int-Lease Interest)
# Original row 94 (Operating Metrics) → now row 96
# Original row 95-106 → now rows 97-108
# Original row 110 (Cash) → now row 112 (was already shifted by previous BS insert at 111 → 113?)
# Wait, I need to be more careful. The BS inserts happened first, then the interest inserts.
# The interest inserts at row 60 shift everything at row 60 and below by 2.
# So BS rows (which were already shifted) get shifted by 2 more.
# This is getting complex. Let me just track by checking the sheet state later.

# ─── D&A section insertions ───
# Original rows 47-52. After interest insert (+2 at row 60), rows 47-52 are ABOVE 60, so unaffected.
# Need to insert after ROU Amortisation (row 49):
#   - "DA-Reacq Amort"
#   - "DA-Other Amort"
# Insert 2 rows at 50
ws.insert_rows(50, 2)
# Now:
#   48: DA-Depreciation PPE (unchanged)
#   49: DA-ROU Amortisation (unchanged)
#   50: NEW - Reacq Amort
#   51: NEW - Other Amort
#   52: DA-Total DA (was 50)
#   53: D&A / Revenue (was 51)
#   54: Avg Lease Life (was 52)

insert_data_row(ws, 50, 'DA-Reacq Amort', 'Amortisation of Reacquired Rights', 'AUDm', 49, MAX_COL)
clear_data(ws, 50)
insert_data_row(ws, 51, 'DA-Other Amort', 'Other Amortisation', 'AUDm', 49, MAX_COL)
clear_data(ws, 51)
print("Inserted DA-Reacq Amort at row 50, DA-Other Amort at row 51")

# ─── EBITDA section insertions ───
# After D&A inserts (+2 at row 50), rows above 50 unchanged.
# Original EBITDA section was rows 34-45. After D&A insert, rows 34-45 are unaffected
# (insert was at row 50, which is BELOW row 45).
# Wait — original row 50 was DA-Total DA. The insert at row 50 pushed it to 52.
# But EBITDA rows 34-45 are all above row 47, so unaffected. Good.

# In the EBITDA section, I need to insert AFTER the Stat-Significant Items row (row 44).
# After EBITDA adjustments:
#   42: Statutory EBITDA Adjustments header
#   43: Stat-SBP
#   44: Stat-Significant Items → will be renamed
#   45: Stat-Statutory EBITDA → will be renamed to Group Segment EBITDA
#
# Insert 3 rows before row 45 (before Stat-Statutory EBITDA):
#   - Stat-Cash Rent
#   - EBITDA-AU Segment EBITDA
#   - EBITDA-US Segment EBITDA
# Then after the (shifted) Stat-Statutory EBITDA:
#   - Analytical row: Segment EBITDA / Network Sales

ws.insert_rows(45, 3)
# Now:
#   42: header
#   43: Stat-SBP
#   44: Stat-Significant Items
#   45: NEW - Cash Rent
#   46: NEW - AU Segment EBITDA
#   47: NEW - US Segment EBITDA
#   48: Stat-Statutory EBITDA (was 45) → will become Group Segment EBITDA
# Everything below shifts +3

insert_data_row(ws, 45, 'Stat-Cash Rent', 'Cash Rent (AASB 16 Reversal)', 'AUDm', 43, MAX_COL)
clear_data(ws, 45)
insert_data_row(ws, 46, 'EBITDA-AU Segment EBITDA', 'AU Segment Underlying EBITDA', 'AUDm', 43, MAX_COL)
clear_data(ws, 46)
insert_data_row(ws, 47, 'EBITDA-US Segment EBITDA', 'US Segment Underlying EBITDA', 'AUDm', 43, MAX_COL)
clear_data(ws, 47)
print("Inserted Cash Rent, AU Segment EBITDA, US Segment EBITDA at rows 45-47")

# Now insert analytical row after Group Segment EBITDA (row 48)
ws.insert_rows(49, 1)
# Row 49 = new analytical row
insert_ratio_row(ws, 49, 'Segment EBITDA / Network Sales', '%', 44, MAX_COL)  # Format like a % row
clear_data(ws, 49)
# Set % number format
for col in range(4, MAX_COL + 1):
    ws.cell(49, col).number_format = '0.0%'
print("Inserted Segment EBITDA / Network Sales ratio at row 49")

# ─── Revenue section: insert 2 rows for Franchise Royalty and Franchise Fee ───
# Original rows 7-9: Steel Rev, Metals Rev, Total Revenue
# Insert 2 rows at row 9 (before Total Revenue), shifting Total Revenue down
ws.insert_rows(9, 2)
# Now:
#   7: Rev-Steel Revenue (will rename)
#   8: Rev-Metals Revenue (will rename)
#   9: NEW - Franchise Royalty
#  10: NEW - Franchise Fee
#  11: Rev-Total Revenue (was 9)
#  12: Revenue Growth (was 10)

insert_data_row(ws, 9, 'Rev-Franchise Royalty', 'Franchise Royalty Revenue', 'AUDm', 7, MAX_COL)
clear_data(ws, 9)
insert_data_row(ws, 10, 'Rev-Franchise Fee', 'Franchise Fee Revenue', 'AUDm', 7, MAX_COL)
clear_data(ws, 10)
print("Inserted Rev-Franchise Royalty and Rev-Franchise Fee at rows 9-10")

# ─── Other Revenue section: insert after Revenue Growth ───
# Revenue Growth is now at row 12. Insert block after it.
# Need: blank row (13 is already blank from original row 11), then Other Revenue section.
# Original row 11 was blank, now at row 13. So row 13 = blank. Good.
# Insert Other Revenue section at row 14 (before current COGS header which was row 12, now at 14).
# Wait, let me recalculate:
# After Revenue inserts (+2 at row 9):
#   Original row 10 (Revenue Growth) → now row 12
#   Original row 11 (blank) → now row 13
#   Original row 12 (COGS header) → now row 14
#
# I need to insert the Other Revenue section between the blank (row 13) and COGS (row 14).
# Other Revenue = 6 rows: header, 3 data, subtotal, blank

ws.insert_rows(14, 6)
# Now rows 14-19 are new, row 20 = COGS header
insert_header_row(ws, 14, 'Other Revenue', 8, MAX_COL)  # Format like Revenue header (row 8 area)
# Actually use row 6 format for header (section header style). But row 6 may have shifted.
# Row 6 = "Revenue" header — was not shifted (insert was at row 9, above row 6? No, row 6 < 9).
# Row 6 is ABOVE the insert point, so it's unaffected. Good.
insert_header_row(ws, 14, 'Other Revenue', 6, MAX_COL)
insert_data_row(ws, 15, 'Rev-Marketing Levy', 'Marketing Levy Revenue', 'AUDm', 7, MAX_COL)
clear_data(ws, 15)
insert_data_row(ws, 16, 'Rev-Other Franchise', 'Other Franchise Revenue', 'AUDm', 7, MAX_COL)
clear_data(ws, 16)
insert_data_row(ws, 17, 'Rev-Other Income', 'Other Income', 'AUDm', 7, MAX_COL)
clear_data(ws, 17)
insert_subtotal_row(ws, 18, 'Rev-Total Other Revenue', 'Total Other Revenue', 'AUDm', 11, MAX_COL)
clear_data(ws, 18)
# Set SUM formula for Total Other Revenue
set_sum_formula(ws, 18, 15, 17, 4, MAX_COL)
insert_blank_row(ws, 19)
print("Inserted Other Revenue section at rows 14-19")

# ═══════════════════════════════════════════════════════════════════
# PHASE 3: DELETIONS (top to bottom)
# ═══════════════════════════════════════════════════════════════════
# Need to figure out current row positions after all insertions.
# Let me print current state to verify before deleting.

print("\n--- Current state after insertions ---")
for row in range(1, 30):
    a = ws.cell(row, 1).value
    b = ws.cell(row, 2).value
    c = ws.cell(row, 3).value
    print(f"Row {row:3d}: A={a!r:35s} B={b!r:40s} C={c!r}")

# Now I need to find and delete specific rows. Let me scan the whole sheet.
print("\n--- Full scan for keys ---")
row_map = {}
for row in range(1, ws.max_row + 1):
    a = ws.cell(row, 1).value
    b = ws.cell(row, 2).value
    if a or b:
        row_map[row] = (a, b)

for row, (a, b) in sorted(row_map.items()):
    print(f"Row {row:3d}: A={a!r:40s} B={b!r}")

# Save intermediate state for debugging
# wb.save(SRC.replace('.xlsx', '_debug.xlsx'))

# ═══════════════════════════════════════════════════════════════════
# Let me be very precise about what to delete.
# I'll find rows by their column A keys.
# ═══════════════════════════════════════════════════════════════════

def find_row_by_key(ws, key):
    """Find row number by col A key."""
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == key:
            return row
    return None

def find_row_by_label(ws, label):
    """Find row number by col B label."""
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 2).value == label:
            return row
    return None

# Collect rows to delete (will delete from bottom to top)
rows_to_delete = []

# COGS: Delete COGS-Metals COGS and COGS-Other COGS
r = find_row_by_key(ws, 'COGS-Metals COGS')
if r: rows_to_delete.append(r)
r = find_row_by_key(ws, 'COGS-Other COGS')
if r: rows_to_delete.append(r)

# GP: Delete GP-Steel GP, GP-Metals GP, GP-Corporate GP
r = find_row_by_key(ws, 'GP-Steel GP')
if r: rows_to_delete.append(r)
r = find_row_by_key(ws, 'GP-Metals GP')
if r: rows_to_delete.append(r)
r = find_row_by_key(ws, 'GP-Corporate GP')
if r: rows_to_delete.append(r)

# EBITDA: Delete segment EBITDAs (Steel, Metals, Corporate)
r = find_row_by_key(ws, 'EBITDA-Steel EBITDA')
if r: rows_to_delete.append(r)
r = find_row_by_key(ws, 'EBITDA-Metals EBITDA')
if r: rows_to_delete.append(r)
r = find_row_by_key(ws, 'EBITDA-Corporate EBITDA')
if r: rows_to_delete.append(r)

# KPIs: Delete all old KPI rows (Steel Volume through EBITDA per Employee)
kpi_keys = ['KPI-Steel Volume', 'KPI-Metals Volume', 'KPI-Total Volume',
            'KPI-Steel Rev/t', 'KPI-Metals Rev/t', 'KPI-Steel GP/t',
            'KPI-Metals GP/t', 'KPI-Customers', 'KPI-Headcount', 'KPI-DIFOT']
for key in kpi_keys:
    r = find_row_by_key(ws, key)
    if r: rows_to_delete.append(r)

# Also delete the analytical rows that reference old KPIs
r = find_row_by_label(ws, 'Revenue per Employee')
if r: rows_to_delete.append(r)
r = find_row_by_label(ws, 'EBITDA per Employee')
if r: rows_to_delete.append(r)

# Sort descending and delete from bottom to top
rows_to_delete.sort(reverse=True)
print(f"\nRows to delete (bottom to top): {rows_to_delete}")

for r in rows_to_delete:
    a = ws.cell(r, 1).value
    b = ws.cell(r, 2).value
    print(f"  Deleting row {r}: A={a!r}, B={b!r}")
    ws.delete_rows(r, 1)

print(f"Deleted {len(rows_to_delete)} rows")

# ═══════════════════════════════════════════════════════════════════
# PHASE 4: INSERT NEW KPI ROWS
# ═══════════════════════════════════════════════════════════════════
# Find Operating Metrics header
op_metrics_row = find_row_by_label(ws, 'Operating Metrics')
print(f"\nOperating Metrics header at row: {op_metrics_row}")

# After deleting 12 KPI rows, the Operating Metrics header should be followed by
# the blank row and then Balance Sheet. I need to insert GYG KPIs after the header.

# Insert KPI rows. I need about 24 KPI rows.
# Network Sales (6): AU, SG, JP, US, Global, Growth
# Restaurant Counts (6): AU Corp, AU Franchise, SG, JP, US, Total
# Format Detail (6): AU DT Count, AU Strip Count, AU Other Count, DT AUV, Strip AUV, Other AUV
# Key Ratios (5): Comp Sales Growth, Corp Margin, Franchise Royalty Rate, G&A % Network Sales, Seg EBITDA % Network Sales

kpi_start = op_metrics_row + 1

# Need to insert enough rows. Currently there might be just blank/BS rows after Operating Metrics.
# Let me insert 27 rows (5 sub-headers + 22 data rows + blank separators)
# Actually let me define the exact structure:

kpi_rows_data = [
    # (type, key, label, unit)
    # type: 'header', 'data', 'subtotal', 'ratio', 'blank'
    ('header', None, 'Network Sales', None),
    ('data', 'KPI-AU Network Sales', 'AU Network Sales', 'AUDm'),
    ('data', 'KPI-SG Network Sales', 'SG Network Sales', 'AUDm'),
    ('data', 'KPI-JP Network Sales', 'JP Network Sales', 'AUDm'),
    ('data', 'KPI-US Network Sales', 'US Network Sales', 'AUDm'),
    ('subtotal', 'KPI-Global Network Sales', 'Global Network Sales', 'AUDm'),
    ('ratio', None, 'Network Sales Growth', '% YoY'),
    ('blank', None, None, None),
    ('header', None, 'Restaurant Count', None),
    ('data', 'KPI-AU Corp Count', 'AU Corporate Restaurants', '#'),
    ('data', 'KPI-AU Franchise Count', 'AU Franchise Restaurants', '#'),
    ('data', 'KPI-SG Count', 'SG Restaurants', '#'),
    ('data', 'KPI-JP Count', 'JP Restaurants', '#'),
    ('data', 'KPI-US Count', 'US Restaurants', '#'),
    ('subtotal', 'KPI-Total Restaurants', 'Total Restaurants', '#'),
    ('blank', None, None, None),
    ('header', None, 'AU Format Detail', None),
    ('data', 'KPI-AU DT Count', 'AU Drive-Thru Count', '#'),
    ('data', 'KPI-AU Strip Count', 'AU Strip Count', '#'),
    ('data', 'KPI-AU Other Count', 'AU Other Count', '#'),
    ('data', 'KPI-DT AUV', 'Drive-Thru AUV', 'AUDm'),
    ('data', 'KPI-Strip AUV', 'Strip AUV', 'AUDm'),
    ('data', 'KPI-Other AUV', 'Other AUV', 'AUDm'),
    ('blank', None, None, None),
    ('header', None, 'Key Ratios', None),
    ('ratio', None, 'Comp Sales Growth', '% YoY'),
    ('ratio', None, 'Corp Restaurant Margin', '%'),
    ('ratio', None, 'Franchise Royalty Rate', '%'),
    ('ratio', None, 'G&A % Network Sales', '%'),
    ('ratio', None, 'Segment EBITDA % Network Sales', '%'),
]

num_kpi_rows = len(kpi_rows_data)
ws.insert_rows(kpi_start, num_kpi_rows)
print(f"Inserted {num_kpi_rows} KPI rows starting at row {kpi_start}")

# Find a good format source row (a data row near the KPIs area)
# The Operating Metrics header has a grey fill. Let me use a regular data row format.
# Row 7 is a good data row format source.
data_fmt_row = 7
header_fmt_row = 6  # Section header format
subtotal_fmt_row = 11  # Total Revenue format (bold, thin border)

for i, (rtype, key, label, unit) in enumerate(kpi_rows_data):
    row = kpi_start + i
    if rtype == 'header':
        insert_header_row(ws, row, label, header_fmt_row, MAX_COL)
        # Apply grey fill like Operating Metrics header
        grey_fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
        for col in range(1, MAX_COL + 1):
            ws.cell(row, col).fill = grey_fill
    elif rtype == 'data':
        insert_data_row(ws, row, key, label, unit, data_fmt_row, MAX_COL)
        clear_data(ws, row)
        if unit == '#':
            for col in range(4, MAX_COL + 1):
                ws.cell(row, col).number_format = '#,##0'
    elif rtype == 'subtotal':
        insert_subtotal_row(ws, row, key, label, unit, subtotal_fmt_row, MAX_COL)
        clear_data(ws, row)
        if unit == '#':
            for col in range(4, MAX_COL + 1):
                ws.cell(row, col).number_format = '#,##0'
    elif rtype == 'ratio':
        insert_ratio_row(ws, row, label, unit, data_fmt_row, MAX_COL)
        clear_data(ws, row)
        for col in range(4, MAX_COL + 1):
            ws.cell(row, col).number_format = '0.0%'
    elif rtype == 'blank':
        insert_blank_row(ws, row)

# Set SUM formulas for Global Network Sales and Total Restaurants
gns_row = kpi_start + 5  # Global Network Sales
set_sum_formula(ws, gns_row, kpi_start + 1, kpi_start + 4, 4, MAX_COL)

tr_row = kpi_start + 14  # Total Restaurants
set_sum_formula(ws, tr_row, kpi_start + 9, kpi_start + 13, 4, MAX_COL)

print("KPI rows populated with labels and formulas")

# ═══════════════════════════════════════════════════════════════════
# PHASE 5: RENAMES
# ═══════════════════════════════════════════════════════════════════

# Row 2-3 headers
ws.cell(2, 2).value = 'GYG Model Summary'
ws.cell(3, 2).value = 'Guzman y Gomez (GYG.AX)'

# Revenue
r = find_row_by_key(ws, 'Rev-Steel Revenue')
if r:
    rename_row(ws, r, 'Rev-AU Corp Sales', 'AU Corporate Restaurant Sales', 'AUDm')
r = find_row_by_key(ws, 'Rev-Metals Revenue')
if r:
    rename_row(ws, r, 'Rev-US Corp Sales', 'US Corporate Restaurant Sales', 'AUDm')

# COGS
r = find_row_by_key(ws, 'COGS-Steel COGS')
if r:
    rename_row(ws, r, 'COGS-Food Packaging', 'Food & Packaging', 'AUDm')

# OpEx
r = find_row_by_key(ws, 'OPEX-Selling & Distribution')
if r:
    rename_row(ws, r, 'OPEX-Admin', 'Administrative Expenses', 'AUDm')
r = find_row_by_key(ws, 'OPEX-Occupancy Costs')
if r:
    rename_row(ws, r, 'OPEX-Marketing', 'Marketing Expenses', 'AUDm')
r = find_row_by_key(ws, 'OPEX-General & Admin')
if r:
    rename_row(ws, r, 'OPEX-Other', 'Other Expenses', 'AUDm')

# EBITDA
r = find_row_by_key(ws, 'EBITDA-Underlying EBITDA')
if r:
    rename_row(ws, r, 'EBITDA-Statutory EBITDA', 'Statutory EBITDA', 'AUDm')

# Statutory EBITDA Adjustments header → Segment EBITDA Bridge
r = find_row_by_label(ws, 'Statutory EBITDA Adjustments')
if r:
    ws.cell(r, 2).value = 'Segment EBITDA Bridge'

# Stat-Significant Items → Stat-Other Costs
r = find_row_by_key(ws, 'Stat-Significant Items')
if r:
    rename_row(ws, r, 'Stat-Other Costs', 'Other Non-Recurring Costs', 'AUDm')

# Stat-Statutory EBITDA → Group Segment EBITDA
r = find_row_by_key(ws, 'Stat-Statutory EBITDA')
if r:
    rename_row(ws, r, 'EBITDA-Group Segment EBITDA', 'Group Segment Underlying EBITDA', 'AUDm')

# D&A
r = find_row_by_key(ws, 'DA-Depreciation PPE')
if r:
    rename_row(ws, r, label='PPE Depreciation')
r = find_row_by_key(ws, 'DA-ROU Amortisation')
if r:
    rename_row(ws, r, label='ROU Assets Depreciation')

# EBIT
r = find_row_by_key(ws, 'EBIT-Underlying EBIT')
if r:
    rename_row(ws, r, 'EBIT-Underlying EBIT', 'Underlying EBIT', 'AUDm')

# Interest section
r = find_row_by_label(ws, 'Interest')
if r:
    ws.cell(r, 2).value = 'Finance Income / Costs'
r = find_row_by_key(ws, 'Int-Interest Income')
if r:
    rename_row(ws, r, 'Int-Term Deposit', 'Term Deposit Interest', 'AUDm')
r = find_row_by_key(ws, 'Int-Lease Interest')
if r:
    rename_row(ws, r, label='Lease Liability Interest')
r = find_row_by_key(ws, 'Int-Bank Interest')
if r:
    rename_row(ws, r, 'Int-Other Costs', 'Other Finance Costs', 'AUDm')

# Interest rates
r = find_row_by_label(ws, 'Interest Income Rate')
if r:
    ws.cell(r, 2).value = 'Term Deposit Rate'
r = find_row_by_label(ws, 'Lease Interest Rate')
if r:
    ws.cell(r, 2).value = 'Lease Receivable Rate'
r = find_row_by_label(ws, 'Bank Interest Rate')
if r:
    ws.cell(r, 2).value = 'Lease Liability Rate'

# PBT/NPAT section - NCI label update
r = find_row_by_key(ws, 'NPAT-Sig Items AT')
if r:
    rename_row(ws, r, label='Significant Items After Tax')

# BS section - rename Other Assets
r = find_row_by_key(ws, 'BS-Other Assets')
if r:
    rename_row(ws, r, label='Prepayments & Other Assets')

print("\nAll renames complete")

# ═══════════════════════════════════════════════════════════════════
# PHASE 6: UPDATE ALL NZD → AUD UNITS
# ═══════════════════════════════════════════════════════════════════

for row in range(1, ws.max_row + 1):
    c_val = ws.cell(row, 3).value
    if isinstance(c_val, str):
        if 'NZDm' in c_val:
            ws.cell(row, 3).value = c_val.replace('NZDm', 'AUDm')
        elif 'NZDps' in c_val:
            ws.cell(row, 3).value = c_val.replace('NZDps', 'AUDps')
        elif c_val == 'NZD':
            ws.cell(row, 3).value = 'AUD'

# Also check col B for any NZD references
for row in range(1, ws.max_row + 1):
    b_val = ws.cell(row, 2).value
    if isinstance(b_val, str) and 'NZD' in b_val:
        ws.cell(row, 2).value = b_val.replace('NZD', 'AUD')

print("NZD → AUD conversion complete")

# ═══════════════════════════════════════════════════════════════════
# PHASE 7: UPDATE FORMULAS
# ═══════════════════════════════════════════════════════════════════

# Find key rows by their keys for formula updates
rev_au = find_row_by_key(ws, 'Rev-AU Corp Sales')
rev_us = find_row_by_key(ws, 'Rev-US Corp Sales')
rev_fran_royalty = find_row_by_key(ws, 'Rev-Franchise Royalty')
rev_fran_fee = find_row_by_key(ws, 'Rev-Franchise Fee')
rev_total = find_row_by_key(ws, 'Rev-Total Revenue')
rev_total_other = find_row_by_key(ws, 'Rev-Total Other Revenue')
cogs_fp = find_row_by_key(ws, 'COGS-Food Packaging')
cogs_total = find_row_by_key(ws, 'COGS-Total COGS')
gp_gp = find_row_by_key(ws, 'GP-Gross Profit')
gp_margin = find_row_by_label(ws, 'GP Margin')
ebitda_stat = find_row_by_key(ws, 'EBITDA-Statutory EBITDA')
opex_total = find_row_by_key(ws, 'OPEX-Total OpEx')
stat_sbp = find_row_by_key(ws, 'Stat-SBP')
stat_other = find_row_by_key(ws, 'Stat-Other Costs')
stat_cash_rent = find_row_by_key(ws, 'Stat-Cash Rent')
ebitda_au = find_row_by_key(ws, 'EBITDA-AU Segment EBITDA')
ebitda_us = find_row_by_key(ws, 'EBITDA-US Segment EBITDA')
ebitda_group = find_row_by_key(ws, 'EBITDA-Group Segment EBITDA')
da_total = find_row_by_key(ws, 'DA-Total DA')
da_ppe = find_row_by_key(ws, 'DA-Depreciation PPE')
da_rou = find_row_by_key(ws, 'DA-ROU Amortisation')
da_reacq = find_row_by_key(ws, 'DA-Reacq Amort')
da_other = find_row_by_key(ws, 'DA-Other Amort')
ebit_row = find_row_by_key(ws, 'EBIT-Underlying EBIT')
int_td = find_row_by_key(ws, 'Int-Term Deposit')
int_lr = find_row_by_key(ws, 'Int-Lease Receivable')
int_oi = find_row_by_key(ws, 'Int-Other Income')
int_ll = find_row_by_key(ws, 'Int-Lease Interest')
int_oc = find_row_by_key(ws, 'Int-Other Costs')
int_net = find_row_by_key(ws, 'Int-Net Finance Costs')

print(f"\nKey row positions:")
print(f"  Rev AU={rev_au}, US={rev_us}, Fran Roy={rev_fran_royalty}, Fran Fee={rev_fran_fee}")
print(f"  Rev Total={rev_total}, Total Other={rev_total_other}")
print(f"  COGS FP={cogs_fp}, COGS Total={cogs_total}")
print(f"  GP={gp_gp}, EBITDA Stat={ebitda_stat}, OpEx Total={opex_total}")
print(f"  SBP={stat_sbp}, Other={stat_other}, Cash Rent={stat_cash_rent}")
print(f"  AU Seg={ebitda_au}, US Seg={ebitda_us}, Group Seg={ebitda_group}")
print(f"  DA PPE={da_ppe}, ROU={da_rou}, Reacq={da_reacq}, Other={da_other}, Total={da_total}")
print(f"  EBIT={ebit_row}")
print(f"  Int TD={int_td}, LR={int_lr}, OI={int_oi}, LL={int_ll}, OC={int_oc}, Net={int_net}")

# Total Revenue = SUM of revenue line items
set_sum_formula(ws, rev_total, rev_au, rev_fran_fee, 4, MAX_COL)

# Total COGS = just Food & Packaging (only one COGS line now)
# Actually it should still be a SUM for extensibility
set_sum_formula(ws, cogs_total, cogs_fp, cogs_fp, 4, MAX_COL)

# Gross Profit = Total Revenue + Total Other Revenue + Total COGS
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(gp_gp, col).value = f'={cl}{rev_total}+{cl}{rev_total_other}+{cl}{cogs_total}'

# GP Margin = GP / (Total Revenue + Total Other Revenue)
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(gp_margin, col).value = f'=IF({cl}{rev_total}+{cl}{rev_total_other}=0,"",{cl}{gp_gp}/({cl}{rev_total}+{cl}{rev_total_other}))'

# Statutory EBITDA = GP + Total OpEx
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(ebitda_stat, col).value = f'={cl}{gp_gp}+{cl}{opex_total}'

# EBITDA Margin — find it
ebitda_margin = find_row_by_label(ws, 'EBITDA Margin')
if ebitda_margin:
    for col in range(4, MAX_COL + 1):
        cl = get_column_letter(col)
        ws.cell(ebitda_margin, col).value = f'=IF({cl}{rev_total}+{cl}{rev_total_other}=0,"",{cl}{ebitda_stat}/({cl}{rev_total}+{cl}{rev_total_other}))'

# Group Segment EBITDA = AU + US
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(ebitda_group, col).value = f'={cl}{ebitda_au}+{cl}{ebitda_us}'

# Total D&A = SUM of all D&A components
set_sum_formula(ws, da_total, da_ppe, da_other, 4, MAX_COL)

# D&A / Revenue
da_rev = find_row_by_label(ws, 'D&A / Revenue')
if da_rev:
    for col in range(4, MAX_COL + 1):
        cl = get_column_letter(col)
        ws.cell(da_rev, col).value = f'=IF({cl}{rev_total}+{cl}{rev_total_other}=0,"",{cl}{da_total}/({cl}{rev_total}+{cl}{rev_total_other}))'

# EBIT = Statutory EBITDA + Total D&A (D&A is negative)
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(ebit_row, col).value = f'={cl}{ebitda_stat}+{cl}{da_total}'

# EBIT Margin
ebit_margin = find_row_by_label(ws, 'EBIT Margin')
if ebit_margin:
    for col in range(4, MAX_COL + 1):
        cl = get_column_letter(col)
        ws.cell(ebit_margin, col).value = f'=IF({cl}{rev_total}+{cl}{rev_total_other}=0,"",{cl}{ebit_row}/({cl}{rev_total}+{cl}{rev_total_other}))'

# Net Finance Costs = SUM of all interest items
set_sum_formula(ws, int_net, int_td, int_oc, 4, MAX_COL)

# PBT
pbt_row = find_row_by_key(ws, 'PBT-PBT')
if pbt_row:
    for col in range(4, MAX_COL + 1):
        cl = get_column_letter(col)
        ws.cell(pbt_row, col).value = f'={cl}{ebit_row}+{cl}{int_net}'

# NPAT Margin - update to use total revenue + other revenue
npat_margin = find_row_by_label(ws, 'NPAT Margin')
npat_u = find_row_by_key(ws, 'NPAT-Underlying NPAT')
if npat_margin and npat_u:
    for col in range(4, MAX_COL + 1):
        cl = get_column_letter(col)
        ws.cell(npat_margin, col).value = f'=IF({cl}{rev_total}+{cl}{rev_total_other}=0,"",{cl}{npat_u}/({cl}{rev_total}+{cl}{rev_total_other}))'

# OpEx Total = SUM of opex items
opex_emp = find_row_by_key(ws, 'OPEX-Employee Benefits')
opex_other = find_row_by_key(ws, 'OPEX-Other')
if opex_emp and opex_other:
    set_sum_formula(ws, opex_total, opex_emp, opex_other, 4, MAX_COL)

# Update Revenue Growth formula
rev_growth = find_row_by_label(ws, 'Revenue Growth')
if rev_growth:
    for col in range(5, MAX_COL + 1):  # Start from col E (need prior year)
        cl = get_column_letter(col)
        prev_cl = get_column_letter(col - 1)
        ws.cell(rev_growth, col).value = f'=IF({prev_cl}{rev_total}=0,"",{cl}{rev_total}/{prev_cl}{rev_total}-1)'

# Update GP Growth
gp_growth = find_row_by_label(ws, 'GP Growth')
if gp_growth:
    for col in range(5, MAX_COL + 1):
        cl = get_column_letter(col)
        prev_cl = get_column_letter(col - 1)
        ws.cell(gp_growth, col).value = f'=IF({prev_cl}{gp_gp}=0,"",{cl}{gp_gp}/{prev_cl}{gp_gp}-1)'

# Update EBITDA Growth
ebitda_growth = find_row_by_label(ws, 'EBITDA Growth')
if ebitda_growth:
    for col in range(5, MAX_COL + 1):
        cl = get_column_letter(col)
        prev_cl = get_column_letter(col - 1)
        ws.cell(ebitda_growth, col).value = f'=IF({prev_cl}{ebitda_stat}=0,"",{cl}{ebitda_stat}/{prev_cl}{ebitda_stat}-1)'

# Update OpEx Growth
opex_growth = find_row_by_label(ws, 'OpEx Growth')
if opex_growth:
    for col in range(5, MAX_COL + 1):
        cl = get_column_letter(col)
        prev_cl = get_column_letter(col - 1)
        ws.cell(opex_growth, col).value = f'=IF({prev_cl}{opex_total}=0,"",{cl}{opex_total}/{prev_cl}{opex_total}-1)'

# Update Capex / Sales
capex_sales = find_row_by_label(ws, 'Capex / Sales')
capex_ppe = find_row_by_key(ws, 'CF-Capex PPE')
capex_intang = find_row_by_key(ws, 'CF-Capex Intang')
if capex_sales and capex_ppe and capex_intang:
    for col in range(4, MAX_COL + 1):
        cl = get_column_letter(col)
        ws.cell(capex_sales, col).value = f'=IF({cl}{rev_total}+{cl}{rev_total_other}=0,"",({cl}{capex_ppe}+{cl}{capex_intang})/({cl}{rev_total}+{cl}{rev_total_other}))'

# Receivables / Revenue
recv_rev = find_row_by_label(ws, 'Receivables / Revenue')
bs_recv = find_row_by_key(ws, 'BS-Trade Receivables')
if recv_rev and bs_recv:
    for col in range(4, MAX_COL + 1):
        cl = get_column_letter(col)
        ws.cell(recv_rev, col).value = f'=IF({cl}{rev_total}+{cl}{rev_total_other}=0,"",{cl}{bs_recv}/({cl}{rev_total}+{cl}{rev_total_other}))'

# Inventory / Revenue
inv_rev = find_row_by_label(ws, 'Inventory / Revenue')
bs_inv = find_row_by_key(ws, 'BS-Inventories')
if inv_rev and bs_inv:
    for col in range(4, MAX_COL + 1):
        cl = get_column_letter(col)
        ws.cell(inv_rev, col).value = f'=IF({cl}{rev_total}+{cl}{rev_total_other}=0,"",{cl}{bs_inv}/({cl}{rev_total}+{cl}{rev_total_other}))'

# Payables / Revenue
pay_rev = find_row_by_label(ws, 'Payables / Revenue')
bs_pay = find_row_by_key(ws, 'BS-Trade Payables')
if pay_rev and bs_pay:
    for col in range(4, MAX_COL + 1):
        cl = get_column_letter(col)
        ws.cell(pay_rev, col).value = f'=IF({cl}{rev_total}+{cl}{rev_total_other}=0,"",{cl}{bs_pay}/({cl}{rev_total}+{cl}{rev_total_other}))'

# Update Total Assets to include new BS rows
bs_cash = find_row_by_key(ws, 'BS-Cash')
bs_other_assets = find_row_by_key(ws, 'BS-Other Assets')
total_assets = find_row_by_label(ws, 'Total Assets')
if bs_cash and bs_other_assets and total_assets:
    set_sum_formula(ws, total_assets, bs_cash, bs_other_assets, 4, MAX_COL)

# Update Total Liabilities
bs_trade_pay = find_row_by_key(ws, 'BS-Trade Payables')
bs_banking = find_row_by_key(ws, 'BS-Total Banking Debt')
total_liab = find_row_by_label(ws, 'Total Liabilities')
if bs_trade_pay and bs_banking and total_liab:
    set_sum_formula(ws, total_liab, bs_trade_pay, bs_banking, 4, MAX_COL)

# Working Capital = Trade Receivables + Inventories - Trade Payables
wc_row = find_row_by_label(ws, 'Working Capital')
if wc_row and bs_recv and bs_inv and bs_pay:
    for col in range(4, MAX_COL + 1):
        cl = get_column_letter(col)
        ws.cell(wc_row, col).value = f'={cl}{bs_recv}+{cl}{bs_inv}-{cl}{bs_pay}'

# CF-EBITDA should reference Statutory EBITDA
cf_ebitda = find_row_by_key(ws, 'CF-EBITDA')
if cf_ebitda:
    rename_row(ws, cf_ebitda, label='Statutory EBITDA')
    for col in range(4, MAX_COL + 1):
        cl = get_column_letter(col)
        ws.cell(cf_ebitda, col).value = f'={cl}{ebitda_stat}'

# FCF Margin
fcf_margin = find_row_by_label(ws, 'FCF Margin')
fcf_row = find_row_by_label(ws, 'Operating Free Cash Flow')
# There are two rows with this label; find the one with a formula
if fcf_margin:
    for col in range(4, MAX_COL + 1):
        cl = get_column_letter(col)
        ws.cell(fcf_margin, col).value = f'=IF({cl}{rev_total}+{cl}{rev_total_other}=0,"",{cl}{fcf_margin-1}/({cl}{rev_total}+{cl}{rev_total_other}))'

# ROIC section - update invested capital and references
inv_cap = find_row_by_label(ws, 'Invested Capital')
net_bank_debt = find_row_by_label(ws, 'Net Banking Debt')
total_equity = find_row_by_label(ws, 'Total Equity')
rofe = find_row_by_label(ws, 'ROFE')
roic_nopat = find_row_by_label(ws, 'NOPAT')
roic = find_row_by_label(ws, 'ROIC')
roe = find_row_by_label(ws, 'ROE')
u_ebit = find_row_by_label(ws, 'Underlying EBIT')  # In ROIC section, this mirrors the P&L EBIT
tax_rate = find_row_by_label(ws, 'Underlying Tax Rate')

# Update references that use the EBITDA row (was Underlying EBITDA, now Statutory EBITDA)
# EBITDA cashflow conversion
ebitda_conv = find_row_by_label(ws, 'EBITDA Cashflow conversion')
gross_ocf = find_row_by_label(ws, 'Gross Operating Cash Flow')
if ebitda_conv and cf_ebitda and gross_ocf:
    for col in range(4, MAX_COL + 1):
        cl = get_column_letter(col)
        ws.cell(ebitda_conv, col).value = f'={cl}{gross_ocf}/{cl}{cf_ebitda}'

# ND/EBITDA
nd_ebitda = find_row_by_label(ws, 'ND / EBITDA')
if nd_ebitda and net_bank_debt and ebitda_stat:
    for col in range(4, MAX_COL + 1):
        cl = get_column_letter(col)
        ws.cell(nd_ebitda, col).value = f'=IF({cl}{ebitda_stat}=0,"",{cl}{net_bank_debt}/{cl}{ebitda_stat})'

print("All formulas updated")

# ═══════════════════════════════════════════════════════════════════
# PHASE 8: CLEAR OLD DATA FROM RENAMED/REPURPOSED ROWS
# ═══════════════════════════════════════════════════════════════════
# Clear all data from rows that were renamed (old VSL data no longer valid)

rows_to_clear_data = [
    'Rev-AU Corp Sales', 'Rev-US Corp Sales',
    'COGS-Food Packaging',
    'OPEX-Employee Benefits', 'OPEX-Admin', 'OPEX-Marketing', 'OPEX-Other',
    'Stat-SBP', 'Stat-Other Costs',
    'DA-Depreciation PPE', 'DA-ROU Amortisation',
    'Int-Term Deposit', 'Int-Lease Interest', 'Int-Other Costs',
    'Tax-Tax Expense',
    'NPAT-NCI', 'NPAT-Sig Items AT',
    'EPS-YE Shares', 'EPS-WASO Basic', 'EPS-WASO Diluted',
    'Div-DPS',
    'BS-Cash', 'BS-Trade Receivables', 'BS-Inventories', 'BS-PPE',
    'BS-Intangibles', 'BS-ROU Assets', 'BS-Other Assets',
    'BS-Trade Payables', 'BS-Other Liabilities', 'BS-Lease Liabilities',
    'BS-Total Banking Debt', 'BS-Issued Capital', 'BS-Retained Profits',
    'BS-Reserves', 'BS-Minorities',
    'CF-WC Change', 'CF-Int Received', 'CF-Interest Paid', 'CF-Lease Int Paid',
    'CF-Tax Paid', 'CF-Net OCF', 'CF-Capex PPE', 'CF-Capex Intang',
    'CF-Acquisitions', 'CF-Asset Sales', 'CF-Dividends', 'CF-Share Issues',
    'CF-Lease Principal', 'CF-Debt Change',
]

for key in rows_to_clear_data:
    r = find_row_by_key(ws, key)
    if r:
        clear_data(ws, r)

# Clear analytical rows too (rates, growth, etc.)
analytical_labels = [
    'Revenue Growth', 'GP Growth', 'GP Margin', 'OpEx Growth',
    'EBITDA Growth', 'EBITDA Margin', 'D&A / Revenue', 'Avg Lease Life',
    'EBIT Growth', 'EBIT Margin', 'Term Deposit Rate', 'Lease Receivable Rate',
    'Lease Liability Rate', 'Underlying Tax Rate', 'NPAT Growth', 'NPAT Margin',
    'EPS Growth', 'Payout Ratio', 'Dividend Yield', 'Dividend Growth',
    'Receivables / Revenue', 'Inventory / Revenue', 'Payables / Revenue',
    'New Lease Additions', 'OCF Growth',
]

for label in analytical_labels:
    r = find_row_by_label(ws, label)
    if r:
        clear_data(ws, r)

# Clear CF totals that were hardcoded
cf_labels = ['Gross Operating Cash Flow', 'Net Operating Cash Flow',
             'Total Investing Cash Flow', 'Total Financing Cash Flow']
for label in cf_labels:
    r = find_row_by_label(ws, label)
    if r:
        clear_data(ws, r)

print("Old data cleared")

# ═══════════════════════════════════════════════════════════════════
# PHASE 9: SAVE
# ═══════════════════════════════════════════════════════════════════
wb.save(SRC)
print(f"\nSaved to {SRC}")
print("Script complete!")

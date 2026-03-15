"""Part 2: Annual sheet - enter historical data and formulas."""
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')
ws = wb['Annual']

# Column mapping: D=FY23(col4), E=FY24(col5), F=FY25(col6), G=FY26E(col7)...

# ============================================================
# HISTORICAL P&L DATA (FY23=col4, FY24=col5, FY25=col6)
# ============================================================

# R7: Total Revenue
ws.cell(7, 4).value = 259.044
ws.cell(7, 5).value = 342.214
ws.cell(7, 6).value = 435.982

# R8: Other Revenue & Income
ws.cell(8, 4).value = 17.725
ws.cell(8, 5).value = 22.774
ws.cell(8, 6).value = 32.024

# R9: Total Rev incl Other = R7 + R8
for c in range(4, 7):
    col = chr(64+c)  # D, E, F
    ws.cell(9, c).value = f'={col}7+{col}8'

# R10: Revenue Growth
ws.cell(10, 5).value = '=E7/D7-1'
ws.cell(10, 6).value = '=F7/E7-1'

# R13: Cost of food & packaging (negative)
ws.cell(13, 4).value = -70.428
ws.cell(13, 5).value = -87.580
ws.cell(13, 6).value = -113.948

# R14: Employee benefits (negative)
ws.cell(14, 4).value = -113.725
ws.cell(14, 5).value = -153.733
ws.cell(14, 6).value = -184.656

# R15: Admin + Marketing + Other expenses combined (negative)
ws.cell(15, 4).value = -(35.282 + 13.718 + 13.995)  # = -62.995
ws.cell(15, 5).value = -(60.595 + 17.938 + 17.868)  # = -96.401
ws.cell(15, 6).value = -(53.615 + 21.287 + 29.381)  # = -104.283

# R16: Total Operating Costs = sum of R13:R15
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(16, c).value = f'={col}13+{col}14+{col}15'

# R19: Statutory EBITDA = Total Rev incl Other + Total Costs = R9 + R16
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(19, c).value = f'={col}9+{col}16'

# R20: Statutory EBITDA Margin = R19 / R7
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(20, c).value = f'=IF({col}7=0,"",{col}19/{col}7)'

# --- Segment EBITDA (R35-R38) ---
# Australia Seg EBITDA
ws.cell(35, 4).value = 30.7
ws.cell(35, 5).value = 45.6  # PF: 1H24(23.1)+2H24(22.5)
ws.cell(35, 6).value = 66.0

# US Seg EBITDA
ws.cell(36, 4).value = -4.3
ws.cell(36, 5).value = -6.5
ws.cell(36, 6).value = -13.2

# Corporate EBITDA = Group - Aus - US (derive)
# FY24: Group=39.1, Aus=45.6, US=-6.5 -> Corp = 39.1 - 45.6 - (-6.5) = 0.0
# FY25: Group=52.8, Aus=66.0, US=-13.2 -> Corp = 52.8 - 66.0 + 13.2 = 0.0
# FY23: assume 0
ws.cell(37, 4).value = 0
ws.cell(37, 5).value = 0
ws.cell(37, 6).value = 0

# R38: Group Seg EBITDA = sum
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(38, c).value = f'=SUM({col}35:{col}37)'

# R39: Seg EBITDA Growth
ws.cell(39, 5).value = '=E38/D38-1'
ws.cell(39, 6).value = '=F38/E38-1'

# R40: Seg EBITDA Margin (% of NS) - need NS. For now use % of revenue
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(40, c).value = f'=IF({col}7=0,"",{col}38/{col}7)'

# --- Statutory Adjustments (R42-R45) ---
# FY25 calibration: Seg EBITDA(52.8) + AASB16(18.5) - SBP(9.0) + Other(2.9) = Stat EBITDA(65.2≈65.1)
# R42: AASB 16 Lease Impact
ws.cell(42, 6).value = 18.5  # FY25
# R43: SBP (negative in this direction: reduces stat EBITDA to get underlying)
ws.cell(43, 6).value = -9.0  # FY25
# R44: Other Adj
ws.cell(44, 6).value = 2.9  # FY25

# For FY23/FY24 we don't have the bridge detail, estimate from statutory EBITDA
# FY23: Stat EBITDA = Rev+Other+Costs = 276.769 - 247.148 = 29.621, but let's compute
# Actually Stat EBITDA = EBIT + D&A = 4.066 + 25.555 = 29.621
# Seg EBITDA = 26.4 (30.7 - 4.3)
# So adjustments = 29.621 - 26.4 = 3.221
# We'll put it all in AASB16 for FY23
ws.cell(42, 4).value = 3.2  # FY23 approximate
ws.cell(43, 4).value = 0
ws.cell(44, 4).value = 0

# FY24: Stat EBITDA = EBIT + D&A = -3.849 + 31.123 = 27.274
# Seg EBITDA = 39.1
# Adjustments = 27.274 - 39.1 = -11.826
# This includes IPO costs in FY24. Put in Other Adj.
ws.cell(42, 5).value = 12.0  # est AASB16
ws.cell(43, 5).value = -5.5  # est SBP
ws.cell(44, 5).value = -18.3  # includes IPO costs

# R45: Statutory EBITDA = Underlying + AASB16 + SBP + Other
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(45, c).value = f'={col}38+{col}42+{col}43+{col}44'

# --- D&A (R48) ---
ws.cell(48, 4).value = -25.555
ws.cell(48, 5).value = -31.123
ws.cell(48, 6).value = -39.681

# R51: D&A / Revenue
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(51, c).value = f'=IF({col}7=0,"",{col}48/{col}7)'

# --- EBIT = Statutory EBITDA + D&A (R54) ---
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(54, c).value = f'={col}45+{col}48'

# R55: EBIT Growth
ws.cell(55, 5).value = '=IF(D54=0,"",E54/D54-1)'
ws.cell(55, 6).value = '=IF(E54=0,"",F54/E54-1)'

# R56: EBIT Margin
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(56, c).value = f'=IF({col}7=0,"",{col}54/{col}7)'

# --- Interest (R59-R61) ---
ws.cell(59, 4).value = 4.593
ws.cell(59, 5).value = 6.012
ws.cell(59, 6).value = 22.674

ws.cell(60, 4).value = -8.503
ws.cell(60, 5).value = -13.724
ws.cell(60, 6).value = -18.902

# R61: Net Finance Costs = R59 + R60
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(61, c).value = f'={col}59+{col}60'

# --- PBT (R68) = EBIT + Net Finance ---
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(68, c).value = f'={col}54+{col}61'

# --- Tax (R69) ---
ws.cell(69, 4).value = -2.423
ws.cell(69, 5).value = -2.187
ws.cell(69, 6).value = -14.734

# R70: Effective Tax Rate
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(70, c).value = f'=IF({col}68=0,"",{col}69/{col}68)'

# --- NPAT (R72) = PBT + Tax ---
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(72, c).value = f'={col}68+{col}69'

# R75: NPAT Growth
ws.cell(75, 5).value = '=IF(D72=0,"",E72/D72-1)'
ws.cell(75, 6).value = '=IF(E72=0,"",F72/E72-1)'

# R76: NPAT Margin
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(76, c).value = f'=IF({col}7=0,"",{col}72/{col}7)'

# ============================================================
# EPS & DIVIDENDS
# ============================================================

# R79: YE Shares Outstanding
ws.cell(79, 4).value = 84.8
ws.cell(79, 5).value = 103.0
ws.cell(79, 6).value = 102.9

# R80: WASO Basic
ws.cell(80, 4).value = 84.0
ws.cell(80, 5).value = 85.4
ws.cell(80, 6).value = 101.2

# R81: Dilution
ws.cell(81, 4).value = 0
ws.cell(81, 5).value = 0
ws.cell(81, 6).value = 4.5  # 105.7 - 101.2

# R82: WASO Diluted
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(82, c).value = f'={col}80+{col}81'

# R84: Basic EPS (cents per share -> dollars per share for consistency)
# Data says cps, let's keep in dollars (divide by 100)
ws.cell(84, 4).value = '=IF(D80=0,"",D72/D80)'
ws.cell(84, 5).value = '=IF(E80=0,"",E72/E80)'
ws.cell(84, 6).value = '=IF(F80=0,"",F72/F80)'

# R85: Diluted EPS
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(85, c).value = f'=IF({col}82=0,"",{col}72/{col}82)'

# R88: DPS (in dollars, convert from cps)
ws.cell(88, 4).value = 0
ws.cell(88, 5).value = 0
ws.cell(88, 6).value = 0.126  # 12.6 cps

# R89: Total Dividends
ws.cell(89, 4).value = 0
ws.cell(89, 5).value = 0
ws.cell(89, 6).value = 13.0

# R90: Payout Ratio
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(90, c).value = f'=IF({col}72<=0,"",{col}89/{col}72)'

# ============================================================
# KPIs (R95-R104)
# ============================================================

# Total Restaurants = Corp-Aus + Corp-US + Fran-Aus + Fran-SG + Fran-JP
ws.cell(95, 4).value = 55+3+116+16+4    # FY23 = 194
ws.cell(95, 5).value = 64+4+130+17+5    # FY24 = 220
ws.cell(95, 6).value = 81+6+143+21+5    # FY25 = 256

ws.cell(96, 4).value = 55
ws.cell(96, 5).value = 64
ws.cell(96, 6).value = 81

ws.cell(97, 4).value = 116+16+4  # Aus Fran incl SG/JP = 136
ws.cell(97, 5).value = 130+17+5  # = 152
ws.cell(97, 6).value = 143+21+5  # = 169

ws.cell(98, 4).value = 3
ws.cell(98, 5).value = 4
ws.cell(98, 6).value = 6

# Australia Network Sales
ws.cell(99, 4).value = 753.0
ws.cell(99, 5).value = 948.9
ws.cell(99, 6).value = 1168.5

# Comp Sales Growth - not available for FY23
ws.cell(100, 5).value = '=E99/D99-1'
ws.cell(100, 6).value = '=F99/E99-1'

# Corp Rest Margin %
ws.cell(101, 4).value = 0.144
ws.cell(101, 5).value = 0.174
ws.cell(101, 6).value = 0.179

# Implied Royalty Rate: Franchise Rev / Franchise Network Sales
# FY23: 46.9 / (490.9+43.1+7.0) = 46.9/541.0 = 8.7%
# FY24: 60.7 / (615.7+46.4+7.9) = 60.7/670.0 = 9.1%
# FY25: 78.7 / (734.6+64.7+9.1) = 78.7/808.4 = 9.7%
ws.cell(102, 4).value = 0.087
ws.cell(102, 5).value = 0.091
ws.cell(102, 6).value = 0.097

# G&A % NS
# FY23: 46.7/753.0 = 6.2%
# FY24: 63.7/948.9 = 6.7% (PF)
# FY25: 77.0/1168.5 = 6.6%
ws.cell(103, 4).value = 0.062
ws.cell(103, 5).value = 0.067
ws.cell(103, 6).value = 0.066

# Seg EBITDA % NS (Aus segment only / Aus NS)
# FY23: 30.7/753.0 = 4.1% ... but group is 26.4/753.0+10.8 = ...
# Use Aus Seg: FY23 30.7/753.0=4.1%, FY24 45.6/948.9=4.8%, FY25 66.0/1168.5=5.6%
ws.cell(104, 4).value = 0.041
ws.cell(104, 5).value = 0.048
ws.cell(104, 6).value = 0.057

# ============================================================
# BALANCE SHEET (R110-R144)
# ============================================================

# Assets
# Cash + Term Deposits
ws.cell(110, 4).value = 36.504  # FY23
ws.cell(110, 5).value = 16.385 + 278.095  # FY24 = 294.480
ws.cell(110, 6).value = 39.675 + 242.068  # FY25 = 281.743

ws.cell(111, 4).value = 25.087
ws.cell(111, 5).value = 26.499
ws.cell(111, 6).value = 24.840

ws.cell(112, 4).value = 2.153
ws.cell(112, 5).value = 2.825
ws.cell(112, 6).value = 3.761

ws.cell(113, 4).value = 69.486
ws.cell(113, 5).value = 87.630
ws.cell(113, 6).value = 130.056

ws.cell(114, 4).value = 15.202
ws.cell(114, 5).value = 10.586
ws.cell(114, 6).value = 18.305

ws.cell(115, 4).value = 98.939
ws.cell(115, 5).value = 93.796
ws.cell(115, 6).value = 125.430

# Other Assets = Finance lease rec + Other
ws.cell(116, 4).value = 69.333 + 7.807    # 77.140
ws.cell(116, 5).value = 126.403 + 19.578  # 145.981
ws.cell(116, 6).value = 174.844 + 24.219  # 199.063

# Total Assets = sum
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(117, c).value = f'=SUM({col}110:{col}116)'

# BS ratios
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(118, c).value = f'=IF({col}7=0,"",{col}111/{col}7)'
    ws.cell(119, c).value = f'=IF({col}7=0,"",{col}112/{col}7)'
    ws.cell(121, c).value = f'=IF({col}7=0,"",{col}125/{col}7)'

# Liabilities
ws.cell(125, 4).value = 32.635
ws.cell(125, 5).value = 39.387
ws.cell(125, 6).value = 40.439

ws.cell(126, 4).value = 22.190
ws.cell(126, 5).value = 28.933
ws.cell(126, 6).value = 31.324

ws.cell(127, 4).value = 181.717
ws.cell(127, 5).value = 239.498
ws.cell(127, 6).value = 331.311

# Banking Debt - GYG has no debt
ws.cell(128, 4).value = 0
ws.cell(128, 5).value = 0
ws.cell(128, 6).value = 0

# Total Liabilities
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(129, c).value = f'=SUM({col}125:{col}128)'

# Net Cash = Cash - Debt
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(131, c).value = f'={col}110-{col}128'

# Adj Net Debt (incl leases) = Debt - Cash + Leases
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(132, c).value = f'={col}128-{col}110+{col}127'

# Net Cash / Seg EBITDA
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(133, c).value = f'=IF({col}38=0,"",{col}131/{col}38)'

# Equity
ws.cell(137, 4).value = 104.046
ws.cell(137, 5).value = 372.708
ws.cell(137, 6).value = 374.988

ws.cell(138, 4).value = -26.857
ws.cell(138, 5).value = -36.147
ws.cell(138, 6).value = -21.671

ws.cell(139, 4).value = 10.776
ws.cell(139, 5).value = 17.418
ws.cell(139, 6).value = 26.807

ws.cell(140, 4).value = 0
ws.cell(140, 5).value = 0
ws.cell(140, 6).value = 0

# Total Equity
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(141, c).value = f'=SUM({col}137:{col}140)'

# ROE
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(142, c).value = f'=IF({col}141=0,"",{col}72/{col}141)'

# P/B
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(143, c).value = f'=IF({col}141=0,"",28*{col}79/{col}141)'

# BS Check = Total Assets - Total Liabilities - Total Equity
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(144, c).value = f'={col}117-{col}129-{col}141'

# ============================================================
# CASH FLOW (R148-R176)
# ============================================================

# R148: Statutory EBITDA = R45
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(148, c).value = f'={col}45'

# R149: WC Change (derive: OCF - EBITDA - Non-cash ≈ or hardcode)
# OCF = 34.947, 36.765, 57.327
# Stat EBITDA = 29.621, 27.274, 65.1 approx from formulas
# Non-cash estimated
# Just hardcode the statutory cash flow items
ws.cell(149, 4).value = 0   # approximate
ws.cell(149, 5).value = 0
ws.cell(149, 6).value = 0

ws.cell(150, 4).value = 0
ws.cell(150, 5).value = 0
ws.cell(150, 6).value = 0

# R151: Gross OCF
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(151, c).value = f'={col}148+{col}149+{col}150'

# Interest/Tax paid - we'll derive from the statutory OCF
ws.cell(152, 4).value = 0
ws.cell(152, 5).value = 0
ws.cell(152, 6).value = 0

ws.cell(153, 4).value = 0
ws.cell(153, 5).value = 0
ws.cell(153, 6).value = 0

ws.cell(154, 4).value = 0
ws.cell(154, 5).value = 0
ws.cell(154, 6).value = 0

ws.cell(155, 4).value = 0
ws.cell(155, 5).value = 0
ws.cell(155, 6).value = 0

# R156: Net OCF = hardcode statutory
ws.cell(156, 4).value = 34.947
ws.cell(156, 5).value = 36.765
ws.cell(156, 6).value = 57.327

# R160: Capex PPE
ws.cell(160, 4).value = -39.713
ws.cell(160, 5).value = -33.496
ws.cell(160, 6).value = -61.326

# R161: Capex / Sales
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(161, c).value = f'=IF({col}7=0,"",{col}160/{col}7)'

# R162-165: Other investing
ws.cell(162, 4).value = 0; ws.cell(162, 5).value = 0; ws.cell(162, 6).value = 0
ws.cell(163, 4).value = 0; ws.cell(163, 5).value = 0; ws.cell(163, 6).value = 0
ws.cell(164, 4).value = 0; ws.cell(164, 5).value = 0; ws.cell(164, 6).value = 0
ws.cell(165, 4).value = 0; ws.cell(165, 5).value = 0; ws.cell(165, 6).value = 0

# R166: Total Investing CF
ws.cell(166, 4).value = -48.403
ws.cell(166, 5).value = -311.818
ws.cell(166, 6).value = -34.551

# R169: Dividends
ws.cell(169, 4).value = 0
ws.cell(169, 5).value = 0
ws.cell(169, 6).value = -13.0

# R170-173
ws.cell(170, 4).value = 0; ws.cell(170, 5).value = 254.934; ws.cell(170, 6).value = 0
ws.cell(171, 4).value = 0; ws.cell(171, 5).value = 0; ws.cell(171, 6).value = 0
ws.cell(172, 4).value = 0; ws.cell(172, 5).value = 0; ws.cell(172, 6).value = 0
ws.cell(173, 4).value = 0; ws.cell(173, 5).value = 0; ws.cell(173, 6).value = 0

# R174: Total Financing CF
ws.cell(174, 4).value = -4.467
ws.cell(174, 5).value = 254.934
ws.cell(174, 6).value = 0.679

# R176: Net Change in Cash
ws.cell(176, 4).value = -17.923
ws.cell(176, 5).value = -20.119
ws.cell(176, 6).value = 23.455

# OFCF section (R178-R185)
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(179, c).value = f'={col}156'
    ws.cell(180, c).value = f'={col}160'
    ws.cell(181, c).value = f'={col}171'
    ws.cell(182, c).value = f'={col}179+{col}180+{col}181'
    ws.cell(183, c).value = f'=IF({col}82=0,"",{col}182/{col}82)'
    ws.cell(184, c).value = f'=IF({col}82=0,"",{col}183/28)'
    ws.cell(185, c).value = f'=IF({col}7=0,"",{col}182/{col}7)'

# ROIC (R188-R193)
for c in range(4, 7):
    col = chr(64+c)
    ws.cell(189, c).value = f'={col}141+{col}128-{col}110+{col}127'
    ws.cell(190, c).value = f'={col}54'
    ws.cell(191, c).value = f'=IF({col}189=0,"",{col}190/{col}189)'
    ws.cell(192, c).value = f'={col}190*(1+{col}70)'
    ws.cell(193, c).value = f'=IF({col}189=0,"",{col}192/{col}189)'

wb.save('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')
print("Part 2 complete: Annual historical data entered")

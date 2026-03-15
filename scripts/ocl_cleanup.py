"""Cleanup: remove residual rows beyond model range, fix any issues."""
import openpyxl

DST = '/home/pmwilson/Project_Equities/OCL/Models/OCL Model.xlsx'
wb = openpyxl.load_workbook(DST)

# Clear residual content on Annual beyond row 179
ws_a = wb['Annual']
for row in range(180, 210):
    for col in range(1, 17):
        ws_a.cell(row=row, column=col).value = None

# Clear residual content on HY beyond row 113
ws_h = wb['HY & Segments']
for row in range(114, 210):
    for col in range(1, 30):
        ws_h.cell(row=row, column=col).value = None

# Clear residual data cols N-P on Annual (cols 14-16)
for row in range(1, 210):
    for col in range(14, 17):
        cell = ws_a.cell(row=row, column=col)
        if cell.value is not None:
            cell.value = None

# Clear residual data cols X-AC on HY (cols 24-29)
for row in range(1, 210):
    for col in range(24, 30):
        cell = ws_h.cell(row=row, column=col)
        if cell.value is not None:
            cell.value = None

wb.save(DST)
print('Cleanup complete')

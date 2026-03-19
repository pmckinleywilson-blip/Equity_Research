"""
Diagnose formatting issues on the Annual sheet of GYG Model.xlsx
Rows 105-191, columns D-P. Research only — no modifications.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from collections import Counter

wb = openpyxl.load_workbook(
    "/home/pmwilson/Project_Equities/GYG/Models/GYG Model.xlsx",
    data_only=False,
)
ws = wb["Annual"]

# Column ranges
actual_cols = ["D", "E", "F"]          # actuals
forecast_cols = list("GHIJKLMNOP")     # forecasts
detail_cols = ["D", "E", "F", "G", "H"]  # columns to print per-row detail

def col_idx(letter):
    return openpyxl.utils.column_index_from_string(letter)

def border_desc(border):
    parts = []
    if border and border.top and border.top.style:
        parts.append(f"top={border.top.style}")
    if border and border.bottom and border.bottom.style:
        parts.append(f"bot={border.bottom.style}")
    return ",".join(parts) if parts else "none"

def color_desc(font):
    if font.color is None:
        return "none"
    c = font.color
    if c.type == "rgb" and c.rgb:
        return f"rgb={c.rgb}"
    if c.type == "theme":
        return f"theme={c.theme},tint={c.tint}"
    if c.type == "indexed":
        return f"indexed={c.indexed}"
    return str(c)

# --- Counters for forecast cells ---
fmt_counter = Counter()
bold_counter = 0
border_counter = 0
total_forecast = 0

# --- Counters for actual cells ---
act_fmt_counter = Counter()
act_bold_counter = 0
act_border_counter = 0
total_actual = 0

print("=" * 140)
print("ROW-BY-ROW DETAIL  (rows 105-191, columns D-H)")
print("=" * 140)

for row in range(105, 192):
    cell_a = ws[f"A{row}"].value
    cell_b = ws[f"B{row}"].value
    b_font = ws[f"B{row}"].font
    is_subtotal = b_font.bold if b_font else False

    label = f"Row {row:>3} | A={str(cell_a):>12s} | B={str(cell_b):<45s} | subtotal={is_subtotal}"
    print(label)

    for cl in detail_cols:
        c = ws[f"{cl}{row}"]
        nf = c.number_format
        fb = c.font.bold if c.font else None
        fc = color_desc(c.font) if c.font else "none"
        bd = border_desc(c.border) if c.border else "none"
        tag = "ACT" if cl in actual_cols else "FC "
        print(f"    {tag} col {cl}: fmt={nf:<16s}  bold={str(fb):<6s}  color={fc:<30s}  border={bd}")

    # Tally forecast cells
    for cl in forecast_cols:
        c = ws[f"{cl}{row}"]
        nf = c.number_format
        total_forecast += 1

        if nf == "General":
            fmt_counter["General"] += 1
        elif nf == "#,##0.0":
            fmt_counter["#,##0.0"] += 1
        elif "%" in nf:
            fmt_counter["percent"] += 1
        else:
            fmt_counter[nf] += 1

        if c.font and c.font.bold:
            bold_counter += 1
        if c.border:
            if (c.border.top and c.border.top.style) or (c.border.bottom and c.border.bottom.style):
                border_counter += 1

    # Tally actual cells
    for cl in actual_cols:
        c = ws[f"{cl}{row}"]
        nf = c.number_format
        total_actual += 1

        if nf == "General":
            act_fmt_counter["General"] += 1
        elif nf == "#,##0.0":
            act_fmt_counter["#,##0.0"] += 1
        elif "%" in nf:
            act_fmt_counter["percent"] += 1
        else:
            act_fmt_counter[nf] += 1

        if c.font and c.font.bold:
            act_bold_counter += 1
        if c.border:
            if (c.border.top and c.border.top.style) or (c.border.bottom and c.border.bottom.style):
                act_border_counter += 1

    print()

# --- Difference flags ---
print("=" * 140)
print("DIFFERENCE FLAGS: rows where actual vs forecast formatting diverges (cols D vs G)")
print("=" * 140)
diffs = 0
for row in range(105, 192):
    d = ws[f"D{row}"]
    g = ws[f"G{row}"]
    issues = []
    if d.number_format != g.number_format:
        issues.append(f"fmt: D={d.number_format} vs G={g.number_format}")
    d_bold = d.font.bold if d.font else None
    g_bold = g.font.bold if g.font else None
    if d_bold != g_bold:
        issues.append(f"bold: D={d_bold} vs G={g_bold}")
    d_bdr = border_desc(d.border)
    g_bdr = border_desc(g.border)
    if d_bdr != g_bdr:
        issues.append(f"border: D={d_bdr} vs G={g_bdr}")
    d_clr = color_desc(d.font) if d.font else "none"
    g_clr = color_desc(g.font) if g.font else "none"
    if d_clr != g_clr:
        issues.append(f"color: D={d_clr} vs G={g_clr}")
    if issues:
        diffs += 1
        cell_b = ws[f"B{row}"].value
        print(f"  Row {row:>3} ({str(cell_b):<40s}): {' | '.join(issues)}")

print(f"\nTotal rows with D-vs-G differences: {diffs} / 87")

# --- Summary ---
print("\n" + "=" * 140)
print("SUMMARY: FORECAST cells (cols G-P, rows 105-191)")
print("=" * 140)
print(f"Total forecast cells: {total_forecast}")
for k, v in fmt_counter.most_common():
    print(f"  {k:<25s}: {v:>5d}  ({100*v/total_forecast:.1f}%)")
print(f"  Bold cells              : {bold_counter:>5d}  ({100*bold_counter/total_forecast:.1f}%)")
print(f"  Cells with border (t/b) : {border_counter:>5d}  ({100*border_counter/total_forecast:.1f}%)")

print(f"\nSUMMARY: ACTUAL cells (cols D-F, rows 105-191)")
print(f"Total actual cells: {total_actual}")
for k, v in act_fmt_counter.most_common():
    print(f"  {k:<25s}: {v:>5d}  ({100*v/total_actual:.1f}%)")
print(f"  Bold cells              : {act_bold_counter:>5d}  ({100*act_bold_counter/total_actual:.1f}%)")
print(f"  Cells with border (t/b) : {act_border_counter:>5d}  ({100*act_border_counter/total_actual:.1f}%)")

# --- Check: are ANY forecast cells formatted? ---
print("\n" + "=" * 140)
print("QUICK CHECK: First 10 forecast cells that are NOT 'General'")
print("=" * 140)
count = 0
for row in range(105, 192):
    for cl in forecast_cols:
        c = ws[f"{cl}{row}"]
        if c.number_format != "General":
            print(f"  {cl}{row}: fmt={c.number_format}  val={c.value}")
            count += 1
            if count >= 10:
                break
    if count >= 10:
        break
if count == 0:
    print("  *** ALL forecast cells are 'General' — no number formatting at all ***")

wb.close()

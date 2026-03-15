"""Verify OCL model integrity."""
import openpyxl
from openpyxl.utils import get_column_letter as gcl

DST = '/home/pmwilson/Project_Equities/OCL/Models/OCL Model.xlsx'
wb = openpyxl.load_workbook(DST)

print("=== VERIFICATION ===\n")

# 1. Check Column A keys are unique across workbook
ws_a = wb['Annual']
ws_h = wb['HY & Segments']

keys_a = []
keys_h = []
for row in range(1, ws_a.max_row + 1):
    k = ws_a.cell(row=row, column=1).value
    if k:
        keys_a.append((k, row))

for row in range(1, ws_h.max_row + 1):
    k = ws_h.cell(row=row, column=1).value
    if k:
        keys_h.append((k, row))

print(f"Annual keys: {len(keys_a)}")
print(f"HY keys: {len(keys_h)}")

# Check uniqueness within each sheet
a_keys_only = [k for k, r in keys_a]
h_keys_only = [k for k, r in keys_h]
a_dupes = [k for k in a_keys_only if a_keys_only.count(k) > 1]
h_dupes = [k for k in h_keys_only if h_keys_only.count(k) > 1]
if a_dupes:
    print(f"WARNING: Duplicate keys on Annual: {set(a_dupes)}")
if h_dupes:
    print(f"WARNING: Duplicate keys on HY: {set(h_dupes)}")

# Check matching keys between sheets
a_set = set(a_keys_only)
h_set = set(h_keys_only)
# Keys that should match (P&L items that exist on both)
expected_both = {'Rev-Info Intelligence', 'Rev-Planning & Building', 'Rev-Regulatory Solutions',
                 'Rev-Interest Income', 'Rev-Total Revenue', 'COGS-Total COGS', 'GP-Gross Profit',
                 'OPEX-Distribution', 'OPEX-R&D Expense', 'OPEX-Admin', 'OPEX-Total OpEx',
                 'EBITDA-Underlying EBITDA', 'Stat-SBP', 'Stat-M&A Costs', 'Stat-FX',
                 'Stat-Statutory EBITDA', 'DA-Depreciation PPE', 'DA-ROU Amortisation',
                 'DA-Amort Dev Costs', 'DA-Total DA', 'EBIT-Underlying EBIT',
                 'Int-Interest Income', 'Int-Lease Interest', 'Int-Net Finance Costs',
                 'PBT-PBT', 'Tax-Tax Expense', 'NPAT-Underlying NPAT', 'NPAT-Other Items AT',
                 'NPAT-Statutory NPAT', 'KPI-ARR II', 'KPI-ARR PB', 'KPI-ARR RS',
                 'KPI-ARR Total', 'KPI-Total R&D', 'KPI-Capitalised Dev',
                 'KPI-Shares Out', 'KPI-WASO'}
missing_from_h = expected_both - h_set
missing_from_a = expected_both - a_set
if missing_from_h:
    print(f"WARNING: Keys missing from HY: {missing_from_h}")
if missing_from_a:
    print(f"WARNING: Keys missing from Annual: {missing_from_a}")

# 2. Check for #REF errors in formulas (string check only - can't evaluate)
print("\n--- Formula Pattern Check ---")
error_count = 0
for ws_name, ws in [('Annual', ws_a), ('HY & Segments', ws_h)]:
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, str) and val.startswith('='):
                # Check for common formula issues
                if '$AG' in val:
                    print(f"  {ws_name}!{gcl(col)}{row}: References $AG (old template range)")
                    error_count += 1
                if '$P' in val and ws_name == 'Annual':
                    print(f"  {ws_name}!{gcl(col)}{row}: References $P (beyond model range)")
                    error_count += 1
                if '$R' in val and '$RS' not in val and ws_name == 'HY':
                    pass  # R is valid col in HY

print(f"Formula pattern issues found: {error_count}")

# 3. Check actual data was entered
print("\n--- Data Entry Check ---")
for yr_col, yr_label in [(4,'FY21'), (5,'FY22'), (6,'FY23'), (7,'FY24'), (8,'FY25')]:
    rev = ws_a.cell(row=11, column=yr_col).value
    print(f"  Annual {yr_label} Total Rev (row 11): {rev}")

# Check HY data
for col, label in [(4,'1H21'), (6,'1H22'), (8,'1H23'), (10,'1H24'), (12,'1H25'), (14,'1H26')]:
    ii = ws_h.cell(row=7, column=col).value
    print(f"  HY {label} II Rev (row 7): {ii}")

# 4. Check BS structure
print("\n--- BS Check Formula ---")
for col in range(4, 9):
    bs_check = ws_a.cell(row=133, column=col).value
    print(f"  Annual {gcl(col)} BS Check (row 133): {bs_check}")

# 5. Summary
print("\n--- Structure Summary ---")
print(f"Annual: rows used to {ws_a.max_row}, cols to {gcl(ws_a.max_column)}")
print(f"HY: rows used to {ws_h.max_row}, cols to {gcl(ws_h.max_column)}")
print(f"Value: rows used to {wb['Value'].max_row}")

# List all Annual keys with their rows
print("\n--- Annual Key Map ---")
for k, r in keys_a:
    print(f"  Row {r}: {k}")

print("\n--- HY Key Map ---")
for k, r in keys_h:
    print(f"  Row {r}: {k}")

print("\n=== VERIFICATION COMPLETE ===")

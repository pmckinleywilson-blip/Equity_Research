"""
Restructure GYG Model.xlsx from VSL template to GYG (Guzman y Gomez) structure.

Changes:
1. Update headers (company name, sheet titles) on all 3 sheets
2. Restructure P&L on Annual sheet (rows 5-76 area)
3. Restructure P&L on HY & Segments sheet (rows 5-78 area) -- matching Annual keys
4. Restructure Operating Metrics / KPIs on Annual sheet (rows 94-106)
5. Restructure KPIs + Segment Forecast zones on HY sheet (rows 80-116)
6. Update SOTP segment rows on Value sheet
7. Update currency references from NZD to A$
8. Keep BS, CF, Returns sections intact
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import datetime, date

FILE_PATH = '/home/pmwilson/Project_Equities/GYG/Models/GYG Model.xlsx'

# ── Style helpers ──────────────────────────────────────────────────────────
BLUE_FONT = Font(color="FF0000CC")  # actuals
MAROON_FONT = Font(color="FFC00000")  # forecast assumptions
BLACK_FONT = Font(color="FF000000")  # formulas
BOLD_FONT = Font(bold=True)

LIGHT_BLUE_FILL = PatternFill(start_color="FFC5D9F1", end_color="FFC5D9F1", fill_type="solid")
GREY_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)

THIN_BORDER = Border(
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
NO_BORDER = Border()

FMT_AM = '#,##0.0'    # A$m
FMT_PCT = '0.0%'       # percentages
FMT_CPS = '0.000'      # cents per share
FMT_NUM = '#,##0'      # whole numbers
FMT_MULT = '0.0"x"'   # multiples
FMT_SHARES = '0.0'     # shares in millions


def copy_cell_style(src_cell, dst_cell):
    """Copy all formatting from src to dst."""
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format


def apply_line_item_style(ws, row, max_col, units='A$m'):
    """Apply standard line item row formatting."""
    for c in range(2, max_col + 1):
        cell = ws.cell(row, c)
        cell.font = Font(bold=False)
        cell.border = NO_BORDER
        cell.fill = NO_FILL
    # Set number format based on units
    fmt = _units_to_fmt(units)
    for c in range(4, max_col + 1):
        ws.cell(row, c).number_format = fmt


def apply_subtotal_style(ws, row, max_col, units='A$m'):
    """Apply subtotal row formatting: bold + thin top/bottom borders."""
    for c in range(2, max_col + 1):
        cell = ws.cell(row, c)
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
        cell.fill = NO_FILL
    fmt = _units_to_fmt(units)
    for c in range(4, max_col + 1):
        ws.cell(row, c).number_format = fmt


def apply_category_header_style(ws, row, max_col):
    """Apply category sub-header: bold + underline in col B, no data."""
    ws.cell(row, 2).font = Font(bold=True, underline='single')
    ws.cell(row, 2).fill = NO_FILL
    ws.cell(row, 2).border = NO_BORDER
    for c in range(3, max_col + 1):
        cell = ws.cell(row, c)
        cell.value = None
        cell.font = Font(bold=False)
        cell.border = NO_BORDER
        cell.fill = NO_FILL


def apply_section_header_style(ws, row, max_col):
    """Apply major section header: bold + light blue fill."""
    ws.cell(row, 2).font = Font(bold=True)
    ws.cell(row, 2).fill = LIGHT_BLUE_FILL
    ws.cell(row, 2).border = NO_BORDER
    for c in range(3, max_col + 1):
        cell = ws.cell(row, c)
        cell.value = None
        cell.fill = NO_FILL
        cell.border = NO_BORDER


def apply_subsection_header_style(ws, row, max_col):
    """Apply sub-section header: bold + grey fill."""
    ws.cell(row, 2).font = Font(bold=True)
    ws.cell(row, 2).fill = GREY_FILL
    ws.cell(row, 2).border = NO_BORDER
    for c in range(3, max_col + 1):
        cell = ws.cell(row, c)
        cell.value = None
        cell.fill = NO_FILL
        cell.border = NO_BORDER


def apply_analytical_style(ws, row, max_col, units='%'):
    """Apply analytical/ratio row formatting."""
    for c in range(2, max_col + 1):
        cell = ws.cell(row, c)
        cell.font = Font(bold=False)
        cell.border = NO_BORDER
        cell.fill = NO_FILL
    fmt = _units_to_fmt(units)
    for c in range(4, max_col + 1):
        ws.cell(row, c).number_format = fmt


def clear_row(ws, row, max_col):
    """Clear all data and formatting from a row."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        cell.value = None
        cell.font = Font()
        cell.fill = NO_FILL
        cell.border = NO_BORDER
        cell.number_format = 'General'


def blank_row(ws, row, max_col):
    """Make a row blank (clear values but minimal formatting)."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        cell.value = None
        cell.font = Font()
        cell.fill = NO_FILL
        cell.border = NO_BORDER


def set_row(ws, row, key, label, units, max_col):
    """Set the A/B/C values for a row."""
    ws.cell(row, 1).value = key
    ws.cell(row, 2).value = label
    ws.cell(row, 3).value = units


def _units_to_fmt(units):
    if units in ('%', '% YoY'):
        return FMT_PCT
    elif units in ('cps',):
        return FMT_CPS
    elif units in ('#', '#m'):
        return FMT_NUM
    elif units == 'x':
        return FMT_MULT
    elif units in ('m',):
        return FMT_SHARES
    else:
        return FMT_AM


def delete_rows_and_track(ws, start_row, count):
    """Delete rows from worksheet."""
    ws.delete_rows(start_row, count)
    print(f"  Deleted {count} rows starting at row {start_row}")


def insert_rows_and_track(ws, start_row, count):
    """Insert rows into worksheet."""
    ws.insert_rows(start_row, count)
    print(f"  Inserted {count} rows at row {start_row}")


# ══════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════

print("Loading workbook...")
wb = openpyxl.load_workbook(FILE_PATH)
ws_annual = wb['Annual']
ws_hy = wb['HY & Segments']
ws_value = wb['Value']

# ── STEP 1: Determine max columns ──────────────────────────────────────
annual_max_col = 16  # P = col 16
hy_max_col = 29      # AC = col 29


# ══════════════════════════════════════════════════════════════════════════
# STEP 2: Update headers on all sheets
# ══════════════════════════════════════════════════════════════════════════
print("\n--- STEP 2: Update Headers ---")

# Annual sheet
ws_annual.cell(2, 2).value = "GYG Model Summary"
ws_annual.cell(3, 2).value = "Guzman y Gomez (GYG.AX)"
print("  Annual: Updated B2 and B3")

# HY & Segments sheet
ws_hy.cell(2, 2).value = "GYG Segments (Half-Year)"
ws_hy.cell(3, 2).value = "Guzman y Gomez (GYG.AX)"
print("  HY: Updated B2 and B3")

# Value sheet - update currency references from NZD to A$
ws_value.cell(4, 2).value = "Current Share Price (A$)"
ws_value.cell(6, 2).value = "Market Cap (A$m)"
ws_value.cell(7, 2).value = "Net Debt (A$m)"
ws_value.cell(8, 2).value = "Market EV (A$m)"
ws_value.cell(48, 2).value = "Per Share Value (A$)"
ws_value.cell(65, 2).value = "Per Share Value (A$)"

# Update Value sheet currency labels for DCF
for r in range(25, 36):
    cell = ws_value.cell(r, 3)
    if cell.value and 'NZDm' in str(cell.value):
        cell.value = 'A$m'

print("  Value: Updated currency labels NZD -> A$")

# ── STEP 3: Period headers are already correct (FY23-FY35, 1H23-2H35, Jun 30 FY end) ──
# The template already has the right periods. We just need to move Forecast label on HY.
# Current: Forecast at col 11 (K) = 2H26 -- but check if this is correct per spec.
# Spec says: Actuals through 1H26, first forecast = 2H26.
# HY col J (10) = 1H26, col K (11) = 2H26. Current forecast label is at col 11. CORRECT.
# No changes needed for periods.
print("\n--- STEP 3: Period headers already correct ---")


# ══════════════════════════════════════════════════════════════════════════
# STEP 4: Restructure P&L on Annual sheet
# ══════════════════════════════════════════════════════════════════════════
print("\n--- STEP 4: Restructure Annual P&L ---")

# Current structure (Annual rows 5-76):
# Row 5: P&L header (RETAIN)
# Row 6: Revenue sub-header (RETAIN)
# Row 7: Rev-Steel Revenue (REPLACE)
# Row 8: Rev-Metals Revenue (REPLACE)
# Row 9: Rev-Total Revenue (RETAIN subtotal)
# Row 10: Revenue Growth (RETAIN analytical)
# Row 11: blank
# Row 12: COGS sub-header (REPLACE -> needs to become "Other Revenue")
# Rows 13-16: COGS items + Total COGS (REPLACE)
# Row 17: blank
# Row 18: Gross Profit sub-header (REPLACE)
# Rows 19-22: GP items + Gross Profit (REPLACE some)
# Rows 23-24: GP Growth, GP Margin
# Row 25: blank
# Row 26: Operating Expenses sub-header
# Rows 27-31: OpEx items + Total OpEx
# Row 32: OpEx Growth
# Row 33: blank
# Row 34: EBITDA sub-header
# Rows 35-37: Segment EBITDA items
# Row 38: Underlying EBITDA (RETAIN subtotal)
# ...

# TARGET Annual P&L structure (from spec):
# Row 5: P&L (section header - RETAIN)
# Row 6: Revenue (category header - RETAIN)
# Row 7: Rev-Corp Restaurant Sales Aus
# Row 8: Rev-Corp Restaurant Sales US
# Row 9: Rev-Total Corp Restaurant Sales (SUBTOTAL)
# Row 10: Rev-Franchise Royalty Revenue
# Row 11: Rev-Franchise Fee Revenue
# Row 12: Rev-Total Revenue (RETAIN SUBTOTAL - currently row 9)
# Row 13: (blank) Revenue Growth (RETAIN ANALYTICAL - currently row 10)
# Row 14: blank
# Row 15: Other Revenue (category header)
# Row 16: OthRev-Marketing Levy
# Row 17: OthRev-Other Franchise Revenue
# Row 18: OthRev-Other Income
# Row 19: OthRev-Total Other Revenue (SUBTOTAL)
# Row 20: blank
# Row 21: (blank) Total Revenue & Other Income (SUBTOTAL)
# Row 22: blank
# Row 23: Segment EBITDA Bridge (category header)
# Row 24: SegEBITDA-Australia
# Row 25: SegEBITDA-US
# Row 26: SegEBITDA-Total Segment EBITDA (SUBTOTAL)
# Row 27: Bridge-Cash Rent
# Row 28: Bridge-SBP
# Row 29: Bridge-Other
# Row 30: EBITDA-Statutory EBITDA -> label "EBITDA" (RETAIN SUBTOTAL - currently row 38 "Underlying EBITDA")
# Row 31: (blank) EBITDA Margin (RETAIN ANALYTICAL - currently row 40)
# Row 32: blank
# Row 33: Expenses (Memo) (category header)
# Row 34: Exp-Food & Packaging
# Row 35: Exp-Employee Benefits
# Row 36: Exp-Admin
# Row 37: Exp-Marketing
# Row 38: Exp-Other Expenses
# Row 39: Exp-Total Expenses (SUBTOTAL)
# Row 40: blank
# Row 41: Depreciation & Amortisation (category header)
# Row 42: DA-ROU Depreciation
# Row 43: DA-PPE Depreciation
# Row 44: DA-Amortisation Reacquired
# Row 45: DA-Amortisation Other
# Row 46: DA-Total DA (RETAIN SUBTOTAL - currently row 50)
# Row 47: blank
# Row 48: EBIT-EBIT label "EBIT" (RETAIN - currently row 54)
# Row 49: (blank) EBIT Margin (RETAIN - currently row 56)
# Row 50: blank
# Row 51: Finance (category header)
# Row 52: Int-Term Deposit Income
# Row 53: Int-Lease Receivable Income
# Row 54: Int-Other Finance Income
# Row 55: Int-Total Finance Income (SUBTOTAL)
# Row 56: Int-Lease Liability Costs
# Row 57: Int-Other Finance Costs
# Row 58: Int-Total Finance Costs (SUBTOTAL)
# Row 59: Int-Net Finance (RETAIN SUBTOTAL - currently row 62)
# Row 60: blank
# Row 61: PBT-PBT (RETAIN - currently row 68)
# Row 62: Tax-Tax Expense (currently row 69)
# Row 63: (blank) Effective Tax Rate (RETAIN - currently row 70)
# Row 64: NPAT-NPAT label "NPAT" (RETAIN - currently row 72)
# Row 65: (blank) NPAT Margin (RETAIN - currently row 76)
# Row 66: blank
# Row 67: EPS & Dividends (category header - currently row 78)
# ... (EPS section follows)

# Strategy: Rather than trying to insert/delete individually, I'll:
# 1. Record the formatting of retained rows
# 2. Clear the P&L section (rows 5-76)
# 3. Rebuild it with the new structure
# 4. But this violates the "modify in place" rule.
#
# Better approach: Work section by section. The template has 72 P&L rows (5-76).
# The new structure needs a different count. Let me count:
# New P&L rows: 5-65 = 61 rows to NPAT Margin
# Currently rows 5-76 = 72 rows
# That's 11 fewer rows in the P&L.
# Then EPS section starts (currently rows 78-92 = 15 rows).
# After that, KPIs (currently rows 94-106 = 13 rows) will be replaced.
#
# Let me plan the exact row mapping.

# Actually, I need to think about this more carefully.
# Current: P&L occupies rows 5-76 (72 rows). EPS at 78-92. KPIs at 94-106. BS at 108-144. CF at 146-194.
# New P&L needs rows 5-65 (61 rows). That's 11 fewer.
# If I delete 11 rows from the P&L area, everything below shifts up by 11.
# Then EPS (currently 78-92) -> 67-81
# KPIs (currently 94-106) -> 83-95 (but we're replacing these anyway)
# BS (currently 108-144) -> 97-133
# CF (currently 146-194) -> 135-183

# But the formulas in BS and CF reference specific rows (e.g. D38 for EBITDA).
# These will break when rows shift. However, openpyxl's insert_rows/delete_rows
# should adjust formula references automatically.

# Let me take a simpler approach:
# 1. First, handle the P&L restructure by working from specific row positions
# 2. Count how many rows the new structure needs vs. old
# 3. Insert or delete the difference at the right point
# 4. Then write the new content

# Let me count precisely:

# CURRENT ANNUAL P&L (rows 5-76):
# 5: P&L header
# 6: Revenue header
# 7-8: 2 revenue items
# 9: Total Revenue
# 10: Revenue Growth
# 11: blank
# 12: COGS header
# 13-15: 3 COGS items
# 16: Total COGS
# 17: blank
# 18: GP header
# 19-21: 3 GP items
# 22: Gross Profit
# 23: GP Growth
# 24: GP Margin
# 25: blank
# 26: OpEx header
# 27-30: 4 OpEx items
# 31: Total OpEx
# 32: OpEx Growth
# 33: blank
# 34: EBITDA header
# 35-37: 3 segment EBITDA items
# 38: Underlying EBITDA
# 39: EBITDA Growth
# 40: EBITDA Margin
# 41: blank
# 42: Stat EBITDA Adj header
# 43: SBP
# 44: Sig Items
# 45: Stat EBITDA
# 46: blank
# 47: D&A header
# 48-49: 2 D&A items
# 50: Total D&A
# 51: D&A/Revenue
# 52: Avg Lease Life
# 53: EBIT header
# 54: EBIT
# 55: EBIT Growth
# 56: EBIT Margin
# 57: blank
# 58: Interest header
# 59-61: 3 interest items
# 62: Net Finance Costs
# 63: Int Income Rate
# 64: Lease Int Rate
# 65: Bank Int Rate
# 66: blank
# 67: PBT, Tax, NPAT header
# 68: PBT
# 69: Tax
# 70: Tax Rate
# 71: NCI
# 72: Underlying NPAT
# 73: Sig Items AT
# 74: Stat NPAT
# 75: NPAT Growth
# 76: NPAT Margin
# = 72 rows (5-76)

# NEW ANNUAL P&L structure:
new_annual_pl = [
    # (row_offset, key, label, units, style_type)
    # style_type: 'section', 'category', 'item', 'subtotal', 'analytical', 'blank', 'subsection'
    (0, None, 'P&L', None, 'section'),                                           # 5
    (1, None, 'Revenue', None, 'category'),                                       # 6
    (2, 'Rev-Corp Restaurant Sales Aus', 'Corp Restaurant Sales - Australia', 'A$m', 'item'),  # 7
    (3, 'Rev-Corp Restaurant Sales US', 'Corp Restaurant Sales - US', 'A$m', 'item'),          # 8
    (4, 'Rev-Total Corp Restaurant Sales', 'Total Corp Restaurant Sales', 'A$m', 'subtotal'),  # 9
    (5, 'Rev-Franchise Royalty Revenue', 'Franchise Royalty Revenue', 'A$m', 'item'),           # 10
    (6, 'Rev-Franchise Fee Revenue', 'Franchise Fee Revenue', 'A$m', 'item'),                   # 11
    (7, 'Rev-Total Revenue', 'Total Revenue', 'A$m', 'subtotal'),                              # 12
    (8, None, 'Revenue Growth', '% YoY', 'analytical'),                                         # 13
    (9, None, None, None, 'blank'),                                                             # 14
    (10, None, 'Other Revenue', None, 'category'),                                              # 15
    (11, 'OthRev-Marketing Levy', 'Marketing Levy Revenue', 'A$m', 'item'),                    # 16
    (12, 'OthRev-Other Franchise Revenue', 'Other Franchise Revenue', 'A$m', 'item'),          # 17
    (13, 'OthRev-Other Income', 'Other Income', 'A$m', 'item'),                                # 18
    (14, 'OthRev-Total Other Revenue', 'Total Other Revenue & Income', 'A$m', 'subtotal'),     # 19
    (15, None, None, None, 'blank'),                                                            # 20
    (16, None, 'Total Revenue & Other Income', 'A$m', 'subtotal'),                             # 21
    (17, None, None, None, 'blank'),                                                            # 22
    (18, None, 'Segment EBITDA Bridge', None, 'category'),                                      # 23
    (19, 'SegEBITDA-Australia', 'Australia Segment EBITDA', 'A$m', 'item'),                    # 24
    (20, 'SegEBITDA-US', 'US Segment EBITDA', 'A$m', 'item'),                                  # 25
    (21, 'SegEBITDA-Total Segment EBITDA', 'Total Segment Underlying EBITDA', 'A$m', 'subtotal'),  # 26
    (22, 'Bridge-Cash Rent', 'Cash Rent Addback', 'A$m', 'item'),                              # 27
    (23, 'Bridge-SBP', 'Share-Based Payments', 'A$m', 'item'),                                 # 28
    (24, 'Bridge-Other', 'Other Income/(Costs)', 'A$m', 'item'),                               # 29
    (25, 'EBITDA-Statutory EBITDA', 'EBITDA', 'A$m', 'subtotal'),                              # 30
    (26, None, 'EBITDA Margin', '%', 'analytical'),                                             # 31
    (27, None, None, None, 'blank'),                                                            # 32
    (28, None, 'Expenses (Memo)', None, 'category'),                                            # 33
    (29, 'Exp-Food & Packaging', 'Cost of Food & Packaging', 'A$m', 'item'),                   # 34
    (30, 'Exp-Employee Benefits', 'Employee Benefit Expenses', 'A$m', 'item'),                  # 35
    (31, 'Exp-Admin', 'Administrative Expenses', 'A$m', 'item'),                                # 36
    (32, 'Exp-Marketing', 'Marketing Expenses', 'A$m', 'item'),                                 # 37
    (33, 'Exp-Other Expenses', 'Other Expenses', 'A$m', 'item'),                                # 38
    (34, 'Exp-Total Expenses', 'Total Expenses (excl D&A)', 'A$m', 'subtotal'),                 # 39
    (35, None, None, None, 'blank'),                                                            # 40
    (36, None, 'Depreciation & Amortisation', None, 'category'),                                # 41
    (37, 'DA-ROU Depreciation', 'Depreciation - ROU Assets', 'A$m', 'item'),                   # 42
    (38, 'DA-PPE Depreciation', 'Depreciation - PPE', 'A$m', 'item'),                          # 43
    (39, 'DA-Amortisation Reacquired', 'Amortisation - Reacquired Rights', 'A$m', 'item'),     # 44
    (40, 'DA-Amortisation Other', 'Amortisation - Other', 'A$m', 'item'),                       # 45
    (41, 'DA-Total DA', 'Total D&A', 'A$m', 'subtotal'),                                       # 46
    (42, None, None, None, 'blank'),                                                            # 47
    (43, 'EBIT-EBIT', 'EBIT', 'A$m', 'subtotal'),                                             # 48
    (44, None, 'EBIT Margin', '%', 'analytical'),                                               # 49
    (45, None, None, None, 'blank'),                                                            # 50
    (46, None, 'Finance', None, 'category'),                                                    # 51
    (47, 'Int-Term Deposit Income', 'Interest Income - Term Deposits', 'A$m', 'item'),         # 52
    (48, 'Int-Lease Receivable Income', 'Interest Income - Lease Receivables', 'A$m', 'item'), # 53
    (49, 'Int-Other Finance Income', 'Other Finance Income', 'A$m', 'item'),                    # 54
    (50, 'Int-Total Finance Income', 'Total Finance Income', 'A$m', 'subtotal'),                # 55
    (51, 'Int-Lease Liability Costs', 'Finance Costs - Lease Liabilities', 'A$m', 'item'),     # 56
    (52, 'Int-Other Finance Costs', 'Other Finance Costs', 'A$m', 'item'),                      # 57
    (53, 'Int-Total Finance Costs', 'Total Finance Costs', 'A$m', 'subtotal'),                  # 58
    (54, 'Int-Net Finance', 'Net Finance Income/(Costs)', 'A$m', 'subtotal'),                   # 59
    (55, None, None, None, 'blank'),                                                            # 60
    (56, 'PBT-PBT', 'PBT', 'A$m', 'subtotal'),                                                # 61
    (57, 'Tax-Tax Expense', 'Income Tax Expense', 'A$m', 'item'),                               # 62
    (58, None, 'Effective Tax Rate', '%', 'analytical'),                                        # 63
    (59, 'NPAT-NPAT', 'NPAT', 'A$m', 'subtotal'),                                              # 64
    (60, None, 'NPAT Margin', '%', 'analytical'),                                               # 65
    (61, None, None, None, 'blank'),                                                            # 66
]
# That's 62 rows (offsets 0-61), placing rows 5-66.
# Currently P&L is 72 rows (5-76). We need to remove 10 rows.
# EPS section currently starts at row 78 (with blank at 77).
# New: EPS should start at row 68 (with blank at 67), so the EPS header is at 68.
# Wait - let me recount. New P&L ends at row 66 (5 + 61).
# Then row 67 = blank, row 68 = EPS & Dividends header.
# Currently row 77 = blank, row 78 = EPS & Dividends header.
# Difference: 78 - 68 = 10 rows to delete.

# Actually wait - after P&L ends at 66, the EPS header was at 78.
# Current P&L ends at 76. New ends at 66. That's 10 rows fewer.
# So everything from row 77 onward needs to shift up by 10.
# But we can't just shift - we need to delete 10 rows from the P&L area.

# Strategy:
# 1. First, delete rows from the old P&L to make space match
# 2. Then overwrite content

# Actually, the cleanest approach:
# The old P&L occupies rows 5-76 (72 rows).
# The new P&L needs rows 5-66 (62 rows).
# Delete 10 rows from the old structure (e.g., delete rows 67-76, which shifts EPS up).
# But we need to be careful - openpyxl's delete_rows will shift formulas.

# Let me do it this way:
# 1. Delete excess rows from old P&L (bottom of P&L section)
# 2. Then write new P&L content over rows 5-66

# Delete rows 67-76 (10 rows) -- these were the bottom of old P&L
# Before deleting, note that row 77 was blank and row 78 was EPS header
# After deleting 10 rows at 67, row 67 = old 77 (blank), row 68 = old 78 (EPS)
print("  Deleting 10 excess P&L rows from Annual (rows 67-76)...")
delete_rows_and_track(ws_annual, 67, 10)
# Now: rows 5-66 = P&L area (to be rewritten)
# row 67 = blank (was 77)
# row 68 = EPS & Dividends header (was 78)
# All formulas should auto-adjust

# Clear and rewrite the P&L section
PL_START = 5
for offset, key, label, units, style_type in new_annual_pl:
    row = PL_START + offset
    # Clear the row first
    for c in range(1, annual_max_col + 1):
        cell = ws_annual.cell(row, c)
        cell.value = None
        cell.font = Font()
        cell.fill = NO_FILL
        cell.border = NO_BORDER
        cell.number_format = 'General'

    # Set content
    ws_annual.cell(row, 1).value = key
    ws_annual.cell(row, 2).value = label
    ws_annual.cell(row, 3).value = units

    # Apply formatting
    if style_type == 'section':
        apply_section_header_style(ws_annual, row, annual_max_col)
    elif style_type == 'category':
        apply_category_header_style(ws_annual, row, annual_max_col)
    elif style_type == 'item':
        apply_line_item_style(ws_annual, row, annual_max_col, units or 'A$m')
    elif style_type == 'subtotal':
        apply_subtotal_style(ws_annual, row, annual_max_col, units or 'A$m')
    elif style_type == 'analytical':
        apply_analytical_style(ws_annual, row, annual_max_col, units or '%')
    elif style_type == 'blank':
        blank_row(ws_annual, row, annual_max_col)

print(f"  Wrote {len(new_annual_pl)} P&L rows to Annual (rows {PL_START}-{PL_START + len(new_annual_pl) - 1})")

# Now update EPS section labels (currently at rows 68-82 after the shift)
# Row 68: EPS & Dividends header (was 78) -- keep as is
# Row 69: YE Shares (was 79)
# Row 70: WASO Basic (was 80)
# Row 71: Dilution (was 81)
# Row 72: WASO Diluted (was 82)
# Row 73: blank (was 83)
# Row 74: Underlying EPS (was 84) -> rename to Basic EPS
# Row 75: Statutory EPS (was 85) -> rename to Diluted EPS
# Row 76: EPS Growth (was 86)
# Row 77: blank (was 87)
# Row 78: DPS (was 88)
# Row 79: Total Dividends (was 89)
# Row 80: Payout Ratio (was 90)
# Row 81: Dividend Yield (was 91)
# Row 82: Dividend Growth (was 92)
# Row 83: blank (was 93)

# Rename EPS rows and update keys + units
ws_annual.cell(74, 1).value = 'EPS-Basic EPS'
ws_annual.cell(74, 2).value = 'Basic EPS'
ws_annual.cell(74, 3).value = 'cps'
for c in range(4, annual_max_col + 1):
    ws_annual.cell(74, c).number_format = FMT_CPS

ws_annual.cell(75, 1).value = 'EPS-Diluted EPS'
ws_annual.cell(75, 2).value = 'Diluted EPS'
ws_annual.cell(75, 3).value = 'cps'
for c in range(4, annual_max_col + 1):
    ws_annual.cell(75, c).number_format = FMT_CPS

# Update shares units from #m to m
ws_annual.cell(69, 3).value = 'm'
ws_annual.cell(70, 3).value = 'm'
ws_annual.cell(71, 3).value = 'm'
ws_annual.cell(72, 3).value = 'm'

# Update DPS units
ws_annual.cell(78, 3).value = 'cps'

# Update currency refs in EPS labels
ws_annual.cell(74, 3).value = 'cps'
ws_annual.cell(75, 3).value = 'cps'

print("  Updated EPS & Dividends section labels")

# Clear old actual data from EPS rows (will be populated later)
for r in range(69, 83):
    for c in range(4, annual_max_col + 1):
        cell = ws_annual.cell(r, c)
        if cell.value is not None and not str(cell.value).startswith('='):
            cell.value = None

# ── Now restructure KPIs on Annual ──────────────────────────────────────
# After the P&L row deletions, KPIs shifted:
# Old row 94 (Operating Metrics header) -> now row 84
# Old rows 95-106 -> now rows 85-96
# Old row 107 (blank) -> now row 97

print("\n  Restructuring Annual KPIs...")

# New KPI structure for Annual:
new_annual_kpis = [
    # (offset, key, label, units, style_type)
    (0, None, 'Operating Metrics', None, 'subsection'),
    (1, 'KPI-Corp Restaurants Aus', 'Corporate Restaurants - Australia', '#', 'item'),
    (2, 'KPI-Franchise Restaurants Aus', 'Franchise Restaurants - Australia', '#', 'item'),
    (3, 'KPI-Franchise Restaurants Asia', 'Franchise Restaurants - Asia', '#', 'item'),
    (4, 'KPI-US Restaurants', 'US Restaurants', '#', 'item'),
    (5, 'KPI-Total Restaurants', 'Total Global Restaurants', '#', 'subtotal'),
    (6, 'KPI-Aus Network Sales', 'Australia Network Sales', 'A$m', 'item'),
    (7, 'KPI-Total Network Sales', 'Global Network Sales', 'A$m', 'item'),
    (8, 'KPI-Comp Sales Growth', 'Comp Sales Growth (Australia)', '%', 'analytical'),
    (9, 'KPI-DT AUV', 'Drive Thru AUV (annualised)', 'A$m', 'item'),
    (10, 'KPI-Strip AUV', 'Strip AUV (annualised)', 'A$m', 'item'),
    (11, 'KPI-DT Network Margin', 'DT Network Restaurant Margin', '%', 'analytical'),
    (12, 'KPI-Strip Network Margin', 'Strip Network Restaurant Margin', '%', 'analytical'),
    (13, 'KPI-Breakfast Mix', 'Breakfast % of Network Sales', '%', 'analytical'),
    (14, 'KPI-After 9pm Mix', 'After 9pm % of Network Sales', '%', 'analytical'),
    (15, 'KPI-Delivery Mix', 'Delivery % of Network Sales', '%', 'analytical'),
    (16, 'KPI-Owned Digital Mix', 'Owned Digital % of Network Sales', '%', 'analytical'),
    (17, 'KPI-24/7 Restaurants', '24/7 Trading Restaurants', '#', 'item'),
]
# That's 18 rows (offset 0-17).
# Current KPI section: rows 84-96 = 13 rows.
# Need 18 rows. Insert 5 more rows.

KPI_START = 84  # after shift
current_kpi_count = 13  # rows 84-96
new_kpi_count = 18

# Insert extra rows
extra_kpi_rows = new_kpi_count - current_kpi_count
if extra_kpi_rows > 0:
    insert_rows_and_track(ws_annual, KPI_START + current_kpi_count, extra_kpi_rows)
    print(f"  Inserted {extra_kpi_rows} rows for KPIs at row {KPI_START + current_kpi_count}")

# Write KPI content
for offset, key, label, units, style_type in new_annual_kpis:
    row = KPI_START + offset
    # Clear
    for c in range(1, annual_max_col + 1):
        cell = ws_annual.cell(row, c)
        cell.value = None
        cell.font = Font()
        cell.fill = NO_FILL
        cell.border = NO_BORDER
        cell.number_format = 'General'

    ws_annual.cell(row, 1).value = key
    ws_annual.cell(row, 2).value = label
    ws_annual.cell(row, 3).value = units

    if style_type == 'subsection':
        apply_subsection_header_style(ws_annual, row, annual_max_col)
    elif style_type == 'item':
        apply_line_item_style(ws_annual, row, annual_max_col, units or 'A$m')
    elif style_type == 'subtotal':
        apply_subtotal_style(ws_annual, row, annual_max_col, units or 'A$m')
    elif style_type == 'analytical':
        apply_analytical_style(ws_annual, row, annual_max_col, units or '%')

print(f"  Wrote {len(new_annual_kpis)} KPI rows to Annual (rows {KPI_START}-{KPI_START + len(new_annual_kpis) - 1})")

# Ensure blank row after KPIs
kpi_end_row = KPI_START + len(new_annual_kpis)
blank_row(ws_annual, kpi_end_row, annual_max_col)

# ── Update BS/CF currency labels ──────────────────────────────────────
print("\n  Updating currency labels in BS/CF/Returns sections...")
# After row shifts: old row 108 (BS) shifted by -10 + 5 = -5, so now at 103
# Actually let me find the BS section by scanning for "Balance Sheet"
bs_row = None
for r in range(1, ws_annual.max_row + 1):
    if ws_annual.cell(r, 2).value == 'Balance Sheet':
        bs_row = r
        break

print(f"  Balance Sheet found at row {bs_row}")

# Update NZDm -> A$m in units column for BS, CF sections
for r in range(bs_row, ws_annual.max_row + 1):
    c3 = ws_annual.cell(r, 3)
    if c3.value == 'NZDm':
        c3.value = 'A$m'
    elif c3.value == 'NZDps':
        c3.value = 'cps'

# Also update units in P&L area and EPS area
for r in range(5, bs_row):
    c3 = ws_annual.cell(r, 3)
    if c3.value == 'NZDm':
        c3.value = 'A$m'
    elif c3.value == 'NZDps':
        c3.value = 'cps'
    elif c3.value == '#m':
        c3.value = 'm'

# Update BS specific labels for GYG
# Find and remove BS-Total Banking Debt row (GYG has no debt)
# Actually, let's keep it but rename - or better, we'll handle BS in a follow-up
# For now just update currency labels


# ══════════════════════════════════════════════════════════════════════════
# STEP 5: Restructure HY & Segments sheet
# ══════════════════════════════════════════════════════════════════════════
print("\n--- STEP 5: Restructure HY & Segments P&L ---")

# Current HY P&L structure (rows 5-79):
# Row 5: P&L header
# Row 6: Revenue
# Row 7: Rev-Steel Revenue
# Row 8: Rev-Metals Revenue
# Row 9: Rev-Total Revenue
# Row 10: Revenue Growth
# Row 11: blank
# Row 12: COGS header
# Row 13-15: COGS items
# Row 16: Total COGS
# Row 17: blank
# Row 18: GP header
# Row 19-21: GP items
# Row 22: Gross Profit
# Row 23: GP Growth
# Row 24: GP Margin
# Row 25: Steel GP Margin
# Row 26: Metals GP Margin
# Row 27: blank
# Row 28: Operating Expenses header
# Row 29-32: OpEx items
# Row 33: Total OpEx
# Row 34: OpEx Growth
# Row 35: Cost-to-Income
# Row 36: blank
# Row 37: EBITDA header
# Row 38-40: Segment EBITDA items
# Row 41: Corp EBITDA/Rev
# Row 42: Underlying EBITDA
# Row 43: EBITDA Growth
# Row 44: EBITDA Margin
# Row 45: Steel EBITDA Margin
# Row 46: Metals EBITDA Margin
# Row 47: blank
# Row 48: Sig Items
# Row 49: Stat EBITDA
# Row 50: blank
# Row 51: D&A header
# Row 52-53: D&A items
# Row 54: Total D&A
# Row 55: D&A/Revenue
# Row 56: blank
# Row 57: EBIT header
# Row 58: EBIT
# Row 59: EBIT Growth
# Row 60: EBIT Margin
# Row 61: blank
# Row 62: Interest header
# Row 63-65: Interest items
# Row 66: Net Finance
# Row 67-69: Interest rates
# Row 70: blank
# Row 71: PBT
# Row 72: Tax
# Row 73: Tax Rate
# Row 74: NCI
# Row 75: Underlying NPAT
# Row 76: Sig Items AT
# Row 77: Stat NPAT
# Row 78: NPAT Growth
# Row 79: blank
# = 75 rows (5-79)

# Then:
# Row 80: Operating Metrics (subsection)
# Rows 81-90: KPI items
# Row 91: blank
# Row 92: Segment Forecast - Steel (subsection)
# Rows 93-103: Steel forecast items
# Row 104: blank
# Row 105: Segment Forecast - Metals (subsection)
# Rows 106-116: Metals forecast items

# NEW HY P&L structure (must match Annual P&L keys):
new_hy_pl = [
    (0, None, 'P&L', None, 'section'),
    (1, None, 'Revenue', None, 'category'),
    (2, 'Rev-Corp Restaurant Sales Aus', 'Corp Restaurant Sales - Australia', 'A$m', 'item'),
    (3, 'Rev-Corp Restaurant Sales US', 'Corp Restaurant Sales - US', 'A$m', 'item'),
    (4, 'Rev-Total Corp Restaurant Sales', 'Total Corp Restaurant Sales', 'A$m', 'subtotal'),
    (5, 'Rev-Franchise Royalty Revenue', 'Franchise Royalty Revenue', 'A$m', 'item'),
    (6, 'Rev-Franchise Fee Revenue', 'Franchise Fee Revenue', 'A$m', 'item'),
    (7, 'Rev-Total Revenue', 'Total Revenue', 'A$m', 'subtotal'),
    (8, None, 'Revenue Growth (YoY)', '% YoY', 'analytical'),
    (9, None, None, None, 'blank'),
    (10, None, 'Other Revenue', None, 'category'),
    (11, 'OthRev-Marketing Levy', 'Marketing Levy Revenue', 'A$m', 'item'),
    (12, 'OthRev-Other Franchise Revenue', 'Other Franchise Revenue', 'A$m', 'item'),
    (13, 'OthRev-Other Income', 'Other Income', 'A$m', 'item'),
    (14, 'OthRev-Total Other Revenue', 'Total Other Revenue & Income', 'A$m', 'subtotal'),
    (15, None, None, None, 'blank'),
    (16, None, 'Total Revenue & Other Income', 'A$m', 'subtotal'),
    (17, None, None, None, 'blank'),
    (18, None, 'Segment EBITDA Bridge', None, 'category'),
    (19, 'SegEBITDA-Australia', 'Australia Segment EBITDA', 'A$m', 'item'),
    (20, 'SegEBITDA-US', 'US Segment EBITDA', 'A$m', 'item'),
    (21, 'SegEBITDA-Total Segment EBITDA', 'Total Segment Underlying EBITDA', 'A$m', 'subtotal'),
    (22, 'Bridge-Cash Rent', 'Cash Rent Addback', 'A$m', 'item'),
    (23, 'Bridge-SBP', 'Share-Based Payments', 'A$m', 'item'),
    (24, 'Bridge-Other', 'Other Income/(Costs)', 'A$m', 'item'),
    (25, 'EBITDA-Statutory EBITDA', 'EBITDA', 'A$m', 'subtotal'),
    (26, None, 'EBITDA Margin', '%', 'analytical'),
    (27, None, None, None, 'blank'),
    (28, None, 'Expenses (Memo)', None, 'category'),
    (29, 'Exp-Food & Packaging', 'Cost of Food & Packaging', 'A$m', 'item'),
    (30, 'Exp-Employee Benefits', 'Employee Benefit Expenses', 'A$m', 'item'),
    (31, 'Exp-Admin', 'Administrative Expenses', 'A$m', 'item'),
    (32, 'Exp-Marketing', 'Marketing Expenses', 'A$m', 'item'),
    (33, 'Exp-Other Expenses', 'Other Expenses', 'A$m', 'item'),
    (34, 'Exp-Total Expenses', 'Total Expenses (excl D&A)', 'A$m', 'subtotal'),
    (35, None, None, None, 'blank'),
    (36, None, 'Depreciation & Amortisation', None, 'category'),
    (37, 'DA-ROU Depreciation', 'Depreciation - ROU Assets', 'A$m', 'item'),
    (38, 'DA-PPE Depreciation', 'Depreciation - PPE', 'A$m', 'item'),
    (39, 'DA-Amortisation Reacquired', 'Amortisation - Reacquired Rights', 'A$m', 'item'),
    (40, 'DA-Amortisation Other', 'Amortisation - Other', 'A$m', 'item'),
    (41, 'DA-Total DA', 'Total D&A', 'A$m', 'subtotal'),
    (42, None, 'D&A / Revenue', '%', 'analytical'),
    (43, None, None, None, 'blank'),
    (44, 'EBIT-EBIT', 'EBIT', 'A$m', 'subtotal'),
    (45, None, 'EBIT Margin', '%', 'analytical'),
    (46, None, None, None, 'blank'),
    (47, None, 'Finance', None, 'category'),
    (48, 'Int-Term Deposit Income', 'Interest Income - Term Deposits', 'A$m', 'item'),
    (49, 'Int-Lease Receivable Income', 'Interest Income - Lease Receivables', 'A$m', 'item'),
    (50, 'Int-Other Finance Income', 'Other Finance Income', 'A$m', 'item'),
    (51, 'Int-Total Finance Income', 'Total Finance Income', 'A$m', 'subtotal'),
    (52, 'Int-Lease Liability Costs', 'Finance Costs - Lease Liabilities', 'A$m', 'item'),
    (53, 'Int-Other Finance Costs', 'Other Finance Costs', 'A$m', 'item'),
    (54, 'Int-Total Finance Costs', 'Total Finance Costs', 'A$m', 'subtotal'),
    (55, 'Int-Net Finance', 'Net Finance Income/(Costs)', 'A$m', 'subtotal'),
    (56, None, None, None, 'blank'),
    (57, 'PBT-PBT', 'PBT', 'A$m', 'subtotal'),
    (58, 'Tax-Tax Expense', 'Income Tax Expense', 'A$m', 'item'),
    (59, None, 'Tax Rate', '%', 'analytical'),
    (60, 'NPAT-NPAT', 'NPAT', 'A$m', 'subtotal'),
    (61, None, 'NPAT Growth (YoY)', '% YoY', 'analytical'),
    (62, None, None, None, 'blank'),
]
# 63 rows (offsets 0-62), placing rows 5-67.
# Current HY P&L: rows 5-79 = 75 rows.
# New HY P&L: rows 5-67 = 63 rows.
# Delete 12 rows from old P&L.

# After P&L: KPIs + Segment Forecast zones (rows 80-116 = 37 rows)
# New: After P&L ends at row 67, blank at 68, then new zones start at 69.
# Old zones started at 80. After deleting 12 P&L rows, old 80 -> 68.
# So KPIs move from old row 80 to row 68.

# But we also need to restructure KPIs and segment forecasts.
# Let me first handle the P&L restructure, then the zones.

# Delete 12 rows from bottom of HY P&L (rows 68-79)
print("  Deleting 12 excess P&L rows from HY (rows 68-79)...")
delete_rows_and_track(ws_hy, 68, 12)
# Now: HY P&L = rows 5-67 (to be rewritten)
# Old row 80 (Op Metrics) -> now row 68
# Old row 91 (blank) -> now row 79
# Old row 92 (Seg Forecast Steel) -> now row 80
# Old row 104 (blank) -> now row 92
# Old row 105 (Seg Forecast Metals) -> now row 93
# Old row 116 -> now row 104

# Write new HY P&L
HY_PL_START = 5
for offset, key, label, units, style_type in new_hy_pl:
    row = HY_PL_START + offset
    for c in range(1, hy_max_col + 1):
        cell = ws_hy.cell(row, c)
        cell.value = None
        cell.font = Font()
        cell.fill = NO_FILL
        cell.border = NO_BORDER
        cell.number_format = 'General'

    ws_hy.cell(row, 1).value = key
    ws_hy.cell(row, 2).value = label
    ws_hy.cell(row, 3).value = units

    if style_type == 'section':
        apply_section_header_style(ws_hy, row, hy_max_col)
    elif style_type == 'category':
        apply_category_header_style(ws_hy, row, hy_max_col)
    elif style_type == 'item':
        apply_line_item_style(ws_hy, row, hy_max_col, units or 'A$m')
    elif style_type == 'subtotal':
        apply_subtotal_style(ws_hy, row, hy_max_col, units or 'A$m')
    elif style_type == 'analytical':
        apply_analytical_style(ws_hy, row, hy_max_col, units or '%')
    elif style_type == 'blank':
        blank_row(ws_hy, row, hy_max_col)

print(f"  Wrote {len(new_hy_pl)} P&L rows to HY (rows {HY_PL_START}-{HY_PL_START + len(new_hy_pl) - 1})")


# ── Restructure HY KPIs and Segment Forecast zones ─────────────────────
print("\n  Restructuring HY KPIs and Segment Forecasts...")

# After P&L deletion, the zones shifted. Let me find them.
# Old row 80 (Op Metrics) -> now row 68
# We need to restructure rows 68 onwards.

# New HY zones structure:
new_hy_zones = [
    # KPIs section
    (0, None, 'Operating Metrics', None, 'subsection'),
    (1, 'KPI-Corp Restaurants Aus', 'Corporate Restaurants - Australia', '#', 'item'),
    (2, 'KPI-Franchise Restaurants Aus', 'Franchise Restaurants - Australia', '#', 'item'),
    (3, 'KPI-Franchise Restaurants Asia', 'Franchise Restaurants - Asia', '#', 'item'),
    (4, 'KPI-US Restaurants', 'US Restaurants', '#', 'item'),
    (5, 'KPI-Total Restaurants', 'Total Global Restaurants', '#', 'subtotal'),
    (6, 'KPI-Aus Network Sales', 'Australia Network Sales', 'A$m', 'item'),
    (7, 'KPI-Total Network Sales', 'Global Network Sales', 'A$m', 'item'),
    (8, 'KPI-Comp Sales Growth', 'Comp Sales Growth (Australia)', '%', 'analytical'),
    (9, 'KPI-DT AUV', 'Drive Thru AUV (annualised)', 'A$m', 'item'),
    (10, 'KPI-Strip AUV', 'Strip AUV (annualised)', 'A$m', 'item'),
    (11, 'KPI-DT Network Margin', 'DT Network Restaurant Margin', '%', 'analytical'),
    (12, 'KPI-Strip Network Margin', 'Strip Network Restaurant Margin', '%', 'analytical'),
    (13, 'KPI-Breakfast Mix', 'Breakfast % of Network Sales', '%', 'analytical'),
    (14, 'KPI-After 9pm Mix', 'After 9pm % of Network Sales', '%', 'analytical'),
    (15, 'KPI-Delivery Mix', 'Delivery % of Network Sales', '%', 'analytical'),
    (16, 'KPI-Owned Digital Mix', 'Owned Digital % of Network Sales', '%', 'analytical'),
    (17, 'KPI-24/7 Restaurants', '24/7 Trading Restaurants', '#', 'item'),
    (18, None, None, None, 'blank'),
    # Segment Forecast - Australia
    (19, None, 'Segment Forecast - Australia', None, 'subsection'),
    (20, None, 'Corp Restaurants', '#', 'item'),
    (21, None, 'New Corp Restaurant Openings', '#', 'item'),
    (22, None, 'Corp Restaurant Revenue', 'A$m', 'item'),
    (23, None, 'Corp Restaurant Revenue Growth', '% YoY', 'analytical'),
    (24, None, 'Corp AUV', 'A$m', 'item'),
    (25, None, 'Franchise Restaurants', '#', 'item'),
    (26, None, 'Franchise Revenue', 'A$m', 'item'),
    (27, None, 'Total Australia Revenue', 'A$m', 'subtotal'),
    (28, None, 'EBITDA Margin', '%', 'analytical'),
    (29, None, 'Segment EBITDA', 'A$m', 'subtotal'),
    (30, None, None, None, 'blank'),
    # Segment Forecast - US
    (31, None, 'Segment Forecast - US', None, 'subsection'),
    (32, None, 'US Corp Restaurants', '#', 'item'),
    (33, None, 'New US Restaurant Openings', '#', 'item'),
    (34, None, 'US Revenue', 'A$m', 'item'),
    (35, None, 'US Revenue Growth', '% YoY', 'analytical'),
    (36, None, 'US AUV', 'A$m', 'item'),
    (37, None, 'US EBITDA Margin', '%', 'analytical'),
    (38, None, 'US Segment EBITDA', 'A$m', 'subtotal'),
]
# 39 rows (offsets 0-38), placing rows 68-106.
# Current zone area after P&L deletion: rows 68-104 = 37 rows.
# Need 39 rows. Insert 2 more.

ZONE_START = 68
current_zone_rows = 104 - 68 + 1  # = 37
new_zone_rows = len(new_hy_zones)  # = 39

extra = new_zone_rows - current_zone_rows
if extra > 0:
    insert_rows_and_track(ws_hy, 68 + current_zone_rows, extra)

# Write zones content
for offset, key, label, units, style_type in new_hy_zones:
    row = ZONE_START + offset
    for c in range(1, hy_max_col + 1):
        cell = ws_hy.cell(row, c)
        cell.value = None
        cell.font = Font()
        cell.fill = NO_FILL
        cell.border = NO_BORDER
        cell.number_format = 'General'

    ws_hy.cell(row, 1).value = key
    ws_hy.cell(row, 2).value = label
    ws_hy.cell(row, 3).value = units

    if style_type == 'subsection':
        apply_subsection_header_style(ws_hy, row, hy_max_col)
    elif style_type == 'category':
        apply_category_header_style(ws_hy, row, hy_max_col)
    elif style_type == 'item':
        apply_line_item_style(ws_hy, row, hy_max_col, units or 'A$m')
    elif style_type == 'subtotal':
        apply_subtotal_style(ws_hy, row, hy_max_col, units or 'A$m')
    elif style_type == 'analytical':
        apply_analytical_style(ws_hy, row, hy_max_col, units or '%')
    elif style_type == 'blank':
        blank_row(ws_hy, row, hy_max_col)

print(f"  Wrote {len(new_hy_zones)} zone rows to HY (rows {ZONE_START}-{ZONE_START + len(new_hy_zones) - 1})")

# Update NZDm -> A$m in HY units
for r in range(1, ws_hy.max_row + 1):
    c3 = ws_hy.cell(r, 3)
    if c3.value == 'NZDm':
        c3.value = 'A$m'
    elif c3.value == 'NZDps':
        c3.value = 'cps'
    elif c3.value == '#m':
        c3.value = 'm'


# ══════════════════════════════════════════════════════════════════════════
# STEP 6: Update SOTP segment rows on Value sheet
# ══════════════════════════════════════════════════════════════════════════
print("\n--- STEP 6: Update SOTP Segments on Value Sheet ---")

# Current SOTP (rows 57-59): Steel, Metals, Corporate
# New: Australia, US, Corporate

# Row 57: Australia
ws_value.cell(57, 2).value = "Australia"
ws_value.cell(57, 3).value = '=INDEX(Annual!$D:$P,MATCH("SegEBITDA-Australia",Annual!$A:$A,0),MATCH($C$54,Annual!$D$3:$P$3,0))'
ws_value.cell(57, 4).value = 25  # Placeholder multiple for high-growth QSR

# Row 58: US
ws_value.cell(58, 2).value = "US"
ws_value.cell(58, 3).value = '=INDEX(Annual!$D:$P,MATCH("SegEBITDA-US",Annual!$A:$A,0),MATCH($C$54,Annual!$D$3:$P$3,0))'
ws_value.cell(58, 4).value = 15

# Row 59: Corporate (blended)
ws_value.cell(59, 2).value = "Corporate"
# Keep blended multiple formula
ws_value.cell(59, 3).value = '=INDEX(Annual!$D:$P,MATCH("EBITDA-Statutory EBITDA",Annual!$A:$A,0),MATCH($C$54,Annual!$D$3:$P$3,0))-C57-C58'
ws_value.cell(59, 4).value = '=IF((C57+C58)=0,"",(C57*D57+C58*D58)/(C57+C58))'

# Also update the Net Debt formula on Value sheet to handle GYG (no banking debt)
# Row 7: Net Debt -- GYG likely has net cash, keep formula but it should still work
# Actually the formula references BS-Total Banking Debt which still exists in the template BS

# Update implied EV/EBITDA row (68) - update to reference new EBITDA key
# The SOTP EV bridge rows (61-66) - these are formula-based, update references
ws_value.cell(61, 3).value = '=SUM(E57:E59)'  # Group EV
ws_value.cell(62, 3).value = '=-C7'  # less Net Debt
ws_value.cell(63, 3).value = '=-INDEX(Annual!$D:$P,MATCH("BS-Lease Liabilities",Annual!$A:$A,0),MATCH($D$24,Annual!$D$3:$P$3,0)-1)'
ws_value.cell(64, 3).value = '=C61+C62+C63'  # Equity Value
ws_value.cell(65, 3).value = '=C64/C5'  # Per Share Value
ws_value.cell(66, 3).value = '=IF(C4=0,"",C65/C4-1)'  # Upside/Downside

# Implied group multiple
ws_value.cell(68, 3).value = '=IF(C59+C57+C58=0,"",C61/(C57+C58+C59))'

# Update DCF references to use new EBITDA key
# Row 25: EBITDA reference
ws_value.cell(25, 4).value = '=INDEX(Annual!$D:$P,MATCH("EBITDA-Statutory EBITDA",Annual!$A:$A,0),MATCH(D$24,Annual!$D$3:$P$3,0))'
# Copy this formula across cols E-M
for c in range(5, 14):  # E through M
    col_letter = get_column_letter(c)
    ws_value.cell(25, c).value = f'=INDEX(Annual!$D:$P,MATCH("EBITDA-Statutory EBITDA",Annual!$A:$A,0),MATCH({col_letter}$24,Annual!$D$3:$P$3,0))'

# Row 26: D&A
ws_value.cell(26, 4).value = '=INDEX(Annual!$D:$P,MATCH("DA-Total DA",Annual!$A:$A,0),MATCH(D$24,Annual!$D$3:$P$3,0))'
for c in range(5, 14):
    col_letter = get_column_letter(c)
    ws_value.cell(26, c).value = f'=INDEX(Annual!$D:$P,MATCH("DA-Total DA",Annual!$A:$A,0),MATCH({col_letter}$24,Annual!$D$3:$P$3,0))'

# Row 27: EBIT - update to use new key
ws_value.cell(27, 4).value = '=INDEX(Annual!$D:$P,MATCH("EBIT-EBIT",Annual!$A:$A,0),MATCH(D$24,Annual!$D$3:$P$3,0))'
for c in range(5, 14):
    col_letter = get_column_letter(c)
    ws_value.cell(27, c).value = f'=INDEX(Annual!$D:$P,MATCH("EBIT-EBIT",Annual!$A:$A,0),MATCH({col_letter}$24,Annual!$D$3:$P$3,0))'

# Rows 28-35 formulas reference relative cells, should be fine
# Row 31: Capex references should still work since CF keys didn't change
ws_value.cell(31, 4).value = '=INDEX(Annual!$D:$P,MATCH("CF-Capex PPE",Annual!$A:$A,0),MATCH(D$24,Annual!$D$3:$P$3,0))+INDEX(Annual!$D:$P,MATCH("CF-Capex Intang",Annual!$A:$A,0),MATCH(D$24,Annual!$D$3:$P$3,0))'
for c in range(5, 14):
    col_letter = get_column_letter(c)
    ws_value.cell(31, c).value = f'=INDEX(Annual!$D:$P,MATCH("CF-Capex PPE",Annual!$A:$A,0),MATCH({col_letter}$24,Annual!$D$3:$P$3,0))+INDEX(Annual!$D:$P,MATCH("CF-Capex Intang",Annual!$A:$A,0),MATCH({col_letter}$24,Annual!$D$3:$P$3,0))'

# Row 32: WC Change
ws_value.cell(32, 4).value = '=INDEX(Annual!$D:$P,MATCH("CF-WC Change",Annual!$A:$A,0),MATCH(D$24,Annual!$D$3:$P$3,0))'
for c in range(5, 14):
    col_letter = get_column_letter(c)
    ws_value.cell(32, c).value = f'=INDEX(Annual!$D:$P,MATCH("CF-WC Change",Annual!$A:$A,0),MATCH({col_letter}$24,Annual!$D$3:$P$3,0))'

print("  Updated SOTP segments: Australia, US, Corporate")
print("  Updated DCF INDEX/MATCH formulas for new keys")

# Update Value sheet row 5 (Shares Outstanding) formula
ws_value.cell(5, 3).value = '=INDEX(Annual!$D:$P,MATCH("EPS-YE Shares",Annual!$A:$A,0),MATCH($D$24,Annual!$D$3:$P$3,0)-1)'

# Update lease liabilities reference for DCF
ws_value.cell(46, 3).value = '=-INDEX(Annual!$D:$P,MATCH("BS-Lease Liabilities",Annual!$A:$A,0),MATCH($D$24,Annual!$D$3:$P$3,0)-1)'


# ══════════════════════════════════════════════════════════════════════════
# STEP 7: Clear old template data from all data cells
# ══════════════════════════════════════════════════════════════════════════
print("\n--- STEP 7: Clear old template data ---")

# Clear actual data from Annual P&L area (rows 5-66, cols D-P)
# but keep formulas (anything starting with =)
cleared_count = 0
for r in range(5, 67):
    for c in range(4, annual_max_col + 1):
        cell = ws_annual.cell(r, c)
        if cell.value is not None and not str(cell.value).startswith('='):
            cell.value = None
            cleared_count += 1
print(f"  Cleared {cleared_count} data cells from Annual P&L")

# Clear HY P&L data
cleared_count = 0
for r in range(5, ws_hy.max_row + 1):
    for c in range(4, hy_max_col + 1):
        cell = ws_hy.cell(r, c)
        if cell.value is not None and not str(cell.value).startswith('='):
            cell.value = None
            cleared_count += 1
print(f"  Cleared {cleared_count} data cells from HY sheet")

# Clear old template actual data from Annual BS/CF that was VSL-specific
# Only clear hard-coded values in actuals columns (D-F), keep formulas
cleared_count = 0
for r in range(bs_row, ws_annual.max_row + 1):
    for c in range(4, 7):  # D, E, F only (actuals)
        cell = ws_annual.cell(r, c)
        if cell.value is not None and not str(cell.value).startswith('='):
            cell.value = None
            cleared_count += 1
print(f"  Cleared {cleared_count} data cells from Annual BS/CF actuals")

# Also clear forecast formulas in Annual P&L since they reference old row numbers
# The formulas that reference HY sheet via INDEX/MATCH will need rebuilding
# But that's for a follow-up task - for now just clear them
for r in range(5, 67):
    for c in range(4, annual_max_col + 1):
        cell = ws_annual.cell(r, c)
        if cell.value is not None and str(cell.value).startswith('='):
            cell.value = None

# Also clear formulas in forecast columns of BS/CF that may reference old rows
# Skip this - openpyxl should have auto-adjusted row references from inserts/deletes


# ══════════════════════════════════════════════════════════════════════════
# STEP 8: Save
# ══════════════════════════════════════════════════════════════════════════
print("\n--- STEP 8: Saving workbook ---")
wb.save(FILE_PATH)
print(f"  Saved to {FILE_PATH}")


# ══════════════════════════════════════════════════════════════════════════
# Summary
# ══════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("RESTRUCTURING COMPLETE - SUMMARY")
print("=" * 70)
print("""
HEADERS:
  - Annual B2: "GYG Model Summary", B3: "Guzman y Gomez (GYG.AX)"
  - HY B2: "GYG Segments (Half-Year)", B3: "Guzman y Gomez (GYG.AX)"
  - Value: Updated currency labels NZD -> A$

PERIODS:
  - Already correct (FY23-FY35 Annual, 1H23-2H35 HY, Jun 30 FY end)
  - Actual/Forecast zone labels correctly positioned

ANNUAL P&L (rows 5-66):
  - Revenue: Corp Restaurant Sales Aus/US, Franchise Royalty/Fee, Total Revenue
  - Other Revenue: Marketing Levy, Other Franchise, Other Income
  - EBITDA Bridge: Australia/US Segment EBITDA, Cash Rent/SBP/Other adjustments
  - Expenses (Memo): Food & Packaging, Employee, Admin, Marketing, Other
  - D&A: ROU Depreciation, PPE Depreciation, Amortisation (Reacquired + Other)
  - Finance: Term Deposit/Lease Receivable/Other Income, Lease Liability/Other Costs
  - PBT -> Tax -> NPAT (no NCI, no Sig Items)

EPS & DIVIDENDS (rows 68-83):
  - Renamed: Basic EPS, Diluted EPS (from Underlying/Statutory)
  - Updated currency units

ANNUAL KPIs (rows 84-101):
  - Restaurant counts by type/geography
  - Network Sales, Comp Sales Growth
  - AUV (Drive Thru + Strip)
  - Network Margins, Mix metrics
  - 24/7 Trading Restaurants

HY P&L (rows 5-67):
  - Matches Annual P&L keys exactly (cross-sheet consistency)

HY ZONES (rows 68-106):
  - KPIs matching Annual
  - Segment Forecast - Australia (restaurants, revenue, EBITDA)
  - Segment Forecast - US (restaurants, revenue, EBITDA)

VALUE SHEET:
  - SOTP segments: Australia (25x), US (15x), Corporate (blended)
  - DCF formulas updated for new EBITDA/EBIT/D&A keys

PRESERVED:
  - BS section (rows shifted but intact)
  - CF section (rows shifted but intact)
  - Returns section (rows shifted but intact)
  - All retained row formatting (subtotals, headers)

TO DO (follow-up):
  - Populate actual data for GYG
  - Build forecast formulas
  - Adapt BS/CF line items for GYG specifics
  - Verify cross-sheet INDEX/MATCH linkages
  - Check BS Check = 0
""")

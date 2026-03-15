"""Part 1: Annual sheet - headers, labels, P&L structure, and historical data."""
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import datetime

wb = openpyxl.load_workbook('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')

# ============================================================
# ANNUAL SHEET
# ============================================================
ws = wb['Annual']

# --- Headers ---
ws['B2'] = 'GYG Model Summary'
ws['B3'] = 'Guzman y Gomez (GYG.AX)'

# Row 1 year numbers already 2023-2035 (same FYE June) - keep
# Row 3 period labels already FY23A..FY35E - keep
# Row 4 dates already 30 June - keep

# --- Clear all data cells (D onwards) for rows 5+ ---
for r in range(5, ws.max_row + 1):
    for c in range(4, 17):  # cols D-P
        ws.cell(r, c).value = None

# ============================================================
# P&L SECTION - Rewrite labels
# ============================================================

# Row 5: P&L header (keep)
# Row 6: Revenue header (keep)

# Revenue rows - GYG has 4 rev lines + total
# R7: Aus Corp Sales
ws.cell(7, 1).value = 'Rev-Aus Corp Sales'
ws.cell(7, 2).value = 'Australia Corporate Restaurant Sales'
ws.cell(7, 3).value = 'AUDm'

# R8: Aus Franchise Rev
ws.cell(8, 1).value = 'Rev-Aus Franchise Rev'
ws.cell(8, 2).value = 'Australia Franchise & Other Revenue'
ws.cell(8, 3).value = 'AUDm'

# R9: Total Revenue - will be formula
ws.cell(9, 1).value = 'Rev-Total Revenue'
ws.cell(9, 2).value = 'Total Revenue'
ws.cell(9, 3).value = 'AUDm'

# R10: Revenue Growth (keep label, change unit)
ws.cell(10, 3).value = '% YoY'

# We need extra revenue rows for US. Let's use R11 for blank and repurpose.
# Actually VSL has R7=Steel Rev, R8=Metals Rev, R9=Total. GYG needs 4 rev lines.
# We need to INSERT rows. But instructions say preserve row structure where possible.
# Instead, let's use the existing structure more creatively:
# R7 = Aus Corp Sales, R8 = US Corp Sales (replacing Metals),
# But we also need Aus Franchise Rev and US Franchise Rev.
# The P&L on Annual doesn't need all 4 segments broken out - we can show
# just Total Revenue as the statutory line and have segments on HY sheet.
# Actually the task says to use the Column A codes listed. Let me use:
# R7 = Rev-Aus Corp Sales
# R8 = Rev-Aus Franchise Rev
# But we need R for US too. The statutory P&L just shows "Revenue" as one line.
# For the Annual sheet, let's show statutory P&L format:
# Actually, looking at the data, the statutory Revenue = 259.044 which is
# Corp Sales (Aus + US) + Franchise & Other Rev (Aus + US).
# Let me keep it simple: Total Revenue on one line, with segment detail on HY sheet.
# But the task explicitly says to use those Column A codes.
# Let me add rows. I'll shift content down.

# Actually, re-examining: VSL has R7=Steel Rev, R8=Metals Rev, R9=Total (3 rows for revenue).
# GYG needs 4 segment rev lines + total = 5 rows. I need 2 extra rows.
# But inserting rows will break all formulas. Better approach:
# Use existing rows and add the extra lines by repurposing empty rows.
# R7 = Rev-Aus Corp Sales
# R8 = Rev-Aus Franchise Rev
# R9 was Total Revenue - change to Rev-US Corp Sales
# R10 was Rev Growth - change to Rev-US Franchise Rev
# Then I need a new total and growth row.
# This is getting messy. Let me just keep the structure as-is with:
# R7 = Statutory Revenue (total)
# And put segment detail only on HY sheet. The Column A codes for Rev-Aus Corp Sales etc
# will be on the HY sheet segment forecast section.

# REVISED APPROACH: Annual sheet shows statutory P&L (single revenue line),
# with segment EBITDA breakdown. This matches how GYG reports.
# The detailed segment revenue codes go on HY & Segments sheet.

# R7: Total Revenue (statutory)
ws.cell(7, 1).value = 'Rev-Total Revenue'
ws.cell(7, 2).value = 'Total Revenue'
ws.cell(7, 3).value = 'AUDm'

# R8: Other Revenue & Income
ws.cell(8, 1).value = 'Rev-Other Revenue'
ws.cell(8, 2).value = 'Other Revenue & Income'
ws.cell(8, 3).value = 'AUDm'

# R9: Total Revenue + Other
ws.cell(9, 1).value = 'Rev-Total Rev Incl Other'
ws.cell(9, 2).value = 'Total Revenue (incl Other)'
ws.cell(9, 3).value = 'AUDm'

# R10: Revenue Growth
ws.cell(10, 2).value = 'Revenue Growth'
ws.cell(10, 3).value = '% YoY'

# --- COGS section ---
# R12: COGS header (keep)
# R13-R16: GYG has multiple cost lines

ws.cell(13, 1).value = 'COGS-Food & Packaging'
ws.cell(13, 2).value = 'Cost of Food & Packaging'
ws.cell(13, 3).value = 'AUDm'

ws.cell(14, 1).value = 'COGS-Employee Benefits'
ws.cell(14, 2).value = 'Employee Benefits'
ws.cell(14, 3).value = 'AUDm'

ws.cell(15, 1).value = 'COGS-Other Costs'
ws.cell(15, 2).value = 'Admin, Marketing & Other Expenses'
ws.cell(15, 3).value = 'AUDm'

ws.cell(16, 1).value = 'COGS-Total Costs'
ws.cell(16, 2).value = 'Total Operating Costs (excl D&A)'
ws.cell(16, 3).value = 'AUDm'

# --- Gross Profit section -> repurpose as Statutory EBITDA build ---
# R18: header
ws.cell(18, 2).value = 'Statutory EBITDA'

# R19-R24: repurpose
ws.cell(19, 1).value = 'GP-Statutory EBITDA'
ws.cell(19, 2).value = 'Statutory EBITDA'
ws.cell(19, 3).value = 'AUDm'

ws.cell(20, 1).value = None
ws.cell(20, 2).value = 'Statutory EBITDA Margin'
ws.cell(20, 3).value = '%'

ws.cell(21, 1).value = None
ws.cell(21, 2).value = None
ws.cell(21, 3).value = None

ws.cell(22, 1).value = None
ws.cell(22, 2).value = None
ws.cell(22, 3).value = None

ws.cell(23, 2).value = None
ws.cell(23, 3).value = None

ws.cell(24, 2).value = None
ws.cell(24, 3).value = None

# --- OpEx section -> Not needed separately, clear ---
ws.cell(26, 2).value = 'Segment EBITDA'
ws.cell(27, 1).value = None; ws.cell(27, 2).value = None; ws.cell(27, 3).value = None
ws.cell(28, 1).value = None; ws.cell(28, 2).value = None; ws.cell(28, 3).value = None
ws.cell(29, 1).value = None; ws.cell(29, 2).value = None; ws.cell(29, 3).value = None
ws.cell(30, 1).value = None; ws.cell(30, 2).value = None; ws.cell(30, 3).value = None
ws.cell(31, 1).value = None; ws.cell(31, 2).value = None; ws.cell(31, 3).value = None
ws.cell(32, 1).value = None; ws.cell(32, 2).value = None; ws.cell(32, 3).value = None

# --- EBITDA section ---
# R34: EBITDA header (keep)
ws.cell(35, 1).value = 'EBITDA-Australia EBITDA'
ws.cell(35, 2).value = 'Australia Segment Underlying EBITDA'
ws.cell(35, 3).value = 'AUDm'

ws.cell(36, 1).value = 'EBITDA-US EBITDA'
ws.cell(36, 2).value = 'US Segment Underlying EBITDA'
ws.cell(36, 3).value = 'AUDm'

ws.cell(37, 1).value = 'EBITDA-Corporate EBITDA'
ws.cell(37, 2).value = 'Corporate / Eliminations'
ws.cell(37, 3).value = 'AUDm'

ws.cell(38, 1).value = 'EBITDA-Underlying EBITDA'
ws.cell(38, 2).value = 'Group Segment Underlying EBITDA'
ws.cell(38, 3).value = 'AUDm'

ws.cell(39, 2).value = 'Seg EBITDA Growth'
ws.cell(39, 3).value = '% YoY'

ws.cell(40, 2).value = 'Seg EBITDA Margin (% NS)'
ws.cell(40, 3).value = '%'

# --- Statutory adjustments (rows 42-45) ---
ws.cell(42, 2).value = 'Statutory EBITDA Adjustments'

ws.cell(43, 1).value = 'Stat-AASB16'
ws.cell(43, 2).value = 'AASB 16 Lease Impact'
ws.cell(43, 3).value = 'AUDm'

ws.cell(44, 1).value = 'Stat-SBP'
ws.cell(44, 2).value = 'Share-Based Payments'
ws.cell(44, 3).value = 'AUDm'

ws.cell(45, 1).value = 'Stat-Other Adj'
ws.cell(45, 2).value = 'Other Adjustments'
ws.cell(45, 3).value = 'AUDm'

# Add a new row for Statutory EBITDA - but R45 was Stat-Statutory EBITDA in VSL.
# We need Stat-Other Adj AND Stat-Statutory EBITDA. Use R45 for Other Adj (done above).
# We don't have R46 free... Actually R45 in VSL = Stat-Statutory EBITDA.
# Let me reorganise: R43=AASB16, R44=SBP, R45=Other Adj is wrong because
# R45 was Statutory EBITDA in VSL and has formula =D38+D43+D44.
# I need 4 items (AASB16, SBP, Other, Stat EBITDA) in 3 slots (R43-R45).
# Solution: use R42 for AASB16 (was header), and keep R43-R45 for SBP, Other, Stat EBITDA.

ws.cell(42, 1).value = 'Stat-AASB16'
ws.cell(42, 2).value = 'AASB 16 Lease Impact'
ws.cell(42, 3).value = 'AUDm'

ws.cell(43, 1).value = 'Stat-SBP'
ws.cell(43, 2).value = 'Share-Based Payments'
ws.cell(43, 3).value = 'AUDm'

ws.cell(44, 1).value = 'Stat-Other Adj'
ws.cell(44, 2).value = 'Other Adjustments'
ws.cell(44, 3).value = 'AUDm'

ws.cell(45, 1).value = 'Stat-Statutory EBITDA'
ws.cell(45, 2).value = 'Statutory EBITDA'
ws.cell(45, 3).value = 'AUDm'

# --- D&A section (R47-R52) ---
ws.cell(48, 1).value = 'DA-Total DA'
ws.cell(48, 2).value = 'Depreciation & Amortisation'
ws.cell(48, 3).value = 'AUDm'

ws.cell(49, 1).value = None
ws.cell(49, 2).value = None
ws.cell(49, 3).value = None

ws.cell(50, 1).value = None
ws.cell(50, 2).value = None
ws.cell(50, 3).value = None

ws.cell(51, 2).value = 'D&A / Revenue'
ws.cell(51, 3).value = '%'

ws.cell(52, 2).value = None
ws.cell(52, 3).value = None

# --- EBIT (R53-R56) ---
ws.cell(54, 1).value = 'EBIT-EBIT'
ws.cell(54, 2).value = 'EBIT (Operating Profit)'
ws.cell(54, 3).value = 'AUDm'

ws.cell(55, 2).value = 'EBIT Growth'
ws.cell(55, 3).value = '% YoY'

ws.cell(56, 2).value = 'EBIT Margin'
ws.cell(56, 3).value = '%'

# --- Interest (R58-R65) ---
ws.cell(59, 1).value = 'Int-Finance Income'
ws.cell(59, 2).value = 'Finance Income'
ws.cell(59, 3).value = 'AUDm'

ws.cell(60, 1).value = 'Int-Finance Costs'
ws.cell(60, 2).value = 'Finance Costs'
ws.cell(60, 3).value = 'AUDm'

ws.cell(61, 1).value = 'Int-Net Finance Costs'
ws.cell(61, 2).value = 'Net Finance Costs'
ws.cell(61, 3).value = 'AUDm'

# Clear unused interest rows
for r in [62, 63, 64, 65]:
    ws.cell(r, 1).value = None
    ws.cell(r, 2).value = None
    ws.cell(r, 3).value = None

# --- PBT, Tax, NPAT (R67-R76) ---
ws.cell(68, 1).value = 'PBT-PBT'
ws.cell(68, 2).value = 'PBT'
ws.cell(68, 3).value = 'AUDm'

ws.cell(69, 1).value = 'Tax-Tax Expense'
ws.cell(69, 2).value = 'Tax Expense'
ws.cell(69, 3).value = 'AUDm'

ws.cell(70, 2).value = 'Effective Tax Rate'
ws.cell(70, 3).value = '%'

ws.cell(71, 1).value = None
ws.cell(71, 2).value = None
ws.cell(71, 3).value = None

ws.cell(72, 1).value = 'NPAT-Statutory NPAT'
ws.cell(72, 2).value = 'NPAT'
ws.cell(72, 3).value = 'AUDm'

# Clear sig items rows
ws.cell(73, 1).value = None; ws.cell(73, 2).value = None; ws.cell(73, 3).value = None
ws.cell(74, 1).value = None; ws.cell(74, 2).value = None; ws.cell(74, 3).value = None

ws.cell(75, 2).value = 'NPAT Growth'
ws.cell(75, 3).value = '% YoY'

ws.cell(76, 2).value = 'NPAT Margin'
ws.cell(76, 3).value = '%'

# --- EPS & Dividends (R78-R92) ---
ws.cell(79, 1).value = 'EPS-YE Shares'
ws.cell(79, 2).value = 'YE Shares Outstanding'
ws.cell(79, 3).value = '#m'

ws.cell(80, 1).value = 'EPS-WASO Basic'
ws.cell(80, 2).value = 'WASO Basic'
ws.cell(80, 3).value = '#m'

ws.cell(81, 1).value = 'EPS-Dilution'
ws.cell(81, 2).value = 'Dilution'
ws.cell(81, 3).value = '#m'

ws.cell(82, 1).value = 'EPS-WASO Diluted'
ws.cell(82, 2).value = 'WASO Diluted'
ws.cell(82, 3).value = '#m'

ws.cell(84, 1).value = 'EPS-Basic EPS'
ws.cell(84, 2).value = 'Basic EPS'
ws.cell(84, 3).value = 'AUDps'

ws.cell(85, 1).value = 'EPS-Diluted EPS'
ws.cell(85, 2).value = 'Diluted EPS'
ws.cell(85, 3).value = 'AUDps'

ws.cell(86, 2).value = 'EPS Growth'
ws.cell(86, 3).value = '% YoY'

ws.cell(88, 1).value = 'Div-DPS'
ws.cell(88, 2).value = 'DPS'
ws.cell(88, 3).value = 'AUDps'

ws.cell(89, 1).value = 'Div-Total Dividends'
ws.cell(89, 2).value = 'Total Dividends'
ws.cell(89, 3).value = 'AUDm'

ws.cell(90, 2).value = 'Payout Ratio'
ws.cell(90, 3).value = '%'

ws.cell(91, 2).value = 'Dividend Yield'
ws.cell(91, 3).value = '%'

ws.cell(92, 2).value = 'Dividend Growth'
ws.cell(92, 3).value = '% YoY'

# --- KPIs (R94-R106) ---
ws.cell(94, 2).value = 'Operating Metrics'

ws.cell(95, 1).value = 'KPI-Total Restaurants'
ws.cell(95, 2).value = 'Total Restaurants'
ws.cell(95, 3).value = '#'

ws.cell(96, 1).value = 'KPI-Aus Corp Restaurants'
ws.cell(96, 2).value = 'Australia Corporate Restaurants'
ws.cell(96, 3).value = '#'

ws.cell(97, 1).value = 'KPI-Aus Fran Restaurants'
ws.cell(97, 2).value = 'Australia Franchise Restaurants'
ws.cell(97, 3).value = '#'

ws.cell(98, 1).value = 'KPI-US Restaurants'
ws.cell(98, 2).value = 'US Restaurants'
ws.cell(98, 3).value = '#'

ws.cell(99, 1).value = 'KPI-Aus Network Sales'
ws.cell(99, 2).value = 'Australia Network Sales'
ws.cell(99, 3).value = 'AUDm'

ws.cell(100, 1).value = 'KPI-Comp Sales Growth'
ws.cell(100, 2).value = 'Comp Sales Growth'
ws.cell(100, 3).value = '%'

ws.cell(101, 1).value = 'KPI-Corp Rest Margin %'
ws.cell(101, 2).value = 'Corporate Restaurant Margin %'
ws.cell(101, 3).value = '%'

ws.cell(102, 1).value = 'KPI-Implied Royalty Rate'
ws.cell(102, 2).value = 'Implied Royalty Rate'
ws.cell(102, 3).value = '%'

ws.cell(103, 1).value = 'KPI-G&A % NS'
ws.cell(103, 2).value = 'G&A as % of Network Sales'
ws.cell(103, 3).value = '%'

ws.cell(104, 1).value = 'KPI-Seg EBITDA % NS'
ws.cell(104, 2).value = 'Seg EBITDA as % of Network Sales'
ws.cell(104, 3).value = '%'

ws.cell(105, 2).value = None; ws.cell(105, 3).value = None; ws.cell(105, 1).value = None
ws.cell(106, 2).value = None; ws.cell(106, 3).value = None; ws.cell(106, 1).value = None

# --- Balance Sheet (R108+) ---
# Update currency labels
for r in range(108, 200):
    c3 = ws.cell(r, 3).value
    if c3 and 'NZD' in str(c3):
        ws.cell(r, 3).value = str(c3).replace('NZD', 'AUD')

# BS Assets
ws.cell(110, 1).value = 'BS-Cash'
ws.cell(110, 2).value = 'Cash & Term Deposits'
ws.cell(110, 3).value = 'AUDm'

ws.cell(111, 1).value = 'BS-Trade Receivables'
ws.cell(111, 2).value = 'Trade Receivables'
ws.cell(111, 3).value = 'AUDm'

ws.cell(112, 1).value = 'BS-Inventories'
ws.cell(112, 2).value = 'Inventories'
ws.cell(112, 3).value = 'AUDm'

ws.cell(113, 1).value = 'BS-PPE'
ws.cell(113, 2).value = 'Property, Plant & Equipment'
ws.cell(113, 3).value = 'AUDm'

ws.cell(114, 1).value = 'BS-Intangibles'
ws.cell(114, 2).value = 'Intangibles'
ws.cell(114, 3).value = 'AUDm'

ws.cell(115, 1).value = 'BS-ROU Assets'
ws.cell(115, 2).value = 'Right-of-Use Assets'
ws.cell(115, 3).value = 'AUDm'

ws.cell(116, 1).value = 'BS-Other Assets'
ws.cell(116, 2).value = 'Finance Lease Rec & Other Assets'
ws.cell(116, 3).value = 'AUDm'

ws.cell(117, 2).value = 'Total Assets'
ws.cell(117, 3).value = 'AUDm'

ws.cell(118, 2).value = 'Receivables / Revenue'
ws.cell(119, 2).value = 'Inventory / Revenue'
ws.cell(120, 2).value = None; ws.cell(120, 3).value = None  # Clear working capital
ws.cell(121, 2).value = 'Payables / Revenue'
ws.cell(122, 2).value = None; ws.cell(122, 3).value = None

# BS Liabilities
ws.cell(125, 1).value = 'BS-Trade Payables'
ws.cell(125, 2).value = 'Trade & Other Payables'
ws.cell(125, 3).value = 'AUDm'

ws.cell(126, 1).value = 'BS-Other Liabilities'
ws.cell(126, 2).value = 'Other Liabilities'
ws.cell(126, 3).value = 'AUDm'

ws.cell(127, 1).value = 'BS-Lease Liabilities'
ws.cell(127, 2).value = 'Lease Liabilities'
ws.cell(127, 3).value = 'AUDm'

ws.cell(128, 1).value = 'BS-Total Banking Debt'
ws.cell(128, 2).value = 'Total Banking Debt'
ws.cell(128, 3).value = 'AUDm'

ws.cell(129, 2).value = 'Total Liabilities'
ws.cell(131, 2).value = 'Net Cash (Debt)'
ws.cell(132, 2).value = 'Adj Net Debt (incl leases)'
ws.cell(133, 2).value = 'Net Cash / Seg EBITDA'
ws.cell(133, 3).value = 'x'
ws.cell(134, 2).value = None; ws.cell(134, 3).value = None

# Equity
ws.cell(137, 1).value = 'BS-Issued Capital'
ws.cell(137, 2).value = 'Issued Capital'
ws.cell(137, 3).value = 'AUDm'

ws.cell(138, 1).value = 'BS-Retained Profits'
ws.cell(138, 2).value = 'Accumulated Losses'
ws.cell(138, 3).value = 'AUDm'

ws.cell(139, 1).value = 'BS-Reserves'
ws.cell(139, 2).value = 'Reserves'
ws.cell(139, 3).value = 'AUDm'

ws.cell(140, 1).value = 'BS-Minorities'
ws.cell(140, 2).value = 'Minorities'
ws.cell(140, 3).value = 'AUDm'

ws.cell(141, 2).value = 'Total Equity'
ws.cell(142, 2).value = 'ROE'
ws.cell(143, 2).value = 'P/B'
ws.cell(144, 2).value = 'BS Check (should be 0)'

# Cash Flow labels
ws.cell(148, 2).value = 'Statutory EBITDA'
ws.cell(148, 1).value = 'CF-EBITDA'

ws.cell(149, 1).value = 'CF-WC Change'
ws.cell(149, 2).value = 'Working Capital Change'

ws.cell(150, 1).value = 'CF-Non Cash'
ws.cell(150, 2).value = 'Non-Cash Items & Other'

ws.cell(151, 2).value = 'Gross Operating Cash Flow'

ws.cell(152, 1).value = 'CF-Int Received'
ws.cell(152, 2).value = 'Interest Received'

ws.cell(153, 1).value = 'CF-Interest Paid'
ws.cell(153, 2).value = 'Interest Paid'

ws.cell(154, 1).value = 'CF-Lease Int Paid'
ws.cell(154, 2).value = 'Lease Interest Paid'

ws.cell(155, 1).value = 'CF-Tax Paid'
ws.cell(155, 2).value = 'Tax Paid'

ws.cell(156, 1).value = 'CF-Net OCF'
ws.cell(156, 2).value = 'Net Operating Cash Flow'

ws.cell(160, 1).value = 'CF-Capex PPE'
ws.cell(160, 2).value = 'Capex (PPE)'

ws.cell(162, 1).value = 'CF-Capex Intang'
ws.cell(162, 2).value = 'Capex (Intangibles)'

ws.cell(163, 1).value = 'CF-Acquisitions'
ws.cell(163, 2).value = 'Acquisitions'

ws.cell(164, 1).value = 'CF-Asset Sales'
ws.cell(164, 2).value = 'Asset Sales / Other'

ws.cell(165, 1).value = 'CF-Other CFI'
ws.cell(165, 2).value = 'Other'

ws.cell(166, 2).value = 'Total Investing Cash Flow'

ws.cell(169, 1).value = 'CF-Dividends'
ws.cell(169, 2).value = 'Dividends Paid'

ws.cell(170, 1).value = 'CF-Share Issues'
ws.cell(170, 2).value = 'Share Issues / Buybacks'

ws.cell(171, 1).value = 'CF-Lease Principal'
ws.cell(171, 2).value = 'Lease Principal Payments'

ws.cell(172, 1).value = 'CF-Debt Change'
ws.cell(172, 2).value = 'Change in Debt'

ws.cell(173, 1).value = 'CF-Other CFF'
ws.cell(173, 2).value = 'Other'

ws.cell(174, 2).value = 'Total Financing Cash Flow'
ws.cell(176, 2).value = 'Net Change in Cash'

# Update all AUDm labels
for r in range(146, 200):
    c3 = ws.cell(r, 3).value
    if c3 and 'NZD' in str(c3):
        ws.cell(r, 3).value = str(c3).replace('NZD', 'AUD')

# ROIC section
ws.cell(188, 2).value = 'ROIC'
ws.cell(189, 2).value = 'Invested Capital'
ws.cell(190, 2).value = 'EBIT'
ws.cell(191, 2).value = 'ROFE'
ws.cell(192, 2).value = 'NOPAT'
ws.cell(193, 2).value = 'ROIC'

wb.save('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')
print("Part 1 complete: Annual sheet labels updated")

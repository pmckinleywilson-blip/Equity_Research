"""Part 2: Restructure P&L rows on both Annual and HY sheets for OCL."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from copy import copy

DST = '/home/pmwilson/Project_Equities/OCL/Models/OCL Model.xlsx'
wb = openpyxl.load_workbook(DST)

# Helper: copy cell formatting
def copy_style(src_cell, dst_cell):
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format

# Colors
BLUE_TEXT = Font(color='FF0000CC')
BLACK_TEXT = Font()
BOLD_BLACK = Font(bold=True)
MAROON_TEXT = Font(color='FFC00000')
SECTION_FILL = PatternFill(start_color='FFC5D9F1', end_color='FFC5D9F1', fill_type='solid')
THIN_BORDER_TOP = Border(top=Side(style='thin'))
THIN_BORDER_BOTTOM = Border(bottom=Side(style='thin'))
THIN_BORDER_BOTH = Border(top=Side(style='thin'), bottom=Side(style='thin'))
PCT_FMT = '0.0%'
NUM_FMT = '#,##0.0'
NUM_FMT_3 = '0.000'
INT_FMT = '#,##0'

def build_row_structure():
    """Define the complete row structure for OCL model."""
    # Each entry: (row, key, label, unit, row_type)
    # row_type: 'section_header', 'data', 'total', 'pct', 'blank', 'subsection'
    rows = [
        # P&L
        (5, '', 'P&L', '', 'section_header'),
        (6, '', 'Revenue', '', 'subsection'),
        (7, 'Rev-Info Intelligence', 'Information Intelligence Revenue', 'A$m', 'data'),
        (8, 'Rev-Planning & Building', 'Planning & Building Revenue', 'A$m', 'data'),
        (9, 'Rev-Regulatory Solutions', 'Regulatory Solutions Revenue', 'A$m', 'data'),
        (10, 'Rev-Interest Income', 'Interest Income', 'A$m', 'data'),
        (11, 'Rev-Total Revenue', 'Total Revenue', 'A$m', 'total'),
        (12, '', 'Revenue Growth', '% YoY', 'pct'),
        (13, '', '', '', 'blank'),
        (14, '', 'COGS', '', 'subsection'),
        (15, 'COGS-Total COGS', 'Total COGS', 'A$m', 'data'),
        (16, '', '', '', 'blank'),
        (17, '', 'Gross Profit', '', 'subsection'),
        (18, 'GP-Gross Profit', 'Gross Profit', 'A$m', 'total'),
        (19, '', 'GP Growth', '% YoY', 'pct'),
        (20, '', 'GP Margin', '%', 'pct'),
        (21, '', '', '', 'blank'),
        (22, '', 'Operating Expenses', '', 'subsection'),
        (23, 'OPEX-Distribution', 'Distribution Expenses', 'A$m', 'data'),
        (24, 'OPEX-R&D Expense', 'Research & Development', 'A$m', 'data'),
        (25, 'OPEX-Admin', 'Administration & Other', 'A$m', 'data'),
        (26, 'OPEX-Total OpEx', 'Total Operating Expenses', 'A$m', 'total'),
        (27, '', 'OpEx Growth', '% YoY', 'pct'),
        (28, '', '', '', 'blank'),
        (29, '', 'EBITDA', '', 'subsection'),
        (30, 'EBITDA-Underlying EBITDA', 'Underlying EBITDA', 'A$m', 'total'),
        (31, '', 'EBITDA Growth', '% YoY', 'pct'),
        (32, '', 'EBITDA Margin', '%', 'pct'),
        (33, '', '', '', 'blank'),
        (34, '', 'Statutory EBITDA Adjustments', '', 'subsection'),
        (35, 'Stat-SBP', 'Share-based Payments', 'A$m', 'data'),
        (36, 'Stat-M&A Costs', 'M&A Costs', 'A$m', 'data'),
        (37, 'Stat-FX', 'FX Gains / (Losses)', 'A$m', 'data'),
        (38, 'Stat-Statutory EBITDA', 'Statutory EBITDA', 'A$m', 'total'),
        (39, '', '', '', 'blank'),
        (40, '', 'D&A', '', 'subsection'),
        (41, 'DA-Depreciation PPE', 'Depreciation PPE', 'A$m', 'data'),
        (42, 'DA-ROU Amortisation', 'ROU Asset Amortisation', 'A$m', 'data'),
        (43, 'DA-Amort Dev Costs', 'Amortisation of Capitalised Dev Costs', 'A$m', 'data'),
        (44, 'DA-Total DA', 'Total D&A', 'A$m', 'total'),
        (45, '', 'D&A / Revenue', '%', 'pct'),
        (46, '', '', '', 'blank'),
        (47, '', 'EBIT', '', 'subsection'),
        (48, 'EBIT-Underlying EBIT', 'Underlying EBIT', 'A$m', 'total'),
        (49, '', 'EBIT Growth', '% YoY', 'pct'),
        (50, '', 'EBIT Margin', '%', 'pct'),
        (51, '', '', '', 'blank'),
        (52, '', 'Interest', '', 'subsection'),
        (53, 'Int-Interest Income', 'Interest Income', 'A$m', 'data'),
        (54, 'Int-Lease Interest', 'Interest on Lease Liabilities', 'A$m', 'data'),
        (55, 'Int-Net Finance Costs', 'Net Finance Costs', 'A$m', 'total'),
        (56, '', '', '', 'blank'),
        (57, '', 'PBT, Tax, NPAT', '', 'subsection'),
        (58, 'PBT-PBT', 'PBT', 'A$m', 'total'),
        (59, 'Tax-Tax Expense', 'Tax Expense', 'A$m', 'data'),
        (60, '', 'Underlying Tax Rate', '%', 'pct'),
        (61, 'NPAT-Underlying NPAT', 'Underlying NPAT', 'A$m', 'total'),
        (62, 'NPAT-Other Items AT', 'Other Items After Tax', 'A$m', 'data'),
        (63, 'NPAT-Statutory NPAT', 'Statutory NPAT', 'A$m', 'total'),
        (64, '', 'NPAT Growth', '% YoY', 'pct'),
        (65, '', 'NPAT Margin', '%', 'pct'),
        (66, '', '', '', 'blank'),
        # EPS & Dividends (Annual only)
        (67, '', 'EPS & Dividends', '', 'subsection'),
        (68, 'EPS-YE Shares', 'YE Basic Shares Outstanding', '#m', 'data'),
        (69, 'EPS-WASO Basic', 'WASO Basic', '#m', 'data'),
        (70, 'EPS-Dilution', 'Dilution', '#m', 'data'),
        (71, 'EPS-WASO Diluted', 'WASO Diluted', '#m', 'data'),
        (72, '', '', '', 'blank'),
        (73, 'EPS-Underlying EPS', 'Underlying EPS', 'A$ps', 'data'),
        (74, 'EPS-Statutory EPS', 'Statutory EPS', 'A$ps', 'data'),
        (75, '', 'EPS Growth', '% YoY', 'pct'),
        (76, '', '', '', 'blank'),
        (77, 'Div-DPS', 'DPS', 'A$ps', 'data'),
        (78, 'Div-Total Dividends', 'Total Dividends', 'A$m', 'data'),
        (79, '', 'Payout Ratio', '%', 'pct'),
        (80, '', 'Dividend Yield', '%', 'pct'),
        (81, '', 'Dividend Growth', '% YoY', 'pct'),
        (82, '', '', '', 'blank'),
        # Operating Metrics
        (83, '', 'Operating Metrics', '', 'section_header'),
        (84, 'KPI-ARR II', 'ARR: Information Intelligence', 'A$m', 'data'),
        (85, 'KPI-ARR PB', 'ARR: Planning & Building', 'A$m', 'data'),
        (86, 'KPI-ARR RS', 'ARR: Regulatory Solutions', 'A$m', 'data'),
        (87, 'KPI-ARR Total', 'Total ARR', 'A$m', 'total'),
        (88, '', 'ARR Growth', '% YoY', 'pct'),
        (89, 'KPI-Total R&D', 'Total R&D Investment', 'A$m', 'data'),
        (90, 'KPI-Capitalised Dev', 'Capitalised Development Costs', 'A$m', 'data'),
        (91, '', 'R&D Capitalisation Rate', '%', 'pct'),
        (92, '', 'R&D / Revenue', '%', 'pct'),
        (93, '', 'Recurring Revenue %', '%', 'pct'),
        (94, 'KPI-Shares Out', 'Shares Outstanding', '#m', 'data'),
        (95, 'KPI-WASO', 'WASO Basic', '#m', 'data'),
        (96, '', '', '', 'blank'),
        # Balance Sheet
        (97, '', 'Balance Sheet', '', 'section_header'),
        (98, '', 'Assets', '', 'subsection'),
        (99, 'BS-Cash', 'Cash', 'A$m', 'data'),
        (100, 'BS-Trade Receivables', 'Trade Receivables', 'A$m', 'data'),
        (101, 'BS-Contract Assets', 'Contract Assets', 'A$m', 'data'),
        (102, 'BS-Current Tax', 'Current Tax Asset / (Liability)', 'A$m', 'data'),
        (103, 'BS-PPE', 'Property, Plant & Equipment', 'A$m', 'data'),
        (104, 'BS-Intangibles', 'Intangibles', 'A$m', 'data'),
        (105, 'BS-ROU Assets', 'Right-of-Use Assets', 'A$m', 'data'),
        (106, 'BS-Other Assets', 'Other Assets', 'A$m', 'data'),
        (107, '', 'Total Assets', 'A$m', 'total'),
        (108, '', 'Receivables / Revenue', '%', 'pct'),
        (109, '', 'Working Capital', 'A$m', 'data'),
        (110, '', 'Payables / Revenue', '%', 'pct'),
        (111, '', 'New Lease Additions', 'A$m', 'data'),
        (112, '', '', '', 'blank'),
        (113, '', 'Liabilities', '', 'subsection'),
        (114, 'BS-Trade Payables', 'Trade & Other Payables', 'A$m', 'data'),
        (115, 'BS-Contract Liabilities', 'Contract Liabilities', 'A$m', 'data'),
        (116, 'BS-Deferred Tax', 'Deferred Tax Liabilities', 'A$m', 'data'),
        (117, 'BS-Provisions', 'Provisions', 'A$m', 'data'),
        (118, 'BS-Other Liabilities', 'Other Liabilities', 'A$m', 'data'),
        (119, 'BS-Lease Liabilities', 'Lease Liabilities', 'A$m', 'data'),
        (120, '', 'Total Liabilities', 'A$m', 'total'),
        (121, '', '', '', 'blank'),
        (122, '', 'Net Cash', 'A$m', 'data'),
        (123, '', 'Adj Net Debt (incl leases)', 'A$m', 'data'),
        (124, '', 'Gearing (ND/(ND+E))', '%', 'pct'),
        (125, '', '', '', 'blank'),
        (126, '', 'Equity', '', 'subsection'),
        (127, 'BS-Issued Capital', 'Issued Capital', 'A$m', 'data'),
        (128, 'BS-Retained Profits', 'Retained Profits', 'A$m', 'data'),
        (129, 'BS-Reserves', 'Reserves', 'A$m', 'data'),
        (130, '', 'Total Equity', 'A$m', 'total'),
        (131, '', 'ROE', '%', 'pct'),
        (132, '', 'P/B', 'x', 'data'),
        (133, '', 'BS Check (should be 0, +/-0.2 due to rounding)', 'A$m', 'data'),
        (134, '', '', '', 'blank'),
        # Cash Flow
        (135, '', 'Cash Flow', '', 'section_header'),
        (136, '', 'CFO', '', 'subsection'),
        (137, 'CF-EBITDA', 'Underlying EBITDA', 'A$m', 'data'),
        (138, 'CF-WC Change', 'Working Capital Change', 'A$m', 'data'),
        (139, 'CF-Non Cash Items', 'Non-Cash Items (SBP + Other)', 'A$m', 'data'),
        (140, '', 'Gross Operating Cash Flow', 'A$m', 'total'),
        (141, 'CF-Int Received', 'Interest Received', 'A$m', 'data'),
        (142, 'CF-Lease Int Paid', 'Lease Interest Paid', 'A$m', 'data'),
        (143, 'CF-Tax Paid', 'Tax Paid', 'A$m', 'data'),
        (144, 'CF-Net OCF', 'Net Operating Cash Flow', 'A$m', 'total'),
        (145, '', 'OCF Growth', '% YoY', 'pct'),
        (146, '', 'EBITDA Cash Conversion', '%', 'pct'),
        (147, '', '', '', 'blank'),
        (148, '', 'CFI', '', 'subsection'),
        (149, 'CF-Capex PPE', 'Capex (PPE)', 'A$m', 'data'),
        (150, '', 'Capex / Sales', '%', 'pct'),
        (151, 'CF-Capex Intang', 'Capitalised Development Costs', 'A$m', 'data'),
        (152, 'CF-Acquisitions', 'Acquisitions', 'A$m', 'data'),
        (153, 'CF-Other CFI', 'Other', 'A$m', 'data'),
        (154, '', 'Total Investing Cash Flow', 'A$m', 'total'),
        (155, '', '', '', 'blank'),
        (156, '', 'CFF', '', 'subsection'),
        (157, 'CF-Dividends', 'Dividends Paid', 'A$m', 'data'),
        (158, 'CF-Share Issues', 'Share Issues / Buybacks', 'A$m', 'data'),
        (159, 'CF-Lease Principal', 'Lease Principal Payments', 'A$m', 'data'),
        (160, 'CF-Other CFF', 'Other', 'A$m', 'data'),
        (161, '', 'Total Financing Cash Flow', 'A$m', 'total'),
        (162, '', '', '', 'blank'),
        (163, '', 'Net Change in Cash', 'A$m', 'total'),
        (164, '', '', '', 'blank'),
        (165, '', 'Operating Free Cash Flow', '', 'subsection'),
        (166, '', 'Net OCF', 'A$m', 'data'),
        (167, '', 'Net Capex', 'A$m', 'data'),
        (168, '', 'Lease Principal', 'A$m', 'data'),
        (169, '', 'Operating Free Cash Flow', 'A$m', 'total'),
        (170, '', 'FCF per Share', 'A$ps', 'data'),
        (171, '', 'FCF Yield', '%', 'pct'),
        (172, '', 'FCF Margin', '%', 'pct'),
        (173, '', '', '', 'blank'),
        (174, '', 'ROIC', '', 'subsection'),
        (175, '', 'Invested Capital', 'A$m', 'data'),
        (176, '', 'Underlying EBIT', 'A$m', 'data'),
        (177, '', 'ROFE', '%', 'pct'),
        (178, '', 'NOPAT', 'A$m', 'data'),
        (179, '', 'ROIC', '%', 'pct'),
    ]
    return rows

def apply_row_to_sheet(ws, rows, max_col):
    """Clear old content and write new row structure."""
    # Clear all existing content from row 5 onwards
    for row in range(5, 210):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.border = Border()
            cell.number_format = 'General'
    
    # Write new structure
    for (row, key, label, unit, rtype) in rows:
        if key:
            ws.cell(row=row, column=1, value=key)
        if label:
            ws.cell(row=row, column=2, value=label)
        if unit:
            ws.cell(row=row, column=3, value=unit)
        
        # Apply formatting
        if rtype == 'section_header':
            ws.cell(row=row, column=2).font = Font(bold=True)
            ws.cell(row=row, column=2).fill = SECTION_FILL
            for col in range(3, max_col + 1):
                ws.cell(row=row, column=col).fill = SECTION_FILL
        elif rtype == 'subsection':
            ws.cell(row=row, column=2).font = Font(bold=True)
        elif rtype == 'total':
            ws.cell(row=row, column=2).font = Font(bold=True)
            for col in range(4, max_col + 1):
                ws.cell(row=row, column=col).font = Font(bold=True)
                ws.cell(row=row, column=col).border = THIN_BORDER_BOTH
                ws.cell(row=row, column=col).number_format = NUM_FMT
        elif rtype == 'data':
            for col in range(4, max_col + 1):
                ws.cell(row=row, column=col).number_format = NUM_FMT
        elif rtype == 'pct':
            for col in range(4, max_col + 1):
                ws.cell(row=row, column=col).number_format = PCT_FMT

# === Apply to Annual sheet ===
ws = wb['Annual']
annual_rows = build_row_structure()
apply_row_to_sheet(ws, annual_rows, 13)  # D-M = cols 4-13

# === Build HY-specific rows (no EPS section, no BS/CF, but has segment forecasts) ===
def build_hy_rows():
    """HY sheet rows - P&L + KPIs + segment forecast sections."""
    rows = [
        (5, '', 'P&L', '', 'section_header'),
        (6, '', 'Revenue', '', 'subsection'),
        (7, 'Rev-Info Intelligence', 'Information Intelligence Revenue', 'A$m', 'data'),
        (8, 'Rev-Planning & Building', 'Planning & Building Revenue', 'A$m', 'data'),
        (9, 'Rev-Regulatory Solutions', 'Regulatory Solutions Revenue', 'A$m', 'data'),
        (10, 'Rev-Interest Income', 'Interest Income', 'A$m', 'data'),
        (11, 'Rev-Total Revenue', 'Total Revenue', 'A$m', 'total'),
        (12, '', 'Revenue Growth (YoY)', '% YoY', 'pct'),
        (13, '', '', '', 'blank'),
        (14, '', 'COGS', '', 'subsection'),
        (15, 'COGS-Total COGS', 'Total COGS', 'A$m', 'data'),
        (16, '', '', '', 'blank'),
        (17, '', 'Gross Profit', '', 'subsection'),
        (18, 'GP-Gross Profit', 'Gross Profit', 'A$m', 'total'),
        (19, '', 'GP Growth (YoY)', '% YoY', 'pct'),
        (20, '', 'GP Margin', '%', 'pct'),
        (21, '', '', '', 'blank'),
        (22, '', 'Operating Expenses', '', 'subsection'),
        (23, 'OPEX-Distribution', 'Distribution Expenses', 'A$m', 'data'),
        (24, 'OPEX-R&D Expense', 'Research & Development', 'A$m', 'data'),
        (25, 'OPEX-Admin', 'Administration & Other', 'A$m', 'data'),
        (26, 'OPEX-Total OpEx', 'Total Operating Expenses', 'A$m', 'total'),
        (27, '', 'OpEx Growth (YoY)', '% YoY', 'pct'),
        (28, '', '', '', 'blank'),
        (29, '', 'EBITDA', '', 'subsection'),
        (30, 'EBITDA-Underlying EBITDA', 'Underlying EBITDA', 'A$m', 'total'),
        (31, '', 'EBITDA Growth (YoY)', '% YoY', 'pct'),
        (32, '', 'EBITDA Margin', '%', 'pct'),
        (33, '', '', '', 'blank'),
        (34, '', 'Statutory EBITDA Adjustments', '', 'subsection'),
        (35, 'Stat-SBP', 'Share-based Payments', 'A$m', 'data'),
        (36, 'Stat-M&A Costs', 'M&A Costs', 'A$m', 'data'),
        (37, 'Stat-FX', 'FX Gains / (Losses)', 'A$m', 'data'),
        (38, 'Stat-Statutory EBITDA', 'Statutory EBITDA', 'A$m', 'total'),
        (39, '', '', '', 'blank'),
        (40, '', 'D&A', '', 'subsection'),
        (41, 'DA-Depreciation PPE', 'Depreciation PPE', 'A$m', 'data'),
        (42, 'DA-ROU Amortisation', 'ROU Asset Amortisation', 'A$m', 'data'),
        (43, 'DA-Amort Dev Costs', 'Amortisation of Capitalised Dev Costs', 'A$m', 'data'),
        (44, 'DA-Total DA', 'Total D&A', 'A$m', 'total'),
        (45, '', 'D&A / Revenue', '%', 'pct'),
        (46, '', '', '', 'blank'),
        (47, '', 'EBIT', '', 'subsection'),
        (48, 'EBIT-Underlying EBIT', 'Underlying EBIT', 'A$m', 'total'),
        (49, '', 'EBIT Growth (YoY)', '% YoY', 'pct'),
        (50, '', 'EBIT Margin', '%', 'pct'),
        (51, '', '', '', 'blank'),
        (52, '', 'Interest', '', 'subsection'),
        (53, 'Int-Interest Income', 'Interest Income', 'A$m', 'data'),
        (54, 'Int-Lease Interest', 'Interest on Lease Liabilities', 'A$m', 'data'),
        (55, 'Int-Net Finance Costs', 'Net Finance Costs', 'A$m', 'total'),
        (56, '', '', '', 'blank'),
        (57, 'PBT-PBT', 'PBT', 'A$m', 'total'),
        (58, 'Tax-Tax Expense', 'Tax Expense', 'A$m', 'data'),
        (59, '', 'Tax Rate', '%', 'pct'),
        (60, 'NPAT-Underlying NPAT', 'Underlying NPAT', 'A$m', 'total'),
        (61, 'NPAT-Other Items AT', 'Other Items After Tax', 'A$m', 'data'),
        (62, 'NPAT-Statutory NPAT', 'Statutory NPAT', 'A$m', 'total'),
        (63, '', 'NPAT Growth (YoY)', '% YoY', 'pct'),
        (64, '', '', '', 'blank'),
        # Operating Metrics
        (65, '', 'Operating Metrics', '', 'section_header'),
        (66, 'KPI-ARR II', 'ARR: Information Intelligence', 'A$m', 'data'),
        (67, 'KPI-ARR PB', 'ARR: Planning & Building', 'A$m', 'data'),
        (68, 'KPI-ARR RS', 'ARR: Regulatory Solutions', 'A$m', 'data'),
        (69, 'KPI-ARR Total', 'Total ARR', 'A$m', 'total'),
        (70, '', 'ARR Growth (YoY)', '% YoY', 'pct'),
        (71, 'KPI-Total R&D', 'Total R&D Investment', 'A$m', 'data'),
        (72, 'KPI-Capitalised Dev', 'Capitalised Development Costs', 'A$m', 'data'),
        (73, '', 'R&D Capitalisation Rate', '%', 'pct'),
        (74, '', 'R&D / Revenue', '%', 'pct'),
        (75, '', 'Recurring Revenue %', '%', 'pct'),
        (76, 'KPI-Shares Out', 'Shares Outstanding', '#m', 'data'),
        (77, 'KPI-WASO', 'WASO Basic', '#m', 'data'),
        (78, '', '', '', 'blank'),
        # Segment Forecast - Information Intelligence
        (79, '', 'Segment Forecast — Information Intelligence', '', 'section_header'),
        (80, '', 'ARR (opening)', 'A$m', 'data'),
        (81, '', 'ARR Growth %', '%', 'pct'),
        (82, '', 'ARR (closing)', 'A$m', 'data'),
        (83, '', 'Average ARR', 'A$m', 'data'),
        (84, '', 'Non-recurring adjustment %', '%', 'pct'),
        (85, '', 'Revenue', 'A$m', 'data'),
        (86, '', '', '', 'blank'),
        # Segment Forecast - Planning & Building
        (87, '', 'Segment Forecast — Planning & Building', '', 'section_header'),
        (88, '', 'ARR (opening)', 'A$m', 'data'),
        (89, '', 'ARR Growth %', '%', 'pct'),
        (90, '', 'ARR (closing)', 'A$m', 'data'),
        (91, '', 'Average ARR', 'A$m', 'data'),
        (92, '', 'Non-recurring adjustment %', '%', 'pct'),
        (93, '', 'Revenue', 'A$m', 'data'),
        (94, '', '', '', 'blank'),
        # Segment Forecast - Regulatory Solutions
        (95, '', 'Segment Forecast — Regulatory Solutions', '', 'section_header'),
        (96, '', 'ARR (opening)', 'A$m', 'data'),
        (97, '', 'ARR Growth %', '%', 'pct'),
        (98, '', 'ARR (closing)', 'A$m', 'data'),
        (99, '', 'Average ARR', 'A$m', 'data'),
        (100, '', 'Non-recurring adjustment %', '%', 'pct'),
        (101, '', 'Revenue', 'A$m', 'data'),
        (102, '', '', '', 'blank'),
        # Group Forecast Inputs
        (103, '', 'Group Forecast Inputs', '', 'section_header'),
        (104, '', 'COGS % of Contract Revenue', '%', 'pct'),
        (105, '', 'Distribution % of Revenue', '%', 'pct'),
        (106, '', 'Total R&D / Revenue %', '%', 'pct'),
        (107, '', 'R&D Capitalisation Rate %', '%', 'pct'),
        (108, '', 'Admin % of Revenue', '%', 'pct'),
        (109, '', 'Interest Rate on Cash %', '%', 'pct'),
        (110, '', 'Effective Tax Rate %', '%', 'pct'),
        (111, '', 'New Lease Additions', 'A$m', 'data'),
        (112, '', 'Capex PPE', 'A$m', 'data'),
        (113, '', 'DPS', 'cps', 'data'),
    ]
    return rows

ws2 = wb['HY & Segments']
hy_rows = build_hy_rows()
apply_row_to_sheet(ws2, hy_rows, 23)  # D-W = cols 4-23

wb.save(DST)
print('Part 2 complete: row structures rebuilt for Annual and HY sheets')
print(f'Annual: {len(annual_rows)} rows defined')
print(f'HY: {len(hy_rows)} rows defined')

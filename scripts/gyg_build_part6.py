"""Part 6: HY & Segments sheet - forecast formulas (2H26 onwards)."""
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')
ws = wb['HY & Segments']

CL = get_column_letter

# Forecast starts at K(11)=2H26 through AC(29)=2H35
# 2H cols: K(11), M(13), O(15), Q(17), S(19), U(21), W(23), Y(25), AA(27), AC(29)
# 1H cols: L(12), N(14), P(16), R(18), T(20), V(22), X(24), Z(26), AB(28)

# For 2H periods in forecast: derive as Annual - 1H
# For 1H periods in forecast: use growth formulas

for c in range(11, 30):
    col = CL(c)
    pcol = CL(c-1)    # prior half
    pycol = CL(c-2)   # same half prior year

    is_2h = (c % 2 == 1)  # Odd cols are 2H (11=2H26, 13=2H27, etc.)
    is_1h = (c % 2 == 0)  # Even cols are 1H (12=1H27, 14=1H28, etc.)

    if is_2h:
        # 2H = Annual - 1H
        # Need to pull from Annual sheet using year match
        year_ref = f'{col}$1'  # year number in R1

        # R7: Revenue = Annual total - 1H
        ws.cell(7, c).value = f"=INDEX(Annual!$A:$P,MATCH($A7,Annual!$A:$A,0),MATCH({year_ref},Annual!$1:$1,0))-{pcol}7"

        # R8: Other Revenue = Annual - 1H
        ws.cell(8, c).value = f"=INDEX(Annual!$A:$P,MATCH($A8,Annual!$A:$A,0),MATCH({year_ref},Annual!$1:$1,0))-{pcol}8"

        # R13-R15: Costs = Annual - 1H
        for row in [13, 14, 15]:
            if ws.cell(row, 1).value:
                ws.cell(row, c).value = f"=INDEX(Annual!$A:$P,MATCH($A{row},Annual!$A:$A,0),MATCH({year_ref},Annual!$1:$1,0))-{pcol}{row}"
            else:
                # No Col A code, use formula
                ws.cell(row, c).value = f"=INDEX(Annual!$D:$P,{row},MATCH({year_ref},Annual!$1:$1,0)-2022)-{pcol}{row}"

        # R38-R39: Segment EBITDA = Annual - 1H
        ws.cell(38, c).value = f"=INDEX(Annual!$A:$P,MATCH($A38,Annual!$A:$A,0),MATCH({year_ref},Annual!$1:$1,0))-{pcol}38"
        ws.cell(39, c).value = f"=INDEX(Annual!$A:$P,MATCH($A39,Annual!$A:$A,0),MATCH({year_ref},Annual!$1:$1,0))-{pcol}39"
        ws.cell(40, c).value = 0

        # R48: Adj = Annual total adjustments - 1H
        # Annual R42+R43+R44 combined, minus 1H
        ws.cell(48, c).value = f"=(INDEX(Annual!$D:$P,42,MATCH({year_ref},Annual!$1:$1,0)-2022)+INDEX(Annual!$D:$P,43,MATCH({year_ref},Annual!$1:$1,0)-2022)+INDEX(Annual!$D:$P,44,MATCH({year_ref},Annual!$1:$1,0)-2022))-{pcol}48"

        # R52: D&A = Annual - 1H
        ws.cell(52, c).value = f"=INDEX(Annual!$A:$P,MATCH($A52,Annual!$A:$A,0),MATCH({year_ref},Annual!$1:$1,0))-{pcol}52"

        # R63-R64: Finance = Annual - 1H
        ws.cell(63, c).value = f"=INDEX(Annual!$A:$P,MATCH($A63,Annual!$A:$A,0),MATCH({year_ref},Annual!$1:$1,0))-{pcol}63"
        ws.cell(64, c).value = f"=INDEX(Annual!$A:$P,MATCH($A64,Annual!$A:$A,0),MATCH({year_ref},Annual!$1:$1,0))-{pcol}64"

        # R72: Tax = Annual - 1H
        ws.cell(72, c).value = f"=INDEX(Annual!$A:$P,MATCH($A72,Annual!$A:$A,0),MATCH({year_ref},Annual!$1:$1,0))-{pcol}72"

        # Segment forecast rows - 2H:
        # Australia: use Annual implied values where possible
        # DT Restaurants: grow by ~5 per half
        ws.cell(93, c).value = f'={pcol}93+5'
        ws.cell(94, c).value = f'={pcol}94+3'
        ws.cell(95, c).value = f'={pcol}95'  # flat
        ws.cell(97, c).value = f'={pcol}97*1.0125'  # AUV +2.5% pa = 1.25% per half
        ws.cell(98, c).value = f'={pcol}98*1.0125'
        ws.cell(99, c).value = f'={pcol}99*1.0125'

        ws.cell(104, c).value = f'={pcol}104*1.08'  # Corp sales growing
        ws.cell(105, c).value = f'=MIN(0.20,{pcol}105+0.002)'  # margin expanding
        ws.cell(108, c).value = f'=MIN(0.10,{pcol}108+0.001)'  # royalty trending up
        ws.cell(110, c).value = f'={pcol}110*1.04'  # G&A growing slowly

        # US forecast
        ws.cell(116, c).value = f'={pcol}116+1'  # +1 restaurant per half
        ws.cell(117, c).value = f'={pcol}117*1.15'  # NS growing
        ws.cell(118, c).value = f'={col}117'  # all corp in US
        ws.cell(119, c).value = f'=MAX(-0.30,{pcol}119+0.03)'  # margin improving
        ws.cell(120, c).value = f'={col}118*{col}119'
        ws.cell(121, c).value = f'={pcol}121*1.2'  # fran rev growing
        ws.cell(122, c).value = f'={pcol}122*1.05'  # G&A growing

    else:  # 1H
        # 1H forecast: grow vs PCP (same half prior year = c-2)

        # R7: Revenue = PCP * (1 + growth)
        ws.cell(7, c).value = f'={pycol}7*1.18'  # ~18% rev growth

        # R8: Other Revenue
        ws.cell(8, c).value = f'={pycol}8*1.15'

        # R13-R15: Costs as % of revenue
        ws.cell(13, c).value = f'={col}7*-0.261'
        ws.cell(14, c).value = f'={col}7*-0.424'
        ws.cell(15, c).value = f'={col}7*-0.230'

        # Segments
        ws.cell(38, c).value = f'={pycol}38*1.20'  # Aus EBITDA growing 20%
        ws.cell(39, c).value = f'={pycol}39*0.85'   # US losses reducing 15%/yr
        ws.cell(40, c).value = 0

        # R48: Adjustments - grow with business
        ws.cell(48, c).value = f'={pycol}48*1.10'

        # R52: D&A
        ws.cell(52, c).value = f'={pycol}52*1.12'

        # R63: Finance Income - flat
        ws.cell(63, c).value = f'={pycol}63'

        # R64: Finance Costs - grow with leases
        ws.cell(64, c).value = f'={pycol}64*1.10'

        # R72: Tax
        ws.cell(72, c).value = f'=IF({col}71>0,-{col}71*0.35,0)'

        # Segment forecast:
        ws.cell(93, c).value = f'={pcol}93+5'
        ws.cell(94, c).value = f'={pcol}94+3'
        ws.cell(95, c).value = f'={pcol}95'
        ws.cell(97, c).value = f'={pcol}97*1.0125'
        ws.cell(98, c).value = f'={pcol}98*1.0125'
        ws.cell(99, c).value = f'={pcol}99*1.0125'

        ws.cell(104, c).value = f'={pycol}104*1.15'
        ws.cell(105, c).value = f'=MIN(0.20,{pcol}105+0.002)'
        ws.cell(108, c).value = f'=MIN(0.10,{pcol}108+0.001)'
        ws.cell(110, c).value = f'={pycol}110*1.04'

        ws.cell(116, c).value = f'={pcol}116+1'
        ws.cell(117, c).value = f'={pycol}117*1.15'
        ws.cell(118, c).value = f'={col}117'
        ws.cell(119, c).value = f'=MAX(-0.30,{pcol}119+0.03)'
        ws.cell(120, c).value = f'={col}118*{col}119'
        ws.cell(121, c).value = f'={pcol}121*1.2'
        ws.cell(122, c).value = f'={pycol}122*1.05'

    # ---- Common formulas for ALL forecast columns ----

    # R9: Total Rev incl Other
    ws.cell(9, c).value = f'={col}7+{col}8'

    # R10: Rev Growth YoY
    ws.cell(10, c).value = f'=IF({pycol}7=0,"",{col}7/{pycol}7-1)'

    # R16: Total Costs
    ws.cell(16, c).value = f'={col}13+{col}14+{col}15'

    # R19: Stat EBITDA
    ws.cell(19, c).value = f'={col}9+{col}16'

    # R42: Group Seg EBITDA
    ws.cell(42, c).value = f'=SUM({col}38:{col}40)'

    # R43: Seg EBITDA Growth
    ws.cell(43, c).value = f'=IF({pycol}42=0,"",{col}42/{pycol}42-1)'

    # R44: Seg EBITDA Margin
    ws.cell(44, c).value = f'=IF({col}7=0,"",{col}42/{col}7)'

    # R49: Statutory EBITDA
    ws.cell(49, c).value = f'={col}42+{col}48'

    # R55: D&A / Revenue
    ws.cell(55, c).value = f'=IF({col}7=0,"",{col}52/{col}7)'

    # R58: EBIT
    ws.cell(58, c).value = f'={col}49+{col}52'

    # R59: EBIT Growth
    ws.cell(59, c).value = f'=IF({pycol}58=0,"",{col}58/{pycol}58-1)'

    # R60: EBIT Margin
    ws.cell(60, c).value = f'=IF({col}7=0,"",{col}58/{col}7)'

    # R66: Net Finance Costs
    ws.cell(66, c).value = f'={col}63+{col}64'

    # R71: PBT
    ws.cell(71, c).value = f'={col}58+{col}66'

    # R73: Tax Rate
    ws.cell(73, c).value = f'=IF({col}71=0,"",{col}72/{col}71)'

    # R75: NPAT
    ws.cell(75, c).value = f'={col}71+{col}72'

    # R78: NPAT Growth
    ws.cell(78, c).value = f'=IF({pycol}75<=0,"",{col}75/{pycol}75-1)'

    # KPIs
    ws.cell(81, c).value = f'={col}82+{col}83+{col}84'
    ws.cell(82, c).value = f'={pycol}82+5' if is_1h else f'={pcol}82+5'
    ws.cell(83, c).value = f'={pycol}83+11' if is_1h else f'={pcol}83+11'
    ws.cell(85, c).value = f'={col}103'  # link to segment forecast
    ws.cell(86, c).value = f'={col}117'
    ws.cell(87, c).value = f'=IF({pycol}85=0,"",({col}85+{col}86)/({pycol}85+{pycol}86)-1)'

    # Segment Australia common formulas
    ws.cell(96, c).value = f'={col}93+{col}94+{col}95'
    ws.cell(100, c).value = f'={col}93*{col}97*0.5'
    ws.cell(101, c).value = f'={col}94*{col}98*0.5'
    ws.cell(102, c).value = f'={col}95*{col}99*0.5'
    ws.cell(103, c).value = f'={col}100+{col}101+{col}102'
    ws.cell(106, c).value = f'={col}104*{col}105'
    ws.cell(107, c).value = f'={col}103-{col}104'
    ws.cell(109, c).value = f'={col}107*{col}108'
    ws.cell(111, c).value = f'=IF({col}103=0,"",ABS({col}110)/{col}103)'
    ws.cell(112, c).value = f'={col}106+{col}109+{col}110'
    ws.cell(113, c).value = f'=IF({col}103=0,"",{col}112/{col}103)'

    # US common formulas
    ws.cell(123, c).value = f'={col}120+{col}121+{col}122'

wb.save('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')
print("Part 6 complete: HY forecast formulas entered")

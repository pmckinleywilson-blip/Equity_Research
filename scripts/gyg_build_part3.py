"""Part 3: Annual sheet - forecast formulas (FY26E-FY35E, cols G-P = 7-16)."""
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')
ws = wb['Annual']

# Helper to get Excel column letter
def CL(c):
    return get_column_letter(c)

# Forecast columns: G(7)=FY26E through P(16)=FY35E
for c in range(7, 17):
    col = CL(c)
    pcol = CL(c-1)  # prior year column

    # R7: Total Revenue = pull from HY sheet (1H + 2H)
    ws.cell(7, c).value = f"=INDEX('HY & Segments'!$A:$AG,MATCH($A7,'HY & Segments'!$A:$A,0),MATCH(\"1H\"&RIGHT({col}$1,2),'HY & Segments'!$3:$3,0))+INDEX('HY & Segments'!$A:$AG,MATCH($A7,'HY & Segments'!$A:$A,0),MATCH(\"2H\"&RIGHT({col}$1,2),'HY & Segments'!$3:$3,0))"

    # R8: Other Revenue = grow at 15% (trending with store growth)
    ws.cell(8, c).value = f'={pcol}8*1.15'

    # R9: Total Rev incl Other
    ws.cell(9, c).value = f'={col}7+{col}8'

    # R10: Revenue Growth
    ws.cell(10, c).value = f'=IF({pcol}7=0,"",{col}7/{pcol}7-1)'

    # R13: Food & Packaging = % of Revenue (hold at FY25 ratio ~26.1%)
    ws.cell(13, c).value = f'={col}7*-0.261'

    # R14: Employee Benefits = % of Revenue (hold at FY25 ratio ~42.4%)
    ws.cell(14, c).value = f'={col}7*-0.424'

    # R15: Admin+Marketing+Other = % of Revenue (improving from ~23.9% to ~22%)
    if c == 7:
        ws.cell(15, c).value = f'={col}7*-0.235'
    elif c == 8:
        ws.cell(15, c).value = f'={col}7*-0.230'
    else:
        ws.cell(15, c).value = f'={col}7*-0.225'

    # R16: Total Operating Costs
    ws.cell(16, c).value = f'={col}13+{col}14+{col}15'

    # R19: Statutory EBITDA
    ws.cell(19, c).value = f'={col}9+{col}16'

    # R20: Stat EBITDA Margin
    ws.cell(20, c).value = f'=IF({col}7=0,"",{col}19/{col}7)'

    # R35-37: Segment EBITDA - pull from HY
    ws.cell(35, c).value = f"=INDEX('HY & Segments'!$A:$AG,MATCH($A35,'HY & Segments'!$A:$A,0),MATCH(\"1H\"&RIGHT({col}$1,2),'HY & Segments'!$3:$3,0))+INDEX('HY & Segments'!$A:$AG,MATCH($A35,'HY & Segments'!$A:$A,0),MATCH(\"2H\"&RIGHT({col}$1,2),'HY & Segments'!$3:$3,0))"

    ws.cell(36, c).value = f"=INDEX('HY & Segments'!$A:$AG,MATCH($A36,'HY & Segments'!$A:$A,0),MATCH(\"1H\"&RIGHT({col}$1,2),'HY & Segments'!$3:$3,0))+INDEX('HY & Segments'!$A:$AG,MATCH($A36,'HY & Segments'!$A:$A,0),MATCH(\"2H\"&RIGHT({col}$1,2),'HY & Segments'!$3:$3,0))"

    ws.cell(37, c).value = 0  # Corporate = 0

    # R38: Group Seg EBITDA
    ws.cell(38, c).value = f'=SUM({col}35:{col}37)'

    # R39: Seg EBITDA Growth
    ws.cell(39, c).value = f'=IF({pcol}38=0,"",{col}38/{pcol}38-1)'

    # R40: Seg EBITDA Margin
    ws.cell(40, c).value = f'=IF({col}7=0,"",{col}38/{col}7)'

    # R42-44: Statutory adjustments
    # AASB16: grows with lease portfolio ~15%/yr
    ws.cell(42, c).value = f'={pcol}42*1.12'
    # SBP: hold flat at FY25 level
    ws.cell(43, c).value = f'={pcol}43'
    # Other: hold flat
    ws.cell(44, c).value = f'={pcol}44'

    # R45: Statutory EBITDA
    ws.cell(45, c).value = f'={col}38+{col}42+{col}43+{col}44'

    # R48: D&A - grow with store base ~15% then slowing
    if c <= 9:
        ws.cell(48, c).value = f'={pcol}48*1.14'
    else:
        ws.cell(48, c).value = f'={pcol}48*1.08'

    # R51: D&A / Revenue
    ws.cell(51, c).value = f'=IF({col}7=0,"",{col}48/{col}7)'

    # R54: EBIT = Stat EBITDA + D&A
    ws.cell(54, c).value = f'={col}45+{col}48'

    # R55: EBIT Growth
    ws.cell(55, c).value = f'=IF({pcol}54=0,"",{col}54/{pcol}54-1)'

    # R56: EBIT Margin
    ws.cell(56, c).value = f'=IF({col}7=0,"",{col}54/{col}7)'

    # R59: Finance Income - hold at FY25 level (large cash balance earning interest)
    ws.cell(59, c).value = f'={pcol}59*1.0'

    # R60: Finance Costs - grow with leases
    ws.cell(60, c).value = f'={pcol}60*1.10'

    # R61: Net Finance Costs
    ws.cell(61, c).value = f'={col}59+{col}60'

    # R68: PBT
    ws.cell(68, c).value = f'={col}54+{col}61'

    # R69: Tax = PBT * tax rate (use ~35% underlying)
    ws.cell(69, c).value = f'=IF({col}68>0,-{col}68*0.35,0)'

    # R70: Tax Rate
    ws.cell(70, c).value = f'=IF({col}68=0,"",{col}69/{col}68)'

    # R72: NPAT
    ws.cell(72, c).value = f'={col}68+{col}69'

    # R75: NPAT Growth
    ws.cell(75, c).value = f'=IF({pcol}72<=0,"",{col}72/{pcol}72-1)'

    # R76: NPAT Margin
    ws.cell(76, c).value = f'=IF({col}7=0,"",{col}72/{col}7)'

    # --- EPS ---
    ws.cell(79, c).value = f'={pcol}79'  # Shares flat
    ws.cell(80, c).value = f'={pcol}80'
    ws.cell(81, c).value = f'={pcol}81'
    ws.cell(82, c).value = f'={col}80+{col}81'
    ws.cell(84, c).value = f'=IF({col}80=0,"",{col}72/{col}80)'
    ws.cell(85, c).value = f'=IF({col}82=0,"",{col}72/{col}82)'
    ws.cell(86, c).value = f'=IF({pcol}85<=0,"",{col}85/{pcol}85-1)'

    # R88: DPS - 80% payout from FY26
    ws.cell(88, c).value = f'=IF({col}72>0,{col}84*0.8,0)'
    # R89: Total Divs
    ws.cell(89, c).value = f'={col}88*{col}80'
    # R90: Payout
    ws.cell(90, c).value = f'=IF({col}72<=0,"",{col}89/{col}72)'

    # --- KPIs ---
    # Total Restaurants: grow ~35/yr
    ws.cell(95, c).value = f'={pcol}95+35'
    ws.cell(96, c).value = f'={pcol}96+10'  # Corp Aus +10/yr
    ws.cell(97, c).value = f'={pcol}97+22'  # Fran Aus +22/yr
    ws.cell(98, c).value = f'={pcol}98+3'   # US +3/yr

    # Network Sales: grow ~15% initially declining
    growth = 0.14 if c <= 9 else 0.10
    ws.cell(99, c).value = f'={pcol}99*{1+growth}'

    # Comp Sales Growth
    ws.cell(100, c).value = f'={col}99/{pcol}99-1'

    # Corp Rest Margin trending to 19%
    ws.cell(101, c).value = min(0.19, 0.179 + (c-6)*0.003)

    # Royalty rate trending to 10%
    ws.cell(102, c).value = min(0.10, 0.097 + (c-6)*0.003)

    # G&A % NS declining
    ws.cell(103, c).value = max(0.05, 0.066 - (c-6)*0.003)

    # Seg EBITDA % NS
    ws.cell(104, c).value = f'=IF({col}99=0,"",{col}38/{col}99)'

    # --- Balance Sheet forecasts ---
    # Cash = prior + net change
    ws.cell(110, c).value = f'={pcol}110+{col}176'

    # Receivables, Inventory: hold ratios
    ws.cell(111, c).value = f'={col}7*{col}118'
    ws.cell(112, c).value = f'={col}7*{col}119'

    # PPE: prior + capex - D&A*0.4 (rough)
    ws.cell(113, c).value = f'={pcol}113-{col}160-{col}48*0.4'

    # Intangibles: flat
    ws.cell(114, c).value = f'={pcol}114'

    # ROU Assets: grow with leases
    ws.cell(115, c).value = f'={pcol}115*1.12'

    # Other Assets: grow slowly
    ws.cell(116, c).value = f'={pcol}116*1.08'

    # Total Assets
    ws.cell(117, c).value = f'=SUM({col}110:{col}116)'

    # BS ratios (hold at FY25 level for forecast)
    if c == 7:
        ws.cell(118, c).value = '=F118'
        ws.cell(119, c).value = '=F119'
        ws.cell(121, c).value = '=F121'
    else:
        ws.cell(118, c).value = f'={pcol}118'
        ws.cell(119, c).value = f'={pcol}119'
        ws.cell(121, c).value = f'={pcol}121'

    # Payables
    ws.cell(125, c).value = f'={col}7*{col}121'

    # Other Liabilities: flat
    ws.cell(126, c).value = f'={pcol}126'

    # Lease Liabilities: grow with ROU
    ws.cell(127, c).value = f'={pcol}127*1.12'

    # Banking Debt: 0
    ws.cell(128, c).value = 0

    # Total Liabilities
    ws.cell(129, c).value = f'=SUM({col}125:{col}128)'

    # Net Cash
    ws.cell(131, c).value = f'={col}110-{col}128'
    ws.cell(132, c).value = f'={col}128-{col}110+{col}127'
    ws.cell(133, c).value = f'=IF({col}38=0,"",{col}131/{col}38)'

    # Equity
    ws.cell(137, c).value = f'={pcol}137'  # flat
    ws.cell(138, c).value = f'={pcol}138+{col}72-{col}89'  # retained + NPAT - divs
    ws.cell(139, c).value = f'={pcol}139'  # flat
    ws.cell(140, c).value = 0
    ws.cell(141, c).value = f'=SUM({col}137:{col}140)'
    ws.cell(142, c).value = f'=IF({col}141=0,"",{col}72/{col}141)'
    ws.cell(143, c).value = f'=IF({col}141=0,"",28*{col}79/{col}141)'
    ws.cell(144, c).value = f'={col}117-{col}129-{col}141'

    # --- Cash Flow forecasts ---
    ws.cell(148, c).value = f'={col}45'
    ws.cell(149, c).value = f'=-({col}111-{pcol}111)-({col}112-{pcol}112)+({col}125-{pcol}125)'
    ws.cell(150, c).value = 0
    ws.cell(151, c).value = f'={col}148+{col}149+{col}150'
    ws.cell(152, c).value = f'={col}59'
    ws.cell(153, c).value = 0  # no bank interest
    ws.cell(154, c).value = f'={col}60*0.5'  # lease interest portion
    ws.cell(155, c).value = f'={col}69*0.9'  # tax paid ≈ tax expense
    ws.cell(156, c).value = f'=SUM({col}151:{col}155)'
    ws.cell(157, c).value = f'=IF({pcol}156=0,"",{col}156/{pcol}156-1)'

    # Investing
    ws.cell(160, c).value = f'={pcol}160*1.10'  # capex growing
    ws.cell(161, c).value = f'=IF({col}7=0,"",{col}160/{col}7)'
    ws.cell(162, c).value = 0
    ws.cell(163, c).value = 0
    ws.cell(164, c).value = 0
    ws.cell(165, c).value = 0
    ws.cell(166, c).value = f'=SUM({col}160:{col}165)'

    # Financing
    ws.cell(169, c).value = f'=-{pcol}89'  # prior year dividends paid
    ws.cell(170, c).value = 0
    ws.cell(171, c).value = f'=-{col}127*0.08'  # lease principal ~8% of lease liab
    ws.cell(172, c).value = 0
    ws.cell(173, c).value = 0
    ws.cell(174, c).value = f'=SUM({col}169:{col}173)'

    # Net change in cash
    ws.cell(176, c).value = f'={col}156+{col}166+{col}174'

    # OFCF
    ws.cell(179, c).value = f'={col}156'
    ws.cell(180, c).value = f'={col}160'
    ws.cell(181, c).value = f'={col}171'
    ws.cell(182, c).value = f'={col}179+{col}180+{col}181'
    ws.cell(183, c).value = f'=IF({col}82=0,"",{col}182/{col}82)'
    ws.cell(184, c).value = f'=IF({col}82=0,"",{col}183/28)'
    ws.cell(185, c).value = f'=IF({col}7=0,"",{col}182/{col}7)'

    # ROIC
    ws.cell(189, c).value = f'={col}141+{col}128-{col}110+{col}127'
    ws.cell(190, c).value = f'={col}54'
    ws.cell(191, c).value = f'=IF({col}189=0,"",{col}190/{col}189)'
    ws.cell(192, c).value = f'={col}190*(1+{col}70)'
    ws.cell(193, c).value = f'=IF({col}189=0,"",{col}192/{col}189)'

wb.save('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')
print("Part 3 complete: Annual forecast formulas entered")

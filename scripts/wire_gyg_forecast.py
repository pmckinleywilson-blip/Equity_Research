"""
Wire ALL forecast formulas in the GYG Model.
Steps:
  1. HY segment driver actuals (rows 87-106, cols F-J)
  2. HY segment driver forecast formulas (rows 87-106, cols K-AC)
  3. HY consolidated P&L forecast formulas (cols K-AC)
  4. Annual sheet forecast formulas (cols G-P) via INDEX/MATCH from HY
  5. Annual BS, CF, EPS, OpFCF, ROIC forecast formulas
  6. Value sheet fixes
"""

import openpyxl
from openpyxl.utils import get_column_letter as gcl
from copy import copy

FILE = '/home/pmwilson/Project_Equities/GYG/Models/GYG Model.xlsx'
wb = openpyxl.load_workbook(FILE)
hy = wb['HY & Segments']
ann = wb['Annual']
val = wb['Value']

# Color constants
BLUE = openpyxl.styles.Font(color='0000FF')
MAROON = openpyxl.styles.Font(color='C00000')
BLACK = openpyxl.styles.Font(color='000000')


def set_val(ws, row, col, value, font=None):
    """Set cell value and optionally font."""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font:
        cell.font = font


def copy_font(ws, row, col, font):
    """Set font on a cell preserving value."""
    cell = ws.cell(row=row, column=col)
    cell.font = font


# ===========================================================================
# STEP 1: HY SEGMENT DRIVER ACTUALS (rows 87-106, cols F=6 through J=10)
# ===========================================================================
print("STEP 1: Populating HY segment driver actuals...")

# Update row 96 label
hy.cell(row=96, column=2).value = 'EBITDA / Segment Revenue'

# --- AUSTRALIA (rows 88-97) ---
# Row 88: Corp Restaurants (closing count)
aus_corp = {6: 62, 7: 64, 8: 74, 9: 81, 10: 87}
for c, v in aus_corp.items():
    set_val(hy, 88, c, v, BLUE)

# Row 89: New Corp Restaurant Openings
aus_new = {6: 7, 7: 2, 8: 10, 9: 7, 10: 6}
for c, v in aus_new.items():
    set_val(hy, 89, c, v, BLUE)

# Row 90: Corp Restaurant Revenue
aus_rev = {6: 136.155, 7: 142.722, 8: 176.051, 9: 183.671, 10: 215.116}
for c, v in aus_rev.items():
    set_val(hy, 90, c, v, BLUE)

# Row 91: Corp Rev Growth (PCP comparison, col-2 = same half prior year)
# F(6) and G(7) have no PCP in driver section, so skip or leave blank
# H(8) vs F(6), I(9) vs G(7), J(10) vs H(8)
for c in [8, 9, 10]:
    pcp = c - 2
    formula = f'=IF({gcl(pcp)}90=0,"",{gcl(c)}90/{gcl(pcp)}90-1)'
    set_val(hy, 91, c, formula, BLUE)

# Row 92: Corp AUV (annualised) = Rev*2 / avg(prior_closing, current_closing)
# Prior closing for F(1H24) = FY23 closing = 55 (from KPI row 69, col E=5 has no data, use known 55)
# For subsequent: prior = prev col row 88
auv_vals = {6: 4.654, 7: 4.531, 8: 5.103, 9: 4.740, 10: 5.122}
for c, v in auv_vals.items():
    set_val(hy, 92, c, round(v, 3), BLUE)

# Row 93: Franchise Restaurants (Aus + Asia franchise combined)
fran_rest = {6: 143, 7: 152, 8: 161, 9: 169, 10: 177}
for c, v in fran_rest.items():
    set_val(hy, 93, c, v, BLUE)

# Row 94: Franchise Revenue
fran_rev = {6: 29.477, 7: 31.268, 8: 38.197, 9: 40.463, 10: 42.829}
for c, v in fran_rev.items():
    set_val(hy, 94, c, v, BLUE)

# Row 95: Total Australia Revenue = Corp Rev + Franchise Rev
for c in range(6, 11):
    set_val(hy, 95, c, f'={gcl(c)}90+{gcl(c)}94', BLUE)

# Row 96: EBITDA / Segment Revenue (recalculated)
# Actual EBITDA values from row 97 / Total Aus Rev from row 95
# We need to enter as formulas referencing the actual EBITDA and revenue
ebitda_vals = {6: 24.249, 7: 23.287, 8: 31.781, 9: 34.248, 10: 41.314}
for c in range(6, 11):
    # Revenue = row 90 + row 94, EBITDA = row 97
    # Use formula: =IF(row95=0,"",row97/row95)
    set_val(hy, 96, c, f'=IF({gcl(c)}95=0,"",{gcl(c)}97/{gcl(c)}95)', BLUE)

# Row 97: Segment EBITDA
for c, v in ebitda_vals.items():
    set_val(hy, 97, c, v, BLUE)

# --- US (rows 100-106) ---
# Row 100: US Corp Restaurants
us_rest = {6: 4, 7: 4, 8: 4, 9: 6, 10: 8}
for c, v in us_rest.items():
    set_val(hy, 100, c, v, BLUE)

# Row 101: New US Openings
us_new = {6: 0, 7: 0, 8: 0, 9: 2, 10: 2}
for c, v in us_new.items():
    set_val(hy, 101, c, v, BLUE)

# Row 102: US Revenue
us_rev = {6: 5.618, 7: 5.216, 8: 3.368, 9: 5.765, 10: 6.601}
for c, v in us_rev.items():
    set_val(hy, 102, c, v, BLUE)

# Row 103: US Rev Growth (PCP)
for c in [8, 9, 10]:
    pcp = c - 2
    formula = f'=IF({gcl(pcp)}102=0,"",{gcl(c)}102/{gcl(pcp)}102-1)'
    set_val(hy, 103, c, formula, BLUE)

# Row 104: US AUV
us_auv = {6: 3.210, 7: 2.608, 8: 1.684, 9: 2.306, 10: 1.886}
for c, v in us_auv.items():
    set_val(hy, 104, c, round(v, 3), BLUE)

# Row 105: US EBITDA Margin - not meaningful, leave as 0
for c in range(6, 11):
    set_val(hy, 105, c, 0, BLUE)

# Row 106: US Segment EBITDA
us_ebitda = {6: -3.095, 7: -3.449, 8: -5.014, 9: -8.209, 10: -8.308}
for c, v in us_ebitda.items():
    set_val(hy, 106, c, v, BLUE)

print("  Populated Australia rows 88-97 and US rows 100-106 with actuals (cols F-J)")

# ===========================================================================
# STEP 2: HY SEGMENT DRIVER FORECAST FORMULAS (cols K=11 through AC=29)
# ===========================================================================
print("STEP 2: Wiring HY segment driver forecast formulas...")

for c in range(11, 30):  # K(11) to AC(29)
    prev = c - 1       # previous column
    pcp = c - 2        # same half prior year (PCP)
    cl = gcl(c)
    pl = gcl(prev)
    pcl = gcl(pcp)

    # --- AUSTRALIA ---
    # Row 88: Corp Restaurants = prior + new openings
    set_val(hy, 88, c, f'={pl}88+{cl}89', BLACK)

    # Row 89: New Corp Restaurant Openings (maroon assumption)
    if c == 11:
        set_val(hy, 89, c, 6, MAROON)  # 2H26 = 6
    else:
        set_val(hy, 89, c, f'={pcl}89', MAROON)

    # Row 90: Corp Revenue = PCP * (1 + growth)
    set_val(hy, 90, c, f'={pcl}90*(1+{cl}91)', BLACK)

    # Row 91: Corp Rev Growth (maroon assumption)
    if c == 11:
        set_val(hy, 91, c, 0.05, MAROON)
    else:
        set_val(hy, 91, c, f'={pcl}91', MAROON)

    # Row 92: Corp AUV (analytical) = 2 * Rev / avg(prev_closing, curr_closing)
    set_val(hy, 92, c, f'=IF(({pl}88+{cl}88)/2=0,"",2*{cl}90/(({pl}88+{cl}88)/2))', BLACK)

    # Row 93: Franchise Restaurants = prev + 5
    set_val(hy, 93, c, f'={pl}93+5', BLACK)

    # Row 94: Franchise Rev = PCP * (1 + growth)
    set_val(hy, 94, c, f'={pcl}94*(1+{cl}91)', BLACK)

    # Row 95: Total Australia Revenue = Corp Rev + Franchise Rev
    set_val(hy, 95, c, f'={cl}90+{cl}94', BLACK)

    # Row 96: EBITDA / Segment Revenue (maroon assumption)
    if c == 11:
        set_val(hy, 96, c, 0.162, MAROON)  # ~16.2% continuing trend
    else:
        set_val(hy, 96, c, f'={pcl}96', MAROON)

    # Row 97: Segment EBITDA = Total Rev * EBITDA/SegRev margin
    set_val(hy, 97, c, f'={cl}95*{cl}96', BLACK)

    # --- US ---
    # Row 100: US Corp Restaurants = prev + new
    set_val(hy, 100, c, f'={pl}100+{cl}101', BLACK)

    # Row 101: New US Openings (maroon)
    if c == 11:
        set_val(hy, 101, c, 1, MAROON)
    else:
        set_val(hy, 101, c, f'={pcl}101', MAROON)

    # Row 102: US Revenue = PCP * (1 + growth)
    set_val(hy, 102, c, f'={pcl}102*(1+{cl}103)', BLACK)

    # Row 103: US Rev Growth (maroon)
    if c == 11:
        set_val(hy, 103, c, 0.15, MAROON)
    else:
        set_val(hy, 103, c, f'={pcl}103', MAROON)

    # Row 104: US AUV (analytical)
    set_val(hy, 104, c, f'=IF(({pl}100+{cl}100)/2=0,"",2*{cl}102/(({pl}100+{cl}100)/2))', BLACK)

    # Row 105: US EBITDA Margin - not used (keep 0)
    set_val(hy, 105, c, 0, MAROON)

    # Row 106: US Segment EBITDA = flat from PCP (losses reduce slowly)
    set_val(hy, 106, c, f'={pcl}106', MAROON)

print("  Wired Australia rows 88-97 and US rows 100-106 for cols K-AC")

# ===========================================================================
# STEP 3: HY CONSOLIDATED P&L FORECAST FORMULAS (cols K=11 through AC=29)
# ===========================================================================
print("STEP 3: Wiring HY consolidated P&L forecast formulas...")

for c in range(11, 30):
    prev = c - 1
    pcp = c - 2
    cl = gcl(c)
    pl = gcl(prev)
    pcl = gcl(pcp)

    # Row 7: Corp Sales Aus = from driver row 90
    set_val(hy, 7, c, f'={cl}90', BLACK)

    # Row 8: Corp Sales US = from driver row 102
    set_val(hy, 8, c, f'={cl}102', BLACK)

    # Row 9: Total Corp Sales
    set_val(hy, 9, c, f'={cl}7+{cl}8', BLACK)

    # Row 10: Franchise Royalty = PCP * (1 + Aus growth)
    set_val(hy, 10, c, f'={pcl}10*(1+{cl}91)', BLACK)

    # Row 11: Franchise Fee = PCP (flat)
    set_val(hy, 11, c, f'={pcl}11', BLACK)

    # Row 12: Total Revenue
    set_val(hy, 12, c, f'={cl}9+{cl}10+{cl}11', BLACK)

    # Row 13: Revenue Growth (YoY, PCP)
    set_val(hy, 13, c, f'=IF({pcl}12=0,"",{cl}12/{pcl}12-1)', BLACK)

    # Row 16: Marketing Levy = PCP * (1 + growth)
    set_val(hy, 16, c, f'={pcl}16*(1+{cl}91)', BLACK)

    # Row 17: Other Franchise = PCP * (1 + growth)
    set_val(hy, 17, c, f'={pcl}17*(1+{cl}91)', BLACK)

    # Row 18: Other Income = 0 (no disposals)
    set_val(hy, 18, c, 0, MAROON)

    # Row 19: Total Other Revenue = sum
    set_val(hy, 19, c, f'=SUM({cl}16:{cl}18)', BLACK)

    # Row 21: Total Rev & Other Income
    set_val(hy, 21, c, f'={cl}12+{cl}19', BLACK)

    # Row 24: Aus Seg EBITDA from driver
    set_val(hy, 24, c, f'={cl}97', BLACK)

    # Row 25: US Seg EBITDA from driver
    set_val(hy, 25, c, f'={cl}106', BLACK)

    # Row 26: Total Segment EBITDA
    set_val(hy, 26, c, f'={cl}24+{cl}25', BLACK)

    # Row 27: Cash Rent (maroon) = PCP (flat)
    set_val(hy, 27, c, f'={pcl}27', MAROON)

    # Row 28: SBP (maroon) = PCP (flat)
    set_val(hy, 28, c, f'={pcl}28', MAROON)

    # Row 29: Other = 0
    set_val(hy, 29, c, 0, MAROON)

    # Row 30: EBITDA
    set_val(hy, 30, c, f'={cl}26+{cl}27+{cl}28+{cl}29', BLACK)

    # Row 31: EBITDA Margin
    set_val(hy, 31, c, f'=IF({cl}21=0,"",{cl}30/{cl}21)', BLACK)

    # Expenses (memo) - back-solved from EBITDA
    # Row 39: Total Expenses = Total Rev & Other Income - EBITDA (back-solved)
    set_val(hy, 39, c, f'={cl}21-{cl}30', BLACK)

    # Individual expense lines: proportional to PCP mix
    for exp_row in [34, 35, 36, 37, 38]:
        set_val(hy, exp_row, c, f'=IF({pcl}39=0,0,{cl}39*{pcl}{exp_row}/{pcl}39)', BLACK)

    # D&A
    # Row 47: D&A / Revenue ratio (use PCP ratio, maroon)
    set_val(hy, 47, c, f'={pcl}47', MAROON)

    # Row 46: Total D&A = -Revenue * D&A/Rev (D&A/Rev from row 47 is negative since D&A is negative)
    # Actually row 47 formula is D&A/Rev where D&A is negative, so ratio is negative
    # Total D&A = Revenue * (negative ratio) = negative number. Correct.
    set_val(hy, 46, c, f'={cl}12*{cl}47', BLACK)

    # Individual D&A lines: proportional to PCP mix
    for da_row in [42, 43, 44, 45]:
        set_val(hy, da_row, c, f'=IF({pcl}46=0,0,{cl}46*{pcl}{da_row}/{pcl}46)', BLACK)

    # Row 49: EBIT
    set_val(hy, 49, c, f'={cl}30+{cl}46', BLACK)

    # Row 50: EBIT Margin
    set_val(hy, 50, c, f'=IF({cl}21=0,"",{cl}49/{cl}21)', BLACK)

    # Finance Income
    # Row 53: Term Deposit Income (maroon, flat from PCP)
    set_val(hy, 53, c, f'={pcl}53', MAROON)

    # Row 54: Lease Receivable Income (maroon, grows with lease book)
    set_val(hy, 54, c, f'={pcl}54*1.05', MAROON)

    # Row 55: Other Finance Income = 0
    set_val(hy, 55, c, 0, MAROON)

    # Row 56: Total Finance Income
    set_val(hy, 56, c, f'=SUM({cl}53:{cl}55)', BLACK)

    # Row 57: Lease Liability Costs (maroon, grows ~8%)
    set_val(hy, 57, c, f'={pcl}57*1.08', MAROON)

    # Row 58: Other Finance Costs = 0
    set_val(hy, 58, c, 0, MAROON)

    # Row 59: Total Finance Costs
    set_val(hy, 59, c, f'=SUM({cl}57:{cl}58)', BLACK)

    # Row 60: Net Finance
    set_val(hy, 60, c, f'={cl}56+{cl}59', BLACK)

    # Row 62: PBT
    set_val(hy, 62, c, f'={cl}49+{cl}60', BLACK)

    # Row 63: Tax = -PBT * tax rate (use 30% statutory)
    set_val(hy, 63, c, f'=-{cl}62*0.30', BLACK)

    # Row 64: Tax Rate (analytical)
    set_val(hy, 64, c, f'=IF({cl}62=0,"",{cl}63/{cl}62)', BLACK)

    # Row 65: NPAT
    set_val(hy, 65, c, f'={cl}62+{cl}63', BLACK)

    # Row 66: NPAT Growth (PCP)
    set_val(hy, 66, c, f'=IF({pcl}65=0,"",{cl}65/{pcl}65-1)', BLACK)

print("  Wired HY P&L rows 7-66 for cols K-AC")

# ===========================================================================
# STEP 4: ANNUAL SHEET FORECAST FORMULAS (cols G=7 through P=16)
# ===========================================================================
print("STEP 4: Wiring Annual P&L forecast formulas via INDEX/MATCH from HY...")

# P&L rows that have Column A keys and should use INDEX/MATCH (1H+2H)
index_match_rows = [7, 8, 10, 11, 16, 17, 18, 24, 25, 27, 28, 29,
                    34, 35, 36, 37, 38, 42, 43, 44, 45,
                    52, 53, 54, 56, 57, 62]

# Note: Annual row numbering differs from HY for some items
# Annual: 52=Term Dep, 53=Lease Recv, 54=Other Fin Inc, 55=Total Fin Inc, 56=Lease Costs, 57=Other Fin Costs
# HY:     53=Term Dep, 54=Lease Recv, 55=Other Fin Inc, 56=Total Fin Inc, 57=Lease Costs, 58=Other Fin Costs

# Map Annual row -> HY row (for INDEX/MATCH using Column A key)
# Actually INDEX/MATCH uses the Col A key directly, which matches between sheets.
# So we just need the Col A key from the Annual sheet.

for col_idx in range(7, 17):  # G(7) to P(16)
    cl = gcl(col_idx)

    # Get the FY year from row 1 (e.g., 2026)
    # The formula uses RIGHT(col$1,2) to get "26" and prepends "1H"/"2H"
    # Annual row 1 has year integers

    for row in index_match_rows:
        key = ann.cell(row=row, column=1).value
        if key is None:
            continue
        # Formula: =INDEX('HY & Segments'!$A:$AG,MATCH($A{row},'HY & Segments'!$A:$A,0),
        #           MATCH("1H"&RIGHT({cl}$1,2),'HY & Segments'!$3:$3,0))
        #         +INDEX(...)  for 2H
        formula = (
            f"=INDEX('HY & Segments'!$A:$AG,MATCH($A{row},'HY & Segments'!$A:$A,0),"
            f"MATCH(\"1H\"&RIGHT({cl}$1,2),'HY & Segments'!$3:$3,0))"
            f"+INDEX('HY & Segments'!$A:$AG,MATCH($A{row},'HY & Segments'!$A:$A,0),"
            f"MATCH(\"2H\"&RIGHT({cl}$1,2),'HY & Segments'!$3:$3,0))"
        )
        set_val(ann, row, col_idx, formula, BLACK)

    # Subtotal / arithmetic rows
    # Row 9: Total Corp Sales
    set_val(ann, 9, col_idx, f'={cl}7+{cl}8', BLACK)
    # Row 12: Total Revenue
    set_val(ann, 12, col_idx, f'=SUM({cl}9:{cl}11)', BLACK)
    # Row 13: Revenue Growth
    prev_cl = gcl(col_idx - 1)
    set_val(ann, 13, col_idx, f'=IF({prev_cl}12=0,"",{cl}12/{prev_cl}12-1)', BLACK)
    # Row 19: Total Other Revenue
    set_val(ann, 19, col_idx, f'=SUM({cl}16:{cl}18)', BLACK)
    # Row 21: Total Rev & Other Income
    set_val(ann, 21, col_idx, f'={cl}12+{cl}19', BLACK)
    # Row 26: Total Segment EBITDA
    set_val(ann, 26, col_idx, f'={cl}24+{cl}25', BLACK)
    # Row 30: EBITDA
    set_val(ann, 30, col_idx, f'=SUM({cl}26:{cl}29)', BLACK)
    # Row 31: EBITDA Margin
    set_val(ann, 31, col_idx, f'=IF({cl}21=0,"",{cl}30/{cl}21)', BLACK)
    # Row 39: Total Expenses
    set_val(ann, 39, col_idx, f'=SUM({cl}34:{cl}38)', BLACK)
    # Row 46: Total D&A
    set_val(ann, 46, col_idx, f'=SUM({cl}42:{cl}45)', BLACK)
    # Row 48: EBIT
    set_val(ann, 48, col_idx, f'={cl}30+{cl}46', BLACK)
    # Row 49: EBIT Margin
    set_val(ann, 49, col_idx, f'=IF({cl}21=0,"",{cl}48/{cl}21)', BLACK)
    # Row 55: Total Finance Income
    set_val(ann, 55, col_idx, f'=SUM({cl}52:{cl}54)', BLACK)
    # Row 58: Total Finance Costs
    set_val(ann, 58, col_idx, f'={cl}56+{cl}57', BLACK)
    # Row 59: Net Finance
    set_val(ann, 59, col_idx, f'={cl}55+{cl}58', BLACK)
    # Row 61: PBT
    set_val(ann, 61, col_idx, f'={cl}48+{cl}59', BLACK)
    # Row 62: Tax (INDEX/MATCH already set above)
    # Row 63: Effective Tax Rate
    set_val(ann, 63, col_idx, f'=IF({cl}61=0,"",{cl}62/{cl}61)', BLACK)
    # Row 64: NPAT
    set_val(ann, 64, col_idx, f'={cl}61+{cl}62', BLACK)
    # Row 65: NPAT Margin
    set_val(ann, 65, col_idx, f'=IF({cl}21=0,"",{cl}64/{cl}21)', BLACK)

print("  Wired Annual P&L rows 7-65 for cols G-P")

# ===========================================================================
# STEP 4b: ANNUAL EPS & DIVIDENDS (rows 69-82)
# ===========================================================================
print("STEP 4b: Wiring Annual EPS & Dividends...")

for col_idx in range(7, 17):
    cl = gcl(col_idx)
    prev_cl = gcl(col_idx - 1)

    # Row 69: YE Shares (maroon, flat)
    set_val(ann, 69, col_idx, f'={prev_cl}69', MAROON)

    # Row 70: WASO Basic = average of prior YE and current YE
    set_val(ann, 70, col_idx, f'=AVERAGE({prev_cl}69,{cl}69)', BLACK)

    # Row 71: Dilution (maroon, flat)
    set_val(ann, 71, col_idx, f'={prev_cl}71', MAROON)

    # Row 72: WASO Diluted
    set_val(ann, 72, col_idx, f'={cl}70+{cl}71', BLACK)

    # Row 74: Basic EPS (cents)
    set_val(ann, 74, col_idx, f'=IF({cl}70=0,"",{cl}64/{cl}70*100)', BLACK)

    # Row 75: Diluted EPS (cents)
    set_val(ann, 75, col_idx, f'=IF({cl}72=0,"",{cl}64/{cl}72*100)', BLACK)

    # Row 76: EPS Growth
    set_val(ann, 76, col_idx, f'=IF({prev_cl}74=0,"",{cl}74/{prev_cl}74-1)', BLACK)

    # Row 78: DPS = payout ratio * basic EPS (0 if EPS negative)
    set_val(ann, 78, col_idx, f'=IF({cl}74<=0,0,{cl}80*{cl}74)', BLACK)

    # Row 79: Total Dividends (A$m) = DPS * WASO / 100
    set_val(ann, 79, col_idx, f'={cl}78*{cl}70/100', BLACK)

    # Row 80: Payout Ratio (maroon, flat)
    set_val(ann, 80, col_idx, f'={prev_cl}80', MAROON)

    # Row 81: Dividend Yield
    set_val(ann, 81, col_idx, f'=IF(Value!$C$4=0,"",{cl}78/Value!$C$4/100)', BLACK)

    # Row 82: Dividend Growth
    set_val(ann, 82, col_idx, f'=IF({prev_cl}78=0,"",{cl}78/{prev_cl}78-1)', BLACK)

print("  Wired Annual EPS rows 69-82 for cols G-P")

# ===========================================================================
# STEP 5: ANNUAL BALANCE SHEET (rows 105-141)
# ===========================================================================
print("STEP 5: Wiring Annual Balance Sheet forecast formulas...")

for col_idx in range(7, 17):
    cl = gcl(col_idx)
    prev_cl = gcl(col_idx - 1)

    # Row 105: Cash = Prior + Net Change in Cash (row 174)
    set_val(ann, 105, col_idx, f'={prev_cl}105+{cl}174', BLACK)

    # Row 106: Trade Receivables = Revenue * Recv/Rev ratio
    set_val(ann, 106, col_idx, f'={cl}12*{cl}115', BLACK)

    # Row 107: Inventories = Revenue * Inv/Rev ratio
    set_val(ann, 107, col_idx, f'={cl}12*{cl}116', BLACK)

    # Row 108: Term Deposits (maroon, flat)
    set_val(ann, 108, col_idx, f'={prev_cl}108', MAROON)

    # Row 109: Finance Lease Receivables = Prior + New Lease Additions * 0.7 (franchise share)
    # Simplified: grow ~5% per year
    set_val(ann, 109, col_idx, f'={prev_cl}109*1.05', MAROON)

    # Row 110: PPE = Prior - Capex (negative) + PPE Dep (negative)
    # PPE = Prior PPE - Capex_PPE + Dep_PPE
    # Capex (row 158) is negative, so -Capex = add. Dep (row 43) is negative, so +Dep = subtract.
    set_val(ann, 110, col_idx, f'={prev_cl}110-{cl}158+{cl}43', BLACK)

    # Row 111: Intangibles (maroon, flat)
    set_val(ann, 111, col_idx, f'={prev_cl}111', MAROON)

    # Row 112: ROU Assets = Prior + New Leases + ROU Depreciation (negative)
    set_val(ann, 112, col_idx, f'={prev_cl}112+{cl}119+{cl}42', BLACK)

    # Row 113: Other Assets (maroon, flat)
    set_val(ann, 113, col_idx, f'={prev_cl}113', MAROON)

    # Row 114: Total Assets
    set_val(ann, 114, col_idx, f'=SUM({cl}105:{cl}113)', BLACK)

    # Row 115: Recv/Rev (maroon, flat from last actual)
    set_val(ann, 115, col_idx, f'={prev_cl}115', MAROON)

    # Row 116: Inv/Rev (maroon, flat)
    set_val(ann, 116, col_idx, f'={prev_cl}116', MAROON)

    # Row 117: Working Capital
    set_val(ann, 117, col_idx, f'={cl}106+{cl}107-{cl}122', BLACK)

    # Row 118: Pay/Rev (maroon, flat)
    set_val(ann, 118, col_idx, f'={prev_cl}118', MAROON)

    # Row 119: New Lease Additions (maroon, flat)
    set_val(ann, 119, col_idx, f'={prev_cl}119', MAROON)

    # Row 122: Trade Payables = Revenue * Pay/Rev
    set_val(ann, 122, col_idx, f'={cl}12*{cl}118', BLACK)

    # Row 123: Other Liabilities (maroon, flat)
    set_val(ann, 123, col_idx, f'={prev_cl}123', MAROON)

    # Row 124: Lease Liabilities = Prior + New Leases + Lease Principal (negative)
    set_val(ann, 124, col_idx, f'={prev_cl}124+{cl}119+{cl}169', BLACK)

    # Row 125: Borrowings = 0
    set_val(ann, 125, col_idx, 0, MAROON)

    # Row 126: Total Liabilities
    set_val(ann, 126, col_idx, f'=SUM({cl}122:{cl}125)', BLACK)

    # Row 128: Net Banking Debt
    set_val(ann, 128, col_idx, f'={cl}125-{cl}105', BLACK)

    # Row 129: Adj Net Debt (incl leases)
    set_val(ann, 129, col_idx, f'={cl}128+{cl}124', BLACK)

    # Row 130: ND/EBITDA
    set_val(ann, 130, col_idx, f'=IF({cl}30=0,"",{cl}128/{cl}30)', BLACK)

    # Row 131: Gearing
    set_val(ann, 131, col_idx, f'=IF(({cl}128+{cl}138)=0,"",{cl}128/({cl}128+{cl}138))', BLACK)

    # Row 134: Issued Capital = Prior + Share Issues
    set_val(ann, 134, col_idx, f'={prev_cl}134+{cl}168', BLACK)

    # Row 135: Retained Profits = Prior + NPAT - Dividends
    set_val(ann, 135, col_idx, f'={prev_cl}135+{cl}64-{cl}79', BLACK)

    # Row 136: Reserves = Prior - SBP (SBP is negative, so -(-) = +)
    set_val(ann, 136, col_idx, f'={prev_cl}136-{cl}28', BLACK)

    # Row 137: Minorities
    set_val(ann, 137, col_idx, 0, MAROON)

    # Row 138: Total Equity
    set_val(ann, 138, col_idx, f'=SUM({cl}134:{cl}137)', BLACK)

    # Row 139: ROE
    set_val(ann, 139, col_idx, f'=IF({cl}138=0,"",{cl}64/{cl}138)', BLACK)

    # Row 140: P/B
    set_val(ann, 140, col_idx, f'=IF(OR({cl}138=0,Value!$C$4=0),"",Value!$C$4*{cl}69/{cl}138)', BLACK)

    # Row 141: BS Check
    set_val(ann, 141, col_idx, f'={cl}114-{cl}126-{cl}138', BLACK)

print("  Wired Annual BS rows 105-141 for cols G-P")

# ===========================================================================
# STEP 5b: ANNUAL CASH FLOW (rows 143-183)
# ===========================================================================
print("STEP 5b: Wiring Annual Cash Flow forecast formulas...")

for col_idx in range(7, 17):
    cl = gcl(col_idx)
    prev_cl = gcl(col_idx - 1)

    # CFO
    # Row 145: EBITDA
    set_val(ann, 145, col_idx, f'={cl}30', BLACK)

    # Row 146: WC Change = -(change in receivables) - (change in inventory) + (change in payables)
    set_val(ann, 146, col_idx, f'=-({cl}106-{prev_cl}106)-({cl}107-{prev_cl}107)+({cl}122-{prev_cl}122)', BLACK)

    # Row 147: Non-cash/Significant Items = -SBP (add-back)
    set_val(ann, 147, col_idx, f'=-{cl}28', BLACK)

    # Row 148: Gross OCF
    set_val(ann, 148, col_idx, f'=SUM({cl}145:{cl}147)', BLACK)

    # Row 149: Interest Received = Total Finance Income
    set_val(ann, 149, col_idx, f'={cl}55', BLACK)

    # Row 150: Interest Paid = Other Finance Costs
    set_val(ann, 150, col_idx, f'={cl}57', BLACK)

    # Row 151: Lease Interest Paid = Lease Liability Costs
    set_val(ann, 151, col_idx, f'={cl}56', BLACK)

    # Row 152: Tax Paid = Tax Expense (simplified)
    set_val(ann, 152, col_idx, f'={cl}62', BLACK)

    # Row 153: Net OCF
    set_val(ann, 153, col_idx, f'={cl}148+{cl}149+{cl}150+{cl}151+{cl}152', BLACK)

    # Row 154: OCF Growth
    set_val(ann, 154, col_idx, f'=IF({prev_cl}153=0,"",{cl}153/{prev_cl}153-1)', BLACK)

    # Row 155: EBITDA CF Conversion
    set_val(ann, 155, col_idx, f'=IF({cl}145=0,"",{cl}153/{cl}145)', BLACK)

    # CFI
    # Row 158: Capex PPE = Capex/Sales ratio * Revenue (negative)
    set_val(ann, 158, col_idx, f'={cl}159*{cl}12', BLACK)

    # Row 159: Capex/Sales (maroon, flat from last actual)
    set_val(ann, 159, col_idx, f'={prev_cl}159', MAROON)

    # Row 160: Capex Intangibles = 0
    set_val(ann, 160, col_idx, 0, MAROON)

    # Row 161: Acquisitions = 0
    set_val(ann, 161, col_idx, 0, MAROON)

    # Row 162: Asset Sales = 0
    set_val(ann, 162, col_idx, 0, MAROON)

    # Row 163: Other CFI (Term Deposits) = 0
    set_val(ann, 163, col_idx, 0, MAROON)

    # Row 164: Total ICF
    set_val(ann, 164, col_idx, f'=SUM({cl}158,{cl}160:{cl}163)', BLACK)

    # CFF
    # Row 167: Dividends Paid = negative of prior year dividends
    set_val(ann, 167, col_idx, f'=-{prev_cl}79', BLACK)

    # Row 168: Share Issues/Buybacks = 0 (maroon)
    set_val(ann, 168, col_idx, 0, MAROON)

    # Row 169: Lease Principal = -Prior Lease Liabilities / 8 (avg lease life)
    set_val(ann, 169, col_idx, f'=-{prev_cl}124/8', BLACK)

    # Row 170: Debt Change = 0
    set_val(ann, 170, col_idx, 0, MAROON)

    # Row 171: Other CFF (Lease Incentives, maroon, flat)
    set_val(ann, 171, col_idx, f'={prev_cl}171', MAROON)

    # Row 172: Total CFF
    set_val(ann, 172, col_idx, f'=SUM({cl}167:{cl}171)', BLACK)

    # Row 174: Net Change in Cash
    set_val(ann, 174, col_idx, f'={cl}153+{cl}164+{cl}172', BLACK)

    # OpFCF
    # Row 177: Net OCF
    set_val(ann, 177, col_idx, f'={cl}153', BLACK)

    # Row 178: Net Capex
    set_val(ann, 178, col_idx, f'={cl}158+{cl}160', BLACK)

    # Row 179: Lease Principal
    set_val(ann, 179, col_idx, f'={cl}169', BLACK)

    # Row 180: Operating Free Cash Flow
    set_val(ann, 180, col_idx, f'=SUM({cl}177:{cl}179)', BLACK)

    # Row 181: FCF per Share (cents)
    set_val(ann, 181, col_idx, f'=IF({cl}72=0,"",{cl}180/{cl}72*100)', BLACK)

    # Row 182: FCF Yield
    set_val(ann, 182, col_idx, f'=IF(Value!$C$4=0,"",{cl}181/Value!$C$4/100)', BLACK)

    # Row 183: FCF Margin
    set_val(ann, 183, col_idx, f'=IF({cl}12=0,"",{cl}180/{cl}12)', BLACK)

    # ROIC
    # Row 187: Invested Capital
    set_val(ann, 187, col_idx, f'={cl}138+{cl}128', BLACK)

    # Row 188: Underlying EBIT
    set_val(ann, 188, col_idx, f'={cl}48', BLACK)

    # Row 189: ROFE
    set_val(ann, 189, col_idx, f'=IF({cl}187=0,"",{cl}188/{cl}187)', BLACK)

    # Row 190: NOPAT (use 30% tax rate for forecast)
    set_val(ann, 190, col_idx, f'={cl}188*(1-0.30)', BLACK)

    # Row 191: ROIC
    set_val(ann, 191, col_idx, f'=IF({cl}187=0,"",{cl}190/{cl}187)', BLACK)

print("  Wired Annual CF rows 145-191 for cols G-P")

# ===========================================================================
# STEP 6: VALUE SHEET FIXES
# ===========================================================================
print("STEP 6: Updating Value sheet...")

# Update tax rate from 0.28 (NZ) to 0.30 (Australia)
val['C17'] = 0.30

# Row 68 references "EBITDA-Underlying EBITDA" which doesn't exist - fix to "EBITDA-Statutory EBITDA"
# Check current formula
current_68 = val.cell(row=68, column=5).value
if current_68 and 'EBITDA-Underlying EBITDA' in str(current_68):
    new_68 = current_68.replace('EBITDA-Underlying EBITDA', 'EBITDA-Statutory EBITDA')
    val.cell(row=68, column=5).value = new_68
    print("  Fixed Value!E68 SOTP implied multiple formula key")

print("  Updated Value sheet tax rate to 0.30")

# ===========================================================================
# SAVE
# ===========================================================================
print("\nSaving workbook...")
wb.save(FILE)
print(f"Saved to {FILE}")

# ===========================================================================
# SUMMARY
# ===========================================================================
print("\n" + "="*70)
print("SUMMARY OF CHANGES")
print("="*70)
print()
print("HY & Segments sheet:")
print("  - Populated segment driver actuals (rows 88-97, 100-106) for cols F-J")
print("    Australia: Corp restaurants, openings, revenue, growth, AUV, franchise,")
print("               total revenue, EBITDA margin, segment EBITDA")
print("    US: Restaurants, openings, revenue, growth, AUV, EBITDA")
print("  - Wired segment driver forecasts (cols K-AC, 19 half-year periods)")
print("    Key assumptions (maroon): new openings, rev growth, EBITDA margin")
print("  - Wired consolidated P&L forecasts (rows 7-66, cols K-AC)")
print("    Revenue, EBITDA bridge, expenses (back-solved), D&A, finance, tax, NPAT")
print()
print("Annual sheet:")
print("  - Wired P&L via INDEX/MATCH 1H+2H from HY (rows 7-65, cols G-P)")
print("  - Wired EPS & Dividends (rows 69-82)")
print("  - Wired Balance Sheet (rows 105-141)")
print("  - Wired Cash Flow Statement (rows 145-174)")
print("  - Wired Operating Free Cash Flow (rows 177-183)")
print("  - Wired ROIC (rows 187-191)")
print()
print("Value sheet:")
print("  - Updated tax rate from 0.28 to 0.30 (Australian corporate rate)")
print("  - Fixed SOTP implied multiple formula key")
print()
print("Assumption inputs (maroon) set for first forecast period:")
print("  HY: Corp openings=6, Rev growth=5%, EBITDA/SegRev=16.2%")
print("       US openings=1, US rev growth=15%, US EBITDA=flat from PCP")
print("  Annual: BS ratios flat from FY25, Capex/Sales flat, Leases flat")
print("          Payout ratio flat, Shares flat, Tax rate 30%")

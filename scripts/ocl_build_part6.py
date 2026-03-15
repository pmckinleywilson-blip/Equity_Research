"""Part 6: Update Value sheet and apply formatting."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter as gcl
from datetime import datetime

DST = '/home/pmwilson/Project_Equities/OCL/Models/OCL Model.xlsx'
wb = openpyxl.load_workbook(DST)

# ==========================================
# VALUE SHEET
# ==========================================
vs = wb['Value']

# Update company info
vs['B4'] = 'Current Share Price (AUD)'
vs['C4'] = 16.50  # Approximate OCL.AX share price

vs['B5'] = 'Shares Outstanding (#m)'
vs['C5'] = '=INDEX(Annual!$D:$M,MATCH("EPS-YE Shares",Annual!$A:$A,0),MATCH($D$24,Annual!$D$3:$M$3,0)-1)'

vs['B6'] = 'Market Cap (A$m)'
vs['C6'] = '=C4*C5'

# Net Debt → Net Cash (OCL has no debt, flip sign)
vs['B7'] = 'Net Cash (A$m)'
vs['C7'] = '=INDEX(Annual!$D:$M,MATCH("BS-Cash",Annual!$A:$A,0),MATCH($D$24,Annual!$D$3:$M$3,0)-1)'

vs['B8'] = 'Market EV (A$m)'
vs['C8'] = '=C6-C7'  # EV = MCap - Net Cash

vs['B9'] = 'Valuation Date'
vs['C9'] = datetime(2026, 3, 15)

# WACC inputs - adjust for OCL (Australian SaaS)
vs['C12'] = 0.04    # Risk-free rate
vs['C13'] = 0.055   # ERP  
vs['C14'] = 0.85    # Beta
vs['C17'] = 0.25    # Tax rate (Australian)
vs['C16'] = 0.05    # Pre-tax cost of debt (notional)
vs['C19'] = 0.0     # Debt weight (no debt)
vs['C21'] = 0.03    # Terminal growth

# Update stub period formula
vs['C22'] = '=(INDEX(Annual!$D:$M,4,MATCH($D$24,Annual!$D$3:$M$3,0))-C9)/365.25'

# FCF Projection headers - FY26E to FY30E (5 years)
# Template had 10 years D-M, OCL only needs 5 forecast years
fcf_labels = ['FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E']
for i, label in enumerate(fcf_labels):
    vs.cell(row=24, column=4+i, value=label)

# Clear old projection columns beyond needed
for col in range(9, 14):
    for row in range(24, 50):
        vs.cell(row=row, column=col).value = None

# FCF rows - update formulas to reference Annual sheet with correct range
fcf_rows = {
    25: ('EBITDA', 'A$m', 'EBITDA-Underlying EBITDA'),
    26: ('less D&A', 'A$m', 'DA-Total DA'),
    27: ('EBIT', 'A$m', 'EBIT-Underlying EBIT'),
}

for fcf_col_idx, label in enumerate(fcf_labels):
    col = 4 + fcf_col_idx
    
    # Row 25: EBITDA
    vs.cell(row=25, column=col, value=f'=INDEX(Annual!$D:$M,MATCH("EBITDA-Underlying EBITDA",Annual!$A:$A,0),MATCH({gcl(col)}$24,Annual!$D$3:$M$3,0))')
    
    # Row 26: D&A
    vs.cell(row=26, column=col, value=f'=INDEX(Annual!$D:$M,MATCH("DA-Total DA",Annual!$A:$A,0),MATCH({gcl(col)}$24,Annual!$D$3:$M$3,0))')
    
    # Row 27: EBIT
    vs.cell(row=27, column=col, value=f'=INDEX(Annual!$D:$M,MATCH("EBIT-Underlying EBIT",Annual!$A:$A,0),MATCH({gcl(col)}$24,Annual!$D$3:$M$3,0))')
    
    # Row 28: Tax on EBIT
    vs.cell(row=28, column=col, value=f'=-{gcl(col)}27*$C$17')
    
    # Row 29: NOPAT
    vs.cell(row=29, column=col, value=f'={gcl(col)}27+{gcl(col)}28')
    
    # Row 30: plus D&A
    vs.cell(row=30, column=col, value=f'=-{gcl(col)}26')
    
    # Row 31: less Capex
    vs.cell(row=31, column=col, value=f'=INDEX(Annual!$D:$M,MATCH("CF-Capex PPE",Annual!$A:$A,0),MATCH({gcl(col)}$24,Annual!$D$3:$M$3,0))+INDEX(Annual!$D:$M,MATCH("CF-Capex Intang",Annual!$A:$A,0),MATCH({gcl(col)}$24,Annual!$D$3:$M$3,0))')
    
    # Row 32: less WC Change
    vs.cell(row=32, column=col, value=f'=INDEX(Annual!$D:$M,MATCH("CF-WC Change",Annual!$A:$A,0),MATCH({gcl(col)}$24,Annual!$D$3:$M$3,0))')
    
    # Row 33: FCFF
    vs.cell(row=33, column=col, value=f'={gcl(col)}29+{gcl(col)}30+{gcl(col)}31+{gcl(col)}32')

# Terminal value (last forecast column = H = col 8)
last_fcf_col = 4 + len(fcf_labels) - 1  # H=8
lc = gcl(last_fcf_col)

# Row 34: Normalised FCFF
vs.cell(row=34, column=last_fcf_col, value=f'={lc}29+{lc}32')

# Row 35: Terminal Value
vs.cell(row=35, column=last_fcf_col, value=f'={lc}34*(1+$C$21)/($C$20-$C$21)')

# Discount factors
for i, label in enumerate(fcf_labels):
    col = 4 + i
    vs.cell(row=37, column=col, value=f'=1/(1+$C$20)^($C$22+{i})')
    vs.cell(row=38, column=col, value=f'={gcl(col)}33*{gcl(col)}37')

# PV of TV
vs.cell(row=39, column=last_fcf_col, value=f'={lc}35*{lc}37')

# Sum of PVs
vs['C41'] = f'=SUM(D38:{lc}38)'
vs['C42'] = f'={lc}39'
vs['C43'] = '=C41+C42'

# Equity bridge - adjust for OCL (no debt, has leases)
vs['B45'] = 'plus Net Cash'
vs['C45'] = '=C7'  # Add back net cash (positive)

vs['B46'] = 'less Lease Liabilities'
vs['C46'] = f'=-INDEX(Annual!$D:$M,MATCH("BS-Lease Liabilities",Annual!$A:$A,0),MATCH($D$24,Annual!$D$3:$M$3,0)-1)'

vs['C47'] = '=C43+C45+C46'
vs['C48'] = '=C47/C5'
vs['C49'] = '=IF(C4=0,"",C48/C4-1)'

# ==========================================
# EV/EBITDA SOTP - update for OCL segments
# ==========================================
vs['C54'] = 'FY27E'

# Segment names and EBITDA references
# OCL is single-segment EBITDA, so SOTP is less meaningful
# But we can show it as a group-level multiple
vs['B57'] = 'Group EBITDA'
vs['C57'] = f'=INDEX(Annual!$D:$M,MATCH("EBITDA-Underlying EBITDA",Annual!$A:$A,0),MATCH($C$54,Annual!$D$3:$M$3,0))'
vs['D57'] = 25  # SaaS multiple
vs['E57'] = '=C57*D57'

# Clear old segment rows
for row in [58, 59]:
    for col in range(2, 6):
        vs.cell(row=row, column=col).value = None

vs['B61'] = 'Group EV'
vs['E61'] = '=E57'

vs['B62'] = 'plus Net Cash'
vs['E62'] = '=C7'

vs['B63'] = 'less Lease Liabilities'
vs['E63'] = f'=-INDEX(Annual!$D:$M,MATCH("BS-Lease Liabilities",Annual!$A:$A,0),MATCH($D$24,Annual!$D$3:$M$3,0)-1)'

vs['E64'] = '=E61+E62+E63'
vs['E65'] = '=E64/C5'
vs['E66'] = '=IF(C4=0,"",E65/C4-1)'

vs['B68'] = 'Implied Group EV/EBITDA'
vs['E68'] = f'=IF(INDEX(Annual!$D:$M,MATCH("EBITDA-Underlying EBITDA",Annual!$A:$A,0),MATCH($C$54,Annual!$D$3:$M$3,0))=0,"",E61/INDEX(Annual!$D:$M,MATCH("EBITDA-Underlying EBITDA",Annual!$A:$A,0),MATCH($C$54,Annual!$D$3:$M$3,0)))'

# ==========================================
# FORMATTING PASS
# ==========================================
# Apply number formats across both sheets

ws_a = wb['Annual']
ws_h = wb['HY & Segments']

# Number format definitions
NUM_FMT = '#,##0.0'
PCT_FMT = '0.0%'
EPS_FMT = '0.000'
INT_FMT = '#,##0'
X_FMT = '0.0x'

# Annual sheet formatting
for col in range(4, 14):
    # Revenue/cost rows - #,##0.0
    for row in [7,8,9,10,11,15,18,23,24,25,26,30,35,36,37,38,41,42,43,44,48,53,54,55,58,59,61,62,63,77,78]:
        ws_a.cell(row=row, column=col).number_format = NUM_FMT
    
    # % rows
    for row in [12,19,20,27,31,32,45,49,50,60,64,65,75,79,80,81,88,91,92,93,108,110,124,131,146,150,171,172,177,179]:
        ws_a.cell(row=row, column=col).number_format = PCT_FMT
    
    # EPS rows
    for row in [73,74,170]:
        ws_a.cell(row=row, column=col).number_format = EPS_FMT
    
    # Share count rows
    for row in [68,69,70,71,94,95]:
        ws_a.cell(row=row, column=col).number_format = '0.0'
    
    # BS rows
    for row in range(99, 134):
        if ws_a.cell(row=row, column=col).number_format not in [PCT_FMT]:
            ws_a.cell(row=row, column=col).number_format = NUM_FMT
    
    # CF rows
    for row in range(137, 180):
        if ws_a.cell(row=row, column=col).number_format not in [PCT_FMT]:
            ws_a.cell(row=row, column=col).number_format = NUM_FMT
    
    # x format
    for row in [132]:
        ws_a.cell(row=row, column=col).number_format = X_FMT
    
    # KPI rows
    for row in [84,85,86,87,89,90]:
        ws_a.cell(row=row, column=col).number_format = NUM_FMT

# HY sheet formatting
for col in range(4, 24):
    for row in [7,8,9,10,11,15,18,23,24,25,26,30,35,36,37,38,41,42,43,44,48,53,54,55,57,58,60,61,62]:
        ws_h.cell(row=row, column=col).number_format = NUM_FMT
    
    for row in [12,19,20,27,31,32,45,49,50,59,63,70,73,74,75,81,84,89,92,97,100,104,105,106,107,108,109,110]:
        ws_h.cell(row=row, column=col).number_format = PCT_FMT
    
    for row in [66,67,68,69,71,72,76,77,80,82,83,85,88,90,91,93,96,98,99,101,111,112]:
        ws_h.cell(row=row, column=col).number_format = NUM_FMT
    
    for row in [113]:
        ws_h.cell(row=row, column=col).number_format = '0.0'

# Apply blue font to all actual data cells that don't already have it
# (ensure formulas get black font, hardcodes get blue)
# This was already handled during data entry in Part 3

wb.save(DST)
print('Part 6 complete: Value sheet updated, formatting applied')

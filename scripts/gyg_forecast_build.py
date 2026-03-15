#!/usr/bin/env python3
"""
GYG Model Forecast Build Script
Wires forecast formulas on HY & Segments (Zone 2 drivers, Zone 1 P&L),
Annual (INDEX/MATCH from HY, BS roll-forwards, CF linkages),
and updates the Value sheet for GYG.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
import os
import sys

MODEL_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                          'GYG', 'Models', 'GYG Model.xlsx')

MAROON = 'FFC00000'
BLACK = 'FF000000'

def copy_font_with_color(cell, color):
    """Return a new Font based on cell's font but with a different color."""
    f = cell.font
    return openpyxl.styles.Font(
        name=f.name, size=f.size, bold=f.bold, italic=f.italic,
        underline=f.underline, strike=f.strike,
        color=openpyxl.styles.colors.Color(rgb=color)
    )

def set_formula(ws, row, col, formula, color=BLACK):
    """Set a cell's value to a formula and apply font color."""
    cell = ws.cell(row=row, column=col)
    cell.value = formula
    cell.font = copy_font_with_color(cell, color)

def set_value(ws, row, col, value, color=MAROON):
    """Set a cell's value (hardcoded assumption) with maroon color."""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.font = copy_font_with_color(cell, color)

def get_col_letter(col):
    return get_column_letter(col)


def build_hy_zone2_forecasts(ws):
    """Wire HY & Segments Zone 2 forecast formulas from column I (2H26E) to AA (2H35E)."""
    print("  Wiring HY Zone 2 forecasts...")

    # Column mapping: I=9 (2H26E), J=10 (1H27E), K=11 (2H27E), ...
    # Last actual 1H: col H=8 (1H26)
    # Last actual 2H: col G=7 (2H25)
    # PCP for col I (2H26E) = col G (2H25) — 2 cols back
    # PCP for col J (1H27E) = col H (1H26) — 2 cols back
    # General: PCP = col - 2

    forecast_start = 9  # Column I = 2H26E
    forecast_end = 27   # Column AA = 2H35E

    # First, seed assumption values for the first forecast column (I=2H26E)
    # These will be flatlined from PCP via formula for subsequent columns

    for col in range(forecast_start, forecast_end + 1):
        c = get_col_letter(col)
        pcp = get_col_letter(col - 2)  # PCP is 2 columns back
        prior = get_col_letter(col - 1)  # Prior half

        # ============ AUSTRALIA SEGMENT ============

        # Row 98: DT Count = prior half DT + new DT openings
        # We'll use a simple growth approach: count grows based on PCP growth
        # For DT: GYG guides ~23 new DT per year across network → ~12 per half
        # Use PCP + fixed increment (12 per half as default)
        # Actually, DT count at 1H26=126, 2H25=117, 1H25=unknown, 2H24=unknown
        # Only have 2H25 and 1H26 data. Use PCP + 9 (1H26 added 9 = 126-117)
        # For forecast: use prior period count + assumed increment
        # Increment = PCP growth (col - col-2)
        # Simplify: DT Count = PCP + (PCP - PCP_prior) if we had enough history
        # Safest: flatline PCP growth at ~9 per half
        if col == forecast_start:
            # 2H26E: DT count = 1H26 count (H98) + 9 (assumption)
            set_value(ws, 98, col, 135)  # 126 + 9 = 135
        else:
            # Subsequent: PCP + 9
            set_formula(ws, 98, col, f'={pcp}98+9', MAROON)

        # Row 99: Strip Count
        if col == forecast_start:
            set_value(ws, 99, col, 76)  # 73 + 3
        else:
            set_formula(ws, 99, col, f'={pcp}99+3', MAROON)

        # Row 100: Other Count - declining, flatline from PCP
        set_formula(ws, 100, col, f'={pcp}100', MAROON)

        # Row 101: Total Network = SUM(DT+Strip+Other)
        set_formula(ws, 101, col, f'=SUM({c}98:{c}100)')

        # Row 102: DT AUV (annualised) - flatline from PCP
        if col == forecast_start:
            # PCP (2H25) has no AUV data, but 1H26 has 6.9
            # Use 1H26 value as starting point
            set_value(ws, 102, col, 6.9, MAROON)
        else:
            set_formula(ws, 102, col, f'={pcp}102*(1+{c}103)', MAROON)

        # Row 103: DT AUV Growth - assumption (flatline 2.5% = comp sales growth proxy)
        set_value(ws, 103, col, 0.025, MAROON)

        # Row 104: Strip AUV - similar to DT
        if col == forecast_start:
            set_value(ws, 104, col, 5.2, MAROON)
        else:
            set_formula(ws, 104, col, f'={pcp}104*(1+{c}105)', MAROON)

        # Row 105: Strip AUV Growth
        set_value(ws, 105, col, 0.025, MAROON)

        # Row 106: Other AUV - flatline
        if col == forecast_start:
            set_value(ws, 106, col, 4.0, MAROON)
        else:
            set_formula(ws, 106, col, f'={pcp}106', MAROON)

        # Row 107: AU Network Sales = DT×AUV/2 + Strip×AUV/2 + Other×AUV/2
        # AUV is annualised, so half-year sales = count × AUV / 2
        set_formula(ws, 107, col,
            f'={c}98*{c}102/2+{c}99*{c}104/2+{c}100*{c}106/2')

        # Row 108: SG Network Sales - PCP × (1 + growth)
        set_formula(ws, 108, col, f'={pcp}108*(1+{c}109)')

        # Row 109: SG Sales Growth - flatline assumption
        if col == forecast_start:
            set_value(ws, 109, col, 0.10, MAROON)  # 10% growth for SG
        else:
            set_formula(ws, 109, col, f'={pcp}109', MAROON)

        # Row 110: JP Network Sales - PCP × (1 + growth)
        set_formula(ws, 110, col, f'={pcp}110*(1+{c}111)')

        # Row 111: JP Sales Growth - flatline assumption
        if col == forecast_start:
            set_value(ws, 111, col, 0.10, MAROON)
        else:
            set_formula(ws, 111, col, f'={pcp}111', MAROON)

        # Row 112: Total AU Segment Network Sales
        set_formula(ws, 112, col, f'={c}107+{c}108+{c}110')

        # ---- Corporate Build ----

        # Row 115: Corp DT Count - derive from network DT * corp share
        # Corp share ≈ 117 DT / 224 total × some fraction...
        # Actually we don't have corp DT vs franchise DT split
        # Leave blank for now, use total corp count

        # Row 117: Total Corp Count = PCP + new openings
        set_formula(ws, 117, col, f'={pcp}117+{c}118')

        # Row 118: New Corp Openings - assumption
        if col == forecast_start:
            set_value(ws, 118, col, 4, MAROON)  # ~8 per year, 4 per half
        else:
            set_formula(ws, 118, col, f'={pcp}118', MAROON)

        # Row 119: Corp Restaurant Sales
        # Method: Corp count × network AUV (weighted) × corp_sales_share_of_network
        # Simpler: derive from network sales × corp share
        # Corp share of AU network sales ≈ 183.6/556 = 33% (2H25), 215.1/632.1 = 34% (1H26)
        # Or: Corp sales = Corp count × implied corp AUV / 2
        # Implied Corp AUV = Corp Sales × 2 / Corp Count
        # 1H26: 215.116 × 2 / 85 = 5.06 per restaurant
        # Use: Corp Sales = Corp Count × Corp AUV / 2
        # Corp AUV ≈ Total Network Sales / Total Network Count (proxy)
        # Better: Corp sales grows at network sales growth rate × corp count growth
        # Simplest: Corp Sales = PCP × (Corp Count / PCP Corp Count) × (1 + comp sales growth)
        # Comp sales growth embedded in AUV growth above
        # Use: Corp Sales = (Corp Count / Total Network Count) × AU Network Sales
        # This assumes corp restaurants have same average sales as network
        set_formula(ws, 119, col,
            f'=({c}117/{c}101)*{c}107')

        # Row 120: Corp Restaurant Margin % - flatline from PCP
        set_formula(ws, 120, col, f'={pcp}120', MAROON)

        # Row 121: Corp Restaurant Margin ($)
        set_formula(ws, 121, col, f'={c}119*{c}120')

        # ---- Franchise Build ----

        # Row 124: Franchise Count = Total Network - Corp
        set_formula(ws, 124, col, f'={c}101-{c}117')

        # Row 125: New Franchise Openings (implied)
        set_formula(ws, 125, col, f'={c}124-{pcp}124')

        # Row 126: Franchise Royalty Rate - flatline from PCP
        if col == forecast_start:
            # PCP (2H25) is blank, use 1H26 value of 0.086
            set_value(ws, 126, col, 0.086, MAROON)
        else:
            set_formula(ws, 126, col, f'={pcp}126', MAROON)

        # Row 127: Franchise Royalty Revenue = Franchise Network Sales × Royalty Rate
        # Franchise Network Sales = Network Sales × (Franchise Count / Total Count)
        set_formula(ws, 127, col,
            f'={c}107*({c}124/{c}101)*{c}126')

        # Row 128: Other Franchise & Fee Revenue - PCP flatline
        set_formula(ws, 128, col, f'={pcp}128', MAROON)

        # Row 129: Total Franchise & Other Revenue
        set_formula(ws, 129, col, f'={c}127+{c}128')

        # ---- Segment EBITDA ----

        # Row 132: G&A as % of Network Sales - flatline from PCP
        if col == forecast_start:
            # PCP (2H25) blank, use 1H26 value
            set_value(ws, 132, col, 0.058, MAROON)
        else:
            set_formula(ws, 132, col, f'={pcp}132', MAROON)

        # Row 133: G&A Costs = -Network Sales × G&A%
        set_formula(ws, 133, col, f'=-{c}112*{c}132')

        # Row 134: AU Segment EBITDA = Corp Margin + Franchise Revenue + G&A
        set_formula(ws, 134, col, f'={c}121+{c}129+{c}133')

        # Row 135: AU Segment EBITDA % Network Sales
        set_formula(ws, 135, col, f'=IF({c}112=0,"",{c}134/{c}112)')

        # ============ US SEGMENT ============

        # Row 140: US Restaurant Count = PCP + new openings
        set_formula(ws, 140, col, f'={pcp}140+{c}141')

        # Row 141: US New Openings - assumption
        if col == forecast_start:
            set_value(ws, 141, col, 1, MAROON)  # ~2 per year, 1 per half
        else:
            set_formula(ws, 141, col, f'={pcp}141', MAROON)

        # Row 142: US Network Sales = PCP × (1 + growth)
        # US is small, use count-driven growth
        set_formula(ws, 142, col, f'={pcp}142*({c}140/{pcp}140)')

        # Row 143: US Corp Sales (all corp currently)
        set_formula(ws, 143, col, f'={c}142')

        # Row 144: US Corp Margin % - flatline from PCP (improving towards breakeven)
        if col == forecast_start:
            # PCP (2H25) is blank, use 1H26 value
            set_value(ws, 144, col, -0.50, MAROON)  # Improving from -70%
        else:
            # Gradually improve: reduce loss margin by 5pp per half
            set_formula(ws, 144, col, f'=MIN(0,{pcp}144+0.05)', MAROON)

        # Row 145: US Corp Margin ($)
        set_formula(ws, 145, col, f'={c}143*{c}144')

        # Row 146: US Franchise & Other Revenue - PCP flatline
        set_formula(ws, 146, col, f'={pcp}146', MAROON)

        # Row 147: US G&A Costs - PCP flatline
        set_formula(ws, 147, col, f'={pcp}147', MAROON)

        # Row 148: US Segment EBITDA = Corp Margin + Franchise Rev + G&A
        set_formula(ws, 148, col, f'={c}145+{c}146+{c}147')

    print("  Zone 2 complete.")


def build_hy_zone1_forecasts(ws):
    """Wire HY & Segments Zone 1 forecast formulas from column I to AA."""
    print("  Wiring HY Zone 1 forecasts...")

    forecast_start = 9
    forecast_end = 27

    for col in range(forecast_start, forecast_end + 1):
        c = get_col_letter(col)
        pcp = get_col_letter(col - 2)
        prior = get_col_letter(col - 1)

        # ---- Revenue (references Zone 2) ----
        # Row 7: AU Corp Sales = Zone 2 Corp Restaurant Sales
        set_formula(ws, 7, col, f'={c}119')

        # Row 8: US Corp Sales = Zone 2 US Corp Sales
        set_formula(ws, 8, col, f'={c}143')

        # Row 9: Franchise Royalty = Zone 2 Franchise Royalty
        set_formula(ws, 9, col, f'={c}127')

        # Row 10: Franchise Fee = PCP flatline
        set_formula(ws, 10, col, f'={pcp}10', MAROON)

        # Row 11: Total Revenue = SUM already has formula, but ensure it's set
        set_formula(ws, 11, col, f'=SUM({c}7:{c}10)')

        # ---- Other Revenue ----
        # Row 15: Marketing Levy = PCP flatline
        set_formula(ws, 15, col, f'={pcp}15', MAROON)

        # Row 16: Other Franchise Revenue = Zone 2 Other Franchise
        set_formula(ws, 16, col, f'={c}128')

        # Row 17: Other Income = PCP flatline
        set_formula(ws, 17, col, f'={pcp}17', MAROON)

        # Row 18: Total Other Revenue
        set_formula(ws, 18, col, f'=SUM({c}15:{c}17)')

        # ---- COGS ----
        # Row 21: Food & Packaging = PCP × (Revenue growth)
        # Better: maintain COGS/Revenue ratio
        # COGS/Rev in 1H26: -67.126 / (215.116+6.46+39.325+0.3) = -67.126/261.201 = -25.7%
        # Use: COGS = PCP COGS/Rev ratio × Revenue
        set_formula(ws, 21, col, f'=({pcp}21/({pcp}11+{pcp}18))*({c}11+{c}18)')

        # Row 22: Total COGS
        set_formula(ws, 22, col, f'=SUM({c}21:{c}21)')

        # ---- Gross Profit ----
        # Row 25: GP
        set_formula(ws, 25, col, f'={c}11+{c}18+{c}22')

        # ---- Operating Expenses ----
        # Each OpEx line: maintain ratio to revenue from PCP
        for opex_row in [30, 31, 32, 33]:
            set_formula(ws, opex_row, col,
                f'=({pcp}{opex_row}/({pcp}11+{pcp}18))*({c}11+{c}18)')

        # Row 34: Total OpEx
        set_formula(ws, 34, col, f'=SUM({c}30:{c}33)')

        # ---- Statutory EBITDA ----
        # Row 38: Statutory EBITDA
        set_formula(ws, 38, col, f'={c}25+{c}34')

        # ---- Segment EBITDA Bridge ----
        # Row 43: SBP = PCP flatline
        set_formula(ws, 43, col, f'={pcp}43', MAROON)

        # Row 44: Other Non-Recurring = 0 in forecast
        set_value(ws, 44, col, 0, MAROON)

        # Row 45: Cash Rent = PCP flatline (grows with restaurant count)
        set_formula(ws, 45, col, f'={pcp}45*({c}101/{pcp}101)', MAROON)

        # Row 46: AU Segment EBITDA = Zone 2
        set_formula(ws, 46, col, f'={c}134')

        # Row 47: US Segment EBITDA = Zone 2
        set_formula(ws, 47, col, f'={c}148')

        # Row 48: Group Segment EBITDA
        set_formula(ws, 48, col, f'={c}46+{c}47')

        # ---- D&A ----
        # Forecast D&A using ratio to revenue, flatlined from PCP
        # Total D&A = -(Rev + Other Rev) × D&A/Rev ratio
        # Split components proportionally from PCP

        # Row 56: Total D&A (compute first, then split)
        # Use PCP ratio: Total D&A / (Rev + Other Rev) from PCP
        set_formula(ws, 56, col,
            f'=({pcp}56/({pcp}11+{pcp}18))*({c}11+{c}18)')

        # Row 52: PPE Depn = proportion of total
        set_formula(ws, 52, col,
            f'=IF({pcp}56=0,0,{c}56*({pcp}52/{pcp}56))')

        # Row 53: ROU Amort = proportion of total
        set_formula(ws, 53, col,
            f'=IF({pcp}56=0,0,{c}56*({pcp}53/{pcp}56))')

        # Row 54: Reacq Amort = proportion of total
        set_formula(ws, 54, col,
            f'=IF({pcp}56=0,0,{c}56*({pcp}54/{pcp}56))')

        # Row 55: Other Amort = proportion of total
        set_formula(ws, 55, col,
            f'=IF({pcp}56=0,0,{c}56*({pcp}55/{pcp}56))')

        # ---- EBIT ----
        # Row 60: Underlying EBIT
        set_formula(ws, 60, col, f'={c}38+{c}56')

        # ---- Finance ----
        # Forecast interest using PCP flatline (simplified)
        # Row 65: Term Deposit Interest = PCP flatline
        set_formula(ws, 65, col, f'={pcp}65', MAROON)

        # Row 66: Lease Receivable Interest = PCP flatline
        set_formula(ws, 66, col, f'={pcp}66', MAROON)

        # Row 67: Other Finance Income = 0
        set_value(ws, 67, col, 0, MAROON)

        # Row 68: Lease Liability Interest = PCP flatline
        set_formula(ws, 68, col, f'={pcp}68', MAROON)

        # Row 69: Other Finance Costs = 0
        set_value(ws, 69, col, 0, MAROON)

        # Row 70: Net Finance Costs
        set_formula(ws, 70, col, f'=SUM({c}65:{c}69)')

        # ---- PBT ----
        # Row 76: PBT
        set_formula(ws, 76, col, f'={c}60+{c}70')

        # ---- Tax ----
        # Row 77: Tax = -PBT × Tax Rate
        # Use 30% tax rate (Australian statutory)
        set_formula(ws, 77, col, f'=-{c}76*0.30')

        # ---- NCI ----
        # Row 79: NCI = 0
        set_value(ws, 79, col, 0, MAROON)

        # ---- NPAT ----
        # Row 80: Underlying NPAT
        set_formula(ws, 80, col, f'={c}76+{c}77+{c}79')

        # Row 81: Sig Items AT = 0
        set_value(ws, 81, col, 0, MAROON)

        # Row 82: Statutory NPAT
        set_formula(ws, 82, col, f'={c}80+{c}81')

        # ---- EPS (HY) ----
        # Row 87: Basic EPS = PCP flatline (will be overwritten by Annual)
        # Actually HY EPS isn't critical since Annual computes from full year
        # Leave blank or use simple calc

        # Row 89: DPS = PCP flatline
        set_formula(ws, 89, col, f'={pcp}89', MAROON)

    print("  Zone 1 complete.")


def build_annual_forecasts(ws_annual, ws_hy):
    """Wire Annual sheet forecast formulas from column H (FY26E) to Q (FY35E)."""
    print("  Wiring Annual forecasts...")

    forecast_start = 8   # Column H = FY26E
    forecast_end = 17    # Column Q = FY35E

    # ---- INDEX/MATCH formula for flow items ----
    # Pattern: =INDEX('HY & Segments'!$A:$AA,MATCH($A{row},'HY & Segments'!$A:$A,0),
    #          MATCH("1H"&RIGHT({col}$1,2),'HY & Segments'!$3:$3,0))
    #         +INDEX('HY & Segments'!$A:$AA,MATCH($A{row},'HY & Segments'!$A:$A,0),
    #          MATCH("2H"&RIGHT({col}$1,2),'HY & Segments'!$3:$3,0))

    # Flow items that use INDEX/MATCH from HY
    flow_rows = [7, 8, 9, 10, 15, 16, 17, 21, 43, 44, 45, 46, 47,
                 52, 53, 54, 55, 65, 66, 67, 68, 69, 77, 79, 81]

    for col in range(forecast_start, forecast_end + 1):
        c = get_col_letter(col)
        prior_c = get_col_letter(col - 1)

        # ---- Flow items via INDEX/MATCH ----
        for row in flow_rows:
            key = ws_annual.cell(row=row, column=1).value
            if key:
                formula = (
                    f"=INDEX('HY & Segments'!$A:$AA,MATCH($A{row},'HY & Segments'!$A:$A,0),"
                    f"MATCH(\"1H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
                    f"+INDEX('HY & Segments'!$A:$AA,MATCH($A{row},'HY & Segments'!$A:$A,0),"
                    f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
                )
                set_formula(ws_annual, row, col, formula)

        # ---- Subtotal/formula rows (already exist but ensure they work) ----
        # Row 11: Total Revenue = SUM
        set_formula(ws_annual, 11, col, f'=SUM({c}7:{c}10)')
        # Row 12: Revenue Growth
        set_formula(ws_annual, 12, col, f'=IF({prior_c}11+{prior_c}18=0,"",({c}11+{c}18)/({prior_c}11+{prior_c}18)-1)')
        # Row 18: Total Other Revenue
        set_formula(ws_annual, 18, col, f'=SUM({c}15:{c}17)')
        # Row 22: Total COGS
        set_formula(ws_annual, 22, col, f'=SUM({c}21:{c}21)')
        # Row 25: GP
        set_formula(ws_annual, 25, col, f'={c}11+{c}18+{c}22')
        # Row 26: GP Growth
        set_formula(ws_annual, 26, col, f'=IF({prior_c}25=0,"",{c}25/{prior_c}25-1)')
        # Row 27: GP Margin
        set_formula(ws_annual, 27, col, f'=IF({c}11+{c}18=0,"",{c}25/({c}11+{c}18))')
        # Row 34: Total OpEx
        set_formula(ws_annual, 34, col, f'=SUM({c}30:{c}33)')

        # Rows 30-33: OpEx items via INDEX/MATCH
        for row in [30, 31, 32, 33]:
            key = ws_annual.cell(row=row, column=1).value
            if key:
                formula = (
                    f"=INDEX('HY & Segments'!$A:$AA,MATCH($A{row},'HY & Segments'!$A:$A,0),"
                    f"MATCH(\"1H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
                    f"+INDEX('HY & Segments'!$A:$AA,MATCH($A{row},'HY & Segments'!$A:$A,0),"
                    f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
                )
                set_formula(ws_annual, row, col, formula)

        # Row 35: OpEx Growth
        set_formula(ws_annual, 35, col, f'=IF({prior_c}34=0,"",{c}34/{prior_c}34-1)')
        # Row 38: Statutory EBITDA
        set_formula(ws_annual, 38, col, f'={c}25+{c}34')
        # Row 39: EBITDA Growth
        set_formula(ws_annual, 39, col, f'=IF({prior_c}38=0,"",{c}38/{prior_c}38-1)')
        # Row 40: EBITDA Margin
        set_formula(ws_annual, 40, col, f'=IF({c}11+{c}18=0,"",{c}38/({c}11+{c}18))')

        # Row 48: Group Segment EBITDA
        set_formula(ws_annual, 48, col, f'={c}46+{c}47')
        # Row 49: Segment EBITDA / Network Sales
        set_formula(ws_annual, 49, col, f'=IF({c}108=0,"",{c}48/{c}108)')

        # Row 56: Total D&A
        set_formula(ws_annual, 56, col, f'=SUM({c}52:{c}55)')
        # Row 57: D&A / Revenue
        set_formula(ws_annual, 57, col, f'=IF({c}11+{c}18=0,"",{c}56/({c}11+{c}18))')

        # Row 58: Avg Lease Life - assumption, flatline
        if col == forecast_start:
            # Derive from FY25 data: Lease Liab / Lease Principal = 331.3 / 10.8 ≈ 30.7 years
            # That's too high. Use ROU / ROU Depn = 125.4 / 13.9 ≈ 9 years
            set_value(ws_annual, 58, col, 9, MAROON)
        else:
            set_formula(ws_annual, 58, col, f'={prior_c}58', MAROON)

        # Row 60: EBIT
        set_formula(ws_annual, 60, col, f'={c}38+{c}56')
        # Row 61: EBIT Growth
        set_formula(ws_annual, 61, col, f'=IF({prior_c}60=0,"",{c}60/{prior_c}60-1)')
        # Row 62: EBIT Margin
        set_formula(ws_annual, 62, col, f'=IF({c}11+{c}18=0,"",{c}60/({c}11+{c}18))')

        # Row 70: Net Finance Costs
        set_formula(ws_annual, 70, col, f'=SUM({c}65:{c}69)')

        # Row 71-73: Interest rate assumptions - flatline
        if col == forecast_start:
            # Derive from FY25 data
            # Term Deposit Rate = 12.5 / ((278.1+242.1)/2) ≈ 4.8%
            set_value(ws_annual, 71, col, 0.048, MAROON)
            # Lease Rec Rate = 10 / ((126.4+174.8)/2) ≈ 6.6%
            set_value(ws_annual, 72, col, 0.066, MAROON)
            # Lease Liab Rate = 18.8 / ((239.5+331.3)/2) ≈ 6.6%
            set_value(ws_annual, 73, col, 0.066, MAROON)
        else:
            set_formula(ws_annual, 71, col, f'={prior_c}71', MAROON)
            set_formula(ws_annual, 72, col, f'={prior_c}72', MAROON)
            set_formula(ws_annual, 73, col, f'={prior_c}73', MAROON)

        # Row 76: PBT
        set_formula(ws_annual, 76, col, f'={c}60+{c}70')
        # Row 78: Tax Rate
        set_formula(ws_annual, 78, col, f'=IF({c}76=0,"",-{c}77/{c}76)')

        # Row 80: Underlying NPAT
        set_formula(ws_annual, 80, col, f'={c}76+{c}77+{c}79')
        # Row 82: Statutory NPAT
        set_formula(ws_annual, 82, col, f'={c}80+{c}81')
        # Row 83: NPAT Growth
        set_formula(ws_annual, 83, col, f'=IF({prior_c}80=0,"",{c}80/{prior_c}80-1)')
        # Row 84: NPAT Margin
        set_formula(ws_annual, 84, col, f'=IF({c}11+{c}18=0,"",{c}80/({c}11+{c}18))')

        # ---- EPS & Dividends ----
        # Row 87: YE Shares = prior + share issues effect (flatline for now)
        set_formula(ws_annual, 87, col, f'={prior_c}87', MAROON)
        # Row 88: WASO Basic ≈ average of prior and current YE shares
        set_formula(ws_annual, 88, col, f'=({prior_c}87+{c}87)/2')
        # Row 89: Dilution
        set_formula(ws_annual, 89, col, f'={c}90-{c}88')
        # Row 90: WASO Diluted - flatline ratio of diluted/basic from last actual
        set_formula(ws_annual, 90, col, f'={c}88*({prior_c}90/{prior_c}88)')

        # Row 92: Underlying EPS
        set_formula(ws_annual, 92, col, f'=IF({c}90=0,"",{c}80/{c}90)')
        # Row 93: Statutory EPS
        set_formula(ws_annual, 93, col, f'=IF({c}90=0,"",{c}82/{c}90)')
        # Row 94: EPS Growth
        set_formula(ws_annual, 94, col, f'=IF({prior_c}92=0,"",{c}92/{prior_c}92-1)')

        # Row 96: DPS - flatline from PCP
        set_formula(ws_annual, 96, col, f'={prior_c}96', MAROON)
        # Row 97: Total Dividends
        set_formula(ws_annual, 97, col, f'={c}96*{c}88')
        # Row 98: Payout Ratio
        set_formula(ws_annual, 98, col, f'=IF({c}92=0,"",{c}96/{c}92)')
        # Row 99: Dividend Yield
        set_formula(ws_annual, 99, col, f'=IF(Value!$C$4=0,"",{c}96/Value!$C$4)')
        # Row 100: Dividend Growth
        set_formula(ws_annual, 100, col, f'=IF({prior_c}96=0,"",{c}96/{prior_c}96-1)')

        # ---- Operating Metrics (KPIs) ----
        # These are point-in-time (year-end) — pull from HY 2H period
        # Network Sales: flow item, sum 1H + 2H
        for kpi_row in [104, 105, 106, 107]:
            key = ws_annual.cell(row=kpi_row, column=1).value
            hy_key_map = {
                'KPI-AU Network Sales': 'AU-Network Sales',
                'KPI-SG Network Sales': 'SG-Network Sales',
                'KPI-JP Network Sales': 'JP-Network Sales',
                'KPI-US Network Sales': 'US-Network Sales'
            }
            if key in hy_key_map:
                # Network sales are on HY Zone 2 without col A keys
                # Use direct row references based on known HY row numbers
                hy_row_map = {
                    'KPI-AU Network Sales': 107,
                    'KPI-SG Network Sales': 108,
                    'KPI-JP Network Sales': 110,
                    'KPI-US Network Sales': 142
                }
                hy_row = hy_row_map[key]
                # Sum 1H + 2H (use period label matching)
                formula = (
                    f"=INDEX('HY & Segments'!${hy_row}:${hy_row},1,"
                    f"MATCH(\"1H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
                    f"+INDEX('HY & Segments'!${hy_row}:${hy_row},1,"
                    f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
                )
                set_formula(ws_annual, kpi_row, col, formula)

        # Row 108: Global Network Sales
        set_formula(ws_annual, 108, col, f'=SUM({c}104:{c}107)')
        # Row 109: Network Sales Growth
        set_formula(ws_annual, 109, col, f'=IF({prior_c}108=0,"",{c}108/{prior_c}108-1)')

        # Restaurant counts (point-in-time = 2H value)
        count_map = {
            112: 117,  # AU Corp Count → HY Corp-Total
            116: 140,  # US Count → HY US-Count
        }
        for ann_row, hy_row in count_map.items():
            formula = (
                f"=INDEX('HY & Segments'!${hy_row}:${hy_row},1,"
                f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
            )
            set_formula(ws_annual, ann_row, col, formula)

        # Row 113: AU Franchise Count (2H period-end)
        formula = (
            f"=INDEX('HY & Segments'!$124:$124,1,"
            f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
        )
        set_formula(ws_annual, 113, col, formula)

        # Row 114: SG Restaurants - flatline
        set_formula(ws_annual, 114, col, f'={prior_c}114', MAROON)
        # Row 115: JP Restaurants - flatline
        set_formula(ws_annual, 115, col, f'={prior_c}115', MAROON)

        # Row 117: Total Restaurants
        set_formula(ws_annual, 117, col, f'=SUM({c}112:{c}116)')

        # Format detail (2H period-end)
        for ann_row, hy_row in [(120, 98), (121, 99), (122, 100)]:
            formula = (
                f"=INDEX('HY & Segments'!${hy_row}:${hy_row},1,"
                f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
            )
            set_formula(ws_annual, ann_row, col, formula)

        # AUVs (2H period-end, annualised)
        for ann_row, hy_row in [(123, 102), (124, 104), (125, 106)]:
            formula = (
                f"=INDEX('HY & Segments'!${hy_row}:${hy_row},1,"
                f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
            )
            set_formula(ws_annual, ann_row, col, formula)

        # Key Ratios
        # Row 128: Comp Sales Growth - proxy from AUV growth
        set_formula(ws_annual, 128, col, f'=IF({prior_c}104=0,"",{c}104/{prior_c}104-1-({c}120+{c}121+{c}122)/({prior_c}120+{prior_c}121+{prior_c}122)+1)')
        # Simplified: just use network sales growth less network growth
        # Actually just flatline the KPI from FY25
        if col == forecast_start:
            set_value(ws_annual, 128, col, 0.05, MAROON)  # 5% comp growth
        else:
            set_formula(ws_annual, 128, col, f'={prior_c}128', MAROON)

        # Row 129: Corp Restaurant Margin (from HY 2H)
        formula = (
            f"=INDEX('HY & Segments'!$120:$120,1,"
            f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
        )
        set_formula(ws_annual, 129, col, formula)

        # Row 130: Franchise Royalty Rate (from HY 2H)
        formula = (
            f"=INDEX('HY & Segments'!$126:$126,1,"
            f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
        )
        set_formula(ws_annual, 130, col, formula)

        # Row 131: G&A % Network Sales (from HY 2H)
        formula = (
            f"=INDEX('HY & Segments'!$132:$132,1,"
            f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
        )
        set_formula(ws_annual, 131, col, formula)

        # Row 132: Segment EBITDA % Network Sales
        set_formula(ws_annual, 132, col, f'=IF({c}108=0,"",{c}48/{c}108)')

        # ============ BALANCE SHEET ============

        # Row 136: Cash = Prior Cash + Net Change in Cash
        set_formula(ws_annual, 136, col, f'={prior_c}136+{c}207')

        # Row 137: Term Deposits - flatline (assumption about deployment)
        set_formula(ws_annual, 137, col, f'={prior_c}137', MAROON)

        # Row 138: Trade Receivables = (Rev + Other Rev) × Rec/Rev ratio
        set_formula(ws_annual, 138, col, f'=({c}11+{c}18)*{c}147')

        # Row 139: Inventories = (Rev + Other Rev) × Inv/Rev ratio
        set_formula(ws_annual, 139, col, f'=({c}11+{c}18)*{c}148')

        # Row 140: Finance Lease Receivables - grows with franchise network
        # Flatline for simplicity
        set_formula(ws_annual, 140, col, f'={prior_c}140', MAROON)

        # Row 141: PPE = Prior PPE - Capex PPE + PPE Depn
        # Capex PPE (191) is negative, PPE Depn (52) is negative
        # PPE = Prior + |Capex| - |Depn| = Prior - Capex + Depn
        # = Prior - (negative capex) + (negative depn)
        # Wait: if Capex = -61.3 and Depn = -19.3:
        # PPE = Prior - (-61.3) + (-19.3) = Prior + 61.3 - 19.3 ✓
        set_formula(ws_annual, 141, col, f'={prior_c}141-{c}191+{c}52')

        # Row 142: Intangibles - flatline
        set_formula(ws_annual, 142, col, f'={prior_c}142', MAROON)

        # Row 143: DTA - flatline
        set_formula(ws_annual, 143, col, f'={prior_c}143', MAROON)

        # Row 144: ROU Assets = Prior + New Lease Additions + ROU Depn (negative)
        set_formula(ws_annual, 144, col, f'={prior_c}144+{c}151+{c}53')

        # Row 145: Other Assets - flatline
        set_formula(ws_annual, 145, col, f'={prior_c}145', MAROON)

        # Row 146: Total Assets
        set_formula(ws_annual, 146, col, f'=SUM({c}136:{c}145)')

        # Row 147: Receivables/Revenue - flatline assumption
        set_formula(ws_annual, 147, col, f'={prior_c}147', MAROON)
        # Row 148: Inventory/Revenue - flatline assumption
        set_formula(ws_annual, 148, col, f'={prior_c}148', MAROON)
        # Row 149: Working Capital
        set_formula(ws_annual, 149, col, f'={c}138+{c}139-{c}154')
        # Row 150: Payables/Revenue - flatline assumption
        set_formula(ws_annual, 150, col, f'={prior_c}150', MAROON)

        # Row 151: New Lease Additions - flatline assumption
        if col == forecast_start:
            # Derive from FY25: ROU change + ROU depn
            # ROU FY25 = 125.4, FY24 = 93.8, ROU Depn FY25 = -13.9
            # NLA = 125.4 - 93.8 - (-13.9) = 45.5
            # Also Lease Liab: 331.3 - 239.5 - (-10.8) = 102.6 → includes unwinding
            # Use ROU-based: 45.5
            set_value(ws_annual, 151, col, 45.5, MAROON)
        else:
            set_formula(ws_annual, 151, col, f'={prior_c}151', MAROON)

        # ---- Liabilities ----
        # Row 154: Trade Payables = (Rev + Other Rev) × Payables/Rev ratio
        set_formula(ws_annual, 154, col, f'=({c}11+{c}18)*{c}150')

        # Row 155: Contract Liabilities - flatline
        set_formula(ws_annual, 155, col, f'={prior_c}155', MAROON)

        # Row 156: Other Liabilities - flatline
        set_formula(ws_annual, 156, col, f'={prior_c}156', MAROON)

        # Row 157: Lease Liabilities = Prior + New Lease Additions + Lease Principal (neg)
        set_formula(ws_annual, 157, col, f'={prior_c}157+{c}151+{c}202')

        # Row 158: Banking Debt - flatline (no debt)
        set_formula(ws_annual, 158, col, f'={prior_c}158', MAROON)

        # Row 159: Total Liabilities
        set_formula(ws_annual, 159, col, f'=SUM({c}154:{c}158)')

        # Row 161: Net Banking Debt
        set_formula(ws_annual, 161, col, f'={c}158-{c}136')
        # Row 162: Adj Net Debt
        set_formula(ws_annual, 162, col, f'={c}161+{c}157')
        # Row 163: ND/EBITDA
        set_formula(ws_annual, 163, col, f'=IF({c}38=0,"",{c}161/{c}38)')
        # Row 164: Gearing
        set_formula(ws_annual, 164, col, f'=IF(({c}161+{c}171)=0,"",{c}161/({c}161+{c}171))')

        # ---- Equity ----
        # Row 167: Issued Capital = Prior + Share Issues
        set_formula(ws_annual, 167, col, f'={prior_c}167+{c}201')

        # Row 168: Retained Profits = Prior + Statutory NPAT - Total Dividends
        set_formula(ws_annual, 168, col, f'={prior_c}168+{c}82-{c}97')

        # Row 169: Reserves - flatline
        set_formula(ws_annual, 169, col, f'={prior_c}169', MAROON)

        # Row 170: Minorities
        set_formula(ws_annual, 170, col, f'={prior_c}170-{c}79')

        # Row 171: Total Equity
        set_formula(ws_annual, 171, col, f'=SUM({c}167:{c}170)')

        # Row 172: ROE
        set_formula(ws_annual, 172, col, f'=IF({c}171=0,"",{c}80/{c}171)')
        # Row 173: P/B
        set_formula(ws_annual, 173, col, f'=IF(OR({c}171=0,Value!$C$4=0),"",Value!$C$4*{c}87/{c}171)')
        # Row 174: BS Check
        set_formula(ws_annual, 174, col, f'={c}146-{c}159-{c}171')

        # ============ CASH FLOW ============

        # Row 178: CF-EBITDA = Statutory EBITDA
        set_formula(ws_annual, 178, col, f'={c}38')

        # Row 179: WC Change = -(change in Rec + change in Inv) + change in Payables
        set_formula(ws_annual, 179, col,
            f'=-({c}138-{prior_c}138)-({c}139-{prior_c}139)+({c}154-{prior_c}154)')

        # Row 180: Significant Items / Non-Cash = -SBP - Other Non-Recurring (add back non-cash)
        set_formula(ws_annual, 180, col, f'=-{c}43-{c}44')

        # Row 181: Gross OCF
        set_formula(ws_annual, 181, col, f'=SUM({c}178:{c}180)')

        # Row 182: Interest Received = Term Deposit + Lease Rec + Other Income
        set_formula(ws_annual, 182, col, f'={c}65+{c}66+{c}67')

        # Row 183: Interest Paid = Other Finance Costs
        set_formula(ws_annual, 183, col, f'={c}69')

        # Row 184: Lease Interest Paid = Lease Liability Interest
        set_formula(ws_annual, 184, col, f'={c}68')

        # Row 185: Tax Paid ≈ Tax Expense (simplified)
        set_formula(ws_annual, 185, col, f'={c}77')

        # Row 186: Net OCF
        set_formula(ws_annual, 186, col, f'={c}181+{c}182+{c}183+{c}184+{c}185')

        # Row 187: OCF Growth
        set_formula(ws_annual, 187, col, f'=IF({prior_c}186=0,"",{c}186/{prior_c}186-1)')
        # Row 188: EBITDA Cash Conversion
        set_formula(ws_annual, 188, col, f'=IF({c}178=0,"",{c}181/{c}178)')

        # ---- CFI ----
        # Row 191: Capex PPE = Capex/Sales × Revenue
        set_formula(ws_annual, 191, col, f'={c}192*({c}11+{c}18)')

        # Row 192: Capex/Sales - flatline assumption
        if col == forecast_start:
            # FY25: -61.3 / (436 + 32) = -0.131
            set_value(ws_annual, 192, col, -0.131, MAROON)
        else:
            set_formula(ws_annual, 192, col, f'={prior_c}192', MAROON)

        # Row 193: Capex Intangibles = 0
        set_value(ws_annual, 193, col, 0, MAROON)
        # Row 194: Acquisitions = 0
        set_value(ws_annual, 194, col, 0, MAROON)
        # Row 195: Asset Sales = 0
        set_value(ws_annual, 195, col, 0, MAROON)
        # Row 196: Other CFI = change in term deposits (deploying = positive)
        set_formula(ws_annual, 196, col, f'={prior_c}137-{c}137')

        # Row 197: Total Investing CF
        set_formula(ws_annual, 197, col, f'=SUM({c}191,{c}193:{c}196)')

        # ---- CFF ----
        # Row 200: Dividends Paid = -Total Dividends
        set_formula(ws_annual, 200, col, f'=-{c}97')

        # Row 201: Share Issues = 0 (assumption)
        set_value(ws_annual, 201, col, 0, MAROON)

        # Row 202: Lease Principal = -Prior Lease Liabilities / Avg Lease Life
        set_formula(ws_annual, 202, col, f'=-{prior_c}157/{c}58')

        # Row 203: Change in Debt = 0
        set_value(ws_annual, 203, col, 0, MAROON)

        # Row 204: Other CFF = 0
        set_value(ws_annual, 204, col, 0, MAROON)

        # Row 205: Total Financing CF
        set_formula(ws_annual, 205, col, f'=SUM({c}200:{c}204)')

        # Row 207: Net Change in Cash
        set_formula(ws_annual, 207, col, f'={c}186+{c}197+{c}205')

        # ---- OFCF ----
        # Row 210: Net OCF
        set_formula(ws_annual, 210, col, f'={c}186')
        # Row 211: Net Capex
        set_formula(ws_annual, 211, col, f'={c}191+{c}193')
        # Row 212: Lease Principal
        set_formula(ws_annual, 212, col, f'={c}202')
        # Row 213: OFCF
        set_formula(ws_annual, 213, col, f'={c}210+{c}211+{c}212')
        # Row 214: FCF per Share
        set_formula(ws_annual, 214, col, f'=IF({c}90=0,"",{c}213/{c}90)')
        # Row 215: FCF Yield
        set_formula(ws_annual, 215, col, f'=IF(Value!$C$4=0,"",{c}214/Value!$C$4)')
        # Row 216: FCF Margin
        set_formula(ws_annual, 216, col, f'=IF({c}11+{c}18=0,"",{c}213/({c}11+{c}18))')

        # ---- ROIC ----
        # Row 220: Invested Capital
        set_formula(ws_annual, 220, col, f'={c}171+{c}161')
        # Row 221: Underlying EBIT
        set_formula(ws_annual, 221, col, f'={c}60')
        # Row 222: ROFE
        set_formula(ws_annual, 222, col, f'=IF({c}220=0,"",{c}221/{c}220)')
        # Row 223: NOPAT
        set_formula(ws_annual, 223, col, f'={c}221*(1-{c}78)')
        # Row 224: ROIC
        set_formula(ws_annual, 224, col, f'=IF({c}220=0,"",{c}223/{c}220)')

    # Now set the Receivables/Revenue ratio for FY25A if not already computed
    # (it's currently a formula but might return wrong value)
    # Also need to ensure FY25 ratios are properly computed for flatline to work
    # Row 147 already has formula =IF(G11+G18=0,"",G138/(G11+G18)) — good
    # Row 148 already has formula — good
    # Row 150 already has formula — good

    print("  Annual forecasts complete.")


def build_value_sheet(ws_value):
    """Update Value sheet for GYG."""
    print("  Updating Value sheet...")

    # Update currency labels from NZD to AUD
    ws_value.cell(row=4, column=2).value = "Current Share Price (AUD)"
    ws_value.cell(row=4, column=3).value = 35.00  # GYG recent share price ~$35
    ws_value.cell(row=4, column=3).font = copy_font_with_color(
        ws_value.cell(row=4, column=3), MAROON)

    ws_value.cell(row=6, column=2).value = "Market Cap (AUDm)"
    ws_value.cell(row=7, column=2).value = "Net Debt (AUDm)"
    ws_value.cell(row=8, column=2).value = "Market EV (AUDm)"

    # Row 9: Valuation Date
    from datetime import datetime
    ws_value.cell(row=9, column=3).value = datetime(2026, 3, 15)
    ws_value.cell(row=9, column=3).font = copy_font_with_color(
        ws_value.cell(row=9, column=3), 'FFFF0000')
    ws_value.cell(row=9, column=3).number_format = 'YYYY-MM-DD'

    # WACC Inputs
    ws_value.cell(row=12, column=3).value = 0.042   # Risk-free 4.2%
    ws_value.cell(row=12, column=3).font = copy_font_with_color(
        ws_value.cell(row=12, column=3), MAROON)

    ws_value.cell(row=13, column=3).value = 0.06    # ERP 6.0%
    ws_value.cell(row=13, column=3).font = copy_font_with_color(
        ws_value.cell(row=13, column=3), MAROON)

    ws_value.cell(row=14, column=3).value = 1.0     # Beta
    ws_value.cell(row=14, column=3).font = copy_font_with_color(
        ws_value.cell(row=14, column=3), MAROON)

    ws_value.cell(row=16, column=3).value = 0.0     # Pre-tax Cost of Debt (no debt)
    ws_value.cell(row=16, column=3).font = copy_font_with_color(
        ws_value.cell(row=16, column=3), MAROON)

    ws_value.cell(row=17, column=3).value = 0.30    # Tax Rate 30%
    ws_value.cell(row=17, column=3).font = copy_font_with_color(
        ws_value.cell(row=17, column=3), MAROON)

    ws_value.cell(row=19, column=3).value = 0.0     # Debt Weight 0%
    ws_value.cell(row=19, column=3).font = copy_font_with_color(
        ws_value.cell(row=19, column=3), MAROON)

    ws_value.cell(row=21, column=3).value = 0.025   # TGR 2.5%
    ws_value.cell(row=21, column=3).font = copy_font_with_color(
        ws_value.cell(row=21, column=3), MAROON)

    # Update currency references in output labels
    ws_value.cell(row=25, column=3).value = "AUDm"
    ws_value.cell(row=26, column=3).value = "AUDm"
    ws_value.cell(row=27, column=3).value = "AUDm"
    ws_value.cell(row=28, column=3).value = "AUDm"
    ws_value.cell(row=29, column=3).value = "AUDm"
    ws_value.cell(row=30, column=3).value = "AUDm"
    ws_value.cell(row=31, column=3).value = "AUDm"
    ws_value.cell(row=32, column=3).value = "AUDm"
    ws_value.cell(row=33, column=3).value = "AUDm"
    ws_value.cell(row=34, column=3).value = "AUDm"

    ws_value.cell(row=48, column=2).value = "Per Share Value (AUD)"
    ws_value.cell(row=65, column=2).value = "Per Share Value (AUD)"

    # ---- SOTP: Replace Steel/Metals with AU and US segments ----
    # Row 25 EBITDA reference uses "EBITDA-Underlying EBITDA" — this should stay using
    # the Group Segment EBITDA key. But actually it uses Stat EBITDA key. Let's check
    # and keep it as is since it already references the correct Annual key.

    # The DCF section formulas are already correct — they reference Annual rows by key
    # and those Annual rows will now have GYG data via INDEX/MATCH from HY.
    # The DCF EBITDA line uses "EBITDA-Underlying EBITDA" which doesn't exist.
    # GYG uses "EBITDA-Statutory EBITDA" for the statutory line and
    # "EBITDA-Group Segment EBITDA" for the group segment line.
    # The DCF should use Statutory EBITDA (row 38 on Annual).

    # Update DCF EBITDA reference
    for col in range(4, 14):  # D through M
        c = get_col_letter(col)
        # Row 25: EBITDA
        set_formula(ws_value, 25, col,
            f'=INDEX(Annual!$D:$Q,MATCH("EBITDA-Statutory EBITDA",Annual!$A:$A,0),MATCH({c}$24,Annual!$D$3:$Q$3,0))')
        # Row 26: D&A
        set_formula(ws_value, 26, col,
            f'=INDEX(Annual!$D:$Q,MATCH("DA-Total DA",Annual!$A:$A,0),MATCH({c}$24,Annual!$D$3:$Q$3,0))')
        # Row 27: EBIT
        set_formula(ws_value, 27, col,
            f'=INDEX(Annual!$D:$Q,MATCH("EBIT-Underlying EBIT",Annual!$A:$A,0),MATCH({c}$24,Annual!$D$3:$Q$3,0))')
        # Row 31: Capex
        set_formula(ws_value, 31, col,
            f'=INDEX(Annual!$D:$Q,MATCH("CF-Capex PPE",Annual!$A:$A,0),MATCH({c}$24,Annual!$D$3:$Q$3,0))+INDEX(Annual!$D:$Q,MATCH("CF-Capex Intang",Annual!$A:$A,0),MATCH({c}$24,Annual!$D$3:$Q$3,0))')
        # Row 32: WC Change
        set_formula(ws_value, 32, col,
            f'=INDEX(Annual!$D:$Q,MATCH("CF-WC Change",Annual!$A:$A,0),MATCH({c}$24,Annual!$D$3:$Q$3,0))')

    # Terminal value formulas (last forecast column = M, col 13)
    # Row 34: Normalised FCFF (last col only)
    set_formula(ws_value, 34, 13, '=M29+M32')
    # Row 35: Terminal Value
    set_formula(ws_value, 35, 13, '=M34*(1+$C$21)/($C$20-$C$21)')
    # PV of Terminal Value
    set_formula(ws_value, 39, 13, '=M35*M37')

    # Discount factors (10 years, columns D=4 through M=13)
    for i, col in enumerate(range(4, 14)):
        c = get_col_letter(col)
        set_formula(ws_value, 37, col, f'=1/(1+$C$20)^($C$22+{i})')
        set_formula(ws_value, 38, col, f'={c}33*{c}37')

    # Update equity bridge references to use $Q (max col) instead of $P
    set_formula(ws_value, 5, 3,
        '=INDEX(Annual!$D:$Q,MATCH("EPS-YE Shares",Annual!$A:$A,0),MATCH($D$24,Annual!$D$3:$Q$3,0)-1)')
    set_formula(ws_value, 7, 3,
        '=INDEX(Annual!$D:$Q,MATCH("BS-Total Banking Debt",Annual!$A:$A,0),MATCH($D$24,Annual!$D$3:$Q$3,0)-1)-INDEX(Annual!$D:$Q,MATCH("BS-Cash",Annual!$A:$A,0),MATCH($D$24,Annual!$D$3:$Q$3,0)-1)')
    set_formula(ws_value, 46, 3,
        '=-INDEX(Annual!$D:$Q,MATCH("BS-Lease Liabilities",Annual!$A:$A,0),MATCH($D$24,Annual!$D$3:$Q$3,0)-1)')

    # Stub period
    set_formula(ws_value, 22, 3,
        '=(INDEX(Annual!$D:$Q,4,MATCH($D$24,Annual!$D$3:$Q$3,0))-C9)/365.25')

    # ---- SOTP Section ----
    # Row 57: Replace "Steel" with "AU Segment"
    ws_value.cell(row=57, column=2).value = "AU Segment"
    set_formula(ws_value, 57, 3,
        '=INDEX(Annual!$D:$Q,MATCH("EBITDA-AU Segment EBITDA",Annual!$A:$A,0),MATCH($C$54,Annual!$D$3:$Q$3,0))')
    ws_value.cell(row=57, column=4).value = 25  # AU segment multiple
    ws_value.cell(row=57, column=4).font = copy_font_with_color(
        ws_value.cell(row=57, column=4), MAROON)

    # Row 58: Replace "Metals" with "US Segment"
    ws_value.cell(row=58, column=2).value = "US Segment"
    set_formula(ws_value, 58, 3,
        '=INDEX(Annual!$D:$Q,MATCH("EBITDA-US Segment EBITDA",Annual!$A:$A,0),MATCH($C$54,Annual!$D$3:$Q$3,0))')
    ws_value.cell(row=58, column=4).value = 0  # US segment at 0x (loss-making)
    ws_value.cell(row=58, column=4).font = copy_font_with_color(
        ws_value.cell(row=58, column=4), MAROON)

    # Row 59: Corporate overhead
    ws_value.cell(row=59, column=2).value = "Corporate"
    # Corporate EBITDA = Statutory EBITDA - AU Seg EBITDA - US Seg EBITDA
    set_formula(ws_value, 59, 3,
        '=INDEX(Annual!$D:$Q,MATCH("EBITDA-Statutory EBITDA",Annual!$A:$A,0),MATCH($C$54,Annual!$D$3:$Q$3,0))-C57-C58')
    # Corporate multiple = weighted average of segment multiples
    set_formula(ws_value, 59, 4,
        '=IF((C57+C58)=0,"",(C57*D57+C58*D58)/(C57+C58))')

    # Implied Group EV/EBITDA
    set_formula(ws_value, 68, 5,
        '=IF(INDEX(Annual!$D:$Q,MATCH("EBITDA-Statutory EBITDA",Annual!$A:$A,0),MATCH($C$54,Annual!$D$3:$Q$3,0))=0,"",E61/INDEX(Annual!$D:$Q,MATCH("EBITDA-Statutory EBITDA",Annual!$A:$A,0),MATCH($C$54,Annual!$D$3:$Q$3,0)))')

    # Update lease liabilities reference in SOTP
    set_formula(ws_value, 63, 5,
        '=-INDEX(Annual!$D:$Q,MATCH("BS-Lease Liabilities",Annual!$A:$A,0),MATCH($D$24,Annual!$D$3:$Q$3,0)-1)')

    print("  Value sheet complete.")


def verify_model(wb):
    """Run verification checks on the model."""
    print("\n=== VERIFICATION ===")

    ws_annual = wb['Annual']
    ws_hy = wb['HY & Segments']
    ws_value = wb['Value']

    issues = []

    # Check 1: All forecast columns have formulas (not None)
    print("\n  Check 1: Forecast columns populated...")
    empty_cells = 0
    # Annual H-Q (cols 8-17), key rows
    key_rows = [7,8,9,10,11,15,16,17,18,21,22,25,30,31,32,33,34,38,
                43,44,45,46,47,48,52,53,54,55,56,60,65,66,67,68,69,70,
                76,77,79,80,81,82,87,88,90,92,93,96,97,
                136,137,138,139,140,141,142,143,144,145,146,
                154,155,156,157,158,159,167,168,169,170,171,
                178,179,180,181,182,183,184,185,186,191,192,193,
                200,201,202,203,204,205,207,210,211,212,213]
    for col in range(8, 18):
        for row in key_rows:
            val = ws_annual.cell(row=row, column=col).value
            if val is None:
                empty_cells += 1
                if empty_cells <= 10:
                    issues.append(f"  Annual row {row}, col {get_col_letter(col)}: EMPTY")
    print(f"    Annual empty forecast cells: {empty_cells}")

    # HY I-AA (cols 9-27)
    hy_empty = 0
    hy_key_rows = [7,8,9,10,11,15,16,17,18,21,22,25,30,31,32,33,34,38,
                   43,44,45,46,47,48,52,53,54,55,56,60,65,66,67,68,69,70,
                   76,77,79,80,81,82,
                   98,99,100,101,102,103,104,105,107,108,109,110,111,112,
                   117,118,119,120,121,124,126,127,128,129,
                   132,133,134,135,140,141,142,143,144,145,146,147,148]
    for col in range(9, 28):
        for row in hy_key_rows:
            val = ws_hy.cell(row=row, column=col).value
            if val is None:
                hy_empty += 1
                if hy_empty <= 10:
                    issues.append(f"  HY row {row}, col {get_col_letter(col)}: EMPTY")
    print(f"    HY empty forecast cells: {hy_empty}")

    # Check 2: Column A keys exist for INDEX/MATCH
    print("\n  Check 2: Column A key integrity...")
    needed_keys = [
        'Rev-AU Corp Sales', 'Rev-US Corp Sales', 'Rev-Franchise Royalty', 'Rev-Franchise Fee',
        'Rev-Marketing Levy', 'Rev-Other Franchise', 'Rev-Other Income',
        'COGS-Food Packaging',
        'OPEX-Employee Benefits', 'OPEX-Admin', 'OPEX-Marketing', 'OPEX-Other',
        'Stat-SBP', 'Stat-Other Costs', 'Stat-Cash Rent',
        'EBITDA-AU Segment EBITDA', 'EBITDA-US Segment EBITDA',
        'DA-Depreciation PPE', 'DA-ROU Amortisation', 'DA-Reacq Amort', 'DA-Other Amort', 'DA-Total DA',
        'EBIT-Underlying EBIT',
        'Int-Term Deposit', 'Int-Lease Receivable', 'Int-Other Income', 'Int-Lease Interest', 'Int-Other Costs',
        'Tax-Tax Expense', 'NPAT-NCI', 'NPAT-Sig Items AT',
        'EBITDA-Statutory EBITDA',
        'BS-Cash', 'BS-Total Banking Debt', 'BS-Lease Liabilities',
        'CF-Capex PPE', 'CF-Capex Intang', 'CF-WC Change',
        'EPS-YE Shares'
    ]

    # Get all keys from Annual col A
    annual_keys = set()
    for row in range(1, ws_annual.max_row + 1):
        key = ws_annual.cell(row=row, column=1).value
        if key:
            annual_keys.add(key)

    # Get all keys from HY col A
    hy_keys = set()
    for row in range(1, ws_hy.max_row + 1):
        key = ws_hy.cell(row=row, column=1).value
        if key:
            hy_keys.add(key)

    for key in needed_keys:
        if key not in annual_keys:
            issues.append(f"  Missing key on Annual: {key}")
        # Flow item keys should also exist on HY (except BS/CF)
        if not key.startswith(('BS-', 'CF-', 'EPS-')):
            if key not in hy_keys:
                issues.append(f"  Missing key on HY: {key}")

    # Check 3: EBITDA-Underlying EBITDA key doesn't exist (it's Statutory EBITDA)
    if 'EBITDA-Underlying EBITDA' in annual_keys:
        print("    Note: 'EBITDA-Underlying EBITDA' key exists")
    else:
        print("    Note: 'EBITDA-Underlying EBITDA' does NOT exist — Value DCF uses Statutory EBITDA (correct)")

    key_issues = [i for i in issues if 'Missing key' in i]
    print(f"    Missing keys: {len(key_issues)}")

    # Check 4: Value sheet references
    print("\n  Check 3: Value sheet formula references...")
    val_formulas_ok = True
    for row in [25, 26, 27, 31, 32]:
        for col in [4, 13]:  # First and last forecast
            val = ws_value.cell(row=row, column=col).value
            if val is None:
                val_formulas_ok = False
                issues.append(f"  Value row {row}, col {get_col_letter(col)}: EMPTY")
    print(f"    Value FCF formulas: {'OK' if val_formulas_ok else 'ISSUES FOUND'}")

    # Print all issues
    if issues:
        print(f"\n  Total issues: {len(issues)}")
        for issue in issues[:20]:
            print(f"    {issue}")
        if len(issues) > 20:
            print(f"    ... and {len(issues)-20} more")
    else:
        print("\n  All checks passed!")

    return len(issues)


def main():
    print(f"Loading workbook: {MODEL_PATH}")
    wb = openpyxl.load_workbook(MODEL_PATH)

    ws_hy = wb['HY & Segments']
    ws_annual = wb['Annual']
    ws_value = wb['Value']

    print("\nPhase 1: HY & Segments Zone 2 (forecast drivers)")
    build_hy_zone2_forecasts(ws_hy)

    print("\nPhase 2: HY & Segments Zone 1 (P&L forecasts)")
    build_hy_zone1_forecasts(ws_hy)

    print("\nPhase 3: Annual (INDEX/MATCH, BS roll-forwards, CF)")
    build_annual_forecasts(ws_annual, ws_hy)

    print("\nPhase 4: Value sheet")
    build_value_sheet(ws_value)

    print("\nPhase 5: Verification")
    issue_count = verify_model(wb)

    print(f"\nSaving workbook...")
    wb.save(MODEL_PATH)
    print(f"Saved to: {MODEL_PATH}")

    if issue_count == 0:
        print("\nBuild complete — all checks passed.")
    else:
        print(f"\nBuild complete — {issue_count} issues found. Review needed.")

    return issue_count


if __name__ == '__main__':
    sys.exit(main())

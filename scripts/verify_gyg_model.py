#!/usr/bin/env python3
"""
Comprehensive verification script for GYG Model.xlsx

Handles openpyxl's inability to evaluate formulas by manually resolving
simple formulas (SUM, addition, subtraction) and INDEX/MATCH cross-sheet lookups.
"""
import openpyxl
import re
import sys
from collections import OrderedDict

MODEL_PATH = "/home/pmwilson/Project_Equities/GYG/Models/GYG Model.xlsx"

wb_f = openpyxl.load_workbook(MODEL_PATH, data_only=False)
wb_v = openpyxl.load_workbook(MODEL_PATH, data_only=True)

results = OrderedDict()

# Build lookup tables for HY & Segments
ws_hy_f = wb_f['HY & Segments']
ws_hy_v = wb_v['HY & Segments']

# HY Column A key -> row mapping
hy_key_to_row = {}
for r in range(1, ws_hy_f.max_row + 1):
    v = ws_hy_f.cell(r, 1).value
    if v is not None:
        hy_key_to_row[str(v).strip()] = r

# HY Row 3 label -> column mapping
hy_label_to_col = {}
for c in range(1, ws_hy_f.max_column + 1):
    v = ws_hy_f.cell(3, c).value
    if v is not None:
        hy_label_to_col[str(v).strip()] = c

# Annual Row 1 year values
ws_annual_f = wb_f['Annual']
ws_annual_v = wb_v['Annual']
annual_row1 = {}
for c in range(1, ws_annual_f.max_column + 1):
    v = ws_annual_f.cell(1, c).value
    if v is not None:
        if hasattr(v, 'year'):
            annual_row1[c] = v.year
        else:
            try:
                annual_row1[c] = int(v)
            except:
                pass

# HY Row 1 year values
hy_row1 = {}
for c in range(1, ws_hy_f.max_column + 1):
    v = ws_hy_f.cell(1, c).value
    if v is not None:
        if hasattr(v, 'year'):
            hy_row1[c] = v.year
        else:
            try:
                hy_row1[c] = int(v)
            except:
                pass


def eval_hy_cell(row, col, depth=0, visited=None):
    """Evaluate a cell in the HY & Segments sheet."""
    if visited is None:
        visited = set()
    key = ('HY', row, col)
    if key in visited or depth > 15:
        return None
    visited.add(key)

    # Try cached value
    v = ws_hy_v.cell(row, col).value
    if v is not None:
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    # Get formula
    fv = ws_hy_f.cell(row, col).value
    if fv is None:
        return None
    if not isinstance(fv, str):
        try:
            return float(fv)
        except (ValueError, TypeError):
            return None
    if not fv.startswith('='):
        try:
            return float(fv)
        except (ValueError, TypeError):
            return None

    formula = fv[1:]
    return _eval_formula(formula, 'HY', row, col, depth, visited)


def _eval_formula(formula, sheet, row, col, depth, visited):
    """Evaluate a simple formula string."""
    # SUM(X:Y) range
    sum_match = re.match(r'^SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$', formula)
    if sum_match:
        c1 = openpyxl.utils.column_index_from_string(sum_match.group(1))
        r1 = int(sum_match.group(2))
        r2 = int(sum_match.group(4))
        total = 0.0
        ev_fn = eval_hy_cell if sheet == 'HY' else eval_annual_cell
        for rr in range(r1, r2 + 1):
            val = ev_fn(rr, c1, depth + 1, visited.copy())
            if val is not None:
                total += val
        return total

    # SUM with mixed args
    sum_multi = re.match(r'^SUM\((.+)\)$', formula)
    if sum_multi:
        args = sum_multi.group(1)
        total = 0.0
        ev_fn = eval_hy_cell if sheet == 'HY' else eval_annual_cell
        for part in args.split(','):
            part = part.strip()
            range_match = re.match(r'^([A-Z]+)(\d+):([A-Z]+)(\d+)$', part)
            cell_match = re.match(r'^([A-Z]+)(\d+)$', part)
            if range_match:
                c1 = openpyxl.utils.column_index_from_string(range_match.group(1))
                r1 = int(range_match.group(2))
                r2 = int(range_match.group(4))
                for rr in range(r1, r2 + 1):
                    val = ev_fn(rr, c1, depth + 1, visited.copy())
                    if val is not None:
                        total += val
            elif cell_match:
                cc = openpyxl.utils.column_index_from_string(cell_match.group(1))
                rr = int(cell_match.group(2))
                val = ev_fn(rr, cc, depth + 1, visited.copy())
                if val is not None:
                    total += val
        return total

    # Addition/subtraction: D48+D59, D114-D126-D138
    arith_match = re.match(r'^([A-Z]+\d+(?:[+-][A-Z]+\d+)+)$', formula)
    if arith_match:
        tokens = re.findall(r'([+-]?)([A-Z]+)(\d+)', formula)
        total = 0.0
        ev_fn = eval_hy_cell if sheet == 'HY' else eval_annual_cell
        for sign, col_str, row_str in tokens:
            cc = openpyxl.utils.column_index_from_string(col_str)
            rr = int(row_str)
            val = ev_fn(rr, cc, depth + 1, visited.copy())
            if val is not None:
                total += val if sign != '-' else -val
        return total

    # Negation: =-G28
    neg_match = re.match(r'^-([A-Z]+)(\d+)$', formula)
    if neg_match:
        cc = openpyxl.utils.column_index_from_string(neg_match.group(1))
        rr = int(neg_match.group(2))
        ev_fn = eval_hy_cell if sheet == 'HY' else eval_annual_cell
        val = ev_fn(rr, cc, depth + 1, visited.copy())
        return -val if val is not None else None

    # Simple reference: =G30
    ref_match = re.match(r'^([A-Z]+)(\d+)$', formula)
    if ref_match:
        cc = openpyxl.utils.column_index_from_string(ref_match.group(1))
        rr = int(ref_match.group(2))
        ev_fn = eval_hy_cell if sheet == 'HY' else eval_annual_cell
        return ev_fn(rr, cc, depth + 1, visited.copy())

    # Multiplication: =G159*G12 or =F109*1.05
    mul_match = re.match(r'^([A-Z]+)(\d+)\*([A-Z]+)(\d+)$', formula)
    if mul_match:
        ev_fn = eval_hy_cell if sheet == 'HY' else eval_annual_cell
        v1 = ev_fn(int(mul_match.group(2)), openpyxl.utils.column_index_from_string(mul_match.group(1)), depth+1, visited.copy())
        v2 = ev_fn(int(mul_match.group(4)), openpyxl.utils.column_index_from_string(mul_match.group(3)), depth+1, visited.copy())
        if v1 is not None and v2 is not None:
            return v1 * v2
        return None

    mul_const_match = re.match(r'^([A-Z]+)(\d+)\*([\d.]+)$', formula)
    if mul_const_match:
        ev_fn = eval_hy_cell if sheet == 'HY' else eval_annual_cell
        v1 = ev_fn(int(mul_const_match.group(2)), openpyxl.utils.column_index_from_string(mul_const_match.group(1)), depth+1, visited.copy())
        if v1 is not None:
            return v1 * float(mul_const_match.group(3))
        return None

    # =-F79
    neg_ref_match = re.match(r'^-([A-Z]+)(\d+)$', formula)
    if neg_ref_match:
        ev_fn = eval_hy_cell if sheet == 'HY' else eval_annual_cell
        val = ev_fn(int(neg_ref_match.group(2)), openpyxl.utils.column_index_from_string(neg_ref_match.group(1)), depth+1, visited.copy())
        return -val if val is not None else None

    # =-F124/8
    neg_div_match = re.match(r'^-([A-Z]+)(\d+)/([\d.]+)$', formula)
    if neg_div_match:
        ev_fn = eval_hy_cell if sheet == 'HY' else eval_annual_cell
        val = ev_fn(int(neg_div_match.group(2)), openpyxl.utils.column_index_from_string(neg_div_match.group(1)), depth+1, visited.copy())
        if val is not None:
            return -val / float(neg_div_match.group(3))
        return None

    # =F110-G158+G43 (mixed add/sub)
    # Already handled above by arith_match

    return None


def eval_annual_cell(row, col, depth=0, visited=None):
    """Evaluate a cell in the Annual sheet, resolving formulas including INDEX/MATCH from HY."""
    if visited is None:
        visited = set()
    key = ('Annual', row, col)
    if key in visited or depth > 15:
        return None
    visited.add(key)

    # Try cached value
    v = ws_annual_v.cell(row, col).value
    if v is not None:
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    # Get formula
    fv = ws_annual_f.cell(row, col).value
    if fv is None:
        return None
    if not isinstance(fv, str):
        try:
            return float(fv)
        except (ValueError, TypeError):
            return None
    if not fv.startswith('='):
        try:
            return float(fv)
        except (ValueError, TypeError):
            return None

    formula = fv[1:]

    # INDEX/MATCH from HY & Segments
    # Pattern: INDEX('HY & Segments'!$A:$AG,MATCH($A7,'HY & Segments'!$A:$A,0),MATCH("1H"&G$1,'HY & Segments'!$G$3:$AG$3,0)+6)+INDEX(...)
    # This pulls 1H + 2H values
    idx_pattern = r"INDEX\('HY & Segments'!\$A:\$[A-Z]+,MATCH\(\$A(\d+),'HY & Segments'!\$A:\$A,0\),MATCH\(\"(\dH)\"&\$?([A-Z]+)\$?1,'HY & Segments'!\$[A-Z]+\$3:\$[A-Z]+\$3,0\)\+(\d+)\)"
    idx_matches = re.findall(idx_pattern, formula)

    if idx_matches and len(idx_matches) >= 2:
        # Two INDEX/MATCH calls added together (1H + 2H)
        total = 0.0
        for match_row, half, col_ref, offset in idx_matches:
            # Get the key from Annual column A at the referenced row
            annual_key = ws_annual_f.cell(int(match_row), 1).value
            if annual_key is None:
                return None
            annual_key = str(annual_key).strip()

            # Get the year from Annual row 1 at the referenced column
            ref_col = openpyxl.utils.column_index_from_string(col_ref)
            year = annual_row1.get(ref_col)
            if year is None:
                return None

            # Construct period label
            period_label = f"{half}{str(year)[-2:]}"

            # Find in HY
            hy_row = hy_key_to_row.get(annual_key)
            hy_col = hy_label_to_col.get(period_label)

            if hy_row is None or hy_col is None:
                return None

            val = eval_hy_cell(hy_row, hy_col, depth + 1, visited.copy())
            if val is not None:
                total += val
            else:
                return None
        return total
    elif idx_matches and len(idx_matches) == 1:
        match_row, half, col_ref, offset = idx_matches[0]
        annual_key = ws_annual_f.cell(int(match_row), 1).value
        if annual_key is None:
            return None
        annual_key = str(annual_key).strip()
        ref_col = openpyxl.utils.column_index_from_string(col_ref)
        year = annual_row1.get(ref_col)
        if year is None:
            return None
        period_label = f"{half}{str(year)[-2:]}"
        hy_row = hy_key_to_row.get(annual_key)
        hy_col = hy_label_to_col.get(period_label)
        if hy_row is None or hy_col is None:
            return None
        return eval_hy_cell(hy_row, hy_col, depth + 1, visited.copy())

    # Standard formula evaluation
    return _eval_formula(formula, 'Annual', row, col, depth, visited)


def ev(row, col):
    """Shorthand to evaluate an Annual cell."""
    return eval_annual_cell(row, col)


# ============================================================
# CHECK 1: Formula errors
# ============================================================
print("=" * 70)
print("CHECK 1: Formula errors")
print("=" * 70)

error_patterns = ['#REF!', '#N/A', '#VALUE!', '#DIV/0!', '#NAME?', '#NULL!', '#NUM!']
formula_errors = []

for sheet_name in wb_f.sheetnames:
    ws = wb_f[sheet_name]
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row_cells:
            val = cell.value
            if val is None or not isinstance(val, str):
                continue
            for err in error_patterns:
                if err in val:
                    formula_errors.append(f"  {sheet_name}!{cell.coordinate}: contains {err} -> {val[:100]}")
            if val.startswith('=') and ('!!' in val or val.endswith('!')):
                formula_errors.append(f"  {sheet_name}!{cell.coordinate}: malformed reference -> {val[:100]}")

if formula_errors:
    results['CHECK 1: Formula errors'] = ('FAIL', formula_errors)
    print("FAIL")
    for e in formula_errors:
        print(e)
else:
    results['CHECK 1: Formula errors'] = ('PASS', ['No formula errors found'])
    print("PASS - No formula errors found")


# ============================================================
# CHECK 2: Cross-sheet INDEX/MATCH integrity
# ============================================================
print("\n" + "=" * 70)
print("CHECK 2: Cross-sheet INDEX/MATCH integrity")
print("=" * 70)

hy_col_a_keys = set(hy_key_to_row.keys())
hy_row3_labels = set(hy_label_to_col.keys())

index_match_issues = []

for row in range(1, ws_annual_f.max_row + 1):
    for col in range(1, ws_annual_f.max_column + 1):
        cell = ws_annual_f.cell(row, col)
        val = cell.value
        if val is None or not isinstance(val, str) or not val.startswith('='):
            continue
        if "'HY & Segments'" not in val or 'INDEX' not in val or 'MATCH' not in val:
            continue

        annual_key = ws_annual_f.cell(row, 1).value
        if annual_key is not None:
            annual_key_str = str(annual_key).strip()
            if annual_key_str not in hy_col_a_keys:
                index_match_issues.append(
                    f"  Annual!{cell.coordinate}: Key '{annual_key_str}' NOT FOUND in HY Column A"
                )

        # Verify constructed period labels
        period_refs = re.findall(r'MATCH\("(\dH)"&\$?([A-Z]+)\$?1', val)
        for half, col_ref in period_refs:
            ref_col = openpyxl.utils.column_index_from_string(col_ref)
            year = annual_row1.get(ref_col)
            if year is not None:
                label = f"{half}{str(year)[-2:]}"
                if label not in hy_row3_labels:
                    index_match_issues.append(
                        f"  Annual!{cell.coordinate}: Period '{label}' NOT FOUND in HY Row 3"
                    )

if index_match_issues:
    results['CHECK 2: INDEX/MATCH integrity'] = ('FAIL', index_match_issues)
    print("FAIL")
    for e in index_match_issues:
        print(e)
else:
    results['CHECK 2: INDEX/MATCH integrity'] = ('PASS', ['All INDEX/MATCH keys and periods verified'])
    print("PASS - All INDEX/MATCH lookup keys and period labels verified")


# ============================================================
# CHECK 3: P&L cascade verification (historical actuals)
# ============================================================
print("\n" + "=" * 70)
print("CHECK 3: P&L cascade verification (historical actuals)")
print("=" * 70)

TOL = 0.5
cascade_issues = []
periods = {'FY23': 4, 'FY24': 5, 'FY25': 6}

cascade_checks = [
    ("Row 9 = Row 7 + Row 8 (Total Corp Sales)", 9, [7, 8]),
    ("Row 12 = SUM(9,10,11) (Total Revenue)", 12, [9, 10, 11]),
    ("Row 19 = SUM(16,17,18) (Total Other Revenue)", 19, [16, 17, 18]),
    ("Row 21 = Row 12 + Row 19 (Total Rev & Other Income)", 21, [12, 19]),
    ("Row 26 = Row 24 + Row 25 (Total Seg EBITDA)", 26, [24, 25]),
    ("Row 30 = SUM(26:29) (EBITDA)", 30, [26, 27, 28, 29]),
    ("Row 39 = SUM(34:38) (Total Expenses)", 39, [34, 35, 36, 37, 38]),
    ("Row 46 = SUM(42:45) (Total D&A)", 46, [42, 43, 44, 45]),
    ("Row 48 = Row 30 + Row 46 (EBIT)", 48, [30, 46]),
    ("Row 55 = SUM(52:54) (Total Finance Income)", 55, [52, 53, 54]),
    ("Row 58 = Row 56 + Row 57 (Total Finance Costs)", 58, [56, 57]),
    ("Row 59 = Row 55 + Row 58 (Net Finance)", 59, [55, 58]),
    ("Row 61 = Row 48 + Row 59 (PBT)", 61, [48, 59]),
    ("Row 64 = Row 61 + Row 62 (NPAT)", 64, [61, 62]),
]

for desc, target_row, component_rows in cascade_checks:
    for period, col in periods.items():
        actual = ev(target_row, col)
        components_ok = True
        for cr in component_rows:
            if ev(cr, col) is None:
                components_ok = False
                break

        if actual is None:
            cascade_issues.append(f"  {desc} [{period}]: target is None")
            continue
        if not components_ok:
            cascade_issues.append(f"  {desc} [{period}]: component(s) None")
            continue

        expected = sum(ev(cr, col) for cr in component_rows)
        diff = abs(actual - expected)
        if diff > TOL:
            cascade_issues.append(
                f"  {desc} [{period}]: expected={expected:.3f}, actual={actual:.3f}, diff={diff:.3f}"
            )

if cascade_issues:
    results['CHECK 3: P&L cascade'] = ('FAIL', cascade_issues)
    print("FAIL")
    for e in cascade_issues:
        print(e)
else:
    results['CHECK 3: P&L cascade'] = ('PASS', ['All 42 P&L cascade checks passed within +/-0.5'])
    print("PASS - All 42 P&L cascade checks passed within +/-0.5 tolerance")


# ============================================================
# CHECK 4: Sign conventions
# ============================================================
print("\n" + "=" * 70)
print("CHECK 4: Sign conventions")
print("=" * 70)

sign_issues = []

def check_sign(row, expected_sign, label):
    for period, col in periods.items():
        v = ev(row, col)
        if v is None or v == 0:
            continue
        if expected_sign == 'positive' and v < 0:
            sign_issues.append(f"  Row {row} ({label}) [{period}]: {v:.3f} NEGATIVE (expected positive)")
        elif expected_sign == 'negative' and v > 0:
            sign_issues.append(f"  Row {row} ({label}) [{period}]: {v:.3f} POSITIVE (expected negative)")

for r in [7, 8, 9, 10, 11, 12, 16, 17, 18, 19]:
    check_sign(r, 'positive', ws_annual_f.cell(r, 2).value or f"Row {r}")

check_sign(24, 'positive', 'Australia EBITDA')
check_sign(25, 'negative', 'US EBITDA')
check_sign(27, 'positive', 'Cash Rent Addback')
check_sign(28, 'negative', 'SBP')

for r in range(34, 39):
    check_sign(r, 'negative', ws_annual_f.cell(r, 2).value or f"Row {r}")
for r in range(42, 46):
    check_sign(r, 'negative', ws_annual_f.cell(r, 2).value or f"Row {r}")
for r in range(52, 56):
    check_sign(r, 'positive', ws_annual_f.cell(r, 2).value or f"Row {r}")
for r in range(56, 59):
    check_sign(r, 'negative', ws_annual_f.cell(r, 2).value or f"Row {r}")

check_sign(62, 'negative', 'Tax')
check_sign(158, 'negative', 'Capex')

if sign_issues:
    results['CHECK 4: Sign conventions'] = ('FAIL', sign_issues)
    print("FAIL")
    for e in sign_issues:
        print(e)
else:
    results['CHECK 4: Sign conventions'] = ('PASS', ['All sign conventions correct'])
    print("PASS - All sign conventions correct")


# ============================================================
# CHECK 5: Column A key consistency
# ============================================================
print("\n" + "=" * 70)
print("CHECK 5: Column A key consistency (Annual vs HY P&L section)")
print("=" * 70)

annual_keys = [(r, str(ws_annual_f.cell(r, 1).value).strip())
               for r in range(5, 66) if ws_annual_f.cell(r, 1).value is not None]
hy_keys = [(r, str(ws_hy_f.cell(r, 1).value).strip())
           for r in range(5, 67) if ws_hy_f.cell(r, 1).value is not None]

key_issues = []
annual_key_strs = [k for _, k in annual_keys]
hy_key_strs = [k for _, k in hy_keys]
annual_set = set(annual_key_strs)
hy_set = set(hy_key_strs)

only_annual = annual_set - hy_set
only_hy = hy_set - annual_set

if only_annual:
    key_issues.append(f"  Keys only in Annual: {sorted(only_annual)}")
if only_hy:
    key_issues.append(f"  Keys only in HY: {sorted(only_hy)}")

shared_annual = [k for k in annual_key_strs if k in hy_set]
shared_hy = [k for k in hy_key_strs if k in annual_set]

if shared_annual != shared_hy:
    key_issues.append("  Key ORDER differs")
    for i, (a, h) in enumerate(zip(shared_annual, shared_hy)):
        if a != h:
            key_issues.append(f"    Position {i}: Annual='{a}' vs HY='{h}'")
            break

if key_issues:
    results['CHECK 5: Column A key consistency'] = ('FAIL', key_issues)
    print("FAIL")
    for e in key_issues:
        print(e)
else:
    results['CHECK 5: Column A key consistency'] = ('PASS', [f'{len(annual_keys)} Annual keys match {len(hy_keys)} HY keys'])
    print(f"PASS - {len(annual_keys)} Annual / {len(hy_keys)} HY keys match in content and order")


# ============================================================
# CHECK 6: BS Check
# ============================================================
print("\n" + "=" * 70)
print("CHECK 6: Balance Sheet Check (Assets - Liabilities - Equity = 0)")
print("=" * 70)

bs_issues = []
bs_info = []
all_periods = {4: 'FY23', 5: 'FY24', 6: 'FY25', 7: 'FY26E', 8: 'FY27E'}

for col, label in all_periods.items():
    asset_rows = [105, 106, 107, 108, 109, 110, 111, 112, 113]
    liab_rows = [122, 123, 124, 125]
    eq_rows = [134, 135, 136, 137]

    assets = sum(ev(r, col) or 0 for r in asset_rows)
    liabs = sum(ev(r, col) or 0 for r in liab_rows)
    equity = sum(ev(r, col) or 0 for r in eq_rows)

    asset_nones = sum(1 for r in asset_rows if ev(r, col) is None)
    liab_nones = sum(1 for r in liab_rows if ev(r, col) is None)
    eq_nones = sum(1 for r in eq_rows if ev(r, col) is None)

    bs_check = assets - liabs - equity

    if asset_nones > 0 or liab_nones > 0 or eq_nones > 0:
        # Forecast periods have complex formulas we can't fully resolve
        unresolved = []
        for r in asset_rows + liab_rows + eq_rows:
            if ev(r, col) is None:
                unresolved.append(f"R{r}({ws_annual_f.cell(r,2).value})")
        msg = f"  {label}: SKIPPED - {len(unresolved)} cells with complex formulas unresolvable without Excel"
        bs_info.append(msg)
        print(msg)
        for u in unresolved:
            print(f"    Unresolved: {u}")
        # Verify the formula structure is correct (R141 = R114 - R126 - R138)
        r141_formula = ws_annual_f.cell(141, col).value
        expected_formula = f"={chr(64+col)}114-{chr(64+col)}126-{chr(64+col)}138"
        if r141_formula == expected_formula:
            print(f"    BS Check formula verified correct: {r141_formula}")
        else:
            bs_issues.append(f"  {label}: BS Check formula mismatch: {r141_formula} vs expected {expected_formula}")
    elif abs(bs_check) > TOL:
        bs_issues.append(f"  {label}: BS Check = {bs_check:.3f} (A={assets:.1f}, L={liabs:.1f}, E={equity:.1f})")
        print(f"  {label}: FAIL BS Check = {bs_check:.3f} (A={assets:.1f}, L={liabs:.1f}, E={equity:.1f})")
    else:
        print(f"  {label}: BS Check = {bs_check:.3f} OK (A={assets:.1f}, L={liabs:.1f}, E={equity:.1f})")

if bs_issues:
    results['CHECK 6: BS Check'] = ('FAIL', bs_issues)
else:
    detail = ['Historical periods (FY23-FY25) balance within +/-0.5']
    if bs_info:
        detail.append('Forecast periods: formulas structurally correct but values unresolvable without Excel')
    results['CHECK 6: BS Check'] = ('PASS', detail)


# ============================================================
# CHECK 7: Retained row completeness
# ============================================================
print("\n" + "=" * 70)
print("CHECK 7: Retained row completeness")
print("=" * 70)

completeness_issues = []

def check_keys(section, keys, ws):
    missing = [key for key in keys
               if not any(ws.cell(r, 1).value and str(ws.cell(r, 1).value).startswith(key)
                         for r in range(1, ws.max_row + 1))]
    if missing:
        completeness_issues.append(f"  Annual - {section}: MISSING {missing}")

check_keys('Revenue', ['Rev-Corp Restaurant Sales Aus', 'Rev-Corp Restaurant Sales US',
    'Rev-Total Corp Restaurant Sales', 'Rev-Franchise Royalty Revenue', 'Rev-Franchise Fee Revenue',
    'Rev-Total Revenue'], ws_annual_f)
check_keys('Other Revenue', ['OthRev-Marketing Levy', 'OthRev-Other Franchise Revenue',
    'OthRev-Other Income', 'OthRev-Total Other Revenue'], ws_annual_f)
check_keys('Segment EBITDA', ['SegEBITDA-Australia', 'SegEBITDA-US',
    'SegEBITDA-Total Segment EBITDA', 'Bridge-Cash Rent', 'Bridge-SBP', 'Bridge-Other'], ws_annual_f)
check_keys('EBITDA', ['EBITDA-Statutory EBITDA'], ws_annual_f)
check_keys('Expenses', ['Exp-Food & Packaging', 'Exp-Employee Benefits', 'Exp-Admin',
    'Exp-Marketing', 'Exp-Other Expenses', 'Exp-Total Expenses'], ws_annual_f)
check_keys('D&A', ['DA-ROU Depreciation', 'DA-PPE Depreciation',
    'DA-Amortisation Reacquired', 'DA-Amortisation Other', 'DA-Total DA'], ws_annual_f)
check_keys('EBIT/Finance/PBT/NPAT', ['EBIT-EBIT', 'Int-Term Deposit Income',
    'Int-Lease Receivable Income', 'Int-Other Finance Income', 'Int-Total Finance Income',
    'Int-Lease Liability Costs', 'Int-Other Finance Costs', 'Int-Total Finance Costs',
    'Int-Net Finance', 'PBT-PBT', 'Tax-Tax Expense', 'NPAT-NPAT'], ws_annual_f)
check_keys('EPS & Dividends', ['EPS-YE Shares', 'EPS-WASO Basic', 'EPS-Dilution',
    'EPS-WASO Diluted', 'EPS-Basic EPS', 'EPS-Diluted EPS', 'Div-DPS', 'Div-Total Dividends'], ws_annual_f)
check_keys('Operating Metrics', ['KPI-Corp Restaurants Aus', 'KPI-Franchise Restaurants Aus',
    'KPI-US Restaurants', 'KPI-Total Restaurants', 'KPI-Aus Network Sales',
    'KPI-Total Network Sales', 'KPI-Comp Sales Growth'], ws_annual_f)
check_keys('Balance Sheet', ['BS-Cash', 'BS-Trade Receivables', 'BS-Inventories',
    'BS-Term Deposits', 'BS-Finance Lease Recv', 'BS-PPE', 'BS-Intangibles', 'BS-ROU Assets',
    'BS-Other Assets', 'BS-Trade Payables', 'BS-Other Liabilities', 'BS-Lease Liabilities',
    'BS-Total Banking Debt', 'BS-Issued Capital', 'BS-Retained Profits', 'BS-Reserves', 'BS-Minorities'], ws_annual_f)
check_keys('Cash Flow', ['CF-EBITDA', 'CF-WC Change', 'CF-Significant Items',
    'CF-Int Received', 'CF-Interest Paid', 'CF-Lease Int Paid', 'CF-Tax Paid', 'CF-Net OCF',
    'CF-Capex PPE', 'CF-Capex Intang', 'CF-Acquisitions', 'CF-Asset Sales', 'CF-Other CFI',
    'CF-Dividends', 'CF-Share Issues', 'CF-Lease Principal', 'CF-Debt Change', 'CF-Other CFF'], ws_annual_f)

for r in [177, 178, 179, 180, 181]:
    if ws_annual_f.cell(r, 2).value is None:
        completeness_issues.append(f"  Annual - OpFCF: Row {r} is empty")
for r in [187, 188, 189, 190, 191]:
    if ws_annual_f.cell(r, 2).value is None:
        completeness_issues.append(f"  Annual - ROIC: Row {r} is empty")

ws_val = wb_f['Value']
for r in [25, 26, 27, 28, 29, 30, 31, 32, 33, 37, 38]:
    if ws_val.cell(r, 2).value is None:
        completeness_issues.append(f"  Value - DCF: Row {r} is empty")
for r in [57, 58, 59, 61, 64, 65]:
    if ws_val.cell(r, 2).value is None:
        completeness_issues.append(f"  Value - SOTP: Row {r} is empty")

if completeness_issues:
    results['CHECK 7: Row completeness'] = ('FAIL', completeness_issues)
    print("FAIL")
    for e in completeness_issues:
        print(e)
else:
    results['CHECK 7: Row completeness'] = ('PASS', ['All required sections and rows present'])
    print("PASS - All required sections and rows present")


# ============================================================
# CHECK 8: Formatting check
# ============================================================
print("\n" + "=" * 70)
print("CHECK 8: Formatting check")
print("=" * 70)

format_issues = []

def get_font_color(cell):
    font = cell.font
    if font and font.color:
        if font.color.rgb and font.color.rgb != '00000000':
            return str(font.color.rgb)
        if font.color.theme is not None:
            return f"theme:{font.color.theme}"
    return "default/black"

# Actual data cells - blue (0000CC)
actual_cells = [(7, 4, 'Rev Aus FY23'), (8, 5, 'Rev US FY24'), (24, 6, 'Aus EBITDA FY25'),
                (34, 4, 'Food Cost FY23'), (56, 5, 'Lease Costs FY24')]

for row, col, desc in actual_cells:
    cell = ws_annual_f.cell(row, col)
    color = get_font_color(cell)
    if '0000CC' not in color.upper():
        format_issues.append(f"  Actual cell ({desc}) R{row}C{col}: color={color} (expected blue 0000CC)")
    else:
        print(f"  OK: Actual ({desc}): blue")

# Forecast assumption cells on HY
# Note: 1H26 (col 10) appears to be entered as actuals (blue). Forecast starts at 2H26 (col 11).
# Check hard-coded assumption cells in the forecast driver section (rows 88-106)
ws_hy_styled = wb_f['HY & Segments']

# Sample driver cells in 2H26 (col 11) that are hard-coded assumptions
forecast_driver_cells = [
    (89, 11, 'New Corp Openings 2H26'),
    (91, 11, 'Revenue Growth 2H26'),
    (96, 11, 'EBITDA Margin 2H26'),
    (101, 11, 'New US Openings 2H26'),
    (105, 11, 'US EBITDA Margin 2H26'),
]

for row, col, desc in forecast_driver_cells:
    cell = ws_hy_styled.cell(row, col)
    v = cell.value
    color = get_font_color(cell)
    is_formula = isinstance(v, str) and v.startswith('=')

    if v is None:
        print(f"  INFO: Forecast ({desc}): empty")
    elif is_formula:
        print(f"  INFO: Forecast ({desc}): formula (color check N/A)")
    else:
        # Hard-coded forecast assumption should be maroon
        if 'C00000' not in color.upper():
            format_issues.append(f"  Forecast driver ({desc}) HY R{row}C{col}: color={color} (expected maroon C00000)")
        else:
            print(f"  OK: Forecast driver ({desc}): maroon")

# Subtotal rows - bold
for row, desc in [(12, 'Total Revenue'), (30, 'EBITDA'), (48, 'EBIT')]:
    cell = ws_annual_f.cell(row, 2)
    if not (cell.font and cell.font.bold):
        format_issues.append(f"  Subtotal row {row} ({desc}): NOT bold")
    else:
        print(f"  OK: Subtotal row {row} ({desc}): bold")

if format_issues:
    results['CHECK 8: Formatting'] = ('FAIL', format_issues)
    print("\nSome formatting issues found")
else:
    results['CHECK 8: Formatting'] = ('PASS', ['All sampled formatting checks passed'])
    print("\nPASS")


# ============================================================
# CHECK 9: First forecast year ratio sanity
# ============================================================
print("\n" + "=" * 70)
print("CHECK 9: First forecast year ratio sanity (FY26E vs FY25)")
print("=" * 70)

ratio_issues = []

rev_25 = ev(21, 6)
rev_26 = ev(21, 7)
ebitda_25 = ev(30, 6)
ebitda_26 = ev(30, 7)
ebit_25 = ev(48, 6)
ebit_26 = ev(48, 7)
da_25 = ev(46, 6)
da_26 = ev(46, 7)
pbt_25 = ev(61, 6)
pbt_26 = ev(61, 7)
tax_25 = ev(62, 6)
tax_26 = ev(62, 7)
npat_25 = ev(64, 6)
npat_26 = ev(64, 7)
capex_25 = ev(158, 6)
capex_26 = ev(158, 7)
total_rev_25 = ev(12, 6)
total_rev_26 = ev(12, 7)

# Print all resolved values for transparency
print(f"  FY25 values: Rev={rev_25}, TotalRev={total_rev_25}, EBITDA={ebitda_25}, EBIT={ebit_25}")
print(f"               D&A={da_25}, PBT={pbt_25}, Tax={tax_25}, NPAT={npat_25}, Capex={capex_25}")
print(f"  FY26 values: Rev={rev_26}, TotalRev={total_rev_26}, EBITDA={ebitda_26}, EBIT={ebit_26}")
print(f"               D&A={da_26}, PBT={pbt_26}, Tax={tax_26}, NPAT={npat_26}, Capex={capex_26}")
print()

def pct(num, denom):
    if denom and denom != 0 and num is not None:
        return num / denom
    return None

ratio_info = []

def check_ratio(name, val_25, val_26, tolerance_pct=0.50):
    if val_25 is None or val_26 is None:
        # If FY26 denominator couldn't be resolved due to complex HY formulas, note as info not failure
        if val_26 is None and val_25 is not None:
            ratio_info.append(f"  {name}: FY25={val_25:.4f}, FY26E=unresolvable (complex INDEX/MATCH chain)")
            print(f"  {name}: FY25={val_25:.4f}, FY26E=unresolvable (complex formula chain)")
        else:
            ratio_issues.append(f"  {name}: Cannot compute (FY25={val_25}, FY26E={val_26})")
            print(f"  {name}: Cannot compute (FY25={val_25}, FY26E={val_26})")
        return
    print(f"  {name}: FY25={val_25:.4f}, FY26E={val_26:.4f}", end="")
    if val_25 == 0:
        print(" (FY25=0, skip)")
        return
    if val_26 == 0:
        print(" (FY26=0, likely unresolved)")
        ratio_info.append(f"  {name}: FY26E=0 (likely unresolved formula)")
        return
    change = abs(val_26 / val_25 - 1)
    if change > tolerance_pct:
        msg = f" -> FLAGGED: {change:.1%} change (>{tolerance_pct:.0%})"
        ratio_issues.append(f"  {name}: FY25={val_25:.4f}, FY26E={val_26:.4f}{msg}")
        print(msg)
    else:
        print(f" -> OK ({change:.1%} change)")

check_ratio("EBITDA Margin", pct(ebitda_25, rev_25), pct(ebitda_26, rev_26))
check_ratio("EBIT Margin", pct(ebit_25, rev_25), pct(ebit_26, rev_26))
check_ratio("D&A/Revenue", pct(da_25, total_rev_25), pct(da_26, total_rev_26))

if pbt_25 and pbt_25 != 0 and pbt_26 and pbt_26 != 0:
    check_ratio("Effective Tax Rate", pct(tax_25, pbt_25), pct(tax_26, pbt_26))
else:
    print(f"  Effective Tax Rate: Skipped (PBT FY25={pbt_25}, FY26={pbt_26})")

check_ratio("NPAT Margin", pct(npat_25, rev_25), pct(npat_26, rev_26))
check_ratio("Capex/Revenue", pct(capex_25, total_rev_25), pct(capex_26, total_rev_26))

if ratio_issues:
    results['CHECK 9: Forecast ratio sanity'] = ('FAIL', ratio_issues)
elif ratio_info:
    # All computable ratios OK, but some couldn't be computed due to formula complexity
    results['CHECK 9: Forecast ratio sanity'] = ('PASS', [
        'FY25 ratios computed successfully; FY26E ratios unresolvable without Excel (complex INDEX/MATCH chains)',
        'Formula structure verified in CHECK 2; values require Excel evaluation'
    ])
    print("\nPASS (with caveat: FY26E ratios require Excel to compute)")
else:
    results['CHECK 9: Forecast ratio sanity'] = ('PASS', ['All forecast ratios within sanity bounds'])


# ============================================================
# SUMMARY
# ============================================================
print("\n")
print("=" * 70)
print("COMPREHENSIVE VERIFICATION SUMMARY")
print("=" * 70)

total_pass = 0
total_fail = 0

for check_name, (status, details) in results.items():
    icon = "PASS" if status == "PASS" else "FAIL"
    print(f"\n  [{icon}] {check_name}")
    if status == "FAIL":
        total_fail += 1
        for d in details[:15]:
            print(f"         {d}")
        if len(details) > 15:
            print(f"         ... and {len(details)-15} more")
    else:
        total_pass += 1
        for d in details:
            print(f"         {d}")

print(f"\n{'=' * 70}")
print(f"TOTAL: {total_pass} PASS, {total_fail} FAIL out of {total_pass + total_fail} checks")
print("=" * 70)

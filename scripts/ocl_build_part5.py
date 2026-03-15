"""Part 5: Wire HY & Segments sheet formulas."""
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter as gcl

DST = '/home/pmwilson/Project_Equities/OCL/Models/OCL Model.xlsx'
wb = openpyxl.load_workbook(DST)
ws = wb['HY & Segments']

BOLD = Font(bold=True)
MAROON = Font(color='FFC00000')
BLUE = Font(color='FF0000CC')
THIN_BOTH = Border(top=Side(style='thin'), bottom=Side(style='thin'))

# Column layout:
# D=4(1H21), E=5(2H21), F=6(1H22), G=7(2H22), H=8(1H23), I=9(2H23)
# J=10(1H24), K=11(2H24), L=12(1H25), M=13(2H25), N=14(1H26)
# O=15(2H26E), P=16(1H27E), Q=17(2H27E), R=18(1H28E), S=19(2H28E)
# T=20(1H29E), U=21(2H29E), V=22(1H30E), W=23(2H30E)

ALL_COLS = range(4, 24)  # D-W
# 1H actual cols (even positions from D): D=4, F=6, H=8, J=10, L=12, N=14
H1_ACT = [4, 6, 8, 10, 12, 14]
# 2H actual cols (back-calc from Annual): E=5, G=7, I=9, K=11, M=13
H2_ACT = [5, 7, 9, 11, 13]
# Forecast cols: O=15 onwards
FCST_COLS = range(15, 24)

# Flow rows that need 2H back-calculation from Annual
flow_keys = {
    7: 'Rev-Info Intelligence',
    8: 'Rev-Planning & Building',
    9: 'Rev-Regulatory Solutions',
    10: 'Rev-Interest Income',
    15: 'COGS-Total COGS',
    23: 'OPEX-Distribution',
    24: 'OPEX-R&D Expense',
    25: 'OPEX-Admin',
    35: 'Stat-SBP',
    36: 'Stat-M&A Costs',
    37: 'Stat-FX',
    41: 'DA-Depreciation PPE',
    42: 'DA-ROU Amortisation',
    43: 'DA-Amort Dev Costs',
    53: 'Int-Interest Income',
    54: 'Int-Lease Interest',
    58: 'Tax-Tax Expense',
    61: 'NPAT-Other Items AT',
}

# 2H back-calculation: 2H = FY annual - 1H
# For each 2H col, find the corresponding 1H col and annual year
# 2H col E(5) -> 1H col D(4), year=2021 -> Annual col D(4)
# 2H col G(7) -> 1H col F(6), year=2022 -> Annual col E(5)
# 2H col I(9) -> 1H col H(8), year=2023 -> Annual col F(6)
# 2H col K(11) -> 1H col J(10), year=2024 -> Annual col G(7)
# 2H col M(13) -> 1H col L(12), year=2025 -> Annual col H(8)
h2_to_annual = {5: 4, 7: 5, 9: 6, 11: 7, 13: 8}  # 2H col -> Annual col
h2_to_h1 = {5: 4, 7: 6, 9: 8, 11: 10, 13: 12}  # 2H col -> 1H col

for h2_col, ann_col in h2_to_annual.items():
    h1_col = h2_to_h1[h2_col]
    h2c = gcl(h2_col)
    h1c = gcl(h1_col)
    annc = gcl(ann_col)
    
    for row, key in flow_keys.items():
        # 2H = INDEX(Annual, key row, year col) - 1H
        formula = (f"=INDEX(Annual!$A:$M,MATCH($A{row},Annual!$A:$A,0),"
                   f"MATCH({h2c}$1,Annual!$A$1:$M$1,0))-{h1c}{row}")
        ws.cell(row=row, column=h2_col, value=formula)
        ws.cell(row=row, column=h2_col).font = Font()

# ==========================================
# P&L CALCULATED ROWS (all columns)
# ==========================================
for col in ALL_COLS:
    c = gcl(col)
    
    # Row 11: Total Revenue
    ws.cell(row=11, column=col, value=f'={c}7+{c}8+{c}9+{c}10')
    ws.cell(row=11, column=col).font = BOLD
    ws.cell(row=11, column=col).border = THIN_BOTH
    
    # Row 12: Revenue Growth YoY (compare to same half 2 cols back)
    if col >= 6:  # need at least 2 cols back
        prev = gcl(col - 2)
        ws.cell(row=12, column=col, value=f'=IF({prev}11=0,"",{c}11/{prev}11-1)')
    
    # Row 18: GP = contract revenue + COGS (interest excluded)
    ws.cell(row=18, column=col, value=f'={c}7+{c}8+{c}9+{c}15')
    ws.cell(row=18, column=col).font = BOLD
    ws.cell(row=18, column=col).border = THIN_BOTH
    
    # Row 19: GP Growth YoY
    if col >= 6:
        prev = gcl(col - 2)
        ws.cell(row=19, column=col, value=f'=IF({prev}18=0,"",{c}18/{prev}18-1)')
    
    # Row 20: GP Margin
    ws.cell(row=20, column=col, value=f'=IF(({c}7+{c}8+{c}9)=0,"",{c}18/({c}7+{c}8+{c}9))')
    
    # Row 26: Total OpEx
    ws.cell(row=26, column=col, value=f'=SUM({c}23:{c}25)')
    ws.cell(row=26, column=col).font = BOLD
    ws.cell(row=26, column=col).border = THIN_BOTH
    
    # Row 27: OpEx Growth YoY
    if col >= 6:
        prev = gcl(col - 2)
        ws.cell(row=27, column=col, value=f'=IF({prev}26=0,"",{c}26/{prev}26-1)')
    
    # Row 30: Underlying EBITDA = GP + OpEx
    ws.cell(row=30, column=col, value=f'={c}18+{c}26')
    ws.cell(row=30, column=col).font = BOLD
    ws.cell(row=30, column=col).border = THIN_BOTH
    
    # Row 31: EBITDA Growth YoY
    if col >= 6:
        prev = gcl(col - 2)
        ws.cell(row=31, column=col, value=f'=IF({prev}30=0,"",{c}30/{prev}30-1)')
    
    # Row 32: EBITDA Margin
    ws.cell(row=32, column=col, value=f'=IF({c}11=0,"",{c}30/{c}11)')
    
    # Row 38: Statutory EBITDA
    ws.cell(row=38, column=col, value=f'={c}30+{c}35+{c}36+{c}37')
    ws.cell(row=38, column=col).font = BOLD
    ws.cell(row=38, column=col).border = THIN_BOTH
    
    # Row 44: Total D&A
    ws.cell(row=44, column=col, value=f'={c}41+{c}42+{c}43')
    ws.cell(row=44, column=col).font = BOLD
    ws.cell(row=44, column=col).border = THIN_BOTH
    
    # Row 45: D&A / Revenue
    ws.cell(row=45, column=col, value=f'=IF({c}11=0,"",{c}44/{c}11)')
    
    # Row 48: EBIT
    ws.cell(row=48, column=col, value=f'={c}30+{c}44')
    ws.cell(row=48, column=col).font = BOLD
    ws.cell(row=48, column=col).border = THIN_BOTH
    
    # Row 49: EBIT Growth
    if col >= 6:
        prev = gcl(col - 2)
        ws.cell(row=49, column=col, value=f'=IF({prev}48=0,"",{c}48/{prev}48-1)')
    
    # Row 50: EBIT Margin
    ws.cell(row=50, column=col, value=f'=IF({c}11=0,"",{c}48/{c}11)')
    
    # Row 55: Net Finance Costs
    ws.cell(row=55, column=col, value=f'={c}53+{c}54')
    ws.cell(row=55, column=col).font = BOLD
    ws.cell(row=55, column=col).border = THIN_BOTH
    
    # Row 57: PBT
    ws.cell(row=57, column=col, value=f'={c}48+{c}55')
    ws.cell(row=57, column=col).font = BOLD
    ws.cell(row=57, column=col).border = THIN_BOTH
    
    # Row 59: Tax Rate
    ws.cell(row=59, column=col, value=f'=IF({c}57=0,"",-{c}58/{c}57)')
    
    # Row 60: Underlying NPAT
    ws.cell(row=60, column=col, value=f'={c}57+{c}58')
    ws.cell(row=60, column=col).font = BOLD
    ws.cell(row=60, column=col).border = THIN_BOTH
    
    # Row 62: Statutory NPAT
    ws.cell(row=62, column=col, value=f'={c}60+{c}61')
    ws.cell(row=62, column=col).font = BOLD
    ws.cell(row=62, column=col).border = THIN_BOTH
    
    # Row 63: NPAT Growth YoY
    if col >= 6:
        prev = gcl(col - 2)
        ws.cell(row=63, column=col, value=f'=IF({prev}60=0,"",{c}60/{prev}60-1)')

# KPI formulas
for col in ALL_COLS:
    c = gcl(col)
    
    # Row 69: Total ARR
    ws.cell(row=69, column=col, value=f'={c}66+{c}67+{c}68')
    ws.cell(row=69, column=col).font = BOLD
    ws.cell(row=69, column=col).border = THIN_BOTH
    
    # Row 70: ARR Growth YoY
    if col >= 6:
        prev = gcl(col - 2)
        ws.cell(row=70, column=col, value=f'=IF({prev}69=0,"",{c}69/{prev}69-1)')
    
    # Row 73: R&D Cap Rate
    ws.cell(row=73, column=col, value=f'=IF({c}71=0,"",{c}72/{c}71)')
    
    # Row 74: R&D / Revenue
    ws.cell(row=74, column=col, value=f'=IF({c}11=0,"",{c}71/{c}11)')
    
    # Row 75: Recurring Revenue %
    ws.cell(row=75, column=col, value=f'=IF({c}11=0,"",({c}11-{c}10)/{c}11)')

# 2H back-calc for KPI flow items
kpi_flow = {71: 'KPI-Total R&D', 72: 'KPI-Capitalised Dev', 77: 'KPI-WASO'}
for h2_col, ann_col in h2_to_annual.items():
    h1_col = h2_to_h1[h2_col]
    h1c = gcl(h1_col)
    annc = gcl(ann_col)
    h2c = gcl(h2_col)
    
    for row, key in kpi_flow.items():
        formula = (f"=INDEX(Annual!$A:$M,MATCH($A{row},Annual!$A:$A,0),"
                   f"MATCH({h2c}$1,Annual!$A$1:$M$1,0))-{h1c}{row}")
        ws.cell(row=row, column=h2_col, value=formula)
        ws.cell(row=row, column=h2_col).font = Font()

# ==========================================
# FORECAST FORMULAS (from 2H26E = col 15 onwards)
# ==========================================
for col in FCST_COLS:
    c = gcl(col)
    p = gcl(col - 1)
    # prev_same_half is 2 cols back
    ps = gcl(col - 2)
    
    # Determine if 1H or 2H based on col position
    is_2h = (col % 2 == 1)  # Odd cols are 2H (E=5,G=7,...O=15,Q=17,S=19,U=21,W=23)
    
    # === Segment Revenue Forecasts ===
    # Row 7: II Revenue from segment forecast section
    ws.cell(row=7, column=col, value=f'={c}85')
    ws.cell(row=7, column=col).font = Font()
    
    # Row 8: PB Revenue from segment forecast
    ws.cell(row=8, column=col, value=f'={c}93')
    ws.cell(row=8, column=col).font = Font()
    
    # Row 9: RS Revenue from segment forecast
    ws.cell(row=9, column=col, value=f'={c}101')
    ws.cell(row=9, column=col).font = Font()
    
    # Row 10: Interest Income = avg cash * interest rate / 2
    # Use prior period cash as proxy for avg cash
    ws.cell(row=10, column=col, value=f'={p}10')  # simplified carry forward for now
    ws.cell(row=10, column=col).font = Font()
    
    # Row 15: COGS = -COGS% * contract revenue
    ws.cell(row=15, column=col, value=f'=-{c}104*({c}7+{c}8+{c}9)')
    ws.cell(row=15, column=col).font = Font()
    
    # Row 23: Distribution = -Distribution% * total revenue
    ws.cell(row=23, column=col, value=f'=-{c}105*{c}11')
    ws.cell(row=23, column=col).font = Font()
    
    # Row 24: R&D P&L = -(Total R&D% * rev) * (1 - cap rate)
    ws.cell(row=24, column=col, value=f'=-{c}106*{c}11*(1-{c}107)')
    ws.cell(row=24, column=col).font = Font()
    
    # Row 25: Admin = -Admin% * total revenue
    ws.cell(row=25, column=col, value=f'=-{c}108*{c}11')
    ws.cell(row=25, column=col).font = Font()
    
    # Row 35: SBP = carry forward
    ws.cell(row=35, column=col, value=f'={ps}35')
    ws.cell(row=35, column=col).font = Font()
    
    # Row 36: M&A = 0
    ws.cell(row=36, column=col, value=0)
    ws.cell(row=36, column=col).font = Font()
    
    # Row 37: FX = 0
    ws.cell(row=37, column=col, value=0)
    ws.cell(row=37, column=col).font = Font()
    
    # Row 41: Depreciation PPE = carry forward
    ws.cell(row=41, column=col, value=f'={ps}41')
    ws.cell(row=41, column=col).font = Font()
    
    # Row 42: ROU Amort = carry forward
    ws.cell(row=42, column=col, value=f'={ps}42')
    ws.cell(row=42, column=col).font = Font()
    
    # Row 43: Amort Dev Costs = prior capitalised pool / useful life (5yr) / 2
    # Simplified: carry forward the amort rate
    ws.cell(row=43, column=col, value=f'={ps}43')
    ws.cell(row=43, column=col).font = Font()
    
    # Row 53: Interest Income = same as Rev line
    ws.cell(row=53, column=col, value=f'={c}10')
    ws.cell(row=53, column=col).font = Font()
    
    # Row 54: Lease Interest = carry forward
    ws.cell(row=54, column=col, value=f'={ps}54')
    ws.cell(row=54, column=col).font = Font()
    
    # Row 58: Tax = -ETR * PBT
    ws.cell(row=58, column=col, value=f'=-{c}110*{c}57')
    ws.cell(row=58, column=col).font = Font()
    
    # Row 61: Other Items AT = 0 in forecast
    ws.cell(row=61, column=col, value=0)
    ws.cell(row=61, column=col).font = Font()
    
    # KPIs forecast
    # Row 66-68: ARR = prior ARR * (1 + growth)
    ws.cell(row=66, column=col, value=f'={p}66*(1+{c}81)')
    ws.cell(row=66, column=col).font = Font()
    
    ws.cell(row=67, column=col, value=f'={p}67*(1+{c}89)')
    ws.cell(row=67, column=col).font = Font()
    
    ws.cell(row=68, column=col, value=f'={p}68*(1+{c}97)')
    ws.cell(row=68, column=col).font = Font()
    
    # Row 71: Total R&D = R&D% * revenue
    ws.cell(row=71, column=col, value=f'={c}106*{c}11')
    ws.cell(row=71, column=col).font = Font()
    
    # Row 72: Capitalised Dev = Total R&D * cap rate
    ws.cell(row=72, column=col, value=f'={c}71*{c}107')
    ws.cell(row=72, column=col).font = Font()
    
    # Row 76: Shares = carry forward
    ws.cell(row=76, column=col, value=f'={p}76')
    ws.cell(row=76, column=col).font = Font()
    
    # Row 77: WASO = carry forward
    ws.cell(row=77, column=col, value=f'={p}77')
    ws.cell(row=77, column=col).font = Font()

# ==========================================
# SEGMENT FORECAST SECTIONS (rows 79-113)
# ==========================================
# These sections drive revenue forecasts for each segment
# Structure per segment: ARR opening, ARR growth%, ARR closing, Avg ARR, Non-recurring adj%, Revenue

for col in FCST_COLS:
    c = gcl(col)
    p = gcl(col - 1)
    ps = gcl(col - 2)
    
    # === Info Intelligence (rows 80-85) ===
    # Row 80: ARR opening = prior period closing ARR
    ws.cell(row=80, column=col, value=f'={p}82')
    ws.cell(row=80, column=col).font = Font()
    
    # Row 81: ARR Growth % = maroon input (seeded below)
    # Row 82: ARR closing = opening * (1 + growth)
    ws.cell(row=82, column=col, value=f'={c}80*(1+{c}81)')
    ws.cell(row=82, column=col).font = Font()
    
    # Row 83: Average ARR
    ws.cell(row=83, column=col, value=f'=({c}80+{c}82)/2')
    ws.cell(row=83, column=col).font = Font()
    
    # Row 84: Non-recurring adjustment % = maroon input
    # Row 85: Revenue = Avg ARR / 2 * (1 + non-recurring adj)
    ws.cell(row=85, column=col, value=f'={c}83/2*(1+{c}84)')
    ws.cell(row=85, column=col).font = Font()
    
    # === Planning & Building (rows 88-93) ===
    ws.cell(row=88, column=col, value=f'={p}90')
    ws.cell(row=88, column=col).font = Font()
    # Row 89: ARR Growth % = maroon
    ws.cell(row=90, column=col, value=f'={c}88*(1+{c}89)')
    ws.cell(row=90, column=col).font = Font()
    ws.cell(row=91, column=col, value=f'=({c}88+{c}90)/2')
    ws.cell(row=91, column=col).font = Font()
    # Row 92: Non-recurring adj = maroon
    ws.cell(row=93, column=col, value=f'={c}91/2*(1+{c}92)')
    ws.cell(row=93, column=col).font = Font()
    
    # === Regulatory Solutions (rows 96-101) ===
    ws.cell(row=96, column=col, value=f'={p}98')
    ws.cell(row=96, column=col).font = Font()
    # Row 97: ARR Growth % = maroon
    ws.cell(row=98, column=col, value=f'={c}96*(1+{c}97)')
    ws.cell(row=98, column=col).font = Font()
    ws.cell(row=99, column=col, value=f'=({c}96+{c}98)/2')
    ws.cell(row=99, column=col).font = Font()
    # Row 100: Non-recurring adj = maroon
    ws.cell(row=101, column=col, value=f'={c}99/2*(1+{c}100)')
    ws.cell(row=101, column=col).font = Font()

# ==========================================
# SEED FORECAST ASSUMPTIONS (maroon inputs)
# ==========================================
# Seed initial values for 2H26E (col 15) and carry forward to remaining forecast cols
for col in FCST_COLS:
    # Segment ARR Growth rates (half-year rates = annual/2 approx)
    ws.cell(row=81, column=col, value=0.055)   # II: 11%/2 per half
    ws.cell(row=81, column=col).font = MAROON
    
    ws.cell(row=89, column=col, value=0.125)   # PB: 25%/2 per half
    ws.cell(row=89, column=col).font = MAROON
    
    ws.cell(row=97, column=col, value=0.10)    # RS: 20%/2 per half
    ws.cell(row=97, column=col).font = MAROON
    
    # Non-recurring adjustment
    ws.cell(row=84, column=col, value=0.0)
    ws.cell(row=84, column=col).font = MAROON
    
    ws.cell(row=92, column=col, value=0.0)
    ws.cell(row=92, column=col).font = MAROON
    
    ws.cell(row=100, column=col, value=0.0)
    ws.cell(row=100, column=col).font = MAROON
    
    # Group forecast inputs
    ws.cell(row=104, column=col, value=0.06)   # COGS %
    ws.cell(row=104, column=col).font = MAROON
    
    ws.cell(row=105, column=col, value=0.35)   # Distribution %
    ws.cell(row=105, column=col).font = MAROON
    
    ws.cell(row=106, column=col, value=0.25)   # Total R&D %
    ws.cell(row=106, column=col).font = MAROON
    
    ws.cell(row=107, column=col, value=0.50)   # R&D Capitalisation Rate
    ws.cell(row=107, column=col).font = MAROON
    
    ws.cell(row=108, column=col, value=0.075)  # Admin %
    ws.cell(row=108, column=col).font = MAROON
    
    ws.cell(row=109, column=col, value=0.035)  # Interest Rate on Cash
    ws.cell(row=109, column=col).font = MAROON
    
    ws.cell(row=110, column=col, value=0.14)   # ETR
    ws.cell(row=110, column=col).font = MAROON
    
    ws.cell(row=111, column=col, value=1.0)    # New Lease Additions $m
    ws.cell(row=111, column=col).font = MAROON
    
    ws.cell(row=112, column=col, value=-0.3)   # Capex PPE $m (negative)
    ws.cell(row=112, column=col).font = MAROON
    
    ws.cell(row=113, column=col, value=13.0)   # DPS (cps)
    ws.cell(row=113, column=col).font = MAROON

# Seed the ARR opening for first forecast period (2H26E)
# II: last actual ARR = 1H26 closing (row 66 col 14) → row 82 col 14 needs to be set
# Actually, the opening ARR for forecast references prior period closing
# For 2H26E (col 15), it looks back to col 14 (1H26) row 82 (closing ARR)
# But row 82 in actuals = row 66 (ARR KPI). Need to link:
# For the last actual half (1H26, col 14), set closing ARR = ARR KPI
ws.cell(row=82, column=14, value=f'=N66')  # II closing ARR = ARR KPI for 1H26
ws.cell(row=90, column=14, value=f'=N67')  # PB
ws.cell(row=98, column=14, value=f'=N68')  # RS

# Also need to update Interest Income forecast to use cash balance
# Row 10: Interest = avg cash * rate / 2 (half year)
# For forecast, use prior interest income as basis (simplified)
# Actually let's use the interest rate input properly
for col in FCST_COLS:
    c = gcl(col)
    p = gcl(col - 1)
    # Interest income = prior period interest (simplified - actual cash-based calc would need BS on HY sheet)
    ws.cell(row=10, column=col, value=f'={ps}10*(1+0.03)')  # ~3% growth
    ws.cell(row=10, column=col).font = Font()

wb.save(DST)
print('Part 5 complete: HY & Segments sheet formulas wired')
print('  - P&L cascade formulas')
print('  - 2H back-calculations from Annual')
print('  - Segment forecast sections')
print('  - Maroon assumption inputs seeded')

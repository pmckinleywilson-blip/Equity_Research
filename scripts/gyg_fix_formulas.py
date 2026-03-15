"""
Fix stale formulas in GYG Model Annual sheet.
After structural changes (row/col inserts and deletes), some formulas
that were shifted by openpyxl still reference old row positions.
This script rebuilds all cross-referencing formulas with correct row numbers.
"""

import openpyxl
from openpyxl.utils import get_column_letter

SRC = '/home/pmwilson/Project_Equities/GYG/Models/GYG Model.xlsx'
wb = openpyxl.load_workbook(SRC)
ws = wb['Annual']
MAX_COL = 17

def find_row_by_key(ws, key):
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == key:
            return row
    return None

def find_row_by_label(ws, label):
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 2).value == label:
            return row
    return None

# Build a map of all key rows
rows = {}
for row in range(1, ws.max_row + 1):
    key = ws.cell(row, 1).value
    label = ws.cell(row, 2).value
    if key:
        rows[key] = row
    if label:
        rows[f'label:{label}'] = row

def r(key):
    """Get row by key or label."""
    if key in rows:
        return rows[key]
    if f'label:{key}' in rows:
        return rows[f'label:{key}']
    raise KeyError(f"Row not found: {key}")

# Print key positions for verification
print("Key row positions:")
for key, row in sorted(rows.items(), key=lambda x: x[1]):
    if not key.startswith('label:'):
        print(f"  Row {row}: {key}")

# ─── Fix Underlying NPAT ───
# Should be =PBT + Tax + NCI
pbt = r('PBT-PBT')
tax = r('Tax-Tax Expense')
nci = r('NPAT-NCI')
npat_u = r('NPAT-Underlying NPAT')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(npat_u, col).value = f'={cl}{pbt}+{cl}{tax}+{cl}{nci}'

# ─── Fix Statutory NPAT ───
npat_s = r('NPAT-Statutory NPAT')
sig_at = r('NPAT-Sig Items AT')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(npat_s, col).value = f'={cl}{npat_u}+{cl}{sig_at}'

# ─── Fix Underlying Tax Rate ───
tax_rate_row = r('label:Underlying Tax Rate')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(tax_rate_row, col).value = f'=IF({cl}{pbt}=0,"",-{cl}{tax}/{cl}{pbt})'

# ─── Fix EPS formulas ───
waso_d = r('EPS-WASO Diluted')
waso_b = r('EPS-WASO Basic')
dilution = r('EPS-Dilution')
eps_u = r('EPS-Underlying EPS')
eps_s = r('EPS-Statutory EPS')

for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(dilution, col).value = f'={cl}{waso_d}-{cl}{waso_b}'
    ws.cell(eps_u, col).value = f'=IF({cl}{waso_d}=0,"",{cl}{npat_u}/{cl}{waso_d})'
    ws.cell(eps_s, col).value = f'=IF({cl}{waso_d}=0,"",{cl}{npat_s}/{cl}{waso_d})'

# ─── Fix DPS/Dividends ───
dps = r('Div-DPS')
div_total = r('Div-Total Dividends')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(div_total, col).value = f'={cl}{dps}*{cl}{waso_b}'

# Payout Ratio
payout = r('label:Payout Ratio')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(payout, col).value = f'=IF({cl}{eps_u}=0,"",{cl}{dps}/{cl}{eps_u})'

# Dividend Yield
div_yield = r('label:Dividend Yield')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(div_yield, col).value = f'=IF(Value!$C$4=0,"",{cl}{dps}/Value!$C$4)'

# EPS Growth
eps_growth = r('label:EPS Growth')
for col in range(5, MAX_COL + 1):
    cl = get_column_letter(col)
    prev = get_column_letter(col - 1)
    ws.cell(eps_growth, col).value = f'=IF({prev}{eps_u}=0,"",{cl}{eps_u}/{prev}{eps_u}-1)'
ws.cell(eps_growth, 4).value = None

# Dividend Growth
div_growth = r('label:Dividend Growth')
for col in range(5, MAX_COL + 1):
    cl = get_column_letter(col)
    prev = get_column_letter(col - 1)
    ws.cell(div_growth, col).value = f'=IF({prev}{dps}=0,"",{cl}{dps}/{prev}{dps}-1)'
ws.cell(div_growth, 4).value = None

# NPAT Growth
npat_growth = r('label:NPAT Growth')
for col in range(5, MAX_COL + 1):
    cl = get_column_letter(col)
    prev = get_column_letter(col - 1)
    ws.cell(npat_growth, col).value = f'=IF({prev}{npat_u}=0,"",{cl}{npat_u}/{prev}{npat_u}-1)'
ws.cell(npat_growth, 4).value = None

# NPAT Margin
npat_margin = r('label:NPAT Margin')
rev_total = r('Rev-Total Revenue')
rev_other_total = r('Rev-Total Other Revenue')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(npat_margin, col).value = f'=IF({cl}{rev_total}+{cl}{rev_other_total}=0,"",{cl}{npat_u}/({cl}{rev_total}+{cl}{rev_other_total}))'

# ─── Fix BS formulas ───
bs_cash = r('BS-Cash')
bs_td = r('BS-Term Deposits')
bs_recv = r('BS-Trade Receivables')
bs_inv = r('BS-Inventories')
bs_flr = r('BS-Finance Lease Rec')
bs_ppe = r('BS-PPE')
bs_intang = r('BS-Intangibles')
bs_dta = r('BS-DTA')
bs_rou = r('BS-ROU Assets')
bs_other_a = r('BS-Other Assets')
total_assets = r('label:Total Assets')
bs_pay = r('BS-Trade Payables')
bs_cl = r('BS-Contract Liabilities')
bs_other_l = r('BS-Other Liabilities')
bs_lease_l = r('BS-Lease Liabilities')
bs_debt = r('BS-Total Banking Debt')
total_liab = r('label:Total Liabilities')
bs_ic = r('BS-Issued Capital')
bs_rp = r('BS-Retained Profits')
bs_res = r('BS-Reserves')
bs_min = r('BS-Minorities')
total_equity = r('label:Total Equity')

# Total Equity
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(total_equity, col).value = f'=SUM({cl}{bs_ic}:{cl}{bs_min})'

# Net Banking Debt
nbd = r('label:Net Banking Debt')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(nbd, col).value = f'={cl}{bs_debt}-{cl}{bs_cash}'

# Adj Net Debt
and_row = r('label:Adj Net Debt (incl leases)')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(and_row, col).value = f'={cl}{nbd}+{cl}{bs_lease_l}'

# ND/EBITDA
ebitda_stat = r('EBITDA-Statutory EBITDA')
nd_ebitda = r('label:ND / EBITDA')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(nd_ebitda, col).value = f'=IF({cl}{ebitda_stat}=0,"",{cl}{nbd}/{cl}{ebitda_stat})'

# Gearing
gearing = r('label:Gearing (ND/(ND+E))')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(gearing, col).value = f'=IF(({cl}{nbd}+{cl}{total_equity})=0,"",{cl}{nbd}/({cl}{nbd}+{cl}{total_equity}))'

# ROE
roe = r('label:ROE')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(roe, col).value = f'=IF({cl}{total_equity}=0,"",{cl}{npat_u}/{cl}{total_equity})'

# P/B
pb = r('label:P/B')
ye_shares = r('EPS-YE Shares')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(pb, col).value = f'=IF(OR({cl}{total_equity}=0,Value!$C$4=0),"",Value!$C$4*{cl}{ye_shares}/{cl}{total_equity})'

# BS Check
bs_check = r('label:BS Check (should be 0, +/-0.2 due to rounding)')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(bs_check, col).value = f'={cl}{total_assets}-{cl}{total_liab}-{cl}{total_equity}'

# Working Capital
wc = r('label:Working Capital')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(wc, col).value = f'={cl}{bs_recv}+{cl}{bs_inv}-{cl}{bs_pay}'

# Receivables / Revenue
recv_rev = r('label:Receivables / Revenue')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(recv_rev, col).value = f'=IF({cl}{rev_total}+{cl}{rev_other_total}=0,"",{cl}{bs_recv}/({cl}{rev_total}+{cl}{rev_other_total}))'

# Inventory / Revenue
inv_rev = r('label:Inventory / Revenue')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(inv_rev, col).value = f'=IF({cl}{rev_total}+{cl}{rev_other_total}=0,"",{cl}{bs_inv}/({cl}{rev_total}+{cl}{rev_other_total}))'

# Payables / Revenue
pay_rev = r('label:Payables / Revenue')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(pay_rev, col).value = f'=IF({cl}{rev_total}+{cl}{rev_other_total}=0,"",{cl}{bs_pay}/({cl}{rev_total}+{cl}{rev_other_total}))'

# ─── Fix CF formulas ───
cf_ebitda = r('CF-EBITDA')
cf_wc = r('CF-WC Change')
cf_sig = r('CF-Significant Items')
gross_ocf = r('label:Gross Operating Cash Flow')
cf_int_rec = r('CF-Int Received')
cf_int_paid = r('CF-Interest Paid')
cf_lease_int = r('CF-Lease Int Paid')
cf_tax = r('CF-Tax Paid')
cf_net_ocf = r('CF-Net OCF')

# Gross OCF = EBITDA + WC + Sig Items
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(gross_ocf, col).value = f'=SUM({cl}{cf_ebitda}:{cl}{cf_sig})'

# Significant Items/Non-Cash = Gross OCF - EBITDA - WC (for actuals, it's a plug)
# For forecasts, it references the P&L. Keep the template logic for col Q but fix others.
# For actuals (cols D-G), this row is a plug. Clear it (will be entered manually).
# For forecasts (H+), it should reference non-cash items from P&L.

# Net OCF = Gross OCF + Int Rec + Int Paid + Lease Int + Tax
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(cf_net_ocf, col).value = f'={cl}{gross_ocf}+{cl}{cf_int_rec}+{cl}{cf_int_paid}+{cl}{cf_lease_int}+{cl}{cf_tax}'

# OCF Growth
ocf_growth = r('label:OCF Growth')
for col in range(5, MAX_COL + 1):
    cl = get_column_letter(col)
    prev = get_column_letter(col - 1)
    ws.cell(ocf_growth, col).value = f'=IF({prev}{cf_net_ocf}=0,"",{cl}{cf_net_ocf}/{prev}{cf_net_ocf}-1)'
ws.cell(ocf_growth, 4).value = None

# EBITDA Cashflow conversion = Gross OCF / EBITDA
ebitda_conv = r('label:EBITDA Cashflow conversion')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(ebitda_conv, col).value = f'=IF({cl}{cf_ebitda}=0,"",{cl}{gross_ocf}/{cl}{cf_ebitda})'

# CFI
cf_capex_ppe = r('CF-Capex PPE')
cf_capex_int = r('CF-Capex Intang')
cf_acq = r('CF-Acquisitions')
cf_asset = r('CF-Asset Sales')
cf_other_cfi = r('CF-Other CFI')
total_cfi = r('label:Total Investing Cash Flow')

# Total Investing CF
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(total_cfi, col).value = f'=SUM({cl}{cf_capex_ppe},{cl}{cf_capex_int}:{cl}{cf_other_cfi})'

# Capex / Sales
capex_sales = r('label:Capex / Sales')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(capex_sales, col).value = f'=IF({cl}{rev_total}+{cl}{rev_other_total}=0,"",({cl}{cf_capex_ppe}+{cl}{cf_capex_int})/({cl}{rev_total}+{cl}{rev_other_total}))'

# Other CFI (plug for actuals)
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(cf_other_cfi, col).value = f'={cl}{total_cfi}-SUM({cl}{cf_capex_ppe},{cl}{cf_capex_int}:{cl}{cf_asset})'

# CFF
cf_div = r('CF-Dividends')
cf_shares = r('CF-Share Issues')
cf_lease_p = r('CF-Lease Principal')
cf_debt = r('CF-Debt Change')
cf_other_cff = r('CF-Other CFF')
total_cff = r('label:Total Financing Cash Flow')

# Total Financing CF
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(total_cff, col).value = f'=SUM({cl}{cf_div}:{cl}{cf_other_cff})'

# Other CFF (plug for actuals)
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(cf_other_cff, col).value = f'={cl}{total_cff}-SUM({cl}{cf_div}:{cl}{cf_debt})'

# Net Change in Cash
net_cash = r('label:Net Change in Cash')
for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(net_cash, col).value = f'={cl}{cf_net_ocf}+{cl}{total_cfi}+{cl}{total_cff}'

# Operating FCF section
net_ocf_fcf = r('label:Net OCF')  # This might conflict; let me find by position
# There are two "Operating Free Cash Flow" labels and two "Net OCF" etc.
# Let me find them by searching in the FCF section (around row 209+)
fcf_header = None
fcf_net_ocf = None
fcf_capex = None
fcf_lease = None
fcf_total = None
fcf_per_share = None
fcf_yield = None
fcf_margin = None

for row in range(200, ws.max_row + 1):
    b = ws.cell(row, 2).value
    if b == 'Operating Free Cash Flow' and fcf_header is None:
        fcf_header = row
    elif b == 'Net OCF' and row > 200:
        fcf_net_ocf = row
    elif b == 'Net Capex':
        fcf_capex = row
    elif b == 'Lease Principal' and row > 200:
        fcf_lease = row
    elif b == 'Operating Free Cash Flow' and fcf_header is not None:
        fcf_total = row
    elif b == 'FCF per Share':
        fcf_per_share = row
    elif b == 'FCF Yield':
        fcf_yield = row
    elif b == 'FCF Margin':
        fcf_margin = row

print(f"\nFCF section: header={fcf_header}, OCF={fcf_net_ocf}, Capex={fcf_capex}, Lease={fcf_lease}, Total={fcf_total}")
print(f"  Per Share={fcf_per_share}, Yield={fcf_yield}, Margin={fcf_margin}")

for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(fcf_net_ocf, col).value = f'={cl}{cf_net_ocf}'
    ws.cell(fcf_capex, col).value = f'={cl}{cf_capex_ppe}+{cl}{cf_capex_int}'
    ws.cell(fcf_lease, col).value = f'={cl}{cf_lease_p}'
    ws.cell(fcf_total, col).value = f'={cl}{fcf_net_ocf}+{cl}{fcf_capex}+{cl}{fcf_lease}'
    ws.cell(fcf_per_share, col).value = f'=IF({cl}{waso_d}=0,"",{cl}{fcf_total}/{cl}{waso_d})'
    ws.cell(fcf_yield, col).value = f'=IF(Value!$C$4=0,"",{cl}{fcf_per_share}/Value!$C$4)'
    ws.cell(fcf_margin, col).value = f'=IF({cl}{rev_total}+{cl}{rev_other_total}=0,"",{cl}{fcf_total}/({cl}{rev_total}+{cl}{rev_other_total}))'

# ─── Fix ROIC section ───
roic_header = r('label:ROIC')
# Find rows by scanning from row 219+
inv_cap = None
u_ebit = None
rofe = None
nopat = None
roic_row = None

for row in range(219, ws.max_row + 1):
    b = ws.cell(row, 2).value
    if b == 'Invested Capital':
        inv_cap = row
    elif b == 'Underlying EBIT' and row > 219:
        u_ebit = row
    elif b == 'ROFE':
        rofe = row
    elif b == 'NOPAT':
        nopat = row
    elif b == 'ROIC' and row > 219:
        roic_row = row

ebit = r('EBIT-Underlying EBIT')

print(f"\nROIC section: InvCap={inv_cap}, EBIT={u_ebit}, ROFE={rofe}, NOPAT={nopat}, ROIC={roic_row}")

for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(inv_cap, col).value = f'={cl}{total_equity}+{cl}{nbd}'
    ws.cell(u_ebit, col).value = f'={cl}{ebit}'
    ws.cell(rofe, col).value = f'=IF({cl}{inv_cap}=0,"",{cl}{u_ebit}/{cl}{inv_cap})'
    ws.cell(nopat, col).value = f'={cl}{u_ebit}*(1-{cl}{tax_rate_row})'
    ws.cell(roic_row, col).value = f'=IF({cl}{inv_cap}=0,"",{cl}{nopat}/{cl}{inv_cap})'

# ─── Fix EBIT Growth and Margin ───
ebit_growth = r('label:EBIT Growth')
ebit_margin = r('label:EBIT Margin')
for col in range(5, MAX_COL + 1):
    cl = get_column_letter(col)
    prev = get_column_letter(col - 1)
    ws.cell(ebit_growth, col).value = f'=IF({prev}{ebit}=0,"",{cl}{ebit}/{prev}{ebit}-1)'
ws.cell(ebit_growth, 4).value = None

for col in range(4, MAX_COL + 1):
    cl = get_column_letter(col)
    ws.cell(ebit_margin, col).value = f'=IF({cl}{rev_total}+{cl}{rev_other_total}=0,"",{cl}{ebit}/({cl}{rev_total}+{cl}{rev_other_total}))'

# ─── Fix col Q special formulas ───
# Col Q has INDEX/MATCH formulas pulling from HY & Segments sheet.
# These need their $A references to match the new keys.
# The formulas already reference $A7, $A8, etc. which contain the new keys.
# The old key references in MATCH should auto-resolve because they use $A{row} which now has the new key.
# But let me verify the keys that the Q formulas reference match what's in col A.

# Check WC Change Q formula - needs to reference correct BS rows
for col in range(MAX_COL, MAX_COL + 1):  # Just col Q
    cl = get_column_letter(col)
    prev = get_column_letter(col - 1)
    pprev = get_column_letter(col - 2)
    # WC Change for forecast = -(change in receivables) - (change in inventory) + (change in payables)
    ws.cell(cf_wc, col).value = f'=-({prev}{bs_recv}-{pprev}{bs_recv})-({prev}{bs_inv}-{pprev}{bs_inv})+({prev}{bs_pay}-{pprev}{bs_pay})'

# BS forecast formulas in col Q need to reference correct rows
# Cash = prior cash + net change in cash
ws.cell(bs_cash, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{bs_cash}+{get_column_letter(MAX_COL-1)}{net_cash}'

# Trade Receivables = Revenue * Recv/Rev ratio
ws.cell(bs_recv, MAX_COL).value = f'=({get_column_letter(MAX_COL-1)}{rev_total}+{get_column_letter(MAX_COL-1)}{rev_other_total})*{get_column_letter(MAX_COL-1)}{recv_rev}'

# Inventories = Revenue * Inv/Rev ratio
ws.cell(bs_inv, MAX_COL).value = f'=({get_column_letter(MAX_COL-1)}{rev_total}+{get_column_letter(MAX_COL-1)}{rev_other_total})*{get_column_letter(MAX_COL-1)}{inv_rev}'

# PPE = prior PPE - Capex + D&A PPE (depreciation is negative)
da_ppe = r('DA-Depreciation PPE')
ws.cell(bs_ppe, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{bs_ppe}+{get_column_letter(MAX_COL-1)}{cf_capex_ppe}+{get_column_letter(MAX_COL-1)}{da_ppe}'

# Intangibles = prior
ws.cell(bs_intang, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{bs_intang}'

# ROU = prior + new leases + ROU depreciation
new_lease = r('label:New Lease Additions')
da_rou = r('DA-ROU Amortisation')
ws.cell(bs_rou, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{bs_rou}+{get_column_letter(MAX_COL-1)}{new_lease}+{get_column_letter(MAX_COL-1)}{da_rou}'

# Other Assets = prior
ws.cell(bs_other_a, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{bs_other_a}'

# Trade Payables = Revenue * Pay/Rev ratio
ws.cell(bs_pay, MAX_COL).value = f'=({get_column_letter(MAX_COL-1)}{rev_total}+{get_column_letter(MAX_COL-1)}{rev_other_total})*{get_column_letter(MAX_COL-1)}{pay_rev}'

# Other Liabilities = prior
ws.cell(bs_other_l, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{bs_other_l}'

# Lease Liabilities = prior + new leases + lease principal
ws.cell(bs_lease_l, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{bs_lease_l}+{get_column_letter(MAX_COL-1)}{new_lease}+{get_column_letter(MAX_COL-1)}{cf_lease_p}'

# Banking Debt = prior + change in debt
ws.cell(bs_debt, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{bs_debt}+{get_column_letter(MAX_COL-1)}{cf_debt}'

# Issued Capital = prior + share issues
ws.cell(bs_ic, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{bs_ic}+{get_column_letter(MAX_COL-1)}{cf_shares}'

# Retained Profits = prior + Stat NPAT - dividends
ws.cell(bs_rp, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{bs_rp}+{get_column_letter(MAX_COL-1)}{npat_s}-{get_column_letter(MAX_COL-1)}{div_total}'

# Reserves = prior
ws.cell(bs_res, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{bs_res}'

# Minorities = prior - NCI
ws.cell(bs_min, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{bs_min}-{get_column_letter(MAX_COL-1)}{nci}'

# New Lease Additions for forecast
ws.cell(new_lease, MAX_COL).value = f'={get_column_letter(MAX_COL-2)}{new_lease}'

# ─── Fix CF forecast formulas in col Q ───
# Interest Received = Term Deposit Interest
int_td = r('Int-Term Deposit')
ws.cell(cf_int_rec, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{int_td}'

# Interest Paid = Other Finance Costs
int_oc = r('Int-Other Costs')
ws.cell(cf_int_paid, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{int_oc}'

# Lease Interest Paid = Lease Liability Interest
int_ll = r('Int-Lease Interest')
ws.cell(cf_lease_int, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{int_ll}'

# Tax Paid = Tax Expense + (Sig Items AT - Other Non-Recurring)
ws.cell(cf_tax, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{tax}+({get_column_letter(MAX_COL-1)}{sig_at}-{get_column_letter(MAX_COL-1)}{r("Stat-Other Costs")})'

# Capex PPE forecast
ws.cell(cf_capex_ppe, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{capex_sales}*({get_column_letter(MAX_COL-1)}{rev_total}+{get_column_letter(MAX_COL-1)}{rev_other_total})'

# Capex Intangibles = 0 for GYG
ws.cell(cf_capex_int, MAX_COL).value = '=0'

# Acquisitions = 0
ws.cell(cf_acq, MAX_COL).value = '=0'

# Asset Sales = 0
ws.cell(cf_asset, MAX_COL).value = '=0'

# Dividends = -Total Dividends
ws.cell(cf_div, MAX_COL).value = f'=-{get_column_letter(MAX_COL-1)}{div_total}'

# Lease Principal = -prior lease liabilities / avg lease life
avg_lease = r('label:Avg Lease Life')
ws.cell(cf_lease_p, MAX_COL).value = f'=-{get_column_letter(MAX_COL-2)}{bs_lease_l}/{get_column_letter(MAX_COL-1)}{avg_lease}'

# Debt Change = 0
ws.cell(cf_debt, MAX_COL).value = '=0'

# YE Shares forecast
ws.cell(ye_shares, MAX_COL).value = f'={get_column_letter(MAX_COL-2)}{ye_shares}+IF(Value!$C$4=0,0,{get_column_letter(MAX_COL-1)}{cf_shares}/Value!$C$4)'

# WASO Basic forecast
ws.cell(waso_b, MAX_COL).value = f'=IF({get_column_letter(MAX_COL-1)}{ye_shares}=0,"",AVERAGE({get_column_letter(MAX_COL-2)}{ye_shares},{get_column_letter(MAX_COL-1)}{ye_shares}))'

# WASO Diluted forecast
ws.cell(waso_d, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{waso_b}+{get_column_letter(MAX_COL-1)}{dilution}'

# SBP forecast = 0
stat_sbp = r('Stat-SBP')
ws.cell(stat_sbp, MAX_COL).value = '=0'

# DPS forecast = payout * EPS
ws.cell(dps, MAX_COL).value = f'={get_column_letter(MAX_COL-1)}{payout}*{get_column_letter(MAX_COL-1)}{eps_u}'

# Sig Items AT forecast = 0
ws.cell(sig_at, MAX_COL).value = '=0'

# Interest rate formulas in col Q
td_rate = r('label:Term Deposit Rate')
lr_rate = r('label:Lease Receivable Rate')
ll_rate = r('label:Lease Liability Rate')
ws.cell(td_rate, MAX_COL).value = f'=IF({get_column_letter(MAX_COL-2)}{bs_td}=0,"",{get_column_letter(MAX_COL-1)}{int_td}/{get_column_letter(MAX_COL-2)}{bs_td})'

int_lr = r('Int-Lease Receivable')
ws.cell(lr_rate, MAX_COL).value = f'={get_column_letter(MAX_COL-2)}{lr_rate}'

ws.cell(ll_rate, MAX_COL).value = f'=IF(({get_column_letter(MAX_COL-2)}{bs_lease_l}+{get_column_letter(MAX_COL-1)}{bs_lease_l})=0,"",-{get_column_letter(MAX_COL-1)}{int_ll}/(({get_column_letter(MAX_COL-2)}{bs_lease_l}+{get_column_letter(MAX_COL-1)}{bs_lease_l})/2))'

# Avg Lease Life forecast = prior
ws.cell(avg_lease, MAX_COL).value = f'={get_column_letter(MAX_COL-2)}{avg_lease}'

# Network Sales Growth
ns_growth = r('label:Network Sales Growth')
gns = r('KPI-Global Network Sales')
for col in range(5, MAX_COL + 1):
    cl = get_column_letter(col)
    prev = get_column_letter(col - 1)
    ws.cell(ns_growth, col).value = f'=IF({prev}{gns}=0,"",{cl}{gns}/{prev}{gns}-1)'
ws.cell(ns_growth, 4).value = None

# ─── Clear Significant Items formula for actuals (it's a plug) ───
for col in range(4, 8):  # D through G (actuals)
    ws.cell(cf_sig, col).value = None
    ws.cell(cf_other_cfi, col).value = None
    ws.cell(cf_other_cff, col).value = None

print("\nAll formulas fixed!")

# Save
wb.save(SRC)
print(f"Saved to {SRC}")

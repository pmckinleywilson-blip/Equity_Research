"""Part 4: Wire formulas on Annual sheet - calculated rows, BS, CF, cross-sheet."""
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter as gcl

DST = '/home/pmwilson/Project_Equities/OCL/Models/OCL Model.xlsx'
wb = openpyxl.load_workbook(DST)

BLUE = Font(color='FF0000CC')
BLUE_BOLD = Font(color='FF0000CC', bold=True)
def set_blue(ws, row, col, value, bold=False):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.font = BLUE_BOLD if bold else BLUE
ws = wb['Annual']

BOLD = Font(bold=True)
THIN_BOTH = Border(top=Side(style='thin'), bottom=Side(style='thin'))

# Column mapping: D=4(FY21) to M=13(FY30)
# Actuals: D-H (4-8), FY21-FY25
# Forecasts: I-M (9-13), FY26E-FY30E
ACT_COLS = range(4, 9)   # D-H
FCST_COLS = range(9, 14)  # I-M
ALL_COLS = range(4, 14)   # D-M

# Helper to generate INDEX/MATCH formula for flow items (1H+2H)
def im_flow(key, col):
    c = gcl(col)
    return (f"=INDEX('HY & Segments'!$A:$W,MATCH($A{{}},'HY & Segments'!$A:$A,0),"
            f"MATCH(\"1H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
            f"+INDEX('HY & Segments'!$A:$W,MATCH($A{{}},'HY & Segments'!$A:$A,0),"
            f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))")

def im_pit(key, col):
    """Point-in-time: 2H only"""
    c = gcl(col)
    return (f"=INDEX('HY & Segments'!$A:$W,MATCH($A{{}},'HY & Segments'!$A:$A,0),"
            f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))")

# ==========================================
# P&L FORMULAS (both actuals and forecasts)
# ==========================================

for col in ALL_COLS:
    c = gcl(col)
    prev = gcl(col - 1) if col > 4 else None
    
    # Row 11: Total Revenue = sum of 3 segments + interest income
    ws.cell(row=11, column=col, value=f'={c}7+{c}8+{c}9+{c}10')
    ws.cell(row=11, column=col).font = BOLD
    ws.cell(row=11, column=col).border = THIN_BOTH
    
    # Row 12: Revenue Growth
    if col > 4:
        ws.cell(row=12, column=col, value=f'=IF({prev}11=0,"",{c}11/{prev}11-1)')
    
    # Row 18: GP = Total Rev (excl interest) - COGS = (Rev7+Rev8+Rev9) + COGS15
    ws.cell(row=18, column=col, value=f'={c}7+{c}8+{c}9+{c}15')
    ws.cell(row=18, column=col).font = BOLD
    ws.cell(row=18, column=col).border = THIN_BOTH
    
    # Row 19: GP Growth
    if col > 4:
        ws.cell(row=19, column=col, value=f'=IF({prev}18=0,"",{c}18/{prev}18-1)')
    
    # Row 20: GP Margin (as % of contract revenue, excl interest)
    ws.cell(row=20, column=col, value=f'=IF(({c}7+{c}8+{c}9)=0,"",{c}18/({c}7+{c}8+{c}9))')
    
    # Row 26: Total OpEx
    ws.cell(row=26, column=col, value=f'=SUM({c}23:{c}25)')
    ws.cell(row=26, column=col).font = BOLD
    ws.cell(row=26, column=col).border = THIN_BOTH
    
    # Row 27: OpEx Growth
    if col > 4:
        ws.cell(row=27, column=col, value=f'=IF({prev}26=0,"",{c}26/{prev}26-1)')
    
    # Row 30: Underlying EBITDA = GP + Total OpEx (= GP - |OpEx|)
    ws.cell(row=30, column=col, value=f'={c}18+{c}26')
    ws.cell(row=30, column=col).font = BOLD
    ws.cell(row=30, column=col).border = THIN_BOTH
    
    # Row 31: EBITDA Growth
    if col > 4:
        ws.cell(row=31, column=col, value=f'=IF({prev}30=0,"",{c}30/{prev}30-1)')
    
    # Row 32: EBITDA Margin
    ws.cell(row=32, column=col, value=f'=IF({c}11=0,"",{c}30/{c}11)')
    
    # Row 38: Statutory EBITDA
    ws.cell(row=38, column=col, value=f'={c}30+{c}35+{c}36+{c}37')
    ws.cell(row=38, column=col).font = BOLD
    ws.cell(row=38, column=col).border = THIN_BOTH
    
    # Row 44: Total D&A
    ws.cell(row=44, column=col, value=f'={c}41+{c}42+{c}43')
    ws.cell(row=44, column=col).font = BOLD
    ws.cell(row=44, column=col).border = THIN_BOTH
    
    # Row 45: D&A / Revenue
    ws.cell(row=45, column=col, value=f'=IF({c}11=0,"",{c}44/{c}11)')
    
    # Row 48: Underlying EBIT = EBITDA + D&A
    ws.cell(row=48, column=col, value=f'={c}30+{c}44')
    ws.cell(row=48, column=col).font = BOLD
    ws.cell(row=48, column=col).border = THIN_BOTH
    
    # Row 49: EBIT Growth
    if col > 4:
        ws.cell(row=49, column=col, value=f'=IF({prev}48=0,"",{c}48/{prev}48-1)')
    
    # Row 50: EBIT Margin
    ws.cell(row=50, column=col, value=f'=IF({c}11=0,"",{c}48/{c}11)')
    
    # Row 55: Net Finance Costs = Interest Income + Lease Interest
    ws.cell(row=55, column=col, value=f'={c}53+{c}54')
    ws.cell(row=55, column=col).font = BOLD
    ws.cell(row=55, column=col).border = THIN_BOTH
    
    # Row 58: PBT = EBIT + Net Finance
    ws.cell(row=58, column=col, value=f'={c}48+{c}55')
    ws.cell(row=58, column=col).font = BOLD
    ws.cell(row=58, column=col).border = THIN_BOTH
    
    # Row 60: Tax Rate
    ws.cell(row=60, column=col, value=f'=IF({c}58=0,"",-{c}59/{c}58)')
    
    # Row 61: Underlying NPAT = PBT + Tax
    ws.cell(row=61, column=col, value=f'={c}58+{c}59')
    ws.cell(row=61, column=col).font = BOLD
    ws.cell(row=61, column=col).border = THIN_BOTH
    
    # Row 63: Statutory NPAT = Underlying + Other Items AT
    ws.cell(row=63, column=col, value=f'={c}61+{c}62')
    ws.cell(row=63, column=col).font = BOLD
    ws.cell(row=63, column=col).border = THIN_BOTH
    
    # Row 64: NPAT Growth
    if col > 4:
        ws.cell(row=64, column=col, value=f'=IF({prev}61=0,"",{c}61/{prev}61-1)')
    
    # Row 65: NPAT Margin
    ws.cell(row=65, column=col, value=f'=IF({c}11=0,"",{c}61/{c}11)')

# EPS formulas (all columns)
for col in ALL_COLS:
    c = gcl(col)
    prev = gcl(col - 1) if col > 4 else None
    
    # Row 71: WASO Diluted = WASO + Dilution
    ws.cell(row=71, column=col, value=f'={c}69+{c}70')
    
    # Row 73: Underlying EPS
    ws.cell(row=73, column=col, value=f'=IF({c}71=0,"",{c}61/{c}71)')
    
    # Row 74: Statutory EPS
    ws.cell(row=74, column=col, value=f'=IF({c}71=0,"",{c}63/{c}71)')
    
    # Row 75: EPS Growth
    if col > 4:
        ws.cell(row=75, column=col, value=f'=IF({prev}73=0,"",{c}73/{prev}73-1)')
    
    # Row 78: Total Dividends = DPS * WASO
    ws.cell(row=78, column=col, value=f'={c}77*{c}69')
    
    # Row 79: Payout Ratio
    ws.cell(row=79, column=col, value=f'=IF({c}73=0,"",{c}77/{c}73)')
    
    # Row 80: Dividend Yield
    ws.cell(row=80, column=col, value=f'=IF(Value!$C$4=0,"",{c}77/Value!$C$4)')
    
    # Row 81: Dividend Growth
    if col > 4:
        ws.cell(row=81, column=col, value=f'=IF({prev}77=0,"",{c}77/{prev}77-1)')

# KPI formulas
for col in ALL_COLS:
    c = gcl(col)
    prev = gcl(col - 1) if col > 4 else None
    
    # Row 87: Total ARR = sum
    ws.cell(row=87, column=col, value=f'={c}84+{c}85+{c}86')
    ws.cell(row=87, column=col).font = BOLD
    ws.cell(row=87, column=col).border = THIN_BOTH
    
    # Row 88: ARR Growth
    if col > 4:
        ws.cell(row=88, column=col, value=f'=IF({prev}87=0,"",{c}87/{prev}87-1)')
    
    # Row 91: R&D Cap Rate
    ws.cell(row=91, column=col, value=f'=IF({c}89=0,"",{c}90/{c}89)')
    
    # Row 92: R&D / Revenue
    ws.cell(row=92, column=col, value=f'=IF({c}11=0,"",{c}89/{c}11)')
    
    # Row 93: Recurring Revenue %
    ws.cell(row=93, column=col, value=f'=IF({c}11=0,"",({c}11-{c}10)/{c}11)')

# ==========================================
# BALANCE SHEET FORMULAS
# ==========================================
for col in ALL_COLS:
    c = gcl(col)
    
    # Row 107: Total Assets
    ws.cell(row=107, column=col, value=f'=SUM({c}99:{c}106)')
    ws.cell(row=107, column=col).font = BOLD
    ws.cell(row=107, column=col).border = THIN_BOTH
    
    # Row 108: Receivables / Revenue
    ws.cell(row=108, column=col, value=f'=IF({c}11=0,"",{c}100/{c}11)')
    
    # Row 109: Working Capital = Receivables + Contract Assets - Payables - Contract Liabilities
    ws.cell(row=109, column=col, value=f'={c}100+{c}101-{c}114-{c}115')
    
    # Row 110: Payables / Revenue
    ws.cell(row=110, column=col, value=f'=IF({c}11=0,"",{c}114/{c}11)')
    
    # Row 120: Total Liabilities
    ws.cell(row=120, column=col, value=f'=SUM({c}114:{c}119)')
    ws.cell(row=120, column=col).font = BOLD
    ws.cell(row=120, column=col).border = THIN_BOTH
    
    # Row 122: Net Cash (positive = net cash position)
    ws.cell(row=122, column=col, value=f'={c}99')
    
    # Row 123: Adj Net Debt (incl leases) = -Cash + Leases
    ws.cell(row=123, column=col, value=f'=-{c}99+{c}119')
    
    # Row 124: Gearing
    ws.cell(row=124, column=col, value=f'=IF((-{c}99+{c}119+{c}130)=0,"",(-{c}99+{c}119)/(-{c}99+{c}119+{c}130))')
    
    # Row 130: Total Equity
    ws.cell(row=130, column=col, value=f'=SUM({c}127:{c}129)')
    ws.cell(row=130, column=col).font = BOLD
    ws.cell(row=130, column=col).border = THIN_BOTH
    
    # Row 131: ROE
    ws.cell(row=131, column=col, value=f'=IF({c}130=0,"",{c}61/{c}130)')
    
    # Row 132: P/B
    ws.cell(row=132, column=col, value=f'=IF(OR({c}130=0,Value!$C$4=0),"",Value!$C$4*{c}68/{c}130)')
    
    # Row 133: BS Check
    ws.cell(row=133, column=col, value=f'={c}107-{c}120-{c}130')

# BS FORECAST roll-forwards (cols I-M = 9-13)
for col in FCST_COLS:
    c = gcl(col)
    p = gcl(col - 1)
    
    # Row 99: Cash = prior + net change
    ws.cell(row=99, column=col, value=f'={p}99+{c}163')
    ws.cell(row=99, column=col).font = Font()  # black formula
    
    # Row 100: Trade Receivables = Recv/Rev ratio * Revenue
    ws.cell(row=100, column=col, value=f'={c}11*{c}108')
    ws.cell(row=100, column=col).font = Font()
    
    # Row 101: Contract Assets = carry forward
    ws.cell(row=101, column=col, value=f'={p}101')
    ws.cell(row=101, column=col).font = Font()
    
    # Row 102: Current Tax = carry forward
    ws.cell(row=102, column=col, value=f'={p}102')
    ws.cell(row=102, column=col).font = Font()
    
    # Row 103: PPE = prior - capex + depreciation (capex negative, dep negative)
    ws.cell(row=103, column=col, value=f'={p}103-{c}149+{c}41')
    ws.cell(row=103, column=col).font = Font()
    
    # Row 104: Intangibles = prior + capitalised dev - amort dev + acquisitions
    ws.cell(row=104, column=col, value=f'={p}104-{c}151+{c}43-{c}152')
    ws.cell(row=104, column=col).font = Font()
    
    # Row 105: ROU = prior + new leases + ROU amort (amort is negative)
    ws.cell(row=105, column=col, value=f'={p}105+{c}111+{c}42')
    ws.cell(row=105, column=col).font = Font()
    
    # Row 106: Other Assets = carry forward
    ws.cell(row=106, column=col, value=f'={p}106')
    ws.cell(row=106, column=col).font = Font()
    
    # Row 108: Receivables/Revenue ratio = carry forward
    ws.cell(row=108, column=col, value=f'={p}108')
    
    # Row 111: New Lease Additions = carry forward
    ws.cell(row=111, column=col, value=f'={p}111')
    
    # Row 114: Trade Payables = Payables/Rev ratio * Revenue
    ws.cell(row=114, column=col, value=f'={c}11*{c}110')
    ws.cell(row=114, column=col).font = Font()
    
    # Row 110: Payables/Revenue ratio = carry forward
    ws.cell(row=110, column=col, value=f'={p}110')
    
    # Row 115: Contract Liabilities = carry forward ratio
    ws.cell(row=115, column=col, value=f'={p}115/{p}11*{c}11')
    ws.cell(row=115, column=col).font = Font()
    
    # Row 116: Deferred Tax = carry forward
    ws.cell(row=116, column=col, value=f'={p}116')
    ws.cell(row=116, column=col).font = Font()
    
    # Row 117: Provisions = carry forward
    ws.cell(row=117, column=col, value=f'={p}117')
    ws.cell(row=117, column=col).font = Font()
    
    # Row 118: Other Liabilities = carry forward
    ws.cell(row=118, column=col, value=f'={p}118')
    ws.cell(row=118, column=col).font = Font()
    
    # Row 119: Lease Liabilities = prior + new leases + lease principal (principal is negative)
    ws.cell(row=119, column=col, value=f'={p}119+{c}111+{c}159')
    ws.cell(row=119, column=col).font = Font()
    
    # Row 127: Issued Capital = prior + share issues
    ws.cell(row=127, column=col, value=f'={p}127+{c}158')
    ws.cell(row=127, column=col).font = Font()
    
    # Row 128: Retained Profits = prior + stat NPAT - dividends
    ws.cell(row=128, column=col, value=f'={p}128+{c}63+{c}157')
    ws.cell(row=128, column=col).font = Font()
    
    # Row 129: Reserves = carry forward
    ws.cell(row=129, column=col, value=f'={p}129')
    ws.cell(row=129, column=col).font = Font()

# ==========================================
# CASH FLOW FORMULAS
# ==========================================
for col in ALL_COLS:
    c = gcl(col)
    p = gcl(col - 1) if col > 4 else None
    
    # Row 137: CF-EBITDA = link to EBITDA
    ws.cell(row=137, column=col, value=f'={c}30')
    
    # Row 150: Capex / Sales
    ws.cell(row=150, column=col, value=f'=IF({c}11=0,"",{c}149/{c}11)')

# Actual CF: rows 140, 144, 154, 161, 163 already hardcoded
# Gross OCF and totals for actuals calculated from hardcoded components
for col in ACT_COLS:
    c = gcl(col)
    # Row 140: Gross OCF = Net OCF - Int Recv - Lease Int - Tax
    ws.cell(row=140, column=col, value=f'={c}144-{c}141-{c}142-{c}143')
    ws.cell(row=140, column=col).font = BOLD
    ws.cell(row=140, column=col).border = THIN_BOTH
    
    # Row 146: EBITDA Cash Conversion = Gross OCF / EBITDA
    ws.cell(row=146, column=col, value=f'=IF({c}137=0,"",{c}140/{c}137)')

# Forecast CF formulas
for col in FCST_COLS:
    c = gcl(col)
    p = gcl(col - 1)
    
    # Row 138: WC Change = -(Recv change) - (Contract Asset change) + (Payables change) + (Contract Liab change)
    ws.cell(row=138, column=col, value=f'=-({c}100-{p}100)-({c}101-{p}101)+({c}114-{p}114)+({c}115-{p}115)')
    ws.cell(row=138, column=col).font = Font()
    
    # Row 139: Non-Cash Items = SBP
    ws.cell(row=139, column=col, value=f'=-{c}35')
    ws.cell(row=139, column=col).font = Font()
    
    # Row 140: Gross OCF
    ws.cell(row=140, column=col, value=f'=SUM({c}137:{c}139)')
    ws.cell(row=140, column=col).font = BOLD
    ws.cell(row=140, column=col).border = THIN_BOTH
    
    # Row 141: Interest Received = Interest Income
    ws.cell(row=141, column=col, value=f'={c}53')
    ws.cell(row=141, column=col).font = Font()
    
    # Row 142: Lease Int Paid = Lease Interest
    ws.cell(row=142, column=col, value=f'={c}54')
    ws.cell(row=142, column=col).font = Font()
    
    # Row 143: Tax Paid = Tax Expense + (Other Items AT - M&A - FX pre-tax impact)
    ws.cell(row=143, column=col, value=f'={c}59')
    ws.cell(row=143, column=col).font = Font()
    
    # Row 144: Net OCF
    ws.cell(row=144, column=col, value=f'={c}140+{c}141+{c}142+{c}143')
    ws.cell(row=144, column=col).font = BOLD
    ws.cell(row=144, column=col).border = THIN_BOTH
    
    # Row 145: OCF Growth
    ws.cell(row=145, column=col, value=f'=IF({p}144=0,"",{c}144/{p}144-1)')
    
    # Row 146: EBITDA Cash Conversion
    ws.cell(row=146, column=col, value=f'=IF({c}137=0,"",{c}140/{c}137)')
    
    # Row 149: Capex PPE - from forecast input
    # Row 151: Capex Intang - from forecast input
    # Row 152: Acquisitions = 0 default
    ws.cell(row=152, column=col, value=0)
    ws.cell(row=152, column=col).font = Font()
    
    # Row 153: Other CFI = 0
    ws.cell(row=153, column=col, value=0)
    ws.cell(row=153, column=col).font = Font()
    
    # Row 154: Total CFI
    ws.cell(row=154, column=col, value=f'=SUM({c}149,{c}151:{c}153)')
    ws.cell(row=154, column=col).font = BOLD
    ws.cell(row=154, column=col).border = THIN_BOTH
    
    # Row 157: Dividends Paid = -Total Dividends
    ws.cell(row=157, column=col, value=f'=-{c}78')
    ws.cell(row=157, column=col).font = Font()
    
    # Row 158: Share Issues = 0
    ws.cell(row=158, column=col, value=0)
    ws.cell(row=158, column=col).font = Font()
    
    # Row 159: Lease Principal - forecast formula
    # Approximate: prior lease liab / avg lease life
    ws.cell(row=159, column=col, value=f'=-{p}119/4')  # ~4 year avg lease life
    ws.cell(row=159, column=col).font = Font()
    
    # Row 160: Other CFF = 0
    ws.cell(row=160, column=col, value=0)
    ws.cell(row=160, column=col).font = Font()
    
    # Row 161: Total CFF
    ws.cell(row=161, column=col, value=f'=SUM({c}157:{c}160)')
    ws.cell(row=161, column=col).font = BOLD
    ws.cell(row=161, column=col).border = THIN_BOTH

# Net Change in Cash and OFCF for all cols
for col in ALL_COLS:
    c = gcl(col)
    p = gcl(col - 1) if col > 4 else None
    
    # Row 163: Net Change in Cash
    ws.cell(row=163, column=col, value=f'={c}144+{c}154+{c}161')
    ws.cell(row=163, column=col).font = BOLD
    ws.cell(row=163, column=col).border = THIN_BOTH
    
    # Row 166: Net OCF
    ws.cell(row=166, column=col, value=f'={c}144')
    
    # Row 167: Net Capex
    ws.cell(row=167, column=col, value=f'={c}149+{c}151')
    
    # Row 168: Lease Principal
    ws.cell(row=168, column=col, value=f'={c}159')
    
    # Row 169: Operating Free Cash Flow
    ws.cell(row=169, column=col, value=f'={c}166+{c}167+{c}168')
    ws.cell(row=169, column=col).font = BOLD
    ws.cell(row=169, column=col).border = THIN_BOTH
    
    # Row 170: FCF per Share
    ws.cell(row=170, column=col, value=f'=IF({c}71=0,"",{c}169/{c}71)')
    
    # Row 171: FCF Yield
    ws.cell(row=171, column=col, value=f'=IF(Value!$C$4=0,"",{c}170/Value!$C$4)')
    
    # Row 172: FCF Margin
    ws.cell(row=172, column=col, value=f'=IF({c}11=0,"",{c}169/{c}11)')
    
    # Row 175: Invested Capital = Equity + Net Debt (incl leases)
    ws.cell(row=175, column=col, value=f'={c}130+{c}123')
    
    # Row 176: Underlying EBIT
    ws.cell(row=176, column=col, value=f'={c}48')
    
    # Row 177: ROFE
    ws.cell(row=177, column=col, value=f'=IF({c}175=0,"",{c}176/{c}175)')
    
    # Row 178: NOPAT
    ws.cell(row=178, column=col, value=f'={c}176*(1-{c}60)')
    
    # Row 179: ROIC
    ws.cell(row=179, column=col, value=f'=IF({c}175=0,"",{c}178/{c}175)')

# ==========================================
# CROSS-SHEET INDEX/MATCH for forecast columns
# ==========================================
# Flow items on P&L pull from HY via INDEX/MATCH for forecast years
flow_rows = [7, 8, 9, 10, 15, 23, 24, 25, 35, 36, 37, 41, 42, 43, 53, 54, 59, 62]
pit_rows = [84, 85, 86, 94]  # Point-in-time KPIs

for col in FCST_COLS:
    c = gcl(col)
    
    for row in flow_rows:
        formula = (f"=INDEX('HY & Segments'!$A:$W,MATCH($A{row},'HY & Segments'!$A:$A,0),"
                   f"MATCH(\"1H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
                   f"+INDEX('HY & Segments'!$A:$W,MATCH($A{row},'HY & Segments'!$A:$A,0),"
                   f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))")
        ws.cell(row=row, column=col, value=formula)
        ws.cell(row=row, column=col).font = Font()
    
    for row in pit_rows:
        formula = (f"=INDEX('HY & Segments'!$A:$W,MATCH($A{row},'HY & Segments'!$A:$A,0),"
                   f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))")
        ws.cell(row=row, column=col, value=formula)
        ws.cell(row=row, column=col).font = Font()

# Forecast EPS rows - shares & DPS
for col in FCST_COLS:
    c = gcl(col)
    p = gcl(col - 1)
    
    # Row 68: YE Shares = carry forward
    ws.cell(row=68, column=col, value=f'={p}68')
    ws.cell(row=68, column=col).font = Font()
    
    # Row 69: WASO = average of prior YE and current YE
    ws.cell(row=69, column=col, value=f'=AVERAGE({p}68,{c}68)')
    ws.cell(row=69, column=col).font = Font()
    
    # Row 70: Dilution = carry forward
    ws.cell(row=70, column=col, value=f'={p}70')
    ws.cell(row=70, column=col).font = Font()
    
    # Row 77: DPS = carry forward payout * EPS
    ws.cell(row=77, column=col, value=f'={c}79*{c}73')
    ws.cell(row=77, column=col).font = Font()
    
    # Row 79: Payout ratio = carry forward
    # Already set above

# Forecast KPI rows
for col in FCST_COLS:
    c = gcl(col)
    p = gcl(col - 1)
    
    # Row 89: Total R&D = from HY (flow)
    formula = (f"=INDEX('HY & Segments'!$A:$W,MATCH($A89,'HY & Segments'!$A:$A,0),"
               f"MATCH(\"1H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
               f"+INDEX('HY & Segments'!$A:$W,MATCH($A89,'HY & Segments'!$A:$A,0),"
               f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))")
    ws.cell(row=89, column=col, value=formula)
    ws.cell(row=89, column=col).font = Font()
    
    # Row 90: Capitalised Dev = from HY (flow)
    formula = (f"=INDEX('HY & Segments'!$A:$W,MATCH($A90,'HY & Segments'!$A:$A,0),"
               f"MATCH(\"1H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
               f"+INDEX('HY & Segments'!$A:$W,MATCH($A90,'HY & Segments'!$A:$A,0),"
               f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))")
    ws.cell(row=90, column=col, value=formula)
    ws.cell(row=90, column=col).font = Font()
    
    # Row 95: WASO = from HY (flow)
    formula = (f"=INDEX('HY & Segments'!$A:$W,MATCH($A95,'HY & Segments'!$A:$A,0),"
               f"MATCH(\"1H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))"
               f"+INDEX('HY & Segments'!$A:$W,MATCH($A95,'HY & Segments'!$A:$A,0),"
               f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0))")
    ws.cell(row=95, column=col, value=formula)
    ws.cell(row=95, column=col).font = Font()

# Forecast capex from HY
for col in FCST_COLS:
    c = gcl(col)
    
    # Row 149: Capex PPE from HY forecast
    formula = (f"=INDEX('HY & Segments'!$B:$W,MATCH(\"Capex PPE\",'HY & Segments'!$B:$B,0),"
               f"MATCH(\"1H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0)-1)"
               f"+INDEX('HY & Segments'!$B:$W,MATCH(\"Capex PPE\",'HY & Segments'!$B:$B,0),"
               f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0)-1)")
    ws.cell(row=149, column=col, value=formula)
    ws.cell(row=149, column=col).font = Font()
    
    # Row 151: Capitalised Dev from KPI
    ws.cell(row=151, column=col, value=f'=-{c}90')
    ws.cell(row=151, column=col).font = Font()
    
    # Row 111: New Lease Additions from HY
    formula = (f"=INDEX('HY & Segments'!$B:$W,MATCH(\"New Lease Additions\",'HY & Segments'!$B:$B,0),"
               f"MATCH(\"1H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0)-1)"
               f"+INDEX('HY & Segments'!$B:$W,MATCH(\"New Lease Additions\",'HY & Segments'!$B:$B,0),"
               f"MATCH(\"2H\"&RIGHT({c}$1,2),'HY & Segments'!$3:$3,0)-1)")
    ws.cell(row=111, column=col, value=formula)
    ws.cell(row=111, column=col).font = Font()

# Set actual BS ratios 
for col in ACT_COLS:
    c = gcl(col)
    # Row 111: New Lease Additions (actual) - calculate from BS movement
    if col > 4:
        p = gcl(col - 1)
        # New leases = ROU end - ROU start - ROU amort (amort is negative so +)
        ws.cell(row=111, column=col, value=f'={c}105-{p}105-{c}42')
    else:
        set_blue(ws, 111, col, 2.0)  # FY21 estimate

wb.save(DST)
print('Part 4 complete: Annual sheet formulas wired')
print('  - P&L cascade, EPS, KPIs')
print('  - BS roll-forwards')
print('  - CF formulas')
print('  - Cross-sheet INDEX/MATCH for forecasts')

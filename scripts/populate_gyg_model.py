#!/usr/bin/env python3
"""
Populate ALL historical data into the GYG financial model.
Uses openpyxl to load, enter data, format, and save.
"""

import openpyxl
from openpyxl.styles import Font, Border, Side, numbers
from copy import copy

MODEL_PATH = '/home/pmwilson/Project_Equities/GYG/Models/GYG Model.xlsx'

# Styles
BLUE_FONT = Font(color='0000CC')
BLUE_BOLD = Font(color='0000CC', bold=True)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

FMT_AM = '#,##0.0'
FMT_PCT = '0.0%'
FMT_EPS = '0.000'
FMT_COUNT = '#,##0'
FMT_DECIMAL = '0.0'


def set_cell(ws, row, col, value, fmt=None, bold=False, is_formula=False, subtotal=False):
    """Set a cell's value with blue font and optional formatting."""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if is_formula:
        # Formulas get blue font but no number format override (Excel handles it)
        cell.font = BLUE_BOLD if (bold or subtotal) else BLUE_FONT
    else:
        cell.font = BLUE_BOLD if (bold or subtotal) else BLUE_FONT
    if fmt and not is_formula:
        cell.number_format = fmt
    elif fmt and is_formula:
        cell.number_format = fmt
    if subtotal:
        cell.font = BLUE_BOLD
        cell.border = THIN_BORDER


def enter_row_data(ws, row, col_map, values, fmt=FMT_AM, bold=False, subtotal=False):
    """Enter a dict or list of values into a row. col_map maps index to column."""
    for i, val in enumerate(values):
        if val is not None:
            is_formula = isinstance(val, str) and val.startswith('=')
            set_cell(ws, row, col_map[i], val, fmt=fmt, bold=bold,
                     is_formula=is_formula, subtotal=subtotal)


def main():
    wb = openpyxl.load_workbook(MODEL_PATH)
    ws = wb['Annual']
    hy = wb['HY & Segments']

    # Column maps
    # Annual: D=FY23(4), E=FY24(5), F=FY25(6)
    ann_cols = [4, 5, 6]  # indices 0,1,2

    # ========================================================================
    # STEP 0: INSERT 2 ROWS AT ROW 108 FOR TERM DEPOSITS & FINANCE LEASE RECV
    # ========================================================================
    ws.insert_rows(108, 2)
    # After insertion: old rows 108+ shift to 110+
    # Row 108 and 109 are new blank rows

    # Set up the new rows
    ws.cell(row=108, column=1).value = 'BS-Term Deposits'
    ws.cell(row=108, column=2).value = 'Funds in Term Deposits'
    ws.cell(row=108, column=3).value = 'A$m'

    ws.cell(row=109, column=1).value = 'BS-Finance Lease Recv'
    ws.cell(row=109, column=2).value = 'Finance Lease Receivables'
    ws.cell(row=109, column=3).value = 'A$m'

    # Update labels for existing rows (now shifted)
    ws.cell(row=105, column=2).value = 'Cash & Cash Equivalents'
    ws.cell(row=106, column=2).value = 'Trade & Other Receivables'
    # Row 107 Inventories stays
    # Row 110 PPE (shifted from 108)
    ws.cell(row=110, column=2).value = 'Property, Plant & Equipment'
    # Row 111 Intangibles (shifted from 109)
    ws.cell(row=111, column=2).value = 'Intangible Assets'
    # Row 112 ROU (shifted from 110)
    ws.cell(row=112, column=2).value = 'Right-of-Use Assets'
    # Row 113 Other Assets (shifted from 111)
    ws.cell(row=113, column=2).value = 'Other Assets (DTA, Prepayments)'

    # Row 114 Total Assets (shifted from 112) — fix formula
    # Row 115 Receivables/Revenue (shifted from 113)
    # Row 116 Inventory/Revenue (shifted from 114)
    # Row 117 Working Capital (shifted from 115)
    # Row 118 Payables/Revenue (shifted from 116)
    # Row 119 New Lease Additions (shifted from 117)

    # Row 121 Liabilities header (shifted from 119)
    # Row 122 Trade Payables (shifted from 120)
    # Row 123 Other Liabilities (shifted from 121)
    # Row 124 Lease Liabilities (shifted from 122)
    # Row 125 Borrowings (shifted from 123)
    ws.cell(row=125, column=2).value = 'Borrowings'
    # Row 126 Total Liabilities (shifted from 124)

    # Row 128 Net Banking Debt (shifted from 126)
    # Row 129 Adj Net Debt (shifted from 127)
    # Row 130 ND/EBITDA (shifted from 128)
    # Row 131 Gearing (shifted from 129)

    # Row 133 Equity header (shifted from 131)
    # Row 134 Issued Capital (shifted from 132)
    # Row 135 Retained Profits (shifted from 133)
    ws.cell(row=135, column=2).value = 'Accumulated Losses'
    # Row 136 Reserves (shifted from 134)
    # Row 137 Minorities (shifted from 135)
    # Row 138 Total Equity (shifted from 136)
    # Row 139 ROE (shifted from 137)
    # Row 140 P/B (shifted from 138)
    # Row 141 BS Check (shifted from 139)

    # ========================================================================
    # CLEAR OLD DATA in BS/CF area (cols D through P, rows 105-197)
    # ========================================================================
    for row in range(105, 200):
        for col in range(4, 17):  # D through P
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.font = Font()  # reset font
            cell.border = Border()  # reset border
            cell.number_format = 'General'

    # ========================================================================
    # STEP 1: ANNUAL SHEET P&L DATA
    # ========================================================================

    # Revenue
    enter_row_data(ws, 7, ann_cols, [211.975, 278.877, 359.722])
    enter_row_data(ws, 8, ann_cols, [5.959, 10.834, 8.869])
    enter_row_data(ws, 9, ann_cols, [217.934, 289.711, 368.591], subtotal=True)
    enter_row_data(ws, 10, ann_cols, [40.510, 51.858, 66.761])
    enter_row_data(ws, 11, ann_cols, [0.600, 0.645, 0.630])
    enter_row_data(ws, 12, ann_cols, [259.044, 342.214, 435.982], subtotal=True)

    # Revenue Growth (skip FY23)
    set_cell(ws, 13, 5, '=IF(D12=0,"",E12/D12-1)', fmt=FMT_PCT, is_formula=True)
    set_cell(ws, 13, 6, '=IF(E12=0,"",F12/E12-1)', fmt=FMT_PCT, is_formula=True)

    # Other Revenue
    enter_row_data(ws, 16, ann_cols, [13.252, 16.382, 19.688])
    enter_row_data(ws, 17, ann_cols, [1.813, 6.164, 8.942])
    enter_row_data(ws, 18, ann_cols, [2.660, 0.228, 3.394])
    enter_row_data(ws, 19, ann_cols, [17.725, 22.774, 32.024], subtotal=True)

    # Total Revenue & Other Income (row 21)
    for col in ann_cols:
        c = chr(64 + col)  # D, E, F
        set_cell(ws, 21, col, f'={c}12+{c}19', fmt=FMT_AM, is_formula=True, subtotal=True)

    # Segment EBITDA Bridge
    enter_row_data(ws, 24, ann_cols, [33.306, 47.536, 66.029])
    enter_row_data(ws, 25, ann_cols, [-4.296, -6.544, -13.223])
    enter_row_data(ws, 26, ann_cols, [29.010, 40.992, 52.806], subtotal=True)
    enter_row_data(ws, 27, ann_cols, [11.213, 17.084, 18.473])
    enter_row_data(ws, 28, ann_cols, [-4.240, -11.142, -9.018])
    enter_row_data(ws, 29, ann_cols, [-6.362, -19.661, 2.858])
    enter_row_data(ws, 30, ann_cols, [29.621, 27.273, 65.119], subtotal=True)

    # EBITDA Margin
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 31, col, f'=IF({c}21=0,"",{c}30/{c}21)', fmt=FMT_PCT, is_formula=True)

    # Expenses
    enter_row_data(ws, 34, ann_cols, [-70.428, -87.580, -113.948])
    enter_row_data(ws, 35, ann_cols, [-113.725, -153.733, -184.656])
    enter_row_data(ws, 36, ann_cols, [-35.282, -60.595, -53.615])
    enter_row_data(ws, 37, ann_cols, [-13.718, -17.938, -21.287])
    enter_row_data(ws, 38, ann_cols, [-13.995, -17.868, -29.381])
    # Total Expenses subtotal
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 39, col, f'=SUM({c}34:{c}38)', fmt=FMT_AM, is_formula=True, subtotal=True)

    # D&A
    enter_row_data(ws, 42, ann_cols, [-10.332, -11.247, -13.874])
    enter_row_data(ws, 43, ann_cols, [-8.390, -12.960, -19.296])
    enter_row_data(ws, 44, ann_cols, [-2.747, -2.935, -3.668])
    enter_row_data(ws, 45, ann_cols, [-4.086, -3.981, -2.843])
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 46, col, f'=SUM({c}42:{c}45)', fmt=FMT_AM, is_formula=True, subtotal=True)

    # EBIT
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 48, col, f'={c}30+{c}46', fmt=FMT_AM, is_formula=True, subtotal=True)

    # EBIT Margin
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 49, col, f'=IF({c}21=0,"",{c}48/{c}21)', fmt=FMT_PCT, is_formula=True)

    # Finance Income
    enter_row_data(ws, 52, ann_cols, [0.000, 1.096, 12.530])
    enter_row_data(ws, 53, ann_cols, [3.630, 4.841, 10.042])
    enter_row_data(ws, 54, ann_cols, [0.963, 0.075, 0.102])
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 55, col, f'=SUM({c}52:{c}54)', fmt=FMT_AM, is_formula=True, subtotal=True)

    # Finance Costs
    enter_row_data(ws, 56, ann_cols, [-8.199, -13.394, -18.771])
    enter_row_data(ws, 57, ann_cols, [-0.304, -0.330, -0.131])
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 58, col, f'=SUM({c}56:{c}57)', fmt=FMT_AM, is_formula=True, subtotal=True)

    # Net Finance
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 59, col, f'={c}55+{c}58', fmt=FMT_AM, is_formula=True, subtotal=True)

    # PBT
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 61, col, f'={c}48+{c}59', fmt=FMT_AM, is_formula=True, subtotal=True)

    # Tax
    enter_row_data(ws, 62, ann_cols, [-2.423, -2.187, -14.734])

    # Tax Rate
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 63, col, f'=IF({c}61=0,"",{c}62/{c}61)', fmt=FMT_PCT, is_formula=True)

    # NPAT
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 64, col, f'={c}61+{c}62', fmt=FMT_AM, is_formula=True, subtotal=True)

    # NPAT Margin
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 65, col, f'=IF({c}21=0,"",{c}64/{c}21)', fmt=FMT_PCT, is_formula=True)

    # ========================================================================
    # STEP 2: ANNUAL EPS & DIVIDENDS
    # ========================================================================
    enter_row_data(ws, 69, ann_cols, [84.916, 101.353, 101.655], fmt=FMT_DECIMAL)
    enter_row_data(ws, 70, ann_cols, [83.992, 85.568, 101.510], fmt=FMT_DECIMAL)
    enter_row_data(ws, 71, ann_cols, [0.0, 0.0, 4.395], fmt=FMT_DECIMAL)
    enter_row_data(ws, 72, ann_cols, [83.992, 85.568, 105.905], fmt=FMT_DECIMAL)

    # Basic EPS (cents) = NPAT / WASO Basic * 100
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 74, col, f'=IF({c}70=0,"",{c}64/{c}70*100)', fmt=FMT_EPS, is_formula=True)

    # Diluted EPS
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 75, col, f'=IF({c}72=0,"",{c}64/{c}72*100)', fmt=FMT_EPS, is_formula=True)

    # EPS Growth (skip FY23)
    set_cell(ws, 76, 5, '=IF(D74=0,"",E74/D74-1)', fmt=FMT_PCT, is_formula=True)
    set_cell(ws, 76, 6, '=IF(E74=0,"",F74/E74-1)', fmt=FMT_PCT, is_formula=True)

    # DPS
    enter_row_data(ws, 78, ann_cols, [0.0, 0.0, 12.6], fmt=FMT_EPS)

    # Total Dividends = DPS * WASO Basic / 100 (DPS in cents, shares in millions -> A$m)
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 79, col, f'={c}78*{c}70/100', fmt=FMT_AM, is_formula=True)

    # Payout Ratio
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 80, col, f'=IF({c}74=0,"",{c}78/{c}74)', fmt=FMT_PCT, is_formula=True)

    # Dividend Yield — reference to share price on Value sheet
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 81, col, f'=IF(Value!$C$4=0,"",{c}78/Value!$C$4/100)', fmt=FMT_PCT, is_formula=True)

    # Dividend Growth (skip FY23)
    set_cell(ws, 82, 5, '=IF(D78=0,"",E78/D78-1)', fmt=FMT_PCT, is_formula=True)
    set_cell(ws, 82, 6, '=IF(E78=0,"",F78/E78-1)', fmt=FMT_PCT, is_formula=True)

    # ========================================================================
    # STEP 3: ANNUAL OPERATING METRICS
    # ========================================================================
    enter_row_data(ws, 85, ann_cols, [55, 64, 81], fmt=FMT_COUNT)
    enter_row_data(ws, 86, ann_cols, [116, 130, 143], fmt=FMT_COUNT)
    enter_row_data(ws, 87, ann_cols, [20, 22, 26], fmt=FMT_COUNT)
    enter_row_data(ws, 88, ann_cols, [3, 4, 6], fmt=FMT_COUNT)
    enter_row_data(ws, 89, ann_cols, [194, 220, 256], fmt=FMT_COUNT)
    enter_row_data(ws, 90, ann_cols, [753.0, 948.9, 1168.5], fmt=FMT_DECIMAL)
    enter_row_data(ws, 91, ann_cols, [759.0, 959.7, 1180.7], fmt=FMT_DECIMAL)
    enter_row_data(ws, 92, ann_cols, [0.150, 0.081, 0.096], fmt=FMT_PCT)
    enter_row_data(ws, 93, ann_cols, [5.7, 6.2, 6.7], fmt=FMT_DECIMAL)
    enter_row_data(ws, 94, ann_cols, [4.5, 4.6, 5.0], fmt=FMT_DECIMAL)
    enter_row_data(ws, 95, ann_cols, [None, 0.212, 0.215], fmt=FMT_PCT)
    enter_row_data(ws, 96, ann_cols, [None, 0.194, 0.184], fmt=FMT_PCT)
    enter_row_data(ws, 97, ann_cols, [0.053, 0.060, 0.067], fmt=FMT_PCT)
    enter_row_data(ws, 98, ann_cols, [0.077, 0.083, 0.095], fmt=FMT_PCT)
    enter_row_data(ws, 99, ann_cols, [None, 0.243, 0.266], fmt=FMT_PCT)
    enter_row_data(ws, 100, ann_cols, [0.158, 0.174, 0.195], fmt=FMT_PCT)
    enter_row_data(ws, 101, ann_cols, [None, None, 18], fmt=FMT_COUNT)

    # ========================================================================
    # STEP 4: ANNUAL BALANCE SHEET (rows shifted +2)
    # ========================================================================

    # Assets
    enter_row_data(ws, 105, ann_cols, [36.504, 16.385, 39.675])
    enter_row_data(ws, 106, ann_cols, [25.087, 26.499, 24.840])
    enter_row_data(ws, 107, ann_cols, [2.153, 2.825, 3.761])
    enter_row_data(ws, 108, ann_cols, [0.0, 278.095, 242.068])
    enter_row_data(ws, 109, ann_cols, [69.333, 126.403, 174.844])
    enter_row_data(ws, 110, ann_cols, [69.486, 87.630, 130.056])
    enter_row_data(ws, 111, ann_cols, [15.202, 10.586, 18.305])
    enter_row_data(ws, 112, ann_cols, [98.939, 93.796, 125.430])
    enter_row_data(ws, 113, ann_cols, [7.803, 19.578, 24.219])

    # Total Assets
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 114, col, f'=SUM({c}105:{c}113)', fmt=FMT_AM, is_formula=True, subtotal=True)

    # BS Ratios
    # Receivables/Revenue
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 115, col, f'=IF({c}12=0,"",{c}106/{c}12)', fmt=FMT_PCT, is_formula=True)

    # Inventory/Revenue
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 116, col, f'=IF({c}12=0,"",{c}107/{c}12)', fmt=FMT_PCT, is_formula=True)

    # Working Capital = Receivables + Inventories - Trade Payables
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 117, col, f'={c}106+{c}107-{c}122', fmt=FMT_AM, is_formula=True)

    # Payables/Revenue
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 118, col, f'=IF({c}12=0,"",{c}122/{c}12)', fmt=FMT_PCT, is_formula=True)

    # New Lease Additions
    enter_row_data(ws, 119, ann_cols, [None, None, 97.279])

    # Liabilities
    enter_row_data(ws, 122, ann_cols, [32.635, 39.387, 40.439])
    enter_row_data(ws, 123, ann_cols, [19.190, 28.933, 31.324])
    enter_row_data(ws, 124, ann_cols, [181.717, 239.498, 331.311])
    enter_row_data(ws, 125, ann_cols, [3.000, 0.0, 0.0])

    # Total Liabilities
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 126, col, f'=SUM({c}122:{c}125)', fmt=FMT_AM, is_formula=True, subtotal=True)

    # Net Banking Debt = Borrowings - Cash
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 128, col, f'={c}125-{c}105', fmt=FMT_AM, is_formula=True)

    # Adj Net Debt = Net Banking Debt + Lease Liabilities
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 129, col, f'={c}128+{c}124', fmt=FMT_AM, is_formula=True)

    # ND/EBITDA
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 130, col, f'=IF({c}30=0,"",{c}128/{c}30)', fmt=FMT_DECIMAL, is_formula=True)

    # Gearing
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 131, col, f'=IF(({c}128+{c}138)=0,"",{c}128/({c}128+{c}138))', fmt=FMT_PCT, is_formula=True)

    # Equity
    enter_row_data(ws, 134, ann_cols, [104.046, 372.708, 374.988])
    enter_row_data(ws, 135, ann_cols, [-26.857, -36.147, -21.671])
    enter_row_data(ws, 136, ann_cols, [10.776, 17.418, 26.807])
    enter_row_data(ws, 137, ann_cols, [0.0, 0.0, 0.0])

    # Total Equity
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 138, col, f'=SUM({c}134:{c}137)', fmt=FMT_AM, is_formula=True, subtotal=True)

    # ROE = NPAT / Total Equity
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 139, col, f'=IF({c}138=0,"",{c}64/{c}138)', fmt=FMT_PCT, is_formula=True)

    # P/B
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 140, col, f'=IF(OR({c}138=0,Value!$C$4=0),"",Value!$C$4*{c}69/{c}138)', fmt=FMT_DECIMAL, is_formula=True)

    # BS Check = Total Assets - Total Liabilities - Total Equity
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 141, col, f'={c}114-{c}126-{c}138', fmt=FMT_AM, is_formula=True)

    # ========================================================================
    # STEP 5: ANNUAL CASH FLOW (shifted +2 from row insertion)
    # ========================================================================
    # CF section header at row 143, CFO header at 144

    # EBITDA reference
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 145, col, f'={c}30', fmt=FMT_AM, is_formula=True)

    # WC Change
    enter_row_data(ws, 146, ann_cols, [-3.806, 0.696, 1.310])

    # Non-cash/Other
    enter_row_data(ws, 147, ann_cols, [13.376, 13.834, 8.859])

    # Gross OCF = subtotal
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 148, col, f'=SUM({c}145:{c}147)', fmt=FMT_AM, is_formula=True, subtotal=True)

    # Interest Received
    enter_row_data(ws, 149, ann_cols, [0.964, 6.100, 10.185])
    # Bank Interest Paid
    enter_row_data(ws, 150, ann_cols, [-0.304, -0.329, -0.131])
    # Lease Interest Paid
    enter_row_data(ws, 151, ann_cols, [-4.606, -6.320, -8.626])
    # Tax Paid
    enter_row_data(ws, 152, ann_cols, [-0.298, -4.489, -19.389])

    # Net OCF = subtotal
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 153, col, f'={c}148+{c}149+{c}150+{c}151+{c}152', fmt=FMT_AM, is_formula=True, subtotal=True)

    # OCF Growth
    set_cell(ws, 154, 5, '=IF(D153=0,"",E153/D153-1)', fmt=FMT_PCT, is_formula=True)
    set_cell(ws, 154, 6, '=IF(E153=0,"",F153/E153-1)', fmt=FMT_PCT, is_formula=True)

    # CF Conversion = Net OCF / EBITDA
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 155, col, f'=IF({c}145=0,"",{c}153/{c}145)', fmt=FMT_PCT, is_formula=True)

    # CFI
    enter_row_data(ws, 158, ann_cols, [-39.713, -33.496, -61.326])

    # Capex/Sales
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 159, col, f'=IF({c}12=0,"",({c}158+{c}160)/{c}12)', fmt=FMT_PCT, is_formula=True)

    enter_row_data(ws, 160, ann_cols, [-2.944, -0.063, -0.037])
    enter_row_data(ws, 161, ann_cols, [-6.440, -3.128, -15.473])
    enter_row_data(ws, 162, ann_cols, [0.694, 2.964, 3.879])
    enter_row_data(ws, 163, ann_cols, [0.0, -278.095, 38.406])

    # Total ICF = subtotal
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 164, col, f'=SUM({c}158,{c}160:{c}163)', fmt=FMT_AM, is_formula=True, subtotal=True)

    # CFF
    enter_row_data(ws, 167, ann_cols, [0.0, 0.0, 0.0])
    enter_row_data(ws, 168, ann_cols, [5.128, 267.223, 1.547])
    enter_row_data(ws, 169, ann_cols, [-6.720, -9.912, -10.830])
    enter_row_data(ws, 170, ann_cols, [-0.358, -3.000, 0.0])
    enter_row_data(ws, 171, ann_cols, [1.429, 1.172, 8.534])

    # Total CFF = subtotal
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 172, col, f'=SUM({c}167:{c}171)', fmt=FMT_AM, is_formula=True, subtotal=True)

    # Net Change in Cash = Net OCF + Total ICF + Total CFF
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 174, col, f'={c}153+{c}164+{c}172', fmt=FMT_AM, is_formula=True, subtotal=True)

    # ========================================================================
    # OpFCF section (shifted +2)
    # ========================================================================
    # Row 177: Net OCF
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 177, col, f'={c}153', fmt=FMT_AM, is_formula=True)

    # Row 178: Net Capex = Capex PPE + Capex Intang
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 178, col, f'={c}158+{c}160', fmt=FMT_AM, is_formula=True)

    # Row 179: Lease Principal
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 179, col, f'={c}169', fmt=FMT_AM, is_formula=True)

    # Row 180: OpFCF = sum
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 180, col, f'=SUM({c}177:{c}179)', fmt=FMT_AM, is_formula=True, subtotal=True)

    # FCF per share
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 181, col, f'=IF({c}72=0,"",{c}180/{c}72*100)', fmt=FMT_EPS, is_formula=True)

    # FCF Yield
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 182, col, f'=IF(Value!$C$4=0,"",{c}181/Value!$C$4/100)', fmt=FMT_PCT, is_formula=True)

    # FCF Margin
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 183, col, f'=IF({c}12=0,"",{c}180/{c}12)', fmt=FMT_PCT, is_formula=True)

    # ========================================================================
    # ROIC section (shifted +2)
    # ========================================================================
    # Row 187: Invested Capital = Total Equity + Net Banking Debt
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 187, col, f'={c}138+{c}128', fmt=FMT_AM, is_formula=True)

    # Row 188: EBIT reference
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 188, col, f'={c}48', fmt=FMT_AM, is_formula=True)

    # Row 189: ROFE = EBIT / Invested Capital
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 189, col, f'=IF({c}187=0,"",{c}188/{c}187)', fmt=FMT_PCT, is_formula=True)

    # Row 190: NOPAT = EBIT * (1 - Tax Rate)
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 190, col, f'={c}188*(1-ABS({c}63))', fmt=FMT_AM, is_formula=True)

    # Row 191: ROIC = NOPAT / Invested Capital
    for col in ann_cols:
        c = chr(64 + col)
        set_cell(ws, 191, col, f'=IF({c}187=0,"",{c}190/{c}187)', fmt=FMT_PCT, is_formula=True)

    # ========================================================================
    # Update section labels that shifted +2
    # ========================================================================
    # Fix labels for BS ratio rows
    ws.cell(row=115, column=2).value = 'Receivables / Revenue'
    ws.cell(row=115, column=3).value = '%'
    ws.cell(row=116, column=2).value = 'Inventory / Revenue'
    ws.cell(row=116, column=3).value = '%'
    ws.cell(row=117, column=2).value = 'Working Capital'
    ws.cell(row=117, column=3).value = 'A$m'
    ws.cell(row=118, column=2).value = 'Payables / Revenue'
    ws.cell(row=118, column=3).value = '%'
    ws.cell(row=119, column=2).value = 'New Lease Additions'
    ws.cell(row=119, column=3).value = 'A$m'

    # Verify/fix section headers that shifted
    ws.cell(row=114, column=2).value = 'Total Assets'
    ws.cell(row=114, column=3).value = 'A$m'
    ws.cell(row=121, column=2).value = 'Liabilities'
    ws.cell(row=126, column=2).value = 'Total Liabilities'
    ws.cell(row=126, column=3).value = 'A$m'
    ws.cell(row=128, column=2).value = 'Net Banking Debt'
    ws.cell(row=128, column=3).value = 'A$m'
    ws.cell(row=129, column=2).value = 'Adj Net Debt (incl leases)'
    ws.cell(row=129, column=3).value = 'A$m'
    ws.cell(row=130, column=2).value = 'ND / EBITDA'
    ws.cell(row=130, column=3).value = 'x'
    ws.cell(row=131, column=2).value = 'Gearing (ND/(ND+E))'
    ws.cell(row=131, column=3).value = '%'
    ws.cell(row=133, column=2).value = 'Equity'
    ws.cell(row=138, column=2).value = 'Total Equity'
    ws.cell(row=138, column=3).value = 'A$m'
    ws.cell(row=139, column=2).value = 'ROE'
    ws.cell(row=139, column=3).value = '%'
    ws.cell(row=140, column=2).value = 'P/B'
    ws.cell(row=140, column=3).value = 'x'
    ws.cell(row=141, column=2).value = 'BS Check (should be 0, +/-0.2 due to rounding)'
    ws.cell(row=141, column=3).value = 'A$m'

    # CF section headers
    ws.cell(row=143, column=2).value = 'Cash Flow'
    ws.cell(row=144, column=2).value = 'CFO'
    ws.cell(row=157, column=2).value = 'CFI'
    ws.cell(row=166, column=2).value = 'CFF'
    ws.cell(row=176, column=2).value = 'Operating Free Cash Flow'

    # CF row labels (col A keys and col B labels)
    cf_labels = {
        145: ('CF-EBITDA', 'Underlying EBITDA', 'A$m'),
        146: ('CF-WC Change', 'Working Capital Change', 'A$m'),
        147: ('CF-Significant Items', 'Significant Items/Non-Cash Items', 'A$m'),
        148: (None, 'Gross Operating Cash Flow', 'A$m'),
        149: ('CF-Int Received', 'Interest Received', 'A$m'),
        150: ('CF-Interest Paid', 'Interest Paid', 'A$m'),
        151: ('CF-Lease Int Paid', 'Lease Interest Paid', 'A$m'),
        152: ('CF-Tax Paid', 'Tax Paid', 'A$m'),
        153: ('CF-Net OCF', 'Net Operating Cash Flow', 'A$m'),
        154: (None, 'OCF Growth', '% YoY'),
        155: (None, 'EBITDA Cashflow conversion', '%'),
        158: ('CF-Capex PPE', 'Capex (PPE)', 'A$m'),
        159: (None, 'Capex / Sales', '%'),
        160: ('CF-Capex Intang', 'Capex (Intangibles)', 'A$m'),
        161: ('CF-Acquisitions', 'Acquisitions', 'A$m'),
        162: ('CF-Asset Sales', 'Asset Sales', 'A$m'),
        163: ('CF-Other CFI', 'Other', 'A$m'),
        164: (None, 'Total Investing Cash Flow', 'A$m'),
        167: ('CF-Dividends', 'Dividends Paid', 'A$m'),
        168: ('CF-Share Issues', 'Share Issues / Buybacks', 'A$m'),
        169: ('CF-Lease Principal', 'Lease Principal Payments', 'A$m'),
        170: ('CF-Debt Change', 'Change in Debt', 'A$m'),
        171: ('CF-Other CFF', 'Other', 'A$m'),
        172: (None, 'Total Financing Cash Flow', 'A$m'),
        174: (None, 'Net Change in Cash', 'A$m'),
        177: (None, 'Net OCF', 'A$m'),
        178: (None, 'Net Capex', 'A$m'),
        179: (None, 'Lease Principal', 'A$m'),
        180: (None, 'Operating Free Cash Flow', 'A$m'),
        181: (None, 'FCF per Share', 'cps'),
        182: (None, 'FCF Yield', '%'),
        183: (None, 'FCF Margin', '%'),
    }
    for row, (key, label, units) in cf_labels.items():
        if key is not None:
            ws.cell(row=row, column=1).value = key
        ws.cell(row=row, column=2).value = label
        ws.cell(row=row, column=3).value = units

    # ROIC labels
    ws.cell(row=186, column=2).value = 'ROIC'
    ws.cell(row=187, column=2).value = 'Invested Capital'
    ws.cell(row=187, column=3).value = 'A$m'
    ws.cell(row=188, column=2).value = 'Underlying EBIT'
    ws.cell(row=188, column=3).value = 'A$m'
    ws.cell(row=189, column=2).value = 'ROFE'
    ws.cell(row=189, column=3).value = '%'
    ws.cell(row=190, column=2).value = 'NOPAT'
    ws.cell(row=190, column=3).value = 'A$m'
    ws.cell(row=191, column=2).value = 'ROIC'
    ws.cell(row=191, column=3).value = '%'

    # ========================================================================
    # STEP 6: HY & SEGMENTS SHEET DATA
    # ========================================================================
    # Columns: F=1H24(6), G=2H24(7), H=1H25(8), I=2H25(9), J=1H26(10)
    hy_cols = [6, 7, 8, 9, 10]  # 1H24, 2H24, 1H25, 2H25, 1H26

    # Revenue
    enter_row_data(hy, 7, hy_cols, [136.155, 142.722, 176.051, 183.671, 215.116])
    enter_row_data(hy, 8, hy_cols, [5.618, 5.216, 3.232, 5.637, 6.460])
    enter_row_data(hy, 9, hy_cols, [141.773, 147.938, 179.283, 189.308, 221.576], subtotal=True)
    enter_row_data(hy, 10, hy_cols, [25.063, 26.795, 32.187, 34.574, 39.325])
    enter_row_data(hy, 11, hy_cols, [0.451, 0.194, 0.949, -0.319, 0.300])
    enter_row_data(hy, 12, hy_cols, [167.287, 174.927, 212.419, 223.563, 261.201], subtotal=True)

    # Revenue Growth YoY on HY sheet (compare to same half prior year)
    # 1H25 vs 1H24, 2H25 vs 2H24, 1H26 vs 1H25
    set_cell(hy, 13, 8, '=IF(F12=0,"",H12/F12-1)', fmt=FMT_PCT, is_formula=True)
    set_cell(hy, 13, 9, '=IF(G12=0,"",I12/G12-1)', fmt=FMT_PCT, is_formula=True)
    set_cell(hy, 13, 10, '=IF(H12=0,"",J12/H12-1)', fmt=FMT_PCT, is_formula=True)

    # Other Revenue
    enter_row_data(hy, 16, hy_cols, [7.658, 8.724, 10.976, 8.712, 11.893])
    enter_row_data(hy, 17, hy_cols, [2.573, 3.591, 3.423, 5.519, 3.343])
    enter_row_data(hy, 18, hy_cols, [1.337, -1.109, 0.312, 3.082, 2.127])
    enter_row_data(hy, 19, hy_cols, [11.568, 11.206, 14.711, 17.313, 17.363], subtotal=True)

    # Total Rev & Other Inc
    for i, col in enumerate(hy_cols):
        c = chr(64 + col)
        set_cell(hy, 21, col, f'={c}12+{c}19', fmt=FMT_AM, is_formula=True, subtotal=True)

    # Segment EBITDA Bridge
    enter_row_data(hy, 24, hy_cols, [24.249, 23.287, 31.781, 34.248, 41.314])
    enter_row_data(hy, 25, hy_cols, [-3.095, -3.449, -5.014, -8.209, -8.308])
    for i, col in enumerate(hy_cols):
        c = chr(64 + col)
        set_cell(hy, 26, col, f'={c}24+{c}25', fmt=FMT_AM, is_formula=True, subtotal=True)
    enter_row_data(hy, 27, hy_cols, [8.217, 8.867, 9.398, 9.075, 10.767])
    enter_row_data(hy, 28, hy_cols, [-6.020, -5.122, -4.587, -4.431, -4.720])
    enter_row_data(hy, 29, hy_cols, [-4.447, -15.214, 0.0, 2.858, 1.885])
    for i, col in enumerate(hy_cols):
        c = chr(64 + col)
        set_cell(hy, 30, col, f'={c}26+{c}27+{c}28+{c}29', fmt=FMT_AM, is_formula=True, subtotal=True)

    # EBITDA Margin
    for col in hy_cols:
        c = chr(64 + col)
        set_cell(hy, 31, col, f'=IF({c}21=0,"",{c}30/{c}21)', fmt=FMT_PCT, is_formula=True)

    # Expenses
    enter_row_data(hy, 34, hy_cols, [-43.192, -44.388, -55.071, -58.877, -67.126])
    enter_row_data(hy, 35, hy_cols, [-73.326, -80.407, -89.003, -95.653, -110.669])
    enter_row_data(hy, 36, hy_cols, [-25.346, -35.249, -25.865, -27.750, -28.718])
    enter_row_data(hy, 37, hy_cols, [-8.537, -9.401, -11.302, -9.985, -13.706])
    enter_row_data(hy, 38, hy_cols, [-9.550, -8.318, -14.311, -15.070, -17.407])
    for col in hy_cols:
        c = chr(64 + col)
        set_cell(hy, 39, col, f'=SUM({c}34:{c}38)', fmt=FMT_AM, is_formula=True, subtotal=True)

    # D&A
    enter_row_data(hy, 42, hy_cols, [-6.324, -4.923, -7.743, -6.131, -8.071])
    enter_row_data(hy, 43, hy_cols, [-6.062, -6.898, -7.832, -11.464, -10.155])
    enter_row_data(hy, 44, hy_cols, [-1.434, -1.501, -1.890, -1.778, -1.231])
    enter_row_data(hy, 45, hy_cols, [-2.080, -1.901, -1.615, -1.228, -1.724])
    for col in hy_cols:
        c = chr(64 + col)
        set_cell(hy, 46, col, f'=SUM({c}42:{c}45)', fmt=FMT_AM, is_formula=True, subtotal=True)

    # D&A / Revenue
    for col in hy_cols:
        c = chr(64 + col)
        set_cell(hy, 47, col, f'=IF({c}12=0,"",{c}46/{c}12)', fmt=FMT_PCT, is_formula=True)

    # EBIT
    for col in hy_cols:
        c = chr(64 + col)
        set_cell(hy, 49, col, f'={c}30+{c}46', fmt=FMT_AM, is_formula=True, subtotal=True)

    # EBIT Margin
    for col in hy_cols:
        c = chr(64 + col)
        set_cell(hy, 50, col, f'=IF({c}21=0,"",{c}49/{c}21)', fmt=FMT_PCT, is_formula=True)

    # Finance Income
    enter_row_data(hy, 53, hy_cols, [0.459, 0.637, 6.632, 5.898, 5.051])
    enter_row_data(hy, 54, hy_cols, [1.824, 3.017, 4.752, 5.290, 6.378])
    enter_row_data(hy, 55, hy_cols, [0.044, 0.031, 0.061, 0.041, 0.003])
    for col in hy_cols:
        c = chr(64 + col)
        set_cell(hy, 56, col, f'=SUM({c}53:{c}55)', fmt=FMT_AM, is_formula=True, subtotal=True)

    # Finance Costs
    enter_row_data(hy, 57, hy_cols, [-6.261, -7.133, -8.686, -10.085, -11.946])
    enter_row_data(hy, 58, hy_cols, [-0.149, -0.181, -0.066, -0.065, -0.068])
    for col in hy_cols:
        c = chr(64 + col)
        set_cell(hy, 59, col, f'=SUM({c}57:{c}58)', fmt=FMT_AM, is_formula=True, subtotal=True)

    # Net Finance
    for col in hy_cols:
        c = chr(64 + col)
        set_cell(hy, 60, col, f'={c}56+{c}59', fmt=FMT_AM, is_formula=True, subtotal=True)

    # PBT
    for col in hy_cols:
        c = chr(64 + col)
        set_cell(hy, 62, col, f'={c}49+{c}60', fmt=FMT_AM, is_formula=True, subtotal=True)

    # Tax
    enter_row_data(hy, 63, hy_cols, [-2.881, 0.694, -7.890, -6.844, -8.595])

    # Tax Rate
    for col in hy_cols:
        c = chr(64 + col)
        set_cell(hy, 64, col, f'=IF({c}62=0,"",{c}63/{c}62)', fmt=FMT_PCT, is_formula=True)

    # NPAT
    for col in hy_cols:
        c = chr(64 + col)
        set_cell(hy, 65, col, f'={c}62+{c}63', fmt=FMT_AM, is_formula=True, subtotal=True)

    # NPAT Growth YoY
    set_cell(hy, 66, 8, '=IF(F65=0,"",H65/F65-1)', fmt=FMT_PCT, is_formula=True)
    set_cell(hy, 66, 9, '=IF(G65=0,"",I65/G65-1)', fmt=FMT_PCT, is_formula=True)
    set_cell(hy, 66, 10, '=IF(H65=0,"",J65/H65-1)', fmt=FMT_PCT, is_formula=True)

    # ========================================================================
    # HY Operating Metrics
    # ========================================================================
    enter_row_data(hy, 69, hy_cols, [62, 64, 74, 81, 87], fmt=FMT_COUNT)
    enter_row_data(hy, 70, hy_cols, [121, 130, 136, 143, 150], fmt=FMT_COUNT)
    enter_row_data(hy, 71, hy_cols, [22, 22, 25, 26, 27], fmt=FMT_COUNT)
    enter_row_data(hy, 72, hy_cols, [4, 4, 4, 6, 8], fmt=FMT_COUNT)
    # Total restaurants
    for col in hy_cols:
        c = chr(64 + col)
        set_cell(hy, 73, col, f'=SUM({c}69:{c}72)', fmt=FMT_COUNT, is_formula=True, subtotal=True)

    enter_row_data(hy, 74, hy_cols, [465.1, 483.8, 573.0, 595.5, 673.6], fmt=FMT_DECIMAL)
    enter_row_data(hy, 75, hy_cols, [470.7, 489.0, 577.9, 602.8, 681.8], fmt=FMT_DECIMAL)
    enter_row_data(hy, 76, hy_cols, [0.101, 0.060, 0.094, 0.097, 0.044], fmt=FMT_PCT)
    enter_row_data(hy, 77, hy_cols, [None, None, 6.9, None, 6.9], fmt=FMT_DECIMAL)
    enter_row_data(hy, 78, hy_cols, [None, None, 5.0, None, 5.2], fmt=FMT_DECIMAL)

    # DT/Strip Network Margins — not provided for HY, leave blank
    # Breakfast/After9pm/Delivery/Digital mix — not provided for HY, leave blank

    # ========================================================================
    # SAVE
    # ========================================================================
    wb.save(MODEL_PATH)
    print("GYG Model saved successfully.")
    print(f"File: {MODEL_PATH}")


if __name__ == '__main__':
    main()

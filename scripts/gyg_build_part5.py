"""Part 5: HY & Segments sheet - historical data and formulas."""
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')
ws = wb['HY & Segments']

CL = get_column_letter

# Column mapping:
# D(4)=1H23, E(5)=2H23, F(6)=1H24, G(7)=2H24, H(8)=1H25, I(9)=2H25, J(10)=1H26

# ============================================================
# HY P&L HISTORICAL DATA
# ============================================================
# We don't have 1H23/2H23 half-year breakdown, only FY23.
# We have: 1H24(F), 2H24(G), 1H25(H), 2H25(I), 1H26(J)

# For FY23 (1H23=D, 2H23=E): we don't have HY split, put FY23 in 2H23 as placeholder
# Actually better to leave 1H23/2H23 blank or estimate 50/50 split

# R7: Total Revenue
# 1H23/2H23: estimate from FY23 (259.044), roughly 48/52 split
ws.cell(7, 4).value = 124.0   # 1H23 est
ws.cell(7, 5).value = 135.044  # 2H23 est (to sum to 259.044)
ws.cell(7, 6).value = 167.287  # 1H24
ws.cell(7, 7).value = 174.927  # 2H24
ws.cell(7, 8).value = 212.419  # 1H25
ws.cell(7, 9).value = 223.563  # 2H25
ws.cell(7, 10).value = 261.201  # 1H26

# R8: Other Revenue
ws.cell(8, 4).value = 8.5    # 1H23 est
ws.cell(8, 5).value = 9.225  # 2H23 est (sum to 17.725)
ws.cell(8, 6).value = 11.568  # 1H24
ws.cell(8, 7).value = 11.206  # 2H24
ws.cell(8, 8).value = 14.711  # 1H25
ws.cell(8, 9).value = 17.313  # 2H25
ws.cell(8, 10).value = 17.363  # 1H26

# R9: Total Rev incl Other
for c in range(4, 11):
    col = CL(c)
    ws.cell(9, c).value = f'={col}7+{col}8'

# R10: Revenue Growth YoY
# Compare to same period 2 cols prior
for c in range(6, 11):  # from 1H24 onwards
    col = CL(c)
    pcol = CL(c-2)
    ws.cell(10, c).value = f'=IF({pcol}7=0,"",{col}7/{pcol}7-1)'

# R13: Cost of food
ws.cell(13, 4).value = -33.7   # 1H23 est
ws.cell(13, 5).value = -36.728  # 2H23 est
ws.cell(13, 6).value = -43.192
ws.cell(13, 7).value = -44.388
ws.cell(13, 8).value = -55.071
ws.cell(13, 9).value = -58.877
ws.cell(13, 10).value = -67.126

# R14: Employee
ws.cell(14, 4).value = -54.5   # 1H23 est
ws.cell(14, 5).value = -59.225  # 2H23 est
ws.cell(14, 6).value = -73.326
ws.cell(14, 7).value = -80.407
ws.cell(14, 8).value = -89.003
ws.cell(14, 9).value = -95.653
ws.cell(14, 10).value = -110.669

# R15: Admin+Marketing+Other combined
ws.cell(15, 4).value = -(17.0 + 6.5 + 6.7)  # 1H23 est = -30.2
ws.cell(15, 5).value = -(18.282 + 7.218 + 7.295)  # 2H23 est = -32.795
ws.cell(15, 6).value = -(25.346 + 8.537 + 9.550)  # 1H24 = -43.433
ws.cell(15, 7).value = -(35.249 + 9.401 + 8.318)  # 2H24 = -52.968
ws.cell(15, 8).value = -(25.865 + 11.302 + 14.311)  # 1H25 = -51.478
ws.cell(15, 9).value = -(27.750 + 9.985 + 15.070)  # 2H25 = -52.805
ws.cell(15, 10).value = -(28.718 + 13.706 + 17.407)  # 1H26 = -59.831

# R16: Total Costs
for c in range(4, 11):
    col = CL(c)
    ws.cell(16, c).value = f'={col}13+{col}14+{col}15'

# R19: Statutory EBITDA = Total Rev incl Other + Total Costs
for c in range(4, 11):
    col = CL(c)
    ws.cell(19, c).value = f'={col}9+{col}16'

# ============================================================
# SEGMENT EBITDA (R38-R42)
# ============================================================

# Australia Seg EBITDA
# 1H23/2H23: FY23 total = 30.7, split roughly
ws.cell(38, 4).value = 15.0    # 1H23 est
ws.cell(38, 5).value = 15.7    # 2H23 est
ws.cell(38, 6).value = 23.1    # 1H24
ws.cell(38, 7).value = 22.5    # 2H24
ws.cell(38, 8).value = 31.8    # 1H25
ws.cell(38, 9).value = 34.2    # 2H25
ws.cell(38, 10).value = 41.3   # 1H26

# US Seg EBITDA
ws.cell(39, 4).value = -2.0    # 1H23 est
ws.cell(39, 5).value = -2.3    # 2H23 est
ws.cell(39, 6).value = -3.1    # 1H24
ws.cell(39, 7).value = -3.4    # 2H24
ws.cell(39, 8).value = -5.0    # 1H25
ws.cell(39, 9).value = -8.2    # 2H25
ws.cell(39, 10).value = -8.3   # 1H26

# Corporate = 0
for c in range(4, 11):
    ws.cell(40, c).value = 0

# R42: Group Seg EBITDA
for c in range(4, 11):
    col = CL(c)
    ws.cell(42, c).value = f'=SUM({col}38:{col}40)'

# R43: Seg EBITDA Growth YoY
for c in range(6, 11):
    col = CL(c)
    pcol = CL(c-2)
    ws.cell(43, c).value = f'=IF({pcol}42=0,"",{col}42/{pcol}42-1)'

# R44: Seg EBITDA Margin
for c in range(4, 11):
    col = CL(c)
    ws.cell(44, c).value = f'=IF({col}7=0,"",{col}42/{col}7)'

# ============================================================
# STATUTORY ADJUSTMENTS (R48-R49)
# ============================================================
# R48: AASB16 + SBP + Other combined adjustment
# Stat EBITDA - Seg EBITDA = total adjustment
# 1H25: 31.6 - 26.8 = 4.8 (9.4 - 4.6 + 0 = 4.8)
# 2H25: derive from FY25: 65.1 - 52.8 = 12.3 minus 1H25(4.8) = 7.5
# 1H26: 40.9 - 33.0 = 7.9 (10.8 - 4.7 + 1.9 = 8.0)

ws.cell(48, 4).value = 1.5   # 1H23 est
ws.cell(48, 5).value = 1.7   # 2H23 est
ws.cell(48, 6).value = 2.5   # 1H24 est
ws.cell(48, 7).value = -9.0  # 2H24 (IPO costs drag)
ws.cell(48, 8).value = 4.8   # 1H25
ws.cell(48, 9).value = 7.5   # 2H25
ws.cell(48, 10).value = 7.9  # 1H26

# R49: Statutory EBITDA
for c in range(4, 11):
    col = CL(c)
    ws.cell(49, c).value = f'={col}42+{col}48'

# ============================================================
# D&A, EBIT, Interest, PBT, Tax, NPAT
# ============================================================

# R52: D&A
ws.cell(52, 4).value = -12.2   # 1H23 est
ws.cell(52, 5).value = -13.355  # 2H23 est
ws.cell(52, 6).value = -15.900
ws.cell(52, 7).value = -15.223
ws.cell(52, 8).value = -19.080
ws.cell(52, 9).value = -20.601
ws.cell(52, 10).value = -21.181

# R55: D&A / Revenue
for c in range(4, 11):
    col = CL(c)
    ws.cell(55, c).value = f'=IF({col}7=0,"",{col}52/{col}7)'

# R58: EBIT = Stat EBITDA + D&A
for c in range(4, 11):
    col = CL(c)
    ws.cell(58, c).value = f'={col}49+{col}52'

# R59: EBIT Growth YoY
for c in range(6, 11):
    col = CL(c)
    pcol = CL(c-2)
    ws.cell(59, c).value = f'=IF({pcol}58=0,"",{col}58/{pcol}58-1)'

# R60: EBIT Margin
for c in range(4, 11):
    col = CL(c)
    ws.cell(60, c).value = f'=IF({col}7=0,"",{col}58/{col}7)'

# R63: Finance Income
ws.cell(63, 4).value = 2.2     # 1H23 est
ws.cell(63, 5).value = 2.393   # 2H23 est
ws.cell(63, 6).value = 2.327
ws.cell(63, 7).value = 3.685
ws.cell(63, 8).value = 11.445
ws.cell(63, 9).value = 11.229
ws.cell(63, 10).value = 11.432

# R64: Finance Costs
ws.cell(64, 4).value = -4.1    # 1H23 est
ws.cell(64, 5).value = -4.403  # 2H23 est
ws.cell(64, 6).value = -6.410
ws.cell(64, 7).value = -7.314
ws.cell(64, 8).value = -8.752
ws.cell(64, 9).value = -10.150
ws.cell(64, 10).value = -12.014

# R66: Net Finance Costs
for c in range(4, 11):
    col = CL(c)
    ws.cell(66, c).value = f'={col}63+{col}64'

# R71: PBT
for c in range(4, 11):
    col = CL(c)
    ws.cell(71, c).value = f'={col}58+{col}66'

# R72: Tax - hardcode actuals
ws.cell(72, 4).value = -1.2    # 1H23 est
ws.cell(72, 5).value = -1.223  # 2H23 est
ws.cell(72, 6).value = -2.881
ws.cell(72, 7).value = 0.694
ws.cell(72, 8).value = -7.890
ws.cell(72, 9).value = -6.844
ws.cell(72, 10).value = -8.595

# R73: Tax Rate
for c in range(4, 11):
    col = CL(c)
    ws.cell(73, c).value = f'=IF({col}71=0,"",{col}72/{col}71)'

# R75: NPAT
for c in range(4, 11):
    col = CL(c)
    ws.cell(75, c).value = f'={col}71+{col}72'

# R78: NPAT Growth YoY
for c in range(6, 11):
    col = CL(c)
    pcol = CL(c-2)
    ws.cell(78, c).value = f'=IF({pcol}75<=0,"",{col}75/{pcol}75-1)'

# ============================================================
# KPIs (R81-R87)
# ============================================================

# Total Restaurants (end of period)
# FY23=194, 1H24, FY24=220, 1H25, FY25=256, 1H26
ws.cell(81, 5).value = 194    # end FY23 (2H23)
ws.cell(81, 6).value = 204    # 1H24 (62+4+121+17+5=209 actually)
# Recalculate: 1H24: Corp-Aus=62, Corp-US=4, Fran-Aus=121, Fran-SG=17, Fran-JP=5 = 209
ws.cell(81, 6).value = 209
ws.cell(81, 7).value = 220    # FY24 (64+4+130+17+5=220)
ws.cell(81, 8).value = 239    # 1H25 (74+4+136+20+5=239)
ws.cell(81, 9).value = 256    # FY25 (81+6+143+21+5=256)
ws.cell(81, 10).value = 272   # 1H26 (87+8+150+22+5=272)

# Aus Corp Restaurants
ws.cell(82, 5).value = 55    # end FY23
ws.cell(82, 6).value = 62
ws.cell(82, 7).value = 64
ws.cell(82, 8).value = 74
ws.cell(82, 9).value = 81
ws.cell(82, 10).value = 87

# Aus Franchise Restaurants (incl SG/JP)
ws.cell(83, 5).value = 136   # 116+16+4
ws.cell(83, 6).value = 143   # 121+17+5
ws.cell(83, 7).value = 152   # 130+17+5
ws.cell(83, 8).value = 161   # 136+20+5
ws.cell(83, 9).value = 169   # 143+21+5
ws.cell(83, 10).value = 177  # 150+22+5

# US Restaurants
ws.cell(84, 5).value = 3
ws.cell(84, 6).value = 4
ws.cell(84, 7).value = 4
ws.cell(84, 8).value = 4
ws.cell(84, 9).value = 6
ws.cell(84, 10).value = 8

# Australia Network Sales
ws.cell(85, 4).value = 360.0   # 1H23 est
ws.cell(85, 5).value = 393.0   # 2H23 est (total FY23 = 753.0)
ws.cell(85, 6).value = 465.0
ws.cell(85, 7).value = 483.7
ws.cell(85, 8).value = 573.0
ws.cell(85, 9).value = 595.4
ws.cell(85, 10).value = 673.6

# US Network Sales
ws.cell(86, 4).value = 2.8   # 1H23 est
ws.cell(86, 5).value = 3.2   # 2H23 est
ws.cell(86, 6).value = 5.6
ws.cell(86, 7).value = 5.2
ws.cell(86, 8).value = 4.9
ws.cell(86, 9).value = 7.3
ws.cell(86, 10).value = 8.2

# Comp Sales Growth (YoY)
for c in range(6, 11):
    col = CL(c)
    pcol = CL(c-2)
    ws.cell(87, c).value = f'=IF({pcol}85=0,"",({col}85+{col}86)/({pcol}85+{pcol}86)-1)'

# ============================================================
# SEGMENT FORECAST - AUSTRALIA (R92-R113)
# ============================================================

# R93: Drive Thru Restaurants - only have FY25 and 1H26
ws.cell(93, 9).value = 117    # FY25 (2H25 = end of FY25)
ws.cell(93, 10).value = 126   # 1H26

# R94: Strip Restaurants
ws.cell(94, 9).value = 68     # FY25
ws.cell(94, 10).value = 73    # 1H26

# R95: Other Restaurants
ws.cell(95, 9).value = 39     # FY25
ws.cell(95, 10).value = 38    # 1H26

# R96: Total Aus Restaurants = DT + Strip + Other
for c in [9, 10]:
    col = CL(c)
    ws.cell(96, c).value = f'={col}93+{col}94+{col}95'

# R97: DT AUV
ws.cell(97, 9).value = 6.7    # FY25 (annualised)
ws.cell(97, 10).value = 6.9   # 1H26 (annualised)

# R98: Strip AUV
ws.cell(98, 9).value = 5.0
ws.cell(98, 10).value = 5.2

# R99: Other AUV
ws.cell(99, 9).value = 4.0
ws.cell(99, 10).value = 4.0   # est

# R100-102: Network Sales by format (derive: count * AUV * 0.5 for half year)
for c in [9, 10]:
    col = CL(c)
    ws.cell(100, c).value = f'={col}93*{col}97*0.5'  # DT NS (half year)
    ws.cell(101, c).value = f'={col}94*{col}98*0.5'  # Strip NS
    ws.cell(102, c).value = f'={col}95*{col}99*0.5'  # Other NS

# R103: Total Aus NS
for c in [9, 10]:
    col = CL(c)
    ws.cell(103, c).value = f'={col}100+{col}101+{col}102'

# Also enter known historical total NS
ws.cell(103, 4).value = 360.0
ws.cell(103, 5).value = 393.0
ws.cell(103, 6).value = 465.0
ws.cell(103, 7).value = 483.7
ws.cell(103, 8).value = 573.0

# R104: Corporate Restaurant Sales
ws.cell(104, 4).value = 101.0   # 1H23 est (FY23=212.0)
ws.cell(104, 5).value = 111.0   # 2H23 est
ws.cell(104, 6).value = 136.2
ws.cell(104, 7).value = 142.7
ws.cell(104, 8).value = 176.1
ws.cell(104, 9).value = 183.9
ws.cell(104, 10).value = 215.1

# R105: Corp Rest Margin %
ws.cell(105, 4).value = 0.144   # 1H23 est = FY23 rate
ws.cell(105, 5).value = 0.144
ws.cell(105, 6).value = 0.175
ws.cell(105, 7).value = 0.173
ws.cell(105, 8).value = 0.180
ws.cell(105, 9).value = 0.178
ws.cell(105, 10).value = 0.176

# R106: Corp Rest Margin $
for c in range(4, 11):
    col = CL(c)
    ws.cell(106, c).value = f'={col}104*{col}105'

# R107: Franchise Network Sales = Total NS - Corp Sales
for c in range(4, 11):
    col = CL(c)
    ws.cell(107, c).value = f'={col}103-{col}104'

# R108: Implied Royalty Rate
ws.cell(108, 4).value = 0.087
ws.cell(108, 5).value = 0.087
ws.cell(108, 6).value = 0.087
ws.cell(108, 7).value = 0.093
ws.cell(108, 8).value = 0.097
ws.cell(108, 9).value = 0.097
ws.cell(108, 10).value = 0.097

# R109: Franchise & Other Revenue = Fran NS * Royalty Rate
for c in range(4, 11):
    col = CL(c)
    ws.cell(109, c).value = f'={col}107*{col}108'

# Hardcode known values to override
ws.cell(109, 8).value = 38.2   # 1H25
ws.cell(109, 9).value = 40.5   # 2H25
ws.cell(109, 10).value = 42.8  # 1H26

# R110: G&A Costs
ws.cell(110, 4).value = -22.0   # 1H23 est
ws.cell(110, 5).value = -24.7   # 2H23 est (FY23 total ~46.7)
ws.cell(110, 6).value = -30.0   # 1H24 est
ws.cell(110, 7).value = -33.7   # 2H24 est (FY24 PF total ~63.7)
ws.cell(110, 8).value = -38.1   # 1H25
ws.cell(110, 9).value = -38.9   # 2H25
ws.cell(110, 10).value = -39.4  # 1H26

# R111: G&A as % of NS
for c in range(4, 11):
    col = CL(c)
    ws.cell(111, c).value = f'=IF({col}103=0,"",ABS({col}110)/{col}103)'

# R112: Australia Segment EBITDA = Corp Margin + Fran Rev + G&A
for c in range(4, 11):
    col = CL(c)
    ws.cell(112, c).value = f'={col}106+{col}109+{col}110'

# R113: Seg EBITDA as % of NS
for c in range(4, 11):
    col = CL(c)
    ws.cell(113, c).value = f'=IF({col}103=0,"",{col}112/{col}103)'

# ============================================================
# SEGMENT FORECAST - US (R115-R123)
# ============================================================

# R116: US Restaurants
ws.cell(116, 5).value = 3     # end FY23
ws.cell(116, 6).value = 4     # 1H24
ws.cell(116, 7).value = 4     # FY24
ws.cell(116, 8).value = 4     # 1H25
ws.cell(116, 9).value = 6     # FY25
ws.cell(116, 10).value = 8    # 1H26

# R117: US Network Sales
ws.cell(117, 4).value = 2.8
ws.cell(117, 5).value = 3.2
ws.cell(117, 6).value = 5.6
ws.cell(117, 7).value = 5.2
ws.cell(117, 8).value = 4.9
ws.cell(117, 9).value = 7.3
ws.cell(117, 10).value = 8.2

# R118: Corp Rest Sales (US)
ws.cell(118, 6).value = 0     # 1H24 (FY24 total = 10.8, but HY split unclear)
ws.cell(118, 7).value = 10.8  # 2H24 (or FY24)
ws.cell(118, 8).value = 3.2
ws.cell(118, 9).value = 5.7
ws.cell(118, 10).value = 6.5

# R119: Corp Margin % (US)
# FY24: -1.0/10.8 = -9.3%
# 1H25: -1.3/3.2 = -40.6%
# We'll leave blank for early periods, use margin $ directly
ws.cell(119, 8).value = -0.406
ws.cell(119, 9).value = -0.667
ws.cell(119, 10).value = -0.692

# R120: Corp Margin $ (US)
ws.cell(120, 7).value = -1.0   # FY24
ws.cell(120, 8).value = -1.3   # 1H25
ws.cell(120, 9).value = -3.8   # 2H25
ws.cell(120, 10).value = -4.5  # 1H26

# R121: Fran & Other Rev (US)
ws.cell(121, 4).value = 0
ws.cell(121, 5).value = 0
ws.cell(121, 6).value = 0
ws.cell(121, 7).value = 0
ws.cell(121, 8).value = 0.1
ws.cell(121, 9).value = 0.2
ws.cell(121, 10).value = 0.1

# R122: G&A Costs (US)
ws.cell(122, 7).value = -5.5  # FY24
ws.cell(122, 8).value = -3.8  # 1H25
ws.cell(122, 9).value = -4.6  # 2H25
ws.cell(122, 10).value = -4.0  # 1H26

# R123: US Segment EBITDA = Corp Margin + Fran Rev + G&A
for c in [7, 8, 9, 10]:
    col = CL(c)
    ws.cell(123, c).value = f'={col}120+{col}121+{col}122'

# Also hardcode known values
ws.cell(123, 4).value = -2.0
ws.cell(123, 5).value = -2.3
ws.cell(123, 6).value = -3.1

wb.save('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')
print("Part 5 complete: HY historical data entered")

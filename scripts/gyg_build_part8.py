"""Part 8: Fix linkage issues and missing formulas."""
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')

# ============================================================
# FIX 1: Annual R35/R36 forecast formulas reference HY Column A codes.
# Annual A35 = 'EBITDA-Australia EBITDA', HY A38 = 'EBITDA-Australia EBITDA'
# The INDEX/MATCH on $A35 will look for the code in HY $A column.
# This should work since HY R38 has the same code. Verified OK.
# ============================================================

# ============================================================
# FIX 2: HY sheet R84 (US Restaurants) missing forecast for 2H26+
# ============================================================
ws = wb['HY & Segments']
CL = get_column_letter

for c in range(11, 30):
    col = CL(c)
    pcol = CL(c-1)
    # R84: US Restaurants - add from HY segment forecast R116
    ws.cell(84, c).value = f'={col}116'

# ============================================================
# FIX 3: Annual R7 forecast pulls Rev-Total Revenue from HY.
# HY R7 has A='Rev-Total Revenue'. But for 1H forecast cols (even),
# we used a growth formula, not a segment build-up.
# The Annual forecast uses INDEX/MATCH which requires HY A7 code.
# Let's verify: Annual R7 G(FY26E) formula:
# =INDEX('HY & Segments'!$A:$AG,MATCH($A7,'HY & Segments'!$A:$A,0),MATCH("1H"&RIGHT(G$1,2),...))
# $A7 on Annual = 'Rev-Total Revenue', matching HY A7 = 'Rev-Total Revenue'. OK.
# But the formula references $A:$AG (cols 1-33) while HY only goes to AC(29).
# AG = col 33. This should be fine since MATCH on row 3 will find the period.
# ============================================================

# ============================================================
# FIX 4: Check that Annual R35/R36 correctly links.
# Annual A35 = 'EBITDA-Australia EBITDA'
# HY A38 = 'EBITDA-Australia EBITDA'
# The MATCH will find it on HY row 38. Then it looks up the column.
# The formula is:
# =INDEX('HY & Segments'!$A:$AG,MATCH($A35,'HY & Segments'!$A:$A,0),MATCH("1H26",'HY & Segments'!$3:$3,0))
# This returns HY R38, col matching "1H26" = col J(10). So it gets J38 = 41.3.
# Similarly for 2H26 it gets K38.
# BUT: For forecast, HY R38 for 1H cols uses =PCP*1.20 (growth from HY).
# And 2H cols use INDEX from Annual. This creates a CIRCULAR REFERENCE!
# 2H26: HY R38 K11 = INDEX(Annual!...) - J38
# Annual R35 G = HY J38 + HY K38 = HY J38 + (Annual G35 - HY J38) = Annual G35
# So it's self-referencing. Need to fix this.
#
# The correct approach: HY segment EBITDA for forecast should NOT reference Annual.
# Instead, it should come from the segment build-up (R112 for Aus, R123 for US).
# Then Annual pulls from HY.
# ============================================================

# FIX: Link HY R38 (Aus Seg EBITDA) to HY R112 (segment forecast calc)
# Link HY R39 (US Seg EBITDA) to HY R123 (US segment forecast calc)
# For ALL forecast columns (11+)

for c in range(11, 30):
    col = CL(c)
    # Aus EBITDA = from segment build
    ws.cell(38, c).value = f'={col}112'
    # US EBITDA = from US segment build
    ws.cell(39, c).value = f'={col}123'

# Also fix 1H forecast columns for HY P&L items.
# Currently 1H uses growth-based forecasts for R7 (revenue), etc.
# But the segment forecast calculates revenue bottom-up.
# For consistency, link HY R7 forecast to Aus Corp Sales + Aus Fran Rev + US Corp Sales + US Fran Rev
# Actually, Rev-Total Revenue is statutory revenue, not segment NS.
# Statutory Revenue = Corp Sales (Aus) + Fran Rev (Aus) + Corp Sales (US) + Fran Rev (US)
# = R104 (on HY) + R109 + R118 + R121
# This is better than arbitrary growth rate.

# But wait - the statutory revenue also includes the HY actual adjustments.
# For forecast: Revenue = Aus Corp Sales + Aus Fran Rev + US Corp Sales + US Fran Rev
for c in range(11, 30):
    col = CL(c)
    ws.cell(7, c).value = f'={col}104+{col}109+{col}118+{col}121'

# Then for 2H, we DON'T use the Annual-1H derivation for revenue either.
# Revenue is directly calculated from segment build. This avoids circularity.

# Similarly fix costs - use % of revenue approach for all forecast cols
for c in range(11, 30):
    col = CL(c)
    ws.cell(13, c).value = f'={col}7*-0.261'
    ws.cell(14, c).value = f'={col}7*-0.424'
    ws.cell(15, c).value = f'={col}7*-0.230'

# Fix Other Revenue for all forecast (was only fixed for 1H)
for c in range(11, 30):
    col = CL(c)
    pcol = CL(c-2)  # PCP
    ws.cell(8, c).value = f'={pcol}8*1.075'  # ~15% annual = 7.5% per half

# Fix D&A for 2H cols - use growth instead of Annual derivation
for c in range(11, 30):
    col = CL(c)
    pcol = CL(c-2)
    ws.cell(52, c).value = f'={pcol}52*1.06'  # ~12% annual = 6% per half

# Fix Finance items for 2H - avoid Annual derivation
for c in range(11, 30):
    col = CL(c)
    pcol = CL(c-2)
    ws.cell(63, c).value = f'={pcol}63*1.0'  # flat
    ws.cell(64, c).value = f'={pcol}64*1.05'  # ~10% annual = 5% per half

# Fix Tax for 2H
for c in range(11, 30):
    col = CL(c)
    ws.cell(72, c).value = f'=IF({col}71>0,-{col}71*0.35,0)'

# Fix R48 adjustments for 2H - avoid Annual derivation
for c in range(11, 30):
    col = CL(c)
    pcol = CL(c-2)
    ws.cell(48, c).value = f'={pcol}48*1.05'  # ~10% annual = 5% per half

# ============================================================
# FIX 5: Annual R7 formula - now that HY builds revenue bottom-up,
# Annual can safely pull 1H+2H from HY without circularity.
# The formula already does this correctly.
# ============================================================

# ============================================================
# FIX 6: Fix HY Corp Sales forecast for 1H columns
# Currently 1H uses =PCP*1.15 for Aus Corp Sales.
# Better: derive from segment forecast (count * AUV * share)
# But we don't have corporate share split. Keep growth formula.
# ============================================================

# ============================================================
# FIX 7: US Franchise Rev forecast
# ============================================================
for c in range(11, 30):
    col = CL(c)
    pcol = CL(c-1)
    # US Fran Rev growing but small
    if ws.cell(121, c).value is None:
        ws.cell(121, c).value = f'={pcol}121*1.1'

# ============================================================
# FIX 8: US G&A for missing 1H forecasts
# ============================================================
for c in range(11, 30):
    col = CL(c)
    pcol = CL(c-2)
    if ws.cell(122, c).value is None or (isinstance(ws.cell(122, c).value, str) and 'pycol' not in str(ws.cell(122, c).value)):
        pass  # Already set in part 6

wb.save('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')
print("Part 8 complete: Linkage issues fixed")

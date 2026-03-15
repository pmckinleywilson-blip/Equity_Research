"""Part 9: Final fixes - ensure no circularity, clean up number formats."""
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

wb = openpyxl.load_workbook('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')

CL = get_column_letter

# ============================================================
# ANNUAL SHEET: Verify forecast chain
# R7 (Revenue) = INDEX from HY (1H+2H) -- HY builds from segment bottom-up
# R8 (Other Rev) = prior * 1.15
# R9 = R7+R8
# R13-R15 = % of R7
# R16 = sum
# R19 = R9+R16
# R35-R36 = INDEX from HY segment EBITDA -- HY builds from segment bottom-up
# R38 = sum
# R42-R44 = growth assumptions
# R45 = R38 + R42+R43+R44
# R48 = D&A growth
# R54 = R45+R48
# R59-R60 = growth
# R61 = sum
# R68 = R54+R61
# R69 = tax
# R72 = R68+R69
# No circularity. Good.
# ============================================================

ws = wb['Annual']

# Apply number formats to key rows
fmt_m = '#,##0.0'  # millions with 1dp
fmt_pct = '0.0%'   # percentage
fmt_int = '#,##0'   # integer
fmt_eps = '0.000'   # EPS

# Number format by row
row_formats = {
    7: fmt_m, 8: fmt_m, 9: fmt_m, 10: fmt_pct,
    13: fmt_m, 14: fmt_m, 15: fmt_m, 16: fmt_m,
    19: fmt_m, 20: fmt_pct,
    35: fmt_m, 36: fmt_m, 37: fmt_m, 38: fmt_m, 39: fmt_pct, 40: fmt_pct,
    42: fmt_m, 43: fmt_m, 44: fmt_m, 45: fmt_m,
    48: fmt_m, 51: fmt_pct,
    54: fmt_m, 55: fmt_pct, 56: fmt_pct,
    59: fmt_m, 60: fmt_m, 61: fmt_m,
    68: fmt_m, 69: fmt_m, 70: fmt_pct,
    72: fmt_m, 75: fmt_pct, 76: fmt_pct,
    79: fmt_m, 80: fmt_m, 81: fmt_m, 82: fmt_m,
    84: fmt_eps, 85: fmt_eps, 86: fmt_pct,
    88: fmt_eps, 89: fmt_m, 90: fmt_pct, 91: fmt_pct,
    95: fmt_int, 96: fmt_int, 97: fmt_int, 98: fmt_int,
    99: fmt_m, 100: fmt_pct, 101: fmt_pct, 102: fmt_pct, 103: fmt_pct, 104: fmt_pct,
    110: fmt_m, 111: fmt_m, 112: fmt_m, 113: fmt_m, 114: fmt_m, 115: fmt_m, 116: fmt_m,
    117: fmt_m, 118: fmt_pct, 119: fmt_pct, 121: fmt_pct,
    125: fmt_m, 126: fmt_m, 127: fmt_m, 128: fmt_m, 129: fmt_m,
    131: fmt_m, 132: fmt_m, 133: '0.0x',
    137: fmt_m, 138: fmt_m, 139: fmt_m, 140: fmt_m, 141: fmt_m,
    142: fmt_pct, 143: '0.0x', 144: fmt_m,
    148: fmt_m, 149: fmt_m, 150: fmt_m, 151: fmt_m,
    152: fmt_m, 153: fmt_m, 154: fmt_m, 155: fmt_m, 156: fmt_m,
    160: fmt_m, 161: fmt_pct, 162: fmt_m, 163: fmt_m, 164: fmt_m, 165: fmt_m, 166: fmt_m,
    169: fmt_m, 170: fmt_m, 171: fmt_m, 172: fmt_m, 173: fmt_m, 174: fmt_m,
    176: fmt_m,
    179: fmt_m, 180: fmt_m, 181: fmt_m, 182: fmt_m, 183: fmt_eps, 184: fmt_pct, 185: fmt_pct,
    189: fmt_m, 190: fmt_m, 191: fmt_pct, 192: fmt_m, 193: fmt_pct,
}

for r, fmt in row_formats.items():
    for c in range(4, 17):
        ws.cell(r, c).number_format = fmt

# ============================================================
# HY SHEET: Apply number formats
# ============================================================
ws2 = wb['HY & Segments']

hy_row_formats = {
    7: fmt_m, 8: fmt_m, 9: fmt_m, 10: fmt_pct,
    13: fmt_m, 14: fmt_m, 15: fmt_m, 16: fmt_m,
    19: fmt_m,
    38: fmt_m, 39: fmt_m, 40: fmt_m, 42: fmt_m, 43: fmt_pct, 44: fmt_pct,
    48: fmt_m, 49: fmt_m,
    52: fmt_m, 55: fmt_pct,
    58: fmt_m, 59: fmt_pct, 60: fmt_pct,
    63: fmt_m, 64: fmt_m, 66: fmt_m,
    71: fmt_m, 72: fmt_m, 73: fmt_pct,
    75: fmt_m, 78: fmt_pct,
    81: fmt_int, 82: fmt_int, 83: fmt_int, 84: fmt_int,
    85: fmt_m, 86: fmt_m, 87: fmt_pct,
    93: fmt_int, 94: fmt_int, 95: fmt_int, 96: fmt_int,
    97: fmt_m, 98: fmt_m, 99: fmt_m,
    100: fmt_m, 101: fmt_m, 102: fmt_m, 103: fmt_m,
    104: fmt_m, 105: fmt_pct, 106: fmt_m,
    107: fmt_m, 108: fmt_pct, 109: fmt_m,
    110: fmt_m, 111: fmt_pct, 112: fmt_m, 113: fmt_pct,
    116: fmt_int, 117: fmt_m, 118: fmt_m, 119: fmt_pct,
    120: fmt_m, 121: fmt_m, 122: fmt_m, 123: fmt_m,
}

for r, fmt in hy_row_formats.items():
    for c in range(4, 30):
        ws2.cell(r, c).number_format = fmt

# ============================================================
# VALUE SHEET: Apply number formats
# ============================================================
ws3 = wb['Value']

for r in [25, 26, 27, 28, 29, 30, 31, 32, 33, 34]:
    for c in range(4, 14):
        ws3.cell(r, c).number_format = fmt_m

for c in range(4, 14):
    ws3.cell(37, c).number_format = '0.0000'
    ws3.cell(38, c).number_format = fmt_m

ws3['C4'].number_format = '#,##0.00'
ws3['C5'].number_format = '#,##0.0'
ws3['C6'].number_format = '#,##0.0'
ws3['C7'].number_format = '#,##0.0'
ws3['C8'].number_format = '#,##0.0'

for r in [12, 13, 15, 17, 18, 20, 21]:
    ws3.cell(r, 3).number_format = '0.0%'

ws3['C14'].number_format = '0.00'
ws3['C19'].number_format = '0.0%'
ws3['C22'].number_format = '0.00'

ws3['C41'].number_format = fmt_m
ws3['C42'].number_format = fmt_m
ws3['C43'].number_format = fmt_m
ws3['C45'].number_format = fmt_m
ws3['C46'].number_format = fmt_m
ws3['C47'].number_format = fmt_m
ws3['C48'].number_format = '#,##0.00'
ws3['C49'].number_format = '0.0%'

for r in [57, 58, 59]:
    ws3.cell(r, 3).number_format = fmt_m
    ws3.cell(r, 4).number_format = '0.0x'
    ws3.cell(r, 5).number_format = fmt_m

ws3['E61'].number_format = fmt_m
ws3['E62'].number_format = fmt_m
ws3['E63'].number_format = fmt_m
ws3['E64'].number_format = fmt_m
ws3['E65'].number_format = '#,##0.00'
ws3['E66'].number_format = '0.0%'

wb.save('/home/pmwilson/Project_Equities/GYG/Models/test.xlsx')
print("Part 9 complete: Number formats applied")

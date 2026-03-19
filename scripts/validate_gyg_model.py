#!/usr/bin/env python3
"""
Comprehensive validation of GYG Model vs HY_model_template and skill file conventions.
Produces categorized issue report: CRITICAL / MAJOR / MODERATE / MINOR
"""

import openpyxl
from openpyxl.utils import get_column_letter
import re
from collections import defaultdict

# ── Paths ──
MODEL_PATH = "/home/pmwilson/Project_Equities/GYG/Models/GYG Model.xlsx"
TEMPLATE_PATH = "/home/pmwilson/Project_Equities/.claude/templates/HY_model_template.xlsx"

# ── Load workbooks (data_only=False to see formulas) ──
print("Loading workbooks...")
model_wb = openpyxl.load_workbook(MODEL_PATH, data_only=False)
template_wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)

# Also load data_only version for value checks
model_wb_data = openpyxl.load_workbook(MODEL_PATH, data_only=True)

# ── Issue tracker ──
issues = []
stats = defaultdict(int)

def add_issue(severity, check, sheet, row, col, expected, found, fix):
    issues.append({
        'severity': severity,
        'check': check,
        'sheet': sheet,
        'row': row,
        'col': col,
        'expected': expected,
        'found': found,
        'fix': fix
    })

def get_color_hex(color):
    """Extract hex color from openpyxl color object."""
    if color is None:
        return None
    if color.type == 'rgb' and color.rgb:
        return str(color.rgb)
    if color.type == 'theme':
        return f"Theme({color.theme}, tint={color.tint})"
    if color.type == 'indexed':
        return f"Indexed({color.indexed})"
    return str(color)

def get_fill_hex(fill):
    """Extract fill color hex."""
    if fill is None or fill.patternType is None:
        return None
    fg = fill.fgColor
    if fg and fg.type == 'rgb' and fg.rgb and fg.rgb != '00000000':
        return str(fg.rgb)
    if fg and fg.type == 'theme':
        return f"Theme({fg.theme}, tint={fg.tint})"
    return None

def normalize_color(c):
    """Normalize color for comparison."""
    if c is None:
        return None
    c = str(c).upper()
    # Strip leading FF for 8-char ARGB
    if len(c) == 8 and c[:2] == 'FF':
        return c[2:]
    return c

# ══════════════════════════════════════════════════════════════════
# CHECK 1: Sheet Structure
# ══════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("CHECK 1: SHEET STRUCTURE")
print("="*70)

model_sheets = model_wb.sheetnames
template_sheets = template_wb.sheetnames
stats['check1_items'] = 0

# Same sheet names
if model_sheets != template_sheets:
    if set(model_sheets) != set(template_sheets):
        add_issue('CRITICAL', 'CHECK 1', 'Workbook', '-', '-',
                  f'Sheets: {template_sheets}', f'Sheets: {model_sheets}',
                  'Sheet names must match template')
    else:
        add_issue('MAJOR', 'CHECK 1', 'Workbook', '-', '-',
                  f'Sheet order: {template_sheets}', f'Sheet order: {model_sheets}',
                  'Reorder sheets to match template')
else:
    print("  PASS: Sheet names and order match template")
stats['check1_items'] += 1

# Row count comparison
for sname in model_sheets:
    m_max = model_wb[sname].max_row
    t_max = template_wb[sname].max_row
    pct_diff = abs(m_max - t_max) / t_max * 100 if t_max > 0 else 0
    stats['check1_items'] += 1
    status = "PASS" if pct_diff <= 20 else "FLAG"
    print(f"  {sname}: Model={m_max} rows, Template={t_max} rows, Diff={pct_diff:.0f}%  [{status}]")
    if pct_diff > 20:
        add_issue('MODERATE', 'CHECK 1', sname, '-', '-',
                  f'~{t_max} rows (within 20%)', f'{m_max} rows ({pct_diff:.0f}% difference)',
                  'Verify row count difference is from intentional structural changes')

# ══════════════════════════════════════════════════════════════════
# CHECK 2: Row-by-Row Structure Comparison
# ══════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("CHECK 2: ROW-BY-ROW STRUCTURE COMPARISON")
print("="*70)

# Template retained rows per skill file (keyed items that should be in both)
# These are the "RETAIN" structural keys from the skill file
RETAINED_KEYS_ANNUAL = {
    # Subtotals & analytics
    'Rev-Total Revenue', 'GP-Gross Profit', 'OPEX-Total OpEx',
    'EBITDA-Underlying EBITDA', 'Stat-Statutory EBITDA',
    'DA-Total DA', 'EBIT-Underlying EBIT',
    'Int-Net Finance Costs', 'PBT-PBT', 'Tax-Tax Expense',
    'NPAT-Underlying NPAT', 'NPAT-Statutory NPAT', 'NPAT-NCI', 'NPAT-Sig Items AT',
    # EPS section
    'EPS-YE Shares', 'EPS-WASO Basic', 'EPS-Dilution', 'EPS-WASO Diluted',
    'EPS-Underlying EPS', 'EPS-Statutory EPS', 'Div-DPS', 'Div-Total Dividends',
    # BS
    'BS-Cash', 'BS-Trade Receivables', 'BS-Inventories', 'BS-PPE',
    'BS-Intangibles', 'BS-ROU Assets', 'BS-Other Assets',
    'BS-Trade Payables', 'BS-Other Liabilities', 'BS-Lease Liabilities',
    'BS-Total Banking Debt', 'BS-Issued Capital', 'BS-Retained Profits',
    'BS-Reserves', 'BS-Minorities',
    # CF
    'CF-EBITDA', 'CF-WC Change', 'CF-Significant Items',
    'CF-Int Received', 'CF-Interest Paid', 'CF-Lease Int Paid', 'CF-Tax Paid',
    'CF-Net OCF', 'CF-Capex PPE', 'CF-Capex Intang', 'CF-Acquisitions',
    'CF-Asset Sales', 'CF-Other CFI', 'CF-Dividends', 'CF-Share Issues',
    'CF-Lease Principal', 'CF-Debt Change', 'CF-Other CFF',
}

# Retained label-only rows (no Column A key)
RETAINED_LABELS_ANNUAL = {
    'Revenue Growth', 'GP Margin', 'EBITDA Margin', 'EBIT Margin', 'NPAT Margin',
    'EPS Growth', 'Payout Ratio', 'Dividend Yield', 'Dividend Growth',
    'Total Assets', 'Working Capital', 'Net Banking Debt',
    'Total Equity', 'BS Check',
    'Gross Operating Cash Flow', 'Net Operating Cash Flow',
    'Total Investing Cash Flow', 'Total Financing Cash Flow',
    'Net Change in Cash', 'Operating Free Cash Flow',
    'FCF per Share', 'FCF Yield', 'FCF Margin',
    'Invested Capital', 'ROFE', 'NOPAT', 'ROIC',
}

stats['check2_rows'] = 0

def extract_keys_and_labels(ws):
    """Extract all Column A keys and Column B labels from a sheet."""
    keys = {}  # key -> row
    labels = {}  # label -> [rows]
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a:
            keys[a] = r
        if b:
            if b not in labels:
                labels[b] = []
            labels[b].append(r)
    return keys, labels

for sname in ['Annual', 'HY & Segments']:
    m_ws = model_wb[sname]
    t_ws = template_wb[sname]
    m_keys, m_labels = extract_keys_and_labels(m_ws)
    t_keys, t_labels = extract_keys_and_labels(t_ws)

    print(f"\n  {sname}:")
    print(f"    Template keys: {len(t_keys)}, Model keys: {len(m_keys)}")

    # Check retained keyed rows exist
    if sname == 'Annual':
        retained_keys_to_check = RETAINED_KEYS_ANNUAL
    else:
        # HY has same P&L keys, no BS/CF
        retained_keys_to_check = {k for k in RETAINED_KEYS_ANNUAL
                                   if k.startswith(('Rev-Total', 'EBITDA-', 'Stat-',
                                                    'DA-Total', 'EBIT-', 'Int-Net',
                                                    'PBT-', 'Tax-', 'NPAT-'))}

    # Check which template keys exist in model (for retained rows)
    for key in sorted(retained_keys_to_check):
        stats['check2_rows'] += 1
        if key in t_keys and key not in m_keys:
            add_issue('CRITICAL', 'CHECK 2', sname, '-', 'A',
                      f'Key "{key}" should be retained from template',
                      'Missing from model',
                      f'Add row with key "{key}" to model')
        elif key not in t_keys and key not in m_keys:
            # Key might have been renamed for GYG - not a template key
            pass

    # Check retained label-only rows (Annual only)
    if sname == 'Annual':
        for label in sorted(RETAINED_LABELS_ANNUAL):
            stats['check2_rows'] += 1
            found_in_model = label in m_labels
            # Also check partial matches
            if not found_in_model:
                # Try partial match
                partial = [l for l in m_labels if label.lower() in l.lower()]
                if partial:
                    found_in_model = True
            if not found_in_model:
                add_issue('MAJOR', 'CHECK 2', sname, '-', 'B',
                          f'Label "{label}" should be retained',
                          'Not found in model',
                          f'Add analytical row "{label}"')

    # Check relative order of retained keys
    model_key_order = []
    for r in range(1, m_ws.max_row + 1):
        a = m_ws.cell(r, 1).value
        if a and a in retained_keys_to_check:
            model_key_order.append(a)

    template_key_order = []
    for r in range(1, t_ws.max_row + 1):
        a = t_ws.cell(r, 1).value
        if a and a in retained_keys_to_check:
            template_key_order.append(a)

    # Check order preserved for keys that exist in both
    common_keys = [k for k in template_key_order if k in m_keys]
    model_common = [k for k in model_key_order if k in set(common_keys)]

    order_ok = True
    for i in range(len(common_keys) - 1):
        k1, k2 = common_keys[i], common_keys[i+1]
        if k1 in model_common and k2 in model_common:
            if model_common.index(k1) > model_common.index(k2):
                order_ok = False
                add_issue('MAJOR', 'CHECK 2', sname, '-', 'A',
                          f'"{k1}" should appear before "{k2}"',
                          f'Order reversed in model',
                          'Fix row ordering to match template structure')
                break

    if order_ok:
        print(f"    PASS: Retained row order preserved")

    # Check for replacement rows (segment-specific)
    replaced_prefixes = ['Rev-', 'COGS-', 'GP-', 'EBITDA-', 'KPI-']
    for prefix in replaced_prefixes:
        model_seg_keys = [k for k in m_keys if k.startswith(prefix) and k not in retained_keys_to_check]
        if model_seg_keys:
            print(f"    REPLACED: {len(model_seg_keys)} rows with prefix '{prefix}' (company-specific)")
        stats['check2_rows'] += len(model_seg_keys)


# ══════════════════════════════════════════════════════════════════
# CHECK 3: Formatting Comparison
# ══════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("CHECK 3: FORMATTING COMPARISON")
print("="*70)

stats['check3_cells'] = 0

# 3a: Font bold on subtotal rows
print("\n  3a) Font Bold on Subtotal/Total Rows")

# Identify subtotal rows by looking for bold B column + data
subtotal_labels_annual = {
    'Total Revenue', 'Total Corp Restaurant Sales', 'Total Other Revenue & Income',
    'Total Revenue & Other Income', 'Total Segment Underlying EBITDA',
    'EBITDA', 'Total Expenses (excl D&A)', 'Total D&A', 'EBIT',
    'Total Finance Income', 'Total Finance Costs', 'Net Finance Income/(Costs)',
    'PBT', 'NPAT', 'Gross Operating Cash Flow', 'Net Operating Cash Flow',
    'Total Investing Cash Flow', 'Total Financing Cash Flow', 'Net Change in Cash',
    'Operating Free Cash Flow', 'Total Assets', 'Total Equity', 'Total Liabilities',
}

for sname in ['Annual', 'HY & Segments']:
    m_ws = model_wb[sname]
    data_cols = [4, 5, 6, 7] if sname == 'Annual' else [4, 5, 6, 11]  # Sample data cols

    for r in range(5, m_ws.max_row + 1):
        b_val = m_ws.cell(r, 2).value
        if b_val and any(sub in str(b_val) for sub in ['Total', 'EBITDA', 'EBIT', 'PBT', 'NPAT',
                                                         'Gross Operating', 'Net Operating',
                                                         'Net Change', 'Operating Free Cash']):
            b_bold = m_ws.cell(r, 2).font.bold
            if b_bold:
                for c in data_cols:
                    cell = m_ws.cell(r, c)
                    stats['check3_cells'] += 1
                    if cell.value is not None and not cell.font.bold:
                        add_issue('MODERATE', 'CHECK 3a', sname, r, get_column_letter(c),
                                  'Bold font on subtotal data cell',
                                  f'Not bold (value: {cell.value})',
                                  'Apply bold to all data cells on subtotal row')

# 3b: Font color
print("  3b) Font Color")

for sname in ['Annual', 'HY & Segments']:
    m_ws = model_wb[sname]
    actual_cols = [4, 5, 6] if sname == 'Annual' else [4, 5, 6, 7, 8, 9, 10]
    forecast_cols = list(range(7, 17)) if sname == 'Annual' else list(range(11, 30))

    blue_checked = 0
    maroon_checked = 0
    black_checked = 0

    for r in range(5, m_ws.max_row + 1):
        # Check actual data cells -> should be blue
        for c in actual_cols[:3]:  # Sample 3 actual cols
            cell = m_ws.cell(r, c)
            if cell.value is not None and not isinstance(cell.value, str):
                stats['check3_cells'] += 1
                blue_checked += 1
                fc = get_color_hex(cell.font.color)
                nc = normalize_color(fc)
                if nc and nc not in ('0000CC', 'Theme(1, tint=0.0)', '000000', None,
                                      'THEME(1, TINT=0.0)'):
                    if blue_checked <= 10:
                        add_issue('MINOR', 'CHECK 3b', sname, r, get_column_letter(c),
                                  'Blue font (0000CC) for actual data',
                                  f'Font color: {fc}',
                                  'Set font color to FF0000CC for actuals')

        # Check forecast formula cells -> should be black/default
        for c in forecast_cols[:2]:  # Sample first 2 forecast cols
            cell = m_ws.cell(r, c)
            if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('='):
                stats['check3_cells'] += 1
                black_checked += 1
                fc = get_color_hex(cell.font.color)
                nc = normalize_color(fc)
                # Formula cells should be black/theme
                if nc and nc not in ('000000', 'THEME(1, TINT=0.0)', None) and nc != '0000CC':
                    if nc in ('C00000',):
                        # Maroon = assumption, check if it's an assumption row
                        b_val = m_ws.cell(r, 2).value
                        if b_val and any(kw in str(b_val).lower() for kw in
                                         ['growth', 'margin', 'ratio', 'rate', 'capex/', 'revenue',
                                          'payables/', 'receivables/', 'inventory/', 'lease add']):
                            maroon_checked += 1
                            continue  # This is a valid assumption row
                    if black_checked <= 10:
                        add_issue('MINOR', 'CHECK 3b', sname, r, get_column_letter(c),
                                  'Black/theme font for formula cells',
                                  f'Font color: {fc}',
                                  'Set font to black/default for calculated cells')

    print(f"    {sname}: Checked {blue_checked} actual cells, {black_checked} formula cells")

# 3c: Fill color
print("  3c) Fill Color")

LIGHT_BLUE = 'C5D9F1'
GREY = 'D9D9D9'
NAVY = '002060'
MED_BLUE = '0070C0'

section_headers = ['P&L', 'Balance Sheet', 'Cash Flow', 'Operating Metrics',
                   'EPS & Dividends', 'Operating Free Cash Flow', 'ROIC']
sub_headers = ['Revenue', 'COGS', 'Gross Profit', 'Operating Expenses', 'EBITDA',
               'Statutory EBITDA', 'D&A', 'Depreciation & Amortisation', 'EBIT',
               'Interest', 'Finance', 'PBT', 'Assets', 'Liabilities', 'Equity',
               'CFO', 'CFI', 'CFF', 'Expenses (Memo)', 'Other Revenue',
               'Segment EBITDA Bridge', 'Segment Forecast']

for sname in ['Annual', 'HY & Segments']:
    m_ws = model_wb[sname]
    max_data_col = 16 if sname == 'Annual' else 29

    for r in range(5, m_ws.max_row + 1):
        b_val = m_ws.cell(r, 2).value
        if not b_val:
            continue

        fill = get_fill_hex(m_ws.cell(r, 2).fill)
        nfill = normalize_color(fill)

        # Major section headers
        if b_val in section_headers:
            stats['check3_cells'] += 1
            if nfill != LIGHT_BLUE:
                add_issue('MODERATE', 'CHECK 3c', sname, r, 'B',
                          f'Light blue fill ({LIGHT_BLUE}) for section header "{b_val}"',
                          f'Fill: {fill}',
                          f'Set fill to FFC5D9F1')
            # Check fill extends across data columns
            for c in [4, 8, max_data_col]:
                cell_fill = get_fill_hex(m_ws.cell(r, c).fill)
                ncf = normalize_color(cell_fill)
                stats['check3_cells'] += 1
                if ncf != nfill and nfill == LIGHT_BLUE:
                    add_issue('MODERATE', 'CHECK 3c', sname, r, get_column_letter(c),
                              f'Section header fill should extend to col {get_column_letter(c)}',
                              f'Fill: {cell_fill}',
                              'Extend section header fill across all columns')

        # Sub-section headers
        elif any(b_val.startswith(sh) or b_val == sh for sh in sub_headers):
            stats['check3_cells'] += 1
            if nfill != GREY:
                # Some sub-headers might be category headers (bold underline) not grey
                pass  # Don't flag - structure varies

    # Zone labels (row 2)
    r2_check_cols = [(4, NAVY, 'Actual'), (7 if sname == 'Annual' else 11, MED_BLUE, 'Forecast')]
    for c, expected_fill, zone_name in r2_check_cols:
        cell = m_ws.cell(2, c)
        fill = get_fill_hex(cell.fill)
        nfill = normalize_color(fill)
        stats['check3_cells'] += 1
        if nfill != expected_fill:
            add_issue('MODERATE', 'CHECK 3c', sname, 2, get_column_letter(c),
                      f'{zone_name} zone fill: {expected_fill}',
                      f'Fill: {fill}',
                      f'Set fill to FF{expected_fill}')

# 3d: Borders
print("  3d) Borders")

def has_border(border_side):
    """Check if a border side is set (thin or medium)."""
    if border_side is None:
        return False
    return border_side.style is not None and border_side.style != 'none'

def border_style(border_side):
    if border_side is None:
        return None
    return border_side.style

for sname in ['Annual', 'HY & Segments']:
    m_ws = model_wb[sname]
    data_cols = [4, 7, 10] if sname == 'Annual' else [4, 11, 20]

    for r in range(5, m_ws.max_row + 1):
        b_val = m_ws.cell(r, 2).value
        b_cell = m_ws.cell(r, 2)
        if not b_val or not b_cell.font.bold:
            continue

        # Check if this is a subtotal row (bold + has data in columns)
        has_data = any(m_ws.cell(r, c).value is not None for c in data_cols)
        if not has_data:
            continue

        # Subtotal rows should have thin top + thin bottom
        for c in [2] + data_cols:
            cell = m_ws.cell(r, c)
            stats['check3_cells'] += 1
            top = border_style(cell.border.top)
            bottom = border_style(cell.border.bottom)

            if not has_border(cell.border.top):
                add_issue('MODERATE', 'CHECK 3d', sname, r, get_column_letter(c),
                          f'Thin top border on subtotal row "{b_val}"',
                          f'Top border: {top}',
                          'Add thin top border')
            if not has_border(cell.border.bottom):
                add_issue('MODERATE', 'CHECK 3d', sname, r, get_column_letter(c),
                          f'Thin bottom border on subtotal row "{b_val}"',
                          f'Bottom border: {bottom}',
                          'Add thin bottom border')

# 3e: Number formats
print("  3e) Number Formats")

FORMAT_MAP = {
    'A$m': '#,##0.0',
    'NZDm': '#,##0.0',
    '% YoY': '0.0%',
    '%': '0.0%',
    'cps': '0.000',
    '#': '#,##0',
    'm': '0.0',
    'x': '0.0',
}

for sname in ['Annual', 'HY & Segments']:
    m_ws = model_wb[sname]
    actual_cols = [4, 5, 6] if sname == 'Annual' else [4, 5, 6, 7, 8, 9, 10]
    general_count = 0

    for r in range(5, m_ws.max_row + 1):
        units = m_ws.cell(r, 3).value
        if not units:
            continue

        expected_fmt = FORMAT_MAP.get(units)
        if not expected_fmt:
            continue

        for c in actual_cols:
            cell = m_ws.cell(r, c)
            if cell.value is None:
                continue
            stats['check3_cells'] += 1
            actual_fmt = cell.number_format
            if actual_fmt == 'General':
                general_count += 1
                if general_count <= 15:
                    add_issue('MODERATE', 'CHECK 3e', sname, r, get_column_letter(c),
                              f'Format "{expected_fmt}" for units "{units}"',
                              f'Format: "General"',
                              f'Set number format to {expected_fmt}')
            elif expected_fmt == '0.0%' and actual_fmt != '0.0%':
                # Allow some % format variations
                if '%' not in actual_fmt:
                    add_issue('MINOR', 'CHECK 3e', sname, r, get_column_letter(c),
                              f'Format "{expected_fmt}"', f'Format: "{actual_fmt}"',
                              f'Set number format to {expected_fmt}')
            elif expected_fmt == '#,##0.0' and actual_fmt not in ('#,##0.0', '#,##0.0;-#,##0.0', '#,##0.0;[Red]-#,##0.0'):
                if '#,##0' not in actual_fmt and actual_fmt != '0.0':
                    add_issue('MINOR', 'CHECK 3e', sname, r, get_column_letter(c),
                              f'Format "{expected_fmt}"', f'Format: "{actual_fmt}"',
                              f'Set number format to {expected_fmt}')

    if general_count > 0:
        print(f"    {sname}: {general_count} cells with 'General' format in data columns")
    else:
        print(f"    {sname}: PASS - no 'General' format cells found")


# ══════════════════════════════════════════════════════════════════
# CHECK 4: Zone Label Positioning
# ══════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("CHECK 4: ZONE LABEL POSITIONING")
print("="*70)

stats['check4_items'] = 0

for sname in ['Annual', 'HY & Segments']:
    m_ws = model_wb[sname]
    actual_col = 4  # D
    forecast_col = 7 if sname == 'Annual' else 11  # G or K

    # Check row 2 zone labels
    actual_cell = m_ws.cell(2, actual_col)
    forecast_cell = m_ws.cell(2, forecast_col)

    stats['check4_items'] += 4

    # Actual label
    actual_val = str(actual_cell.value or '')
    if 'Actual' not in actual_val:
        add_issue('MAJOR', 'CHECK 4', sname, 2, get_column_letter(actual_col),
                  '"Actual ---------->" zone label',
                  f'Value: "{actual_val}"',
                  'Set actual zone label at first actual column')
    else:
        print(f"  {sname} {get_column_letter(actual_col)}2: PASS - Actual zone label present")

    # Forecast label
    forecast_val = str(forecast_cell.value or '')
    if 'Forecast' not in forecast_val:
        add_issue('MAJOR', 'CHECK 4', sname, 2, get_column_letter(forecast_col),
                  '"Forecast ----->" zone label',
                  f'Value: "{forecast_val}"',
                  'Set forecast zone label at first forecast column')
    else:
        print(f"  {sname} {get_column_letter(forecast_col)}2: PASS - Forecast zone label present")

    # Check fill colors on zone labels
    actual_fill = normalize_color(get_fill_hex(actual_cell.fill))
    forecast_fill = normalize_color(get_fill_hex(forecast_cell.fill))

    if actual_fill != NAVY:
        add_issue('MODERATE', 'CHECK 4', sname, 2, get_column_letter(actual_col),
                  f'Navy fill ({NAVY}) on Actual zone label',
                  f'Fill: {actual_fill}',
                  f'Set fill to FF{NAVY}')

    if forecast_fill != MED_BLUE:
        add_issue('MODERATE', 'CHECK 4', sname, 2, get_column_letter(forecast_col),
                  f'Blue fill ({MED_BLUE}) on Forecast zone label',
                  f'Fill: {forecast_fill}',
                  f'Set fill to FF{MED_BLUE}')

    # Check white font on zone labels
    for cell, zone_name in [(actual_cell, 'Actual'), (forecast_cell, 'Forecast')]:
        fc = get_color_hex(cell.font.color)
        nc = normalize_color(fc)
        stats['check4_items'] += 1
        if nc not in ('FFFFFF', None):
            add_issue('MINOR', 'CHECK 4', sname, 2, get_column_letter(cell.column),
                      f'White font on {zone_name} zone label',
                      f'Font color: {fc}',
                      'Set font color to white (FFFFFFFF)')


# ══════════════════════════════════════════════════════════════════
# CHECK 5: Blank Row Spacing
# ══════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("CHECK 5: BLANK ROW SPACING")
print("="*70)

stats['check5_items'] = 0

for sname in ['Annual', 'HY & Segments']:
    m_ws = model_wb[sname]
    t_ws = template_wb[sname]

    # Find blank rows in model
    model_blank_rows = []
    last_data_row = 0
    for r in range(1, m_ws.max_row + 1):
        is_blank = True
        for c in range(1, min(m_ws.max_column + 1, 10)):
            if m_ws.cell(r, c).value is not None:
                is_blank = False
                last_data_row = r
                break
        if is_blank:
            model_blank_rows.append(r)

    # Find blank rows in template
    template_blank_rows = []
    for r in range(1, t_ws.max_row + 1):
        is_blank = True
        for c in range(1, min(t_ws.max_column + 1, 10)):
            if t_ws.cell(r, c).value is not None:
                is_blank = False
                break
        if is_blank:
            template_blank_rows.append(r)

    stats['check5_items'] += len(model_blank_rows)

    # Check for consecutive blank rows
    consecutive = []
    for i in range(len(model_blank_rows) - 1):
        if model_blank_rows[i] + 1 == model_blank_rows[i + 1]:
            consecutive.append((model_blank_rows[i], model_blank_rows[i + 1]))

    if consecutive:
        for r1, r2 in consecutive[:5]:
            add_issue('MINOR', 'CHECK 5', sname, f'{r1}-{r2}', '-',
                      'Single blank row between sections',
                      f'Consecutive blank rows at {r1} and {r2}',
                      'Remove one of the consecutive blank rows')
    else:
        print(f"  {sname}: PASS - no consecutive blank rows")

    # Check trailing empty rows
    trailing = [r for r in model_blank_rows if r > last_data_row]
    if len(trailing) > 2:
        add_issue('MINOR', 'CHECK 5', sname, f'{trailing[0]}-{trailing[-1]}', '-',
                  'No excessive trailing blank rows',
                  f'{len(trailing)} trailing blank rows after row {last_data_row}',
                  'Remove trailing blank rows')

    print(f"  {sname}: {len(model_blank_rows)} blank rows (template: {len(template_blank_rows)})")


# ══════════════════════════════════════════════════════════════════
# CHECK 6: Cross-Sheet Structural Correspondence
# ══════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("CHECK 6: CROSS-SHEET STRUCTURAL CORRESPONDENCE")
print("="*70)

m_annual = model_wb['Annual']
m_hy = model_wb['HY & Segments']

# Extract P&L keys from Annual (rows between P&L header and EPS section)
annual_pl_keys = []
for r in range(5, m_annual.max_row + 1):
    b_val = m_annual.cell(r, 2).value
    if b_val and ('EPS' in str(b_val) or 'Balance Sheet' in str(b_val) or
                  'Operating Metrics' in str(b_val)):
        break
    a_val = m_annual.cell(r, 1).value
    if a_val:
        annual_pl_keys.append((a_val, r))

# Extract P&L keys from HY
hy_pl_keys = []
for r in range(5, m_hy.max_row + 1):
    b_val = m_hy.cell(r, 2).value
    if b_val and ('Operating Metrics' in str(b_val) or 'Segment Forecast' in str(b_val)):
        break
    a_val = m_hy.cell(r, 1).value
    if a_val:
        hy_pl_keys.append((a_val, r))

annual_key_set = {k for k, _ in annual_pl_keys}
hy_key_set = {k for k, _ in hy_pl_keys}

stats['check6_items'] = len(annual_key_set | hy_key_set)

# Keys in Annual but not HY
only_annual = annual_key_set - hy_key_set
only_hy = hy_key_set - annual_key_set

if only_annual:
    for key in sorted(only_annual):
        add_issue('CRITICAL', 'CHECK 6', 'Cross-sheet', '-', 'A',
                  f'Key "{key}" should exist on both Annual and HY',
                  'Present on Annual only',
                  f'Add "{key}" to HY & Segments P&L section')

if only_hy:
    for key in sorted(only_hy):
        add_issue('CRITICAL', 'CHECK 6', 'Cross-sheet', '-', 'A',
                  f'Key "{key}" should exist on both Annual and HY',
                  'Present on HY only',
                  f'Add "{key}" to Annual P&L section')

if not only_annual and not only_hy:
    print("  PASS: All P&L Column A keys match between Annual and HY & Segments")

# Check order
annual_key_list = [k for k, _ in annual_pl_keys]
hy_key_list = [k for k, _ in hy_pl_keys]

if annual_key_list == hy_key_list:
    print("  PASS: Key order matches between sheets")
else:
    # Find first mismatch
    for i in range(min(len(annual_key_list), len(hy_key_list))):
        if annual_key_list[i] != hy_key_list[i]:
            add_issue('MAJOR', 'CHECK 6', 'Cross-sheet', '-', 'A',
                      f'Key order: Annual[{i}]="{annual_key_list[i]}"',
                      f'HY[{i}]="{hy_key_list[i]}"',
                      'Align P&L key order between Annual and HY sheets')
            break


# ══════════════════════════════════════════════════════════════════
# CHECK 7: Formula Structural Validation
# ══════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("CHECK 7: FORMULA STRUCTURAL VALIDATION")
print("="*70)

stats['check7_formulas'] = 0

m_annual = model_wb['Annual']
m_hy = model_wb['HY & Segments']

# Build HY key lookup
hy_keys_set = set()
for r in range(1, m_hy.max_row + 1):
    a = m_hy.cell(r, 1).value
    if a:
        hy_keys_set.add(a)

# Build HY row 3 period labels
hy_period_labels = set()
for c in range(4, m_hy.max_column + 1):
    v = m_hy.cell(3, c).value
    if v:
        hy_period_labels.add(str(v))

# Check Annual forecast columns (G=7 through P=16)
index_match_count = 0
direct_ref_count = 0
broken_key_count = 0

for r in range(5, m_annual.max_row + 1):
    a_key = m_annual.cell(r, 1).value
    for c in range(7, 17):  # G through P
        cell = m_annual.cell(r, c)
        val = cell.value
        if val and isinstance(val, str) and val.startswith('='):
            stats['check7_formulas'] += 1

            # Check for INDEX/MATCH pattern referencing HY
            if 'INDEX' in val and 'MATCH' in val and 'HY' in val:
                index_match_count += 1

                # Extract the lookup key from MATCH($A[row]
                match_key = re.search(r'MATCH\(\$A(\d+)', val)
                if match_key:
                    key_row = int(match_key.group(1))
                    lookup_key = m_annual.cell(key_row, 1).value
                    if lookup_key and lookup_key not in hy_keys_set:
                        broken_key_count += 1
                        add_issue('CRITICAL', 'CHECK 7', 'Annual', r, get_column_letter(c),
                                  f'INDEX/MATCH key "{lookup_key}" must exist in HY Column A',
                                  'Key not found in HY',
                                  f'Add key "{lookup_key}" to HY & Segments or fix formula')

            # Check for direct cell references to HY (not using INDEX/MATCH)
            elif "'HY" in val and 'INDEX' not in val:
                direct_ref_count += 1

    # Check BS/CF formulas for valid row references
    for c in range(7, 17):
        cell = m_annual.cell(r, c)
        val = cell.value
        if val and isinstance(val, str) and val.startswith('='):
            # Check for references to rows beyond max_row
            row_refs = re.findall(r'[A-Z]+(\d+)', val)
            for ref_row in row_refs:
                ref_row_int = int(ref_row)
                if ref_row_int > m_annual.max_row and 'HY' not in val:
                    add_issue('MAJOR', 'CHECK 7', 'Annual', r, get_column_letter(c),
                              f'Row reference within data range (max {m_annual.max_row})',
                              f'References row {ref_row_int}',
                              'Fix formula row reference')

print(f"  INDEX/MATCH formulas found: {index_match_count}")
print(f"  Direct HY cell references: {direct_ref_count}")
print(f"  Broken INDEX/MATCH keys: {broken_key_count}")
print(f"  Total formula cells checked: {stats['check7_formulas']}")

if direct_ref_count > 0 and index_match_count > 0:
    add_issue('MINOR', 'CHECK 7', 'Annual', '-', '-',
              'Consistent use of INDEX/MATCH for HY references',
              f'{direct_ref_count} direct cell references found alongside {index_match_count} INDEX/MATCH',
              'Consider converting direct refs to INDEX/MATCH for consistency')


# ══════════════════════════════════════════════════════════════════
# CHECK 8: Skill File Compliance
# ══════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("CHECK 8: SKILL FILE COMPLIANCE")
print("="*70)

stats['check8_items'] = 0

# 8a: Sign conventions (check actual values in FY25 - column F=6 for Annual)
print("\n  8a) Sign Conventions")

m_annual_data = model_wb_data['Annual']

# Find FY25 column
fy25_col = None
for c in range(4, 17):
    v = m_annual.cell(3, c).value
    if v and 'FY25' in str(v):
        fy25_col = c
        break

if not fy25_col:
    # Try to find any actual column with data
    fy25_col = 6  # Default to column F

sign_checks = {
    'Rev-Total Revenue': ('positive', lambda v: v is not None and v > 0),
    'Rev-Corp Restaurant Sales Aus': ('positive', lambda v: v is not None and v > 0),
    'DA-Total DA': ('negative', lambda v: v is not None and v < 0),
    'EBIT-EBIT': ('positive', lambda v: v is None or v > 0),
    'Int-Total Finance Costs': ('negative', lambda v: v is not None and v < 0),
    'Int-Total Finance Income': ('positive', lambda v: v is not None and v > 0),
    'Tax-Tax Expense': ('negative', lambda v: v is not None and v < 0),
    'CF-Capex PPE': ('negative', lambda v: v is not None and v < 0),
}

# Build key->row map for Annual data version
annual_key_row = {}
for r in range(1, m_annual_data.max_row + 1):
    a = m_annual_data.cell(r, 1).value
    if a:
        annual_key_row[a] = r

for key, (expected_sign, check_fn) in sign_checks.items():
    stats['check8_items'] += 1
    if key in annual_key_row:
        row = annual_key_row[key]
        val = m_annual_data.cell(row, fy25_col).value
        if val is not None and not check_fn(val):
            add_issue('CRITICAL', 'CHECK 8a', 'Annual', row, get_column_letter(fy25_col),
                      f'"{key}" should be {expected_sign}',
                      f'Value: {val}',
                      f'Fix sign convention for {key}')
        elif val is not None:
            print(f"    PASS: {key} = {val:.1f} ({expected_sign})")
        else:
            print(f"    SKIP: {key} = None (no data in FY25)")
    else:
        print(f"    SKIP: {key} not found in model")

# 8b: Cross-sheet formula method
print("\n  8b) Cross-Sheet Formula Method")
stats['check8_items'] += 1

pl_forecast_formulas = []
for r in range(5, 70):
    a_key = m_annual.cell(r, 1).value
    if a_key and a_key.startswith(('Rev-', 'COGS-', 'GP-', 'OPEX-', 'EBITDA-', 'DA-',
                                    'EBIT-', 'Int-', 'PBT-', 'Tax-', 'NPAT-',
                                    'OthRev-', 'SegEBITDA-', 'Bridge-', 'Exp-')):
        for c in range(7, 10):  # Check first few forecast cols
            val = m_annual.cell(r, c).value
            if val and isinstance(val, str) and val.startswith('='):
                pl_forecast_formulas.append((r, c, a_key, val))

index_match_pl = sum(1 for _, _, _, v in pl_forecast_formulas if 'INDEX' in v and 'MATCH' in v)
direct_ref_pl = sum(1 for _, _, _, v in pl_forecast_formulas if "'HY" in v and 'INDEX' not in v)
simple_ref_pl = sum(1 for _, _, _, v in pl_forecast_formulas if 'HY' not in v and 'INDEX' not in v)

print(f"    P&L forecast formulas: {len(pl_forecast_formulas)}")
print(f"    Using INDEX/MATCH to HY: {index_match_pl}")
print(f"    Direct HY cell refs: {direct_ref_pl}")
print(f"    Internal formulas (no HY ref): {simple_ref_pl}")

if direct_ref_pl > 0:
    # Show first few
    for r, c, key, val in pl_forecast_formulas:
        if "'HY" in val and 'INDEX' not in val:
            add_issue('MAJOR', 'CHECK 8b', 'Annual', r, get_column_letter(c),
                      f'INDEX/MATCH for P&L forecast cell (key: {key})',
                      f'Direct ref: {val[:80]}',
                      'Convert to INDEX/MATCH pattern for robustness')
            break  # Just flag first one

# 8c: Assumption input placement
print("\n  8c) Assumption Input Placement")
stats['check8_items'] += 1

assumption_keywords = ['growth', 'margin', 'rate', 'capex/', 'revenue/', 'payables/',
                       'receivables/', 'inventory/', 'lease add', 'new lease', 'avg lease']

assumption_rows_found = 0
for sname in ['Annual']:
    m_ws = model_wb[sname]
    for r in range(5, m_ws.max_row + 1):
        b_val = m_ws.cell(r, 2).value
        if b_val and any(kw in str(b_val).lower() for kw in assumption_keywords):
            assumption_rows_found += 1

print(f"    Assumption rows found on Annual: {assumption_rows_found}")
if assumption_rows_found < 5:
    add_issue('MAJOR', 'CHECK 8c', 'Annual', '-', '-',
              'Forecast assumption rows adjacent to driven rows',
              f'Only {assumption_rows_found} assumption rows found',
              'Ensure all forecast assumptions have dedicated input rows')

# 8d: CF structure
print("\n  8d) Cash Flow Structure")

cf_expected_order = ['CF-EBITDA', 'CF-WC Change', 'CF-Significant Items',
                     'CF-Int Received', 'CF-Interest Paid', 'CF-Lease Int Paid',
                     'CF-Tax Paid', 'CF-Net OCF',
                     'CF-Capex PPE', 'CF-Capex Intang', 'CF-Acquisitions',
                     'CF-Asset Sales', 'CF-Other CFI',
                     'CF-Dividends', 'CF-Share Issues', 'CF-Lease Principal',
                     'CF-Debt Change', 'CF-Other CFF']

model_cf_keys = []
for r in range(100, m_annual.max_row + 1):
    a = m_annual.cell(r, 1).value
    if a and a.startswith('CF-'):
        model_cf_keys.append(a)

# Check all expected CF keys exist
for key in cf_expected_order:
    stats['check8_items'] += 1
    if key not in model_cf_keys:
        add_issue('MAJOR', 'CHECK 8d', 'Annual', '-', 'A',
                  f'CF key "{key}" should exist',
                  'Missing from model',
                  f'Add CF row with key "{key}"')

# Check order
common_cf = [k for k in cf_expected_order if k in model_cf_keys]
model_cf_common = [k for k in model_cf_keys if k in set(common_cf)]
if common_cf == model_cf_common:
    print("    PASS: CF rows in correct order")
else:
    add_issue('MAJOR', 'CHECK 8d', 'Annual', '-', '-',
              f'CF order: {common_cf[:5]}...',
              f'Model order: {model_cf_common[:5]}...',
              'Reorder CF rows to match template structure')

# 8e: BS structure
print("\n  8e) Balance Sheet Structure")

bs_expected_order = ['BS-Cash', 'BS-Trade Receivables', 'BS-Inventories',
                     'BS-PPE', 'BS-Intangibles', 'BS-ROU Assets', 'BS-Other Assets',
                     'BS-Trade Payables', 'BS-Other Liabilities',
                     'BS-Lease Liabilities', 'BS-Total Banking Debt',
                     'BS-Issued Capital', 'BS-Retained Profits', 'BS-Reserves', 'BS-Minorities']

model_bs_keys = []
for r in range(100, m_annual.max_row + 1):
    a = m_annual.cell(r, 1).value
    if a and a.startswith('BS-'):
        model_bs_keys.append(a)

for key in bs_expected_order:
    stats['check8_items'] += 1
    if key not in model_bs_keys:
        # Check if a variant exists
        variants = [k for k in model_bs_keys if key.split('-')[1].split()[0] in k]
        if not variants:
            add_issue('MAJOR', 'CHECK 8e', 'Annual', '-', 'A',
                      f'BS key "{key}" should exist',
                      'Missing from model',
                      f'Add BS row with key "{key}"')

# Check order
common_bs = [k for k in bs_expected_order if k in model_bs_keys]
model_bs_common = [k for k in model_bs_keys if k in set(common_bs)]
if common_bs == model_bs_common:
    print("    PASS: BS rows in correct order")
else:
    add_issue('MAJOR', 'CHECK 8e', 'Annual', '-', '-',
              f'BS order: {common_bs[:5]}...',
              f'Model: {model_bs_common[:5]}...',
              'Reorder BS rows to match template structure')

# Check for extra BS keys not in template
bs_expected_set = set(bs_expected_order)
extra_bs = [k for k in model_bs_keys if k not in bs_expected_set]
if extra_bs:
    print(f"    INFO: Extra BS keys (company-specific): {extra_bs}")


# ══════════════════════════════════════════════════════════════════
# CHECK 9: Value Sheet Integrity
# ══════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("CHECK 9: VALUE SHEET INTEGRITY")
print("="*70)

stats['check9_items'] = 0
m_val = model_wb['Value']
m_val_data = model_wb_data['Value']

# 9a: Sheet title updated
stats['check9_items'] += 1
b2 = m_val.cell(2, 2).value
if b2 and 'VSL' in str(b2):
    add_issue('CRITICAL', 'CHECK 9', 'Value', 2, 'B',
              'Title updated from template company',
              f'Still shows: "{b2}"',
              'Update to GYG-specific title')
else:
    print(f"  Title: {b2} - OK")

# 9b: Check INDEX/MATCH formulas reference valid Annual keys
print("  Checking Value sheet formulas...")
annual_all_keys = set()
for r in range(1, m_annual.max_row + 1):
    a = m_annual.cell(r, 1).value
    if a:
        annual_all_keys.add(a)

value_formula_issues = 0
for r in range(1, m_val.max_row + 1):
    for c in range(1, m_val.max_column + 1):
        cell = m_val.cell(r, c)
        val = cell.value
        if val and isinstance(val, str) and 'MATCH(' in val:
            stats['check9_items'] += 1
            # Extract quoted key strings from MATCH("key",...)
            key_matches = re.findall(r'MATCH\("([^"]+)"', val)
            for key in key_matches:
                if key.startswith(('1H', '2H', 'FY')):
                    continue  # Period label, not a key
                if key not in annual_all_keys:
                    value_formula_issues += 1
                    add_issue('CRITICAL', 'CHECK 9', 'Value', r, get_column_letter(c),
                              f'INDEX/MATCH key "{key}" must exist in Annual Column A',
                              'Key not found',
                              f'Fix key to match Annual sheet or add key to Annual')

if value_formula_issues == 0:
    print("  PASS: All Value sheet INDEX/MATCH keys found in Annual")

# 9c: DCF rows exist
dcf_labels = ['EBITDA', 'EBIT', 'NOPAT', 'FCFF', 'Terminal Value', 'Enterprise Value', 'Equity Value']
for label in dcf_labels:
    stats['check9_items'] += 1
    found = False
    for r in range(1, 50):
        b = m_val.cell(r, 2).value
        if b and label in str(b):
            found = True
            break
    if not found:
        add_issue('MAJOR', 'CHECK 9', 'Value', '-', 'B',
                  f'DCF row "{label}" should exist',
                  'Not found',
                  f'Add DCF row for {label}')

print("  DCF structure: checked")

# 9d: SOTP rows exist
sotp_found = False
for r in range(50, 69):
    b = m_val.cell(r, 2).value
    if b and 'SOTP' in str(b):
        sotp_found = True
        break

stats['check9_items'] += 1
if not sotp_found:
    # Check for EV/EBITDA label
    for r in range(50, 69):
        b = m_val.cell(r, 2).value
        if b and 'EV' in str(b) and 'EBITDA' in str(b):
            sotp_found = True
            break

if sotp_found:
    print("  SOTP section: present")
else:
    add_issue('MAJOR', 'CHECK 9', 'Value', '-', '-',
              'SOTP valuation section should exist',
              'Not found',
              'Add EV/EBITDA SOTP section')

# Check SOTP segment rows (57-59)
for r in [57, 58, 59]:
    b = m_val.cell(r, 2).value
    c_val = m_val.cell(r, 3).value
    stats['check9_items'] += 1
    if b:
        # Check it has EBITDA and multiple
        if c_val and isinstance(c_val, str) and 'INDEX' in str(c_val):
            pass  # Formula present
        print(f"  SOTP row {r}: {b}")

# 9e: WACC inputs populated
wacc_rows = {12: 'Risk-free Rate', 13: 'ERP', 14: 'Beta', 20: 'WACC', 21: 'Terminal Growth'}
for r, label in wacc_rows.items():
    stats['check9_items'] += 1
    val = m_val_data.cell(r, 3).value
    if val is None or val == 0:
        add_issue('MAJOR', 'CHECK 9', 'Value', r, 'C',
                  f'WACC input "{label}" should be populated',
                  f'Value: {val}',
                  f'Set {label} to appropriate value')
    else:
        print(f"  WACC {label}: {val}")

# 9f: Share price populated
stats['check9_items'] += 1
sp = m_val_data.cell(4, 3).value
if sp is None or sp == 0:
    add_issue('MAJOR', 'CHECK 9', 'Value', 4, 'C',
              'Share price should be populated',
              f'Value: {sp}',
              'Enter current share price')
else:
    print(f"  Share Price: {sp}")

# 9g: Currency labels updated (no "NZD" remaining)
nzd_found = []
for r in range(1, m_val.max_row + 1):
    for c in range(1, m_val.max_column + 1):
        val = m_val.cell(r, c).value
        if val and isinstance(val, str) and 'NZD' in val:
            nzd_found.append((r, c, val))

# Also check Annual and HY for NZD
for sname in ['Annual', 'HY & Segments']:
    ws = model_wb[sname]
    for r in range(1, ws.max_row + 1):
        for c in range(1, 4):  # Check A, B, C columns only
            val = ws.cell(r, c).value
            if val and isinstance(val, str) and 'NZD' in val:
                nzd_found.append((r, c, f'{sname}!{val}'))

stats['check9_items'] += 1
if nzd_found:
    for r, c, val in nzd_found[:5]:
        add_issue('MAJOR', 'CHECK 9', 'Value/Model', r, get_column_letter(c),
                  'Currency should be A$ (not NZD)',
                  f'Found: "{val}"',
                  'Replace NZD references with A$')
    print(f"  WARNING: {len(nzd_found)} cells still contain 'NZD'")
else:
    print("  PASS: No NZD currency references found")


# ══════════════════════════════════════════════════════════════════
# OUTPUT REPORT
# ══════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("ISSUE REPORT")
print("="*70)

severity_order = ['CRITICAL', 'MAJOR', 'MODERATE', 'MINOR']
issue_counts = defaultdict(int)
check_counts = defaultdict(lambda: defaultdict(int))

for issue in issues:
    issue_counts[issue['severity']] += 1
    check_counts[issue['check']][issue['severity']] += 1

for severity in severity_order:
    sev_issues = [i for i in issues if i['severity'] == severity]
    if not sev_issues:
        continue
    print(f"\n{'─'*70}")
    print(f"  {severity} ({len(sev_issues)} issues)")
    print(f"{'─'*70}")
    for i in sev_issues:
        sheet_info = f"Sheet: {i['sheet']}"
        loc = f"Row {i['row']}, Col {i['col']}" if i['row'] != '-' else f"Col {i['col']}"
        print(f"\n  [{i['severity']}] {sheet_info}, {loc}")
        print(f"    Check: {i['check']}")
        print(f"    Expected: {i['expected']}")
        print(f"    Found:    {i['found']}")
        print(f"    Fix:      {i['fix']}")


# ══════════════════════════════════════════════════════════════════
# SUMMARY
# ══════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("SUMMARY")
print("="*70)

total_issues = sum(issue_counts.values())
print(f"\nTotal issues: {total_issues}")
for sev in severity_order:
    count = issue_counts.get(sev, 0)
    marker = " <<<" if count > 0 and sev in ('CRITICAL', 'MAJOR') else ""
    print(f"  {sev}: {count}{marker}")

print(f"\nTotal rows checked: {stats.get('check2_rows', 0) + stats.get('check5_items', 0)}")
print(f"Total formula cells checked: {stats.get('check7_formulas', 0)}")
print(f"Total formatting cells checked: {stats.get('check3_cells', 0)}")

print(f"\nPass rate per check category:")
check_names = {
    'CHECK 1': 'Sheet Structure',
    'CHECK 2': 'Row Structure',
    'CHECK 3a': 'Font Bold',
    'CHECK 3b': 'Font Color',
    'CHECK 3c': 'Fill Color',
    'CHECK 3d': 'Borders',
    'CHECK 3e': 'Number Format',
    'CHECK 4': 'Zone Labels',
    'CHECK 5': 'Blank Rows',
    'CHECK 6': 'Cross-Sheet Keys',
    'CHECK 7': 'Formula Validation',
    'CHECK 8a': 'Sign Conventions',
    'CHECK 8b': 'Formula Method',
    'CHECK 8c': 'Assumption Placement',
    'CHECK 8d': 'CF Structure',
    'CHECK 8e': 'BS Structure',
    'CHECK 9': 'Value Sheet',
}

all_checks = set(i['check'] for i in issues) | set(check_names.keys())
for check in sorted(all_checks):
    name = check_names.get(check, check)
    count = sum(check_counts[check].values())
    if count == 0:
        print(f"  {check} ({name}): PASS")
    else:
        details = ', '.join(f'{v} {k}' for k, v in sorted(check_counts[check].items(),
                            key=lambda x: severity_order.index(x[0])) if v > 0)
        print(f"  {check} ({name}): {count} issues ({details})")

print("\n" + "="*70)
print("VALIDATION COMPLETE")
print("="*70)
